# -*- coding: utf-8 -*-
"""네이버 플레이스 블로그 자동 포스팅 — PySide6 GUI"""
APP_VERSION = "2.4.6"

import os
import sys
import faulthandler
import traceback
from datetime import datetime

# ── 크래시 로그 (frozen GUI segfault/예외 진단용) ──
_app_data = os.environ.get('LOCALAPPDATA') or os.environ.get('APPDATA') or os.path.expanduser("~")
_log_dir = os.path.join(_app_data, 'BlogMaster')
try:
    os.makedirs(_log_dir, exist_ok=True)
    _crash_log = os.path.join(_log_dir, 'crash_log.txt')
    _crash_file = open(_crash_log, 'a', encoding='utf-8')
    _crash_file.write(f"\n--- App Started v{APP_VERSION} at {datetime.now()} ---\n")
    _crash_file.flush()
    faulthandler.enable(file=_crash_file)

    def _handle_exception(exc_type, exc_value, exc_tb):
        if issubclass(exc_type, KeyboardInterrupt):
            sys.__excepthook__(exc_type, exc_value, exc_tb)
            return
        with open(_crash_log, 'a', encoding='utf-8') as _f:
            _f.write(f"\n--- Uncaught Exception at {datetime.now()} ---\n")
            traceback.print_exception(exc_type, exc_value, exc_tb, file=_f)
    sys.excepthook = _handle_exception
except Exception:
    pass

if sys.platform == "win32":
    os.environ["PYTHONUTF8"] = "1"

# frozen exe에선 cwd 변경 위험 → 개발 모드에서만 변경
if not getattr(sys, "frozen", False):
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QSpinBox, QPushButton, QComboBox, QTextEdit,
    QFrame, QMessageBox, QDialog, QScrollArea, QTableWidget,
    QTableWidgetItem, QHeaderView, QSplitter, QCheckBox,
    QTreeWidget, QTreeWidgetItem, QAbstractButton,
    QRadioButton, QButtonGroup
)
from PySide6.QtCore import Qt, QThread, Signal, Slot, QObject
from PySide6.QtGui import QFont, QColor, QTextCursor, QShortcut, QKeySequence

import threading
import datetime
import json

from config import load_config, save_config
from content_generator import generate_content
from image_handler import download_images
from places_crawler import crawl_places, crawl_places_parallel, save_results, load_results, expand_keyword_to_districts
from naver_poster import NaverBlogPoster
from usage_tracker import get_remaining, add_usage, add_post


class ImageReplaceWorker(QThread):
    done = Signal(str, bool, str)  # img_path, success, message

    def __init__(self, img_path, api_key, keyword, translator=None):
        super().__init__()
        self.img_path = img_path
        self.api_key = api_key
        self.keyword = keyword
        self.translator = translator

    def run(self):
        try:
            import image_handler as _ih
            import shutil
            if not self.api_key:
                self.done.emit(self.img_path, False, "API 키 없음")
                return
            if not self.keyword:
                self.done.emit(self.img_path, False, "키워드 없음")
                return
            tmp_list = _ih.download_images(self.api_key, self.keyword, 1, translator=self.translator)
            if not tmp_list:
                self.done.emit(self.img_path, False, "새 이미지 없음")
                return
            shutil.copyfile(tmp_list[0], self.img_path)
            self.done.emit(self.img_path, True, "")
        except Exception as e:
            self.done.emit(self.img_path, False, str(e))


STYLE = """
QMainWindow, QDialog { background: #f8fafc; }
QWidget { font-family: 'Malgun Gothic', '맑은 고딕', sans-serif; color: #475569; }
QLabel { color: #1e293b; background: transparent; }

/* ── 헤더 ── */
#header {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #4f46e5, stop:1 #6366f1);
}
#header QLabel { color: white; font-size: 16px; font-weight: bold; letter-spacing: 0px; }
#header QPushButton {
    background: rgba(255,255,255,0.15); color: white; border: 1px solid rgba(255,255,255,0.25);
    border-radius: 7px; padding: 5px 10px; font-weight: bold; font-size: 11px;
}
#header QPushButton:hover { background: rgba(255,255,255,0.28); }

/* ── 좌측 패널 ── */
#leftPanel {
    background: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 12px;
}

#sectionLabel { font-size: 12px; font-weight: bold; color: #6366f1; letter-spacing: 0.5px; }
#subLabel { font-size: 11px; color: #94a3b8; }

/* ── 입력 필드 ── */
QLineEdit, QSpinBox, QComboBox {
    background: #ffffff; border: 1px solid #e2e8f0; border-radius: 8px;
    padding: 8px 12px; font-size: 13px; color: #1e293b;
    selection-background-color: #4f46e5; selection-color: #ffffff;
}
QLineEdit:focus, QSpinBox:focus, QComboBox:focus { border: 1.5px solid #4f46e5; }
QLineEdit:hover, QSpinBox:hover, QComboBox:hover { border-color: #cbd5e1; }
QComboBox::drop-down { border: none; width: 22px; }
QComboBox::down-arrow { width: 10px; height: 10px; }
QComboBox QAbstractItemView {
    background: #ffffff; border: 1px solid #e2e8f0; border-radius: 8px; padding: 4px;
    selection-background-color: #eef2ff; selection-color: #1e293b; outline: none;
}

/* ── 버튼 ── */
#btnCrawl {
    background: #4f46e5; color: white; border: none; border-radius: 9px;
    font-size: 13px; font-weight: bold; padding: 9px 12px;
}
#btnCrawl:hover { background: #4338ca; }
#btnCrawl:pressed { background: #3730a3; }

#btnStop {
    background: #f1f5f9; color: #64748b; border: 1.5px solid #e2e8f0;
    border-radius: 9px; font-size: 12px; padding: 4px 8px;
}
#btnStop:hover { background: #e2e8f0; color: #475569; }

#btnPost {
    background: #4f46e5; color: white; border: none; border-radius: 9px;
    font-size: 13px; font-weight: bold; padding: 9px 12px;
}
#btnPost:hover { background: #4338ca; }
#btnPost:pressed { background: #3730a3; }

#btnResult {
    background: #f8fafc; color: #475569; border: 1.5px solid #e2e8f0;
    border-radius: 9px; font-size: 12px; padding: 4px 8px;
}
#btnResult:hover { background: #f1f5f9; border-color: #cbd5e1; color: #334155; }

/* ── 우측 로그 패널 ── */
#logPanel {
    background: #ffffff; border: 1px solid #e2e8f0; border-radius: 12px;
}

#logHeader { font-size: 13px; font-weight: bold; color: #1e293b; }

QTextEdit {
    background: #f8fafc; border: none; font-size: 11px; color: #334155;
    border-radius: 6px; padding: 8px;
}

/* ── 스크롤바 ── */
QScrollBar:vertical {
    background: #f1f5f9; width: 6px; margin: 0;
}
QScrollBar::handle:vertical {
    background: #cbd5e1; border-radius: 3px; min-height: 20px;
}
QScrollBar::handle:vertical:hover { background: #94a3b8; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }

QScrollBar:horizontal {
    background: #f1f5f9; height: 6px; margin: 0;
}
QScrollBar::handle:horizontal {
    background: #cbd5e1; border-radius: 3px; min-width: 20px;
}
QScrollBar::handle:horizontal:hover { background: #94a3b8; }
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width: 0; }

/* ── 체크박스 ── */
QCheckBox { font-size: 12px; color: #475569; spacing: 5px; }
QCheckBox::indicator {
    width: 15px; height: 15px; border-radius: 4px;
    border: 1.5px solid #cbd5e1; background: white;
}
QCheckBox::indicator:hover { border-color: #4f46e5; }
QCheckBox::indicator:checked { background: #4f46e5; border-color: #4f46e5; }

/* ── 라디오 ── */
QRadioButton { font-size: 12px; color: #475569; spacing: 6px; }
QRadioButton::indicator { width: 16px; height: 16px; border-radius: 8px; border: 1.5px solid #cbd5e1; background: white; }
QRadioButton::indicator:hover { border-color: #4f46e5; }
QRadioButton::indicator:checked { background: #4f46e5; border: 4px solid #4f46e5; }

/* ── 테이블 / 트리 ── */
QTableWidget, QTableView, QTreeWidget, QTreeView {
    background: #ffffff; border: 1px solid #e2e8f0; border-radius: 10px;
    gridline-color: #f1f5f9; font-size: 12px; color: #475569;
    alternate-background-color: #f8fafc; outline: none;
}
QTableWidget::item, QTreeWidget::item, QTableView::item, QTreeView::item { padding: 6px 8px; }
QTableWidget::item:selected, QTreeWidget::item:selected,
QTableView::item:selected, QTreeView::item:selected { background: #eef2ff; color: #1e293b; }
QHeaderView::section {
    background: #1e293b; color: #ffffff; padding: 9px 8px; border: none;
    border-right: 1px solid #334155; font-size: 12px; font-weight: bold;
}
QHeaderView::section:last { border-right: none; }
QTableCornerButton::section { background: #1e293b; border: none; }

/* ── 탭 ── */
QTabWidget::pane { border: 1px solid #e2e8f0; border-radius: 10px; top: -1px; background: #ffffff; }
QTabBar::tab {
    background: #f1f5f9; color: #64748b; padding: 8px 18px; margin-right: 4px;
    border-top-left-radius: 8px; border-top-right-radius: 8px; font-size: 12px; font-weight: bold;
}
QTabBar::tab:selected { background: #4f46e5; color: #ffffff; }
QTabBar::tab:hover:!selected { background: #e2e8f0; color: #334155; }

/* ── 그룹박스 (카드) ── */
QGroupBox {
    background: #ffffff; border: 1px solid #e2e8f0; border-radius: 12px;
    margin-top: 14px; padding: 14px 12px 12px 12px; font-size: 13px; font-weight: bold; color: #1e293b;
}
QGroupBox::title { subcontrol-origin: margin; left: 14px; top: 2px; padding: 0 4px; color: #4f46e5; }

/* ── 기본 버튼 (objectName 미지정) ── */
QPushButton {
    background: #ffffff; color: #334155; border: 1px solid #e2e8f0;
    border-radius: 8px; padding: 8px 14px; font-size: 12px;
}
QPushButton:hover { background: #f8fafc; border-color: #cbd5e1; }
QPushButton:pressed { background: #f1f5f9; }
QPushButton:disabled { background: #f1f5f9; color: #cbd5e1; border-color: #f1f5f9; }

/* ── 프로그레스바 ── */
QProgressBar { background: #f1f5f9; border: none; border-radius: 6px; text-align: center; color: #1e293b; font-size: 11px; }
QProgressBar::chunk { background: #4f46e5; border-radius: 6px; }

/* ── 툴팁 ── */
QToolTip { background: #1e293b; color: #ffffff; border: none; padding: 6px 9px; border-radius: 6px; font-size: 12px; }
"""


class UpdateProgressDialog(QDialog):
    """프리미엄 모던 디자인 업데이트 진행 다이얼로그 (frameless + 그라디언트 + 라운드)"""
    def __init__(self, parent=None):
        super().__init__(parent)
        from PySide6.QtWidgets import QProgressBar, QGraphicsDropShadowEffect
        from PySide6.QtGui import QColor
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setFixedSize(520, 280)
        self._cancelled = False

        # 컨테이너 (그림자 + 둥근 모서리)
        container = QFrame(self)
        container.setObjectName("container")
        container.setGeometry(10, 10, 500, 260)
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(30)
        shadow.setColor(QColor(0, 0, 0, 80))
        shadow.setOffset(0, 4)
        container.setGraphicsEffect(shadow)

        container.setStyleSheet("""
            QFrame#container {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ffffff, stop:1 #f8fafc);
                border: 1px solid #e2e8f0;
                border-radius: 16px;
            }
            QLabel#title { font-size: 18px; font-weight: 800; color: #0f172a; }
            QLabel#sub { font-size: 12px; color: #64748b; }
            QLabel#pct {
                font-size: 56px; font-weight: 800;
                color: #4a6cf7;
                letter-spacing: -2px;
            }
            QLabel#status { font-size: 13px; color: #475569; }
            QProgressBar {
                border: none;
                border-radius: 6px;
                background: #e2e8f0;
                min-height: 12px;
                max-height: 12px;
            }
            QProgressBar::chunk {
                border-radius: 6px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #6366f1, stop:0.5 #4a6cf7, stop:1 #22c55e);
            }
        """)

        layout = QVBoxLayout(container)
        layout.setContentsMargins(32, 24, 32, 24)
        layout.setSpacing(8)

        title_row = QHBoxLayout()
        title_row.setSpacing(10)
        icon_lbl = QLabel("⬇")
        icon_lbl.setStyleSheet("font-size: 22px; color: #4a6cf7;")
        title_row.addWidget(icon_lbl)
        title = QLabel("업데이트 다운로드")
        title.setObjectName("title")
        title_row.addWidget(title)
        title_row.addStretch()
        layout.addLayout(title_row)

        sub = QLabel("새 버전을 받아오는 중입니다. 잠시만 기다려주세요.")
        sub.setObjectName("sub")
        layout.addWidget(sub)
        layout.addSpacing(12)

        self.pct_lbl = QLabel("0%")
        self.pct_lbl.setObjectName("pct")
        self.pct_lbl.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.pct_lbl)

        self.bar = QProgressBar()
        self.bar.setRange(0, 100)
        self.bar.setTextVisible(False)
        layout.addWidget(self.bar)
        layout.addSpacing(8)

        self.status_lbl = QLabel("준비 중...")
        self.status_lbl.setObjectName("status")
        self.status_lbl.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.status_lbl)

    def setProgress(self, pct: int, label: str = ""):
        self.pct_lbl.setText(f"{pct}%")
        self.bar.setValue(pct)
        if label:
            self.status_lbl.setText(label)

    def setLabelText(self, label: str):
        self.status_lbl.setText(label)

    def wasCanceled(self):
        return self._cancelled


class ToggleSwitch(QAbstractButton):
    """슬라이드 ON/OFF 토글 (녹색 ON / 회색 OFF, 하얀 thumb, 텍스트 표시)"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setCheckable(True)
        self.setFixedSize(64, 28)
        self.setCursor(Qt.PointingHandCursor)

    def paintEvent(self, _):
        from PySide6.QtGui import QPainter, QColor, QBrush, QPen, QFont
        from PySide6.QtCore import QRectF
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        h = self.height(); w = self.width()
        track = QColor("#22c55e") if self.isChecked() else QColor("#9ca3af")
        p.setBrush(QBrush(track))
        p.setPen(Qt.NoPen)
        p.drawRoundedRect(QRectF(0, 0, w, h), h/2, h/2)
        # 텍스트
        p.setPen(QPen(QColor("#ffffff")))
        f = QFont(); f.setBold(True); f.setPointSize(8)
        p.setFont(f)
        if self.isChecked():
            p.drawText(QRectF(0, 0, w - h, h), Qt.AlignCenter, "ON")
        else:
            p.drawText(QRectF(h, 0, w - h, h), Qt.AlignCenter, "OFF")
        # thumb
        d = h - 4
        x = w - d - 2 if self.isChecked() else 2
        p.setBrush(QBrush(QColor("#ffffff")))
        p.drawEllipse(QRectF(x, 2, d, d))


class MainWindow(QMainWindow):
    log_signal = Signal(str)
    post_log_signal = Signal(str)
    status_signal = Signal(str, str)
    crawl_count_signal = Signal(str)
    post_count_signal = Signal(str)
    publish_count_signal = Signal(str)
    resv_signal = Signal()  # 예약 현황 라벨 갱신(메인스레드)
    acct_status_signal = Signal(str)  # 계정 점검 결과(메인스레드)
    _app_quit_signal = Signal()

    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"블로그마스터 v{APP_VERSION}")
        self.setMinimumSize(900, 600)
        self.resize(1200, 780)

        self.cfg = load_config()
        self.crawled_data = []
        self._generated_posts = []
        self.is_crawling = False
        self.is_posting = False
        self.stop_flag = False
        self.result_file = ""
        self.last_keyword = self.cfg.get("last_keyword", "")
        self._active_posters = []  # 백그라운드 시 함께 최소화할 포스터 추적

        self.log_signal.connect(self._append_log)
        self.post_log_signal.connect(self._append_post_log)
        self.status_signal.connect(self._update_status)
        self.crawl_count_signal.connect(self._update_crawl_count)
        self.post_count_signal.connect(self._update_post_count)
        self.publish_count_signal.connect(self._update_publish_count)
        self.resv_signal.connect(self._update_reservation_label)
        self.acct_status_signal.connect(self._show_acct_status)
        self._app_quit_signal.connect(lambda: __import__('PySide6.QtWidgets', fromlist=['QApplication']).QApplication.instance().quit())

        self._build_ui()
        self._migrate_legacy_logs()

        # 단축키
        QShortcut(QKeySequence("F5"), self).activated.connect(self._start_crawl)
        QShortcut(QKeySequence("F6"), self).activated.connect(self._stop)
        QShortcut(QKeySequence("F7"), self).activated.connect(self._generate_posts)
        QShortcut(QKeySequence("F8"), self).activated.connect(self._view_generated_posts)

        self._emit_log("프로그램 시작됨")
        self._emit_log("설정에서 API키와 계정 정보를 먼저 입력해주세요")

        # 시스템 트레이 아이콘 구성
        self._force_quit = False
        self._init_tray_icon()

        # 키워드마스터 연동 — 시작 직후 핸드오프 파일 있으면 키워드 가져오기
        try:
            from PySide6.QtCore import QTimer as _QT_km
            _QT_km.singleShot(1200, self._import_from_keyword_master)
        except Exception:
            pass

    def _init_tray_icon(self):
        self.tray = None
        try:
            from PySide6.QtWidgets import QSystemTrayIcon, QMenu, QStyle
            from PySide6.QtGui import QIcon, QAction
            if not QSystemTrayIcon.isSystemTrayAvailable():
                return
            base_dirs = []
            if getattr(sys, 'frozen', False):
                base_dirs.append(os.path.dirname(sys.executable))
                base_dirs.append(getattr(sys, '_MEIPASS', ''))
            base_dirs.append(os.path.dirname(os.path.abspath(__file__)))
            icon = QIcon()
            for d in base_dirs:
                if not d:
                    continue
                p = os.path.join(d, "icon.ico")
                if os.path.exists(p):
                    icon = QIcon(p)
                    break
            if icon.isNull():
                try:
                    icon = self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon)
                except Exception:
                    icon = QIcon()
            self.tray = QSystemTrayIcon(icon, self)
            self.tray.setToolTip("블로그마스터")
            menu = QMenu()
            act_show = QAction("창 열기", self)
            act_show.triggered.connect(self._tray_show)
            act_quit = QAction("종료", self)
            act_quit.triggered.connect(self._tray_quit)
            menu.addAction(act_show)
            menu.addSeparator()
            menu.addAction(act_quit)
            self.tray.setContextMenu(menu)
            self.tray.activated.connect(self._on_tray_activated)
            self.tray.show()
        except Exception:
            self.tray = None

    def _on_tray_activated(self, reason):
        from PySide6.QtWidgets import QSystemTrayIcon
        if reason in (QSystemTrayIcon.Trigger, QSystemTrayIcon.DoubleClick):
            self._tray_show()

    def _tray_show(self):
        self.showNormal()
        self.raise_()
        self.activateWindow()
        # 최소화해둔 포스팅 브라우저도 복원
        for p in list(getattr(self, "_active_posters", [])):
            try:
                if getattr(p, "driver", None):
                    p.driver.execute_script("")  # ping
                    p.driver.set_window_rect(x=100, y=100, width=900, height=700)
            except Exception:
                pass

    def _tray_quit(self):
        self._force_quit = True
        QApplication.quit()

    def _clear_own_session(self):
        user = getattr(self, "current_user", None)
        if user and user.get("role") != "admin":
            try:
                from users import clear_session
                clear_session(user.get("username", ""), user.get("_session_id", ""))
            except Exception:
                pass

    def closeEvent(self, event):
        if getattr(self, "_force_quit", False):
            self._clear_own_session()
            event.accept()
            return
        # X 버튼 → 3지선다 (종료/백그라운드/취소)
        if getattr(self, "tray", None) is not None:
            box = QMessageBox(self)
            box.setWindowTitle("종료 확인")
            box.setText("프로그램을 어떻게 하시겠습니까?")
            box.setIcon(QMessageBox.Question)
            btn_quit = box.addButton("종료", QMessageBox.DestructiveRole)
            btn_bg = box.addButton("백그라운드 실행", QMessageBox.AcceptRole)
            btn_cancel = box.addButton("취소", QMessageBox.RejectRole)
            box.setDefaultButton(btn_bg)
            box.exec()
            clicked = box.clickedButton()
            if clicked is btn_quit:
                self._clear_own_session()
                self._force_quit = True
                event.accept()
                try:
                    QApplication.quit()
                except Exception:
                    pass
            elif clicked is btn_bg:
                event.ignore()
                self.hide()
                # 실행 중인 포스팅 브라우저도 함께 최소화
                for p in list(getattr(self, "_active_posters", [])):
                    try:
                        if getattr(p, "driver", None):
                            p.driver.minimize_window()
                    except Exception:
                        pass
                try:
                    self.tray.showMessage(
                        "블로그마스터",
                        "백그라운드에서 실행 중입니다. 트레이 아이콘을 우클릭하면 종료할 수 있습니다.",
                        msecs=3000,
                    )
                except Exception:
                    pass
            else:
                # 취소
                event.ignore()
        else:
            # 트레이 지원 안 되는 환경은 단순 확인
            reply = QMessageBox.question(self, "종료 확인", "프로그램을 종료하시겠습니까?",
                                          QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                event.accept()
            else:
                event.ignore()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root_layout = QVBoxLayout(central)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        # ── 헤더 ──
        header = QFrame()
        header.setObjectName("header")
        header.setFixedHeight(55)
        h_layout = QHBoxLayout(header)
        h_layout.setContentsMargins(12, 0, 10, 0)
        h_layout.setSpacing(5)

        title = QLabel("블로그마스터")
        h_layout.addWidget(title)
        ver_lbl = QLabel(f"v{APP_VERSION}")
        ver_lbl.setStyleSheet("font-size: 12px; color: #94a3b8; margin-left: 6px;")
        h_layout.addWidget(ver_lbl)

        # ── 카카오톡 오픈채팅: 버튼 클릭 시 QR + 바로입장 팝업 ──
        import webbrowser as _wb
        from PySide6.QtGui import QPixmap as _QPixmap
        _KAKAO_URL = "https://open.kakao.com/o/gZvid0Ai"
        try:
            from app_paths import get_bundle_dir as _gbd
            _qr_path = os.path.join(_gbd(), "kakao_qr.png")
        except Exception:
            _qr_path = os.path.join(os.path.dirname(__file__), "kakao_qr.png")

        def _open_kakao_qr_popup():
            d = QDialog(self)
            d.setWindowTitle("카카오톡 오픈채팅 입장")
            _v = QVBoxLayout(d)
            if os.path.exists(_qr_path):
                _big = QLabel()
                _big.setPixmap(_QPixmap(_qr_path).scaled(280, 280, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                _big.setAlignment(Qt.AlignCenter)
                _v.addWidget(_big)
            _cap = QLabel("휴대폰 카메라로 QR을 찍거나\n아래 버튼으로 바로 입장하세요")
            _cap.setAlignment(Qt.AlignCenter)
            _v.addWidget(_cap)
            _enter = QPushButton("💬 바로입장")
            _enter.setCursor(Qt.PointingHandCursor)
            _enter.setStyleSheet("background: #FEE500; color: #3C1E1E; border: none; border-radius: 8px; padding: 10px 18px; font-weight: bold;")
            _enter.clicked.connect(lambda: _wb.open(_KAKAO_URL))
            _v.addWidget(_enter)
            d.exec()

        btn_kakao = QPushButton("💬 오픈채팅")
        btn_kakao.setCursor(Qt.PointingHandCursor)
        btn_kakao.setStyleSheet("background: #FEE500; color: #3C1E1E; border: none; border-radius: 6px; padding: 5px 9px; font-weight: bold; margin-left: 10px;")
        btn_kakao.clicked.connect(_open_kakao_qr_popup)
        h_layout.addWidget(btn_kakao)

        # ── 사용법 (노션 가이드) ──
        _NOTION_URL = "https://shared-rise-9e5.notion.site/37df3feefb738039ae90e6ac73a9e2ea"
        btn_manual = QPushButton("📖 사용법")
        btn_manual.setCursor(Qt.PointingHandCursor)
        btn_manual.setStyleSheet("background: #0f172a; color: white; border: none; border-radius: 6px; padding: 5px 9px; font-weight: bold; margin-left: 8px;")
        btn_manual.clicked.connect(lambda: _wb.open(_NOTION_URL))
        h_layout.addWidget(btn_manual)

        h_layout.addStretch()

        self.dash_expires = QLabel("")
        self.dash_expires.setStyleSheet("font-size: 20px; font-weight: bold; color: #000;")
        h_layout.addWidget(self.dash_expires)

        btn_payment = QPushButton("💳 구독 연장")
        btn_payment.setCursor(Qt.PointingHandCursor)
        btn_payment.setStyleSheet("background: #6366f1; color: white; border: none; border-radius: 6px; padding: 5px 9px; font-weight: bold;")
        btn_payment.clicked.connect(self._open_payment_dialog)
        h_layout.addWidget(btn_payment)

        btn_prompts = QPushButton("프롬프트")
        btn_prompts.setCursor(Qt.PointingHandCursor)
        btn_prompts.clicked.connect(self._open_prompt_editor)
        h_layout.addWidget(btn_prompts)

        btn_engage = QPushButton("💬 댓글·이웃")
        btn_engage.setCursor(Qt.PointingHandCursor)
        btn_engage.setStyleSheet("background: #10b981; color: white; border: none; border-radius: 6px; padding: 5px 9px; font-weight: bold;")
        btn_engage.clicked.connect(self._open_engage_dialog)
        h_layout.addWidget(btn_engage)

        self.btn_admin = QPushButton("관리자")
        self.btn_admin.setStyleSheet("background: #f59e0b; color: white; border: none; border-radius: 6px; padding: 6px 14px;")
        self.btn_admin.setCursor(Qt.PointingHandCursor)
        self.btn_admin.clicked.connect(self._open_admin_dialog)
        self.btn_admin.setVisible(False)  # 기본 숨김, admin 로그인 시 노출
        h_layout.addWidget(self.btn_admin)

        btn_settings = QPushButton("설정")
        btn_settings.setCursor(Qt.PointingHandCursor)
        btn_settings.clicked.connect(self._open_settings)
        h_layout.addWidget(btn_settings)

        btn_logout = QPushButton("로그아웃")
        btn_logout.setCursor(Qt.PointingHandCursor)
        btn_logout.setStyleSheet("background: #ef4444; color: white; border: none; border-radius: 6px; padding: 6px 14px;")
        btn_logout.clicked.connect(self._logout)
        h_layout.addWidget(btn_logout)

        root_layout.addWidget(header)

        # ── 메인 ──
        main_widget = QWidget()
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(12)

        # ── 좌측 패널 ──
        left = QFrame()
        left.setObjectName("leftPanel")
        left.setFixedWidth(300)
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(18, 18, 18, 18)
        left_layout.setSpacing(5)

        # 크롤링 모드 선택
        mode_row = QHBoxLayout()
        mode_row.setSpacing(6)
        self._radio_keyword = QRadioButton("지역 기반")
        self._radio_url = QRadioButton("키워드 기반")
        self._radio_keyword.setChecked(True)
        self._crawl_mode_group = QButtonGroup(self)
        self._crawl_mode_group.addButton(self._radio_keyword, 0)
        self._crawl_mode_group.addButton(self._radio_url, 1)
        mode_row.addWidget(self._radio_keyword)
        mode_row.addWidget(self._radio_url)
        mode_row.addStretch()
        left_layout.addLayout(mode_row)

        # URL 입력창 (URL 기반 선택 시만 표시)
        self._crawl_url_input = QLineEdit()
        self._crawl_url_input.setPlaceholderText("네이버 지도 검색 URL 붙여넣기")
        self._crawl_url_input.setVisible(False)
        left_layout.addWidget(self._crawl_url_input)

        self._radio_keyword.toggled.connect(self._on_crawl_mode_changed)

        # 키워드 (히스토리 드롭다운)
        lbl = QLabel("키워드 입력")
        lbl.setObjectName("sectionLabel")
        left_layout.addWidget(lbl)

        self.keyword_input = QComboBox()
        self.keyword_input.setEditable(True)
        self.keyword_input.setInsertPolicy(QComboBox.NoInsert)
        self.keyword_input.lineEdit().setCompleter(None)
        self.keyword_input.lineEdit().setPlaceholderText("키워드를 입력하세요")
        self.keyword_input.lineEdit().returnPressed.connect(self._start_crawl)
        self._load_keyword_history()
        left_layout.addWidget(self.keyword_input)

        # 픽사베이 검색어: 입력칸 제거 — 메인 키워드 하나로 완전 자동 큐레이션
        # (gpt-4o-mini 자동 추출 → 하드코딩 테이블 폴백). 수동 오버라이드 미사용.

        # 지역설정 버튼 (드롭다운 트리 다이얼로그)
        self._selected_regions = list(self.cfg.get("selected_regions", []) or [])
        self.btn_regions = QPushButton()
        self.btn_regions.setCursor(Qt.PointingHandCursor)
        self.btn_regions.setStyleSheet(
            "text-align: left; padding: 7px 11px; background: #f8fafc; "
            "border: 1.5px solid #e2e8f0; border-radius: 8px; font-size: 12px; color: #334155;"
        )
        self._refresh_regions_button()
        self.btn_regions.clicked.connect(self._open_regions_dialog)
        left_layout.addWidget(self.btn_regions)

        # 제외 키워드 버튼
        self._exclude_keywords_by_biz = dict(self.cfg.get("exclude_keywords_by_biz", {}) or {})
        self.btn_excludes = QPushButton()
        self.btn_excludes.setCursor(Qt.PointingHandCursor)
        self.btn_excludes.setStyleSheet(
            "text-align: left; padding: 7px 11px; background: #f8fafc; "
            "border: 1.5px solid #e2e8f0; border-radius: 8px; font-size: 12px; color: #334155;"
        )
        self._refresh_excludes_button()
        self.btn_excludes.clicked.connect(self._open_excludes_dialog)
        left_layout.addWidget(self.btn_excludes)

        # 크롤링 결과보기 (지역설정/제외키워드 밑으로 이동)
        btn_result = QPushButton("크롤링 결과보기")
        btn_result.setObjectName("btnResult")
        btn_result.setCursor(Qt.PointingHandCursor)
        btn_result.clicked.connect(self._show_results)
        left_layout.addWidget(btn_result)

        left_layout.addSpacing(5)
        self._add_divider(left_layout)

        # 크롤링 설정
        lbl = QLabel("크롤링 설정")
        lbl.setObjectName("sectionLabel")
        left_layout.addWidget(lbl)

        sub = QLabel("수집 개수")
        sub.setObjectName("subLabel")
        left_layout.addWidget(sub)

        self.count_spin = QSpinBox()
        self.count_spin.setRange(1, 10000)
        self.count_spin.setValue(5000)
        left_layout.addWidget(self.count_spin)

        btn_start = QPushButton("크롤링 시작 (F5)")
        btn_start.setObjectName("btnCrawl")
        btn_start.setCursor(Qt.PointingHandCursor)
        btn_start.setMinimumHeight(32)
        btn_start.clicked.connect(self._start_crawl)
        left_layout.addWidget(btn_start)

        btn_stop = QPushButton("중단 (F6)")
        btn_stop.setObjectName("btnStop")
        btn_stop.setCursor(Qt.PointingHandCursor)
        btn_stop.setMinimumHeight(30)
        btn_stop.clicked.connect(self._stop)
        left_layout.addWidget(btn_stop)

        left_layout.addSpacing(5)
        self._add_divider(left_layout)

        # 포스팅 설정
        lbl = QLabel("포스팅 설정")
        lbl.setObjectName("sectionLabel")
        left_layout.addWidget(lbl)

        sub = QLabel("발행 간격")
        sub.setObjectName("subLabel")
        left_layout.addWidget(sub)

        iv_widget = QWidget()
        iv_layout = QHBoxLayout(iv_widget)
        iv_layout.setContentsMargins(0, 0, 0, 0)

        self.interval_hour = QComboBox()
        self.interval_hour.addItems([str(h) for h in range(25)])
        self.interval_hour.setCurrentText("2")
        iv_layout.addWidget(self.interval_hour)
        iv_layout.addWidget(QLabel("시간"))

        self.interval_min = QComboBox()
        self.interval_min.addItems([str(m) for m in range(0, 60, 3)])
        self.interval_min.setCurrentText("0")
        iv_layout.addWidget(self.interval_min)
        iv_layout.addWidget(QLabel("분"))

        # 발행간격 옆 랜덤 체크박스 — 체크 시 ±10분 랜덤
        self.interval_random = QCheckBox("랜덤(±20분)")
        self.interval_random.setStyleSheet("font-size: 12px; color: #334155;")
        self.interval_random.setToolTip("체크 시 설정 간격에서 10분 단위로 ±20분 랜덤 (5단계 — 네이버 예약 10분 단위 제약)")
        iv_layout.addWidget(self.interval_random)
        iv_layout.addStretch()

        left_layout.addWidget(iv_widget)

        # 발행 간격 변경 시 자동 저장
        self.interval_hour.currentIndexChanged.connect(self._save_interval_settings)
        self.interval_min.currentIndexChanged.connect(self._save_interval_settings)
        self.interval_random.stateChanged.connect(self._save_interval_settings)

        # (v1.6.6) 예약발행 남아있음 체크박스 제거 — 포스팅 시 항상 자동 감지하여 이어쓰기

        left_layout.addSpacing(8)

        btn_generate = QPushButton("포스트 글쓰기 (F7)")
        btn_generate.setObjectName("btnPost")
        btn_generate.setCursor(Qt.PointingHandCursor)
        btn_generate.setMinimumHeight(32)
        btn_generate.clicked.connect(self._generate_posts)
        left_layout.addWidget(btn_generate)

        btn_stop_generating = QPushButton("글쓰기 중단")
        btn_stop_generating.setStyleSheet(
            "background: qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #fb923c,stop:1 #f97316);"
            "color: white; border: none; border-radius: 9px; "
            "font-size: 13px; font-weight: bold; padding: 4px 11px;"
        )
        btn_stop_generating.setMinimumHeight(32)
        btn_stop_generating.setCursor(Qt.PointingHandCursor)
        btn_stop_generating.clicked.connect(self._stop_generating)
        left_layout.addWidget(btn_stop_generating)

        btn_view_posts = QPushButton("생성된 포스트 보기 및 업로드 (F8)")
        btn_view_posts.setStyleSheet(
            "background: qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #60a5fa,stop:1 #3b82f6);"
            "color: white; border: none; border-radius: 9px; "
            "font-size: 13px; font-weight: bold; padding: 4px 11px;"
        )
        btn_view_posts.setMinimumHeight(32)
        btn_view_posts.setCursor(Qt.PointingHandCursor)
        btn_view_posts.clicked.connect(self._view_generated_posts)
        left_layout.addWidget(btn_view_posts)

        btn_stop_posting = QPushButton("포스팅 중단")
        btn_stop_posting.setStyleSheet(
            "background: qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #f87171,stop:1 #ef4444);"
            "color: white; border: none; border-radius: 9px; "
            "font-size: 13px; font-weight: bold; padding: 4px 11px;"
        )
        btn_stop_posting.setMinimumHeight(32)
        btn_stop_posting.setCursor(Qt.PointingHandCursor)
        btn_stop_posting.clicked.connect(self._stop_posting_only)
        left_layout.addWidget(btn_stop_posting)

        btn_gallery = QPushButton("이미지 갤러리 보기")
        btn_gallery.setObjectName("btnResult")
        btn_gallery.setCursor(Qt.PointingHandCursor)
        btn_gallery.setMinimumHeight(30)
        btn_gallery.clicked.connect(self._open_image_gallery)
        left_layout.addWidget(btn_gallery)

        left_layout.addStretch()
        main_layout.addWidget(left)

        # ── 우측 패널 ──
        right = QFrame()
        right.setObjectName("logPanel")
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(15, 12, 15, 15)
        right_layout.setSpacing(5)

        # 대시보드
        dash = QFrame()
        dash.setStyleSheet("background: transparent; border: none;")
        dash_layout = QHBoxLayout(dash)
        dash_layout.setContentsMargins(4, 4, 4, 4)

        from PySide6.QtWidgets import QSizePolicy
        from PySide6.QtGui import QPixmap, QPainter, QPolygon
        from PySide6.QtCore import QPoint
        # 드롭다운 화살표 이미지 생성 (1회)
        _arrow_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_down_arrow.png").replace("\\", "/")
        if not os.path.exists(_arrow_path):
            pm = QPixmap(10, 6)
            pm.fill(Qt.transparent)
            p = QPainter(pm)
            p.setBrush(Qt.black)
            p.setPen(Qt.NoPen)
            p.drawPolygon(QPolygon([QPoint(0, 0), QPoint(10, 0), QPoint(5, 6)]))
            p.end()
            pm.save(_arrow_path)

        self.account_combo = QComboBox()
        self.account_combo.setCursor(Qt.PointingHandCursor)
        self.account_combo.setSizeAdjustPolicy(QComboBox.AdjustToMinimumContentsLengthWithIcon)
        self.account_combo.setMinimumContentsLength(6)
        self.account_combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.account_combo.setMaximumWidth(260)
        self.account_combo.setStyleSheet(f"""
            QComboBox {{
                font-size: 12px;
                color: #000;
                background-color: #ffffff;
                padding: 1px 4px;
                border: 1px solid #7f9db9;
                border-radius: 0px;
                min-height: 20px;
            }}
            QComboBox:hover {{
                background-color: #f0f8ff;
            }}
            QComboBox:focus {{
                border: 1px solid #316ac5;
            }}
            QComboBox::drop-down {{
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 18px;
                border-left: 1px solid #7f9db9;
                background-color: #ece9d8;
            }}
            QComboBox::down-arrow {{
                image: url({_arrow_path});
                width: 10px;
                height: 6px;
            }}
            QComboBox QAbstractItemView {{
                border: 1px solid #7f9db9;
                background-color: #ffffff;
                color: #000;
                selection-background-color: #316ac5;
                selection-color: #ffffff;
                font-size: 12px;
                padding: 0px;
                outline: none;
            }}
            QComboBox QAbstractItemView::item {{
                min-height: 20px;
                padding: 2px 6px;
                color: #000;
                background-color: #ffffff;
            }}
            QComboBox QAbstractItemView::item:hover {{
                background-color: #316ac5;
                color: #ffffff;
            }}
            QComboBox QAbstractItemView::item:selected {{
                background-color: #ffffff;
                color: #000;
            }}
            QComboBox QAbstractItemView::item:selected:hover {{
                background-color: #316ac5;
                color: #ffffff;
            }}
        """)
        self._refresh_account_combo()
        self.account_combo.currentIndexChanged.connect(self._on_account_changed)

        dash_layout.addWidget(QLabel("계정:"))
        dash_layout.addWidget(self.account_combo)
        # 예약 발행 현황 (계정 옆) — 전환 시 캐시에서 자동 표시 + 조회 버튼으로 실시간 갱신
        self.btn_resv = QPushButton("📅 예약조회")
        self.btn_resv.setCursor(Qt.PointingHandCursor)
        self.btn_resv.setStyleSheet("background:#0ea5e9; color:white; border:none; border-radius:6px; padding:5px 10px; font-size:11px; font-weight:bold;")
        self.btn_resv.clicked.connect(self._check_reservations)
        dash_layout.addWidget(self.btn_resv)
        # 계정 보호조치 점검 버튼
        self.btn_acct_check = QPushButton("🛡️ 계정점검")
        self.btn_acct_check.setCursor(Qt.PointingHandCursor)
        self.btn_acct_check.setStyleSheet("background:#64748b; color:white; border:none; border-radius:6px; padding:5px 10px; font-size:11px; font-weight:bold; margin-left:4px;")
        self.btn_acct_check.clicked.connect(self._check_accounts_status)
        dash_layout.addWidget(self.btn_acct_check)
        # 날짜별 발행 일정표 (자동 기록된 발행/예약 주제를 날짜별로 조회)
        self.btn_schedule = QPushButton("🗓️ 발행일정표")
        self.btn_schedule.setCursor(Qt.PointingHandCursor)
        self.btn_schedule.setStyleSheet("background:#8b5cf6; color:white; border:none; border-radius:6px; padding:5px 10px; font-size:11px; font-weight:bold; margin-left:4px;")
        self.btn_schedule.clicked.connect(self._show_schedule_calendar)
        dash_layout.addWidget(self.btn_schedule)
        self.resv_label = QLabel("📅 예약: -")
        self.resv_label.setStyleSheet("color:#475569; font-size:11px; margin-left:6px;")
        dash_layout.addWidget(self.resv_label)
        dash_layout.addStretch()
        right_layout.addWidget(dash)

        self._update_dashboard()
        self._update_reservation_label()
        self._refresh_status_panes()
        from PySide6.QtCore import QTimer as _QT
        self._status_pane_timer = _QT(self)
        self._status_pane_timer.setInterval(2000)
        self._status_pane_timer.timeout.connect(self._refresh_status_panes)
        self._status_pane_timer.start()

        # 로그 헤더
        log_hdr = QWidget()
        log_hdr_layout = QHBoxLayout(log_hdr)
        log_hdr_layout.setContentsMargins(0, 0, 0, 0)

        self.indicator = QLabel("●")
        self.indicator.setStyleSheet("color: #22c55e; font-size: 14px;")
        log_hdr_layout.addWidget(self.indicator)

        lbl = QLabel("실행 로그")
        lbl.setObjectName("logHeader")
        log_hdr_layout.addWidget(lbl)

        # 봇 개수 선택 (1/2) — 크롤링 동시 워커 수. 기본 2 (throttle 안전값, 최대 2봇)
        bot_lbl = QLabel("  |  봇")
        bot_lbl.setStyleSheet("color: #64748b; font-size: 12px; margin-left: 8px;")
        log_hdr_layout.addWidget(bot_lbl)

        self.bot_count_group = QButtonGroup(self)
        saved_bots = int(self.cfg.get("bot_count", 2) or 2)
        if saved_bots not in (1, 2):
            saved_bots = 2
        for n in (1, 2):
            rb = QRadioButton(str(n))
            rb.setStyleSheet("font-size: 12px; color: #334155; padding: 0 2px;")
            rb.setCursor(Qt.PointingHandCursor)
            rb.setChecked(n == saved_bots)
            self.bot_count_group.addButton(rb, n)
            log_hdr_layout.addWidget(rb)
        self.bot_count_group.buttonClicked.connect(self._save_bot_count)

        log_hdr_layout.addStretch()

        self.crawl_count_label = QLabel("크롤링 0개")
        self.crawl_count_label.setStyleSheet("color: #f59e0b; font-weight: bold; font-size: 12px;")
        log_hdr_layout.addWidget(self.crawl_count_label)

        _sep1 = QLabel("  |  ")
        _sep1.setStyleSheet("color: #cbd5e1; font-size: 12px;")
        log_hdr_layout.addWidget(_sep1)

        self.post_count_label = QLabel("포스트생성 0개")
        self.post_count_label.setStyleSheet("color: #8b5cf6; font-weight: bold; font-size: 12px;")
        log_hdr_layout.addWidget(self.post_count_label)

        _sep2 = QLabel("  |  ")
        _sep2.setStyleSheet("color: #cbd5e1; font-size: 12px;")
        log_hdr_layout.addWidget(_sep2)

        self.publish_count_label = QLabel("발행 0개")
        self.publish_count_label.setStyleSheet("color: #22c55e; font-weight: bold; font-size: 12px;")
        log_hdr_layout.addWidget(self.publish_count_label)

        self.status_label = QLabel("")
        self.status_label.hide()
        log_hdr_layout.addWidget(self.status_label)

        right_layout.addWidget(log_hdr)

        # 크롤링 현황 / 포스트 현황 (위아래 분할)
        from PySide6.QtWidgets import QSplitter
        status_split = QSplitter(Qt.Vertical)

        crawl_box = QFrame()
        crawl_box.setStyleSheet("background: #f8fafc; border: 1.5px solid #e2e8f0; border-radius: 10px;")
        crawl_lay = QVBoxLayout(crawl_box)
        crawl_lay.setContentsMargins(10, 8, 10, 8)
        self.crawl_status_label = QLabel("크롤링 현황")
        self.crawl_status_label.setStyleSheet("font-weight: bold; font-size: 12px; color: #6366f1; letter-spacing: 0.3px;")
        crawl_lay.addWidget(self.crawl_status_label)

        # 워커별 로그 2개 (병렬 크롤 시각화 — 최대 2봇)
        from PySide6.QtWidgets import QSplitter as _QS
        _worker_split = _QS(Qt.Horizontal)
        _worker_split.setChildrenCollapsible(False)
        self.crawl_worker_logs = []
        self.crawl_worker_region_labels = []
        for i in range(2):
            col = QFrame()
            col.setStyleSheet("background: white; border: 1px solid #e2e8f0; border-radius: 8px;")
            col.setMinimumWidth(120)
            cl = QVBoxLayout(col)
            cl.setContentsMargins(6, 4, 6, 6)
            cl.setSpacing(2)
            hdr_row = QHBoxLayout()
            hdr_row.setContentsMargins(0, 0, 0, 0)
            hdr = QLabel(f"봇 {i+1}")
            hdr.setStyleSheet("font-weight: bold; font-size: 10px; color: #6366f1; background: #eef2ff; border-radius: 4px; padding: 1px 5px;")
            hdr_row.addWidget(hdr)
            region_lbl = QLabel("")
            region_lbl.setStyleSheet("font-size: 11px; color: #334155; font-weight: bold; padding-left: 4px;")
            region_lbl.setMinimumWidth(80)
            hdr_row.addWidget(region_lbl, 1)
            cl.addLayout(hdr_row)
            te = QTextEdit()
            te.setReadOnly(True)
            te.setStyleSheet("font-size: 11px; border: none; background: white; color: #475569;")
            te.setMinimumHeight(120)
            cl.addWidget(te, 1)
            _worker_split.addWidget(col)
            self.crawl_worker_logs.append(te)
            self.crawl_worker_region_labels.append(region_lbl)
        _worker_split.setSizes([300, 300])
        crawl_lay.addWidget(_worker_split, 1)

        # 공용 로그 — 숨겨진 백업 (레이아웃 차지 안 함)
        self.crawl_log = QTextEdit()
        self.crawl_log.setReadOnly(True)
        self.crawl_log.hide()

        status_split.addWidget(crawl_box)

        post_box = QFrame()
        post_box.setStyleSheet("background: #f8fafc; border: 1.5px solid #e2e8f0; border-radius: 10px;")
        post_lay = QVBoxLayout(post_box)
        post_lay.setContentsMargins(10, 8, 10, 8)
        self.post_status_label = QLabel("포스트 현황")
        self.post_status_label.setStyleSheet("font-weight: bold; font-size: 12px; color: #8b5cf6; letter-spacing: 0.3px;")
        post_lay.addWidget(self.post_status_label)
        self.post_log = QTextEdit()
        self.post_log.setReadOnly(True)
        self.post_log.setStyleSheet("background: white; border: 1px solid #e2e8f0; border-radius: 8px; font-size: 11px; color: #475569; padding: 8px;")
        post_lay.addWidget(self.post_log)
        status_split.addWidget(post_box)

        status_split.setSizes([400, 400])
        right_layout.addWidget(status_split, stretch=1)

        # 통합 로그 (보이지 않는 백업 — 기존 _append_log 호환용)
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.hide()
        right_layout.addWidget(self.log_text)

        main_layout.addWidget(right, stretch=1)
        root_layout.addWidget(main_widget, stretch=1)

    def _add_divider(self, layout):
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setStyleSheet("background: #f1f5f9; border: none; max-height: 1px; margin: 2px 0;")
        layout.addWidget(line)
        layout.addSpacing(4)

    # ── 대시보드 ──
    def _update_dashboard(self):
        cfg = load_config()
        provider = cfg.get("ai_provider", "gpt").lower()
        info = get_remaining(provider)
        pass  # 대시보드 사용량 표시 제거됨

    # ── 로그 ──
    def _emit_log(self, msg):
        self.log_signal.emit(msg)

    def _emit_post_log(self, msg):
        self.post_log_signal.emit(msg)

    def _append_post_log(self, msg):
        now = datetime.datetime.now().strftime("%H:%M:%S")
        line = f"<span style='color:#9ca3af'>[{now}]</span>  {msg}"
        if hasattr(self, "post_log"):
            self.post_log.append(line)
            self.post_log.moveCursor(QTextCursor.End)

    def _append_log(self, msg):
        now = datetime.datetime.now().strftime("%H:%M:%S")
        _err_keywords = ("실패", "오류", "에러", "error", "Error", "Exception", "!!!", "크래시", "중단됨")
        if not msg.startswith("<span") and any(k in msg for k in _err_keywords):
            msg = f"<span style='color:#ef4444;font-weight:bold'>{msg}</span>"
        line = f"<span style='color:#9ca3af'>[{now}]</span>  {msg}"
        self.log_text.append(line)
        self.log_text.moveCursor(QTextCursor.End)
        # _emit_log → 크롤링 패널로만 라우팅 (포스팅 패널에는 절대 안 감)
        try:
            import re as _re
            m = _re.search(r"\[봇(\d+)\]", msg)
            if m and hasattr(self, "crawl_worker_logs"):
                idx = int(m.group(1)) - 1
                if 0 <= idx < len(self.crawl_worker_logs):
                    self.crawl_worker_logs[idx].append(line)
                    self.crawl_worker_logs[idx].moveCursor(QTextCursor.End)
                    m2 = _re.search(r"\[봇\d+\]\s*(.+?)\s*크롤 시작", msg)
                    if m2 and hasattr(self, "crawl_worker_region_labels") and idx < len(self.crawl_worker_region_labels):
                        self.crawl_worker_region_labels[idx].setText(m2.group(1).strip())
                    return
            # 봇 태그 없는 일반 메시지 → 크롤링 첫 번째 패널에 표시
            if hasattr(self, "crawl_worker_logs") and self.crawl_worker_logs:
                self.crawl_worker_logs[0].append(line)
                self.crawl_worker_logs[0].moveCursor(QTextCursor.End)
        except Exception:
            pass

    def _emit_status(self, text, color="#22c55e"):
        self.status_signal.emit(text, color)

    def _update_status(self, text, color):
        self.status_label.setText(text)
        self.status_label.setStyleSheet(f"color: {color}; font-weight: bold; font-size: 12px;")
        self.indicator.setStyleSheet(f"color: {color}; font-size: 14px;")

    def _emit_crawl_count(self, text):
        self.crawl_count_signal.emit(text)

    def _update_crawl_count(self, text):
        if hasattr(self, "crawl_count_label"):
            self.crawl_count_label.setText(f"크롤링 {text}")

    def _emit_post_count(self, text):
        self.post_count_signal.emit(text)

    def _update_post_count(self, text):
        if hasattr(self, "post_count_label"):
            self.post_count_label.setText(f"포스트생성 {text}")

    def _emit_publish_count(self, text):
        self.publish_count_signal.emit(text)

    def _update_publish_count(self, text):
        if hasattr(self, "publish_count_label"):
            self.publish_count_label.setText(f"발행 {text}")

    def _save_interval_settings(self):
        try:
            cfg = load_config()
            cfg["posting_interval_hour"] = self.interval_hour.currentText()
            cfg["posting_interval_min"] = self.interval_min.currentText()
            cfg["posting_interval_random"] = self.interval_random.isChecked()
            save_config(cfg)
            self.cfg = cfg
        except Exception:
            pass

    def _load_interval_settings(self):
        try:
            h = str(self.cfg.get("posting_interval_hour", "2"))
            m = str(self.cfg.get("posting_interval_min", "0"))
            r = bool(self.cfg.get("posting_interval_random", False))
            self.interval_hour.blockSignals(True)
            self.interval_min.blockSignals(True)
            self.interval_random.blockSignals(True)
            if h in [self.interval_hour.itemText(i) for i in range(self.interval_hour.count())]:
                self.interval_hour.setCurrentText(h)
            if m in [self.interval_min.itemText(i) for i in range(self.interval_min.count())]:
                self.interval_min.setCurrentText(m)
            self.interval_random.setChecked(r)
            self.interval_hour.blockSignals(False)
            self.interval_min.blockSignals(False)
            self.interval_random.blockSignals(False)
        except Exception:
            pass

    def _get_interval_seconds(self):
        # 발행 간격 = 시간/분 드롭다운 값. 랜덤 체크 시 ±10분 적용.
        base = int(self.interval_hour.currentText()) * 3600 + int(self.interval_min.currentText()) * 60
        base = max(600, base)  # 최소 10분 (네이버 예약 최소 단위)
        try:
            if getattr(self, "interval_random", None) and self.interval_random.isChecked():
                import random as _r
                # 네이버 예약은 10분 단위만 가능 → ±20분을 10분 단위 5슬롯에서 매번 독립 랜덤 선택
                # (예: 2시간 기준 → 1:40 / 1:50 / 2:00 / 2:10 / 2:20)
                return max(600, base + _r.choice([-1200, -600, 0, 600, 1200]))
        except Exception:
            pass
        return base

    # ── 계정 선택 ──
    # ── 예약 발행 현황 (계정별 캐시 + 실시간 조회) ──
    def _resv_cache_file(self):
        return os.path.join(os.path.dirname(__file__), "reservation_status.json")

    def _active_blog_id(self):
        try:
            cfg = load_config()
            idx = cfg.get("active_account", 0)
            accs = cfg.get("accounts", [])
            return (accs[idx].get("blog_id", "") if idx < len(accs) else "") or f"acc{idx}"
        except Exception:
            return "default"

    def _load_resv_cache(self):
        try:
            with open(self._resv_cache_file(), "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}

    # ── 날짜별 발행 일정표 ──────────────────────────────────────────────────
    def _schedule_file(self):
        name = f"posting_schedule_{self._account_key()}.json"
        try:
            from app_paths import data_file as _df
            return _df(name)
        except Exception:
            return os.path.join(os.path.dirname(os.path.abspath(__file__)), name)

    def _append_schedule_record(self, blog_id, schedule_time, place_name, keyword, title):
        """발행/예약 1건을 날짜별 일정표 파일에 자동 기록. (기존 발행 로직과 무관, 실패해도 무시)"""
        try:
            import datetime as _dt
            if schedule_time:
                dt_str, status = schedule_time, "예약"
            else:
                dt_str, status = _dt.datetime.now().strftime("%Y-%m-%d %H:%M"), "즉시발행"
            rec = {"datetime": dt_str, "blog_id": blog_id or "", "keyword": keyword or "",
                   "place": place_name or "", "title": title or "", "status": status,
                   "recorded_at": _dt.datetime.now().strftime("%Y-%m-%d %H:%M")}
            path = self._schedule_file()
            data = []
            if os.path.exists(path):
                try:
                    with open(path, "r", encoding="utf-8") as f:
                        data = json.load(f) or []
                except Exception:
                    data = []
            data.append(rec)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=1)
        except Exception:
            pass

    def _show_schedule_calendar(self):
        """발행 일정표 — 기록을 날짜별로 묶어 보여줌."""
        from PySide6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QTreeWidget,
                                       QTreeWidgetItem, QPushButton, QLabel, QComboBox)
        from collections import defaultdict
        path = self._schedule_file()
        data = []
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f) or []
            except Exception:
                data = []

        dlg = QDialog(self)
        dlg.setWindowTitle("🗓️ 발행 일정표 (날짜별)")
        dlg.resize(800, 640)
        v = QVBoxLayout(dlg)

        top = QHBoxLayout()
        head = QLabel(f"총 {len(data)}건")
        head.setStyleSheet("font-weight:bold; font-size:13px;")
        top.addWidget(head)
        top.addWidget(QLabel("  명의:"))
        cmb = QComboBox()
        ids = ["전체"] + sorted({r.get("blog_id", "") for r in data if r.get("blog_id")})
        cmb.addItems(ids)
        top.addWidget(cmb)
        top.addStretch()
        v.addLayout(top)

        tree = QTreeWidget()
        tree.setHeaderLabels(["날짜 / 주제(키워드)", "개수"])
        tree.setColumnWidth(0, 470)
        tree.setColumnWidth(1, 90)
        v.addWidget(tree)

        def rebuild():
            tree.clear()
            sel = cmb.currentText()
            rows = [r for r in data if sel == "전체" or r.get("blog_id") == sel]
            groups = defaultdict(list)
            for r in rows:
                dtv = r.get("datetime", "")
                groups[dtv[:10] if len(dtv) >= 10 else "미정"].append(r)
            for date in sorted(groups.keys(), reverse=True):
                recs = groups[date]
                # 그 날 주제(키워드)별로 몇 개 올렸는지만 대략 집계
                by_kw = defaultdict(int)
                for r in recs:
                    by_kw[(r.get("keyword", "") or "(주제 없음)")] += 1
                parent = QTreeWidgetItem([f"📅 {date}", f"총 {len(recs)}개"])
                fnt = parent.font(0); fnt.setBold(True)
                parent.setFont(0, fnt); parent.setFont(1, fnt)
                for kw in sorted(by_kw.keys()):
                    parent.addChild(QTreeWidgetItem([f"    {kw}", f"{by_kw[kw]}개"]))
                tree.addTopLevelItem(parent)
                parent.setExpanded(True)

        cmb.currentIndexChanged.connect(rebuild)
        rebuild()

        row = QHBoxLayout()
        row.addStretch()
        b_close = QPushButton("닫기")
        b_close.clicked.connect(dlg.accept)
        row.addWidget(b_close)
        v.addLayout(row)
        dlg.exec()

    def _save_resv_status(self, blog_id, count, latest_str):
        """예약 현황을 계정별로 파일에 저장(파일만 — 라벨 갱신은 호출부에서 메인스레드로)."""
        if not blog_id:
            return
        try:
            import datetime as _dt
            c = self._load_resv_cache()
            c[blog_id] = {"count": int(count), "latest": latest_str or "",
                          "checked": _dt.datetime.now().strftime("%m-%d %H:%M")}
            with open(self._resv_cache_file(), "w", encoding="utf-8") as f:
                json.dump(c, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _update_reservation_label(self):
        if not hasattr(self, "resv_label"):
            return
        c = self._load_resv_cache().get(self._active_blog_id(), {})
        if not c:
            self.resv_label.setText("📅 예약: 조회 전")
            self.resv_label.setToolTip("옆 '예약조회' 버튼으로 확인하세요.")
            return
        import datetime as _dt
        cnt = c.get("count", 0)
        latest = c.get("latest", "")
        txt = f"📅 예약 {cnt}건"
        if latest:
            try:
                lt = _dt.datetime.strptime(latest, "%Y-%m-%d %H:%M")
                secs = (lt - _dt.datetime.now()).total_seconds()
                if secs > 0:
                    d = int(secs // 86400); h = int((secs % 86400) // 3600)
                    txt += f" · 마지막 {latest} ({d}일 {h}시간 남음)"
                else:
                    txt += f" · 마지막 {latest} (지남)"
            except Exception:
                txt += f" · 마지막 {latest}"
        if c.get("checked"):
            txt += f"   ({c['checked']} 기준)"
        self.resv_label.setText(txt)
        self.resv_label.setToolTip("")

    def _kill_profile_chrome(self, blog_id):
        """해당 계정 프로필을 쓰는 고아 크롬 프로세스 종료 (프로필 잠금/충돌 방지)."""
        if not blog_id:
            return
        try:
            import subprocess
            ps = ("Get-CimInstance Win32_Process -Filter \"Name='chrome.exe'\" | "
                  "Where-Object { $_.CommandLine -like '*chrome_profile*" + blog_id + "*' } | "
                  "ForEach-Object { Stop-Process -Id $_.ProcessId -Force -ErrorAction SilentlyContinue }")
            subprocess.run(["powershell", "-NoProfile", "-Command", ps],
                           timeout=15, capture_output=True)
        except Exception:
            pass

    def _check_reservations(self):
        if getattr(self, "_resv_checking", False):
            return
        if self.is_crawling or getattr(self, "is_posting", False):
            QMessageBox.information(self, "안내", "크롤링/발행 중에는 예약조회를 할 수 없습니다. 끝난 뒤 눌러주세요.")
            return
        cfg = load_config()
        idx = cfg.get("active_account", 0)
        accs = cfg.get("accounts", [])
        if idx >= len(accs):
            QMessageBox.warning(self, "안내", "계정 정보가 없습니다.")
            return
        account = accs[idx]
        if not account.get("naver_id") or not account.get("naver_pw"):
            QMessageBox.warning(self, "안내", "네이버 아이디/비밀번호를 먼저 설정해주세요.")
            return
        if self._block_if_expired():
            return
        self._resv_checking = True
        self.btn_resv.setEnabled(False)
        self.resv_label.setText("📅 예약 조회 중... (구석 로그인 창 — 캡챠 뜨면 앞으로 나옵니다)")
        bid = account.get("blog_id", "") or self._active_blog_id()

        from PySide6.QtCore import QObject, Signal as _Sig
        class _RS(QObject):
            done = _Sig(int, str, str)  # count, latest, error
        sig = _RS()

        def _on_done(count, latest, err):
            self._resv_checking = False
            self.btn_resv.setEnabled(True)
            if err:
                self.resv_label.setText("📅 예약 조회 실패")
                QMessageBox.warning(self, "예약 조회", f"조회 실패:\n{err}")
                return
            self._save_resv_status(bid, count, latest)
            self._update_reservation_label()

        sig.done.connect(_on_done)

        # 로그인 창을 계정점검과 동일한 정상 크기로 (작은 창이 봇탐지→캡챠 유발 가능성 → 동일하게 맞춤)
        _cx, _cy, _cw, _ch = 80, 60, 1180, 820

        def _worker():
            poster = None
            try:
                import time as _t
                self._kill_profile_chrome(bid)  # 시작 전 고아 크롬 정리 (잠금/충돌 방지)
                for _attempt in range(3):
                    try:
                        poster = NaverBlogPoster(
                            naver_id=account["naver_id"], naver_pw=account["naver_pw"],
                            blog_id=account["blog_id"],
                            window_x=_cx, window_y=_cy,   # 구석에서 로그인 (캡챠 뜨면 login()이 앞으로)
                            window_w=_cw, window_h=_ch,
                            stop_flag=lambda: False, speed_factor=0.7)  # 포스팅과 동일(봇탐지 회피)
                        poster.start_browser()
                        break
                    except Exception:
                        try:
                            if poster:
                                poster.close()
                        except Exception:
                            pass
                        poster = None
                        _t.sleep(2.0)
                if poster is None:
                    sig.done.emit(0, "", "브라우저를 시작하지 못했습니다. 잠시 후 다시 시도해주세요.")
                    return
                self._active_posters.append(poster)  # 참조 유지 → 워커 종료 시 드라이버 GC로 인한 크래시 방지
                if not poster.login():
                    sig.done.emit(0, "", "네이버 로그인 실패 (캡챠가 떴다면 창에서 풀고 다시 눌러주세요)")
                    return
                # 로그인 끝 → 창을 화면 밖으로 (수집 과정은 안 보이게)
                try:
                    poster.driver.set_window_position(-32000, -32000)
                except Exception:
                    pass
                existing = poster.peek_reservations() or []
                total = getattr(poster, "reservation_total", -1)
                count = total if (isinstance(total, int) and total > len(existing)) else len(existing)
                latest = max(existing).strftime("%Y-%m-%d %H:%M") if existing else ""
                sig.done.emit(int(count), latest, "")
            except Exception as e:
                sig.done.emit(0, "", str(e))
            finally:
                try:
                    if poster:
                        try:
                            self._active_posters.remove(poster)
                        except Exception:
                            pass
                        poster.close()
                except Exception:
                    pass
                try:
                    self._kill_profile_chrome(bid)  # 종료 후 잔류 크롬 정리
                except Exception:
                    pass

        import threading
        threading.Thread(target=_worker, daemon=True).start()

    def _show_acct_status(self, txt):
        self.btn_acct_check.setEnabled(True)
        self.btn_acct_check.setText("🛡️ 계정점검")
        QMessageBox.information(self, "계정 상태 점검 결과", txt)

    def _check_accounts_status(self):
        if getattr(self, "_acct_checking", False):
            return
        if self.is_crawling or getattr(self, "is_posting", False):
            QMessageBox.information(self, "안내", "크롤링/발행 중에는 계정점검을 할 수 없습니다.")
            return
        cfg = load_config()
        accs = cfg.get("accounts", [])
        targets = [a for a in accs if a.get("naver_id") and a.get("naver_pw")]
        if not targets:
            QMessageBox.warning(self, "안내", "네이버 아이디/비밀번호가 설정된 계정이 없습니다.")
            return
        reply = QMessageBox.question(
            self, "계정 점검",
            f"{len(targets)}개 계정의 보호조치 여부를 점검합니다.\n계정마다 로그인해서 확인하므로 몇 분 걸릴 수 있어요. 진행할까요?",
            QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return
        self._acct_checking = True
        self.btn_acct_check.setEnabled(False)
        self.btn_acct_check.setText("점검 중...")

        def _worker():
            results = []
            try:
                for a in targets:
                    bid = a.get("blog_id", "") or a.get("naver_id", "")
                    self._emit_post_log(f"[계정점검] {bid} 확인 중...")
                    poster = None
                    try:
                        self._kill_profile_chrome(a.get("blog_id", ""))
                        poster = NaverBlogPoster(
                            naver_id=a["naver_id"], naver_pw=a["naver_pw"],
                            blog_id=a.get("blog_id", "") or a.get("naver_id", ""),
                            window_x=80, window_y=60,   # 보이게 — 캡챠 뜨면 처리 가능(포스팅과 동일)
                            window_w=1180, window_h=820, stop_flag=lambda: False, speed_factor=0.7)
                        poster.start_browser()
                        status = poster.check_status()
                    except Exception as e:
                        status = f"확인불가({str(e)[:20]})"
                    finally:
                        try:
                            if poster:
                                poster.close()
                        except Exception:
                            pass
                        try:
                            self._kill_profile_chrome(a.get("blog_id", ""))
                        except Exception:
                            pass
                    icon = {"정상": "✅", "보호조치": "🛑", "캡챠필요": "🔐",
                            "로그인실패": "❌", "확인불가": "⚠️"}.get(status, "⚠️")
                    results.append(f"{icon} {bid} : {status}")
                    self._emit_post_log(f"[계정점검] {bid} → {status}")
            except Exception as e:
                results.append(f"오류: {e}")
            finally:
                self._acct_checking = False
            self.acct_status_signal.emit("\n".join(results) if results else "결과 없음")

        import threading
        threading.Thread(target=_worker, daemon=True).start()

    def _refresh_account_combo(self):
        accounts = self.cfg.get("accounts", [])
        active = self.cfg.get("active_account", 0)
        self.account_combo.blockSignals(True)
        self.account_combo.clear()
        for i in range(9):
            acc = accounts[i] if i < len(accounts) else {}
            nid = (acc.get("naver_id") or "").strip()
            label = f"{i+1}. {nid}" if nid else f"{i+1}. (미설정)"
            self.account_combo.addItem(label)
        if 0 <= active < self.account_combo.count():
            self.account_combo.setCurrentIndex(active)
        self.account_combo.blockSignals(False)

    def _on_account_changed(self, idx):
        if idx < 0:
            return
        prev_idx = self.cfg.get("active_account", 0)
        # 작업 중엔 계정 전환 금지 — 수집/발행 데이터가 다른 계정으로 섞이는 것 방지
        if self.is_crawling or getattr(self, 'is_posting', False) or getattr(self, 'is_generating', False):
            self.account_combo.blockSignals(True)
            self.account_combo.setCurrentIndex(prev_idx)
            self.account_combo.blockSignals(False)
            QMessageBox.warning(self, "안내",
                "크롤링·발행·생성이 진행 중일 때는 계정을 전환할 수 없습니다.\n작업이 끝난 뒤 전환해 주세요.")
            return
        if idx != prev_idx:
            accounts = self.cfg.get("accounts", [])
            new_bid = (accounts[idx].get("blog_id", "") if idx < len(accounts) else "") or f"acc{idx}"
            ret = QMessageBox.question(self, "계정 전환", f"계정을 {new_bid}(으)로 바꾸시겠습니까?",
                                       QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if ret != QMessageBox.Yes:
                self.account_combo.blockSignals(True)
                self.account_combo.setCurrentIndex(prev_idx)
                self.account_combo.blockSignals(False)
                return
        cfg = load_config()
        cfg["active_account"] = idx
        save_config(cfg)
        self.cfg = cfg
        # 계정 전환 시 메모리 상태 초기화 (크롤/생성 데이터는 계정별 파일에서 다시 로드됨)
        self.crawled_data = []
        self._generated_posts = []
        self.result_file = ""
        self.keyword_input.clear()
        self._load_keyword_history()
        bid = cfg.get('accounts', [{}])[idx].get('blog_id', '') or '(미설정)'
        self._emit_log(f"계정 전환: {idx+1}번 ({bid})")
        try:
            self._update_reservation_label()  # 전환된 계정의 예약 현황 캐시 표시
        except Exception:
            pass
        # 전환된 명의의 만료 상태를 봇1에 즉시 반영
        try:
            if self._is_user_expired():
                self._show_expired_bot1()
            else:
                self._clear_expired_bot1()
        except Exception:
            pass

    def _current_blog_id(self) -> str:
        try:
            idx = self.cfg.get("active_account", 0)
            return self.cfg.get("accounts", [])[idx].get("blog_id", "") or f"acc{idx}"
        except Exception:
            return "default"

    # ── 키워드 히스토리 (엑셀 저장, 계정별) ──
    def _get_history_file(self):
        # app_user 기준 단일 파일 — 같은 로그인 사용자의 모든 네이버 슬롯 공유
        return os.path.join(os.path.dirname(__file__), f"search_history_{self._account_key()}.xlsx")

    def _load_keyword_history(self):
        try:
            import openpyxl
        except Exception:
            self.keyword_input.setCurrentText("")
            return
        filepath = self._get_history_file()
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    self.keyword_input.addItem(str(row[0]))
            wb.close()
        except Exception:
            pass
        # 입력칸은 비우고 드롭다운에만 히스토리 유지
        self.keyword_input.setCurrentText("")

    def _save_keyword_history(self, keyword: str):
        try:
            import openpyxl
        except Exception:
            return
        filepath = self._get_history_file()
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
        except Exception:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "검색 히스토리"
            ws.append(["키워드", "검색일시", "수집 결과수"])

        ws.append([keyword, now, ""])
        wb.save(filepath)
        wb.close()

        # 드롭다운 갱신 (중복 제거)
        existing = set()
        items = []
        for i in range(self.keyword_input.count()):
            items.append(self.keyword_input.itemText(i))
            existing.add(self.keyword_input.itemText(i))
        if keyword not in existing:
            self.keyword_input.insertItem(0, keyword)
        self.keyword_input.setCurrentText(keyword)

    # ── 키워드마스터 연동 (핸드오프 수신) ──
    def _km_handoff_path(self):
        base = os.environ.get("LOCALAPPDATA") or os.environ.get("APPDATA") or os.path.expanduser("~")
        return os.path.join(base, "BlogMaster", "selected_keywords.json")

    def _append_keyword_history_only(self, keyword: str):
        """히스토리 엑셀에만 추가 (현재 입력칸/드롭다운은 호출부에서 처리)."""
        try:
            import openpyxl
        except Exception:
            return
        filepath = self._get_history_file()
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            wb = openpyxl.load_workbook(filepath); ws = wb.active
        except Exception:
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "검색 히스토리"
            ws.append(["키워드", "검색일시", "수집 결과수"])
        ws.append([keyword, now, "(키워드마스터)"])
        wb.save(filepath); wb.close()

    def _import_from_keyword_master(self):
        """키워드마스터가 남긴 selected_keywords.json → 키워드 드롭다운/히스토리에 반영."""
        try:
            import json as _json
            path = self._km_handoff_path()
            if not os.path.exists(path):
                return
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = _json.load(f)
            except Exception:
                return
            kws = [str(k).strip() for k in (data.get("place_keywords") or []) if str(k).strip()]
            # 파일 소비 — 재처리 방지 (읽자마자 .done 으로 이동)
            try:
                done = path + ".done"
                if os.path.exists(done):
                    os.remove(done)
                os.replace(path, done)
            except Exception:
                try:
                    os.remove(path)
                except Exception:
                    pass
            if not kws:
                return
            existing = set(self.keyword_input.itemText(i) for i in range(self.keyword_input.count()))
            added = 0
            for kw in kws:
                if kw not in existing:
                    self.keyword_input.insertItem(0, kw)
                    existing.add(kw)
                    try:
                        self._append_keyword_history_only(kw)
                    except Exception:
                        pass
                    added += 1
            self.keyword_input.setCurrentText("")
            tgt_blog = data.get("target_blog_id") or ""
            self._emit_log(f"키워드마스터에서 키워드 {added}개 수신" + (f" (대상 명의: {tgt_blog})" if tgt_blog else ""))
            try:
                QMessageBox.information(
                    self, "키워드마스터 연동",
                    f"키워드마스터에서 키워드 {added}개를 받았습니다."
                    + (f"\n대상 명의: {tgt_blog}" if tgt_blog else "")
                    + "\n\n'키워드 입력' 드롭다운에서 골라 크롤을 시작하세요.")
            except Exception:
                pass
        except Exception:
            pass

    def _update_history_result_count(self, keyword: str, count: int):
        try:
            import openpyxl
        except Exception:
            return
        filepath = self._get_history_file()
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            # 마지막 행에서 해당 키워드 찾아서 결과수 업데이트
            for row in reversed(list(ws.iter_rows(min_row=2))):
                if row[0].value == keyword and not row[2].value:
                    row[2].value = count
                    break
            wb.save(filepath)
            wb.close()
        except Exception:
            pass


    # ── 자동 업데이트 ──
    @Slot()
    def _show_update_dialog(self):
        reply = QMessageBox.question(self, "업데이트",
            "새 버전이 있습니다.\n지금 다운로드하고 설치하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return
        self._emit_log("업데이트 다운로드 중...")

        def _download():
            try:
                import requests as _req
                import tempfile, zipfile, shutil, subprocess
                url = "https://github.com/kingth0506/BlogMaster/releases/latest/download/BlogMaster_files.zip"
                r = _req.get(url, stream=True, timeout=300)
                tmp_zip = os.path.join(tempfile.gettempdir(), "BlogMaster_update.zip")
                with open(tmp_zip, "wb") as f:
                    for chunk in r.iter_content(chunk_size=65536):
                        f.write(chunk)
                self._emit_log("다운로드 완료. 업데이트 적용 중...")
                # 현재 exe 위치 기준으로 교체
                app_dir = os.path.dirname(os.path.abspath(sys.executable))
                tmp_extract = os.path.join(tempfile.gettempdir(), "BlogMaster_update")
                if os.path.exists(tmp_extract):
                    shutil.rmtree(tmp_extract)
                with zipfile.ZipFile(tmp_zip, "r") as zf:
                    zf.extractall(tmp_extract)
                # 배치 스크립트로 현재 프로세스 종료 후 파일 교체 → 재시작
                bat = os.path.join(tempfile.gettempdir(), "blogmaster_update.bat")
                exe_name = os.path.basename(sys.executable)
                with open(bat, "w", encoding="utf-8") as f:
                    f.write(f"@echo off\n")
                    f.write(f"timeout /t 2 /nobreak >nul\n")
                    f.write(f"xcopy /E /Y /I \"{tmp_extract}\" \"{app_dir}\"\n")
                    f.write(f"start \"\" \"{os.path.join(app_dir, exe_name)}\"\n")
                    f.write(f"del \"%~f0\"\n")
                subprocess.Popen(["cmd", "/c", bat], shell=False)
                self._app_quit_signal.emit()
            except Exception as e:
                self._emit_log(f"업데이트 실패: {e}")

        import threading
        threading.Thread(target=_download, daemon=True).start()

    # ── 관리자 메뉴 ──
    def _open_admin_dialog(self):
        dlg = AdminDialog(self)
        dlg.exec()

    def apply_user_session(self, user: dict):
        """로그인 성공 후 호출 — 역할에 따라 관리자 버튼 노출 + 계정 슬롯 제한"""
        self.current_user = user
        # 로그인 시 로컬 프롬프트를 '기본 + 이 사용자 것'만으로 정리 (이전 사용자 프롬프트 제거)
        try:
            self._scope_prompts_to_user(user)
        except Exception:
            pass
        try:
            self.btn_admin.setVisible(user.get("role") == "admin")
        except Exception:
            pass
        # 계정 슬롯 제한 (max_accounts: 3/6/9), 관리자는 9개 모두 허용
        if user.get("role") == "admin":
            max_acc = 9
        else:
            max_acc = int(user.get("max_accounts", 3))
        try:
            # 콤보박스 항목 수 제한
            while self.account_combo.count() > max_acc:
                self.account_combo.removeItem(self.account_combo.count() - 1)
        except Exception:
            pass
        # 관리자: Firebase users.api_keys → 로컬 config 동기화 (다중PC 자동 적용)
        if user.get("role") == "admin":
            self._sync_admin_api_keys_from_firebase(user)
        self._update_expires_label()
        self.cfg = load_config()
        self._refresh_account_combo()
        self._load_interval_settings()
        # 고아 이미지 폴더 자동 정리 (현재 생성 포스트 없는 폴더 제거)
        try:
            self._cleanup_orphan_image_folders()
        except Exception:
            pass
        # 세션 하트비트 (5분마다 갱신)
        _sid = user.get("_session_id", "")
        _uname = user.get("username", "")
        if _sid and _uname and user.get("role") != "admin":
            import threading as _th_hb, time as _time_hb
            def _heartbeat_loop(uname=_uname, sid=_sid):
                while True:
                    _time_hb.sleep(5 * 60)
                    if getattr(self, "current_user", {}).get("_session_id") != sid:
                        break
                    try:
                        from users import heartbeat_session
                        heartbeat_session(uname, sid)
                    except Exception:
                        pass
            _th_hb.Thread(target=_heartbeat_loop, daemon=True).start()

        # 만료 시 결제창 자동 팝업 (admin 제외, 00시 기준)
        if user.get("role") != "admin":
            try:
                import datetime as _dt
                _expires = (user.get("expires") or "").strip()
                if _expires:
                    _y, _m, _d = map(int, _expires.split("-"))
                    _days_left = (_dt.date(_y, _m, _d) - _dt.date.today()).days
                    if _days_left < 0:
                        from PySide6.QtCore import QTimer
                        QTimer.singleShot(300, self._open_payment_dialog)
            except Exception:
                pass

        # 만료 사용자: 진입 즉시 봇1에 '이용기간 만료' 표시
        if self._is_user_expired():
            self._show_expired_bot1()

        # 개인정보 수집·이용 동의 없으면 동의 창 (필수)
        self._require_privacy_consent()

        # 공용키 유예 안내 — 기존 공용키 사용자는 유예 종료일까지 매 로그인 시 안내
        try:
            self._notify_shared_key_grace(user)
        except Exception:
            pass

    def _notify_shared_key_grace(self, user: dict):
        """공용(부여) API 키 유예 대상에게 '며칠 후 사라짐' 안내 (매 로그인 1회).
        유예 종료 후엔 '만료됨 → 본인 키 입력' 안내."""
        if not user or user.get("role") == "admin":
            return
        g = (user.get("shared_api_grace_until") or "").strip()
        if not g or not user.get("shared_api_keys_admin_granted"):
            return
        import datetime as _dt
        try:
            y, m, d = map(int, g.split("-"))
            left = (_dt.date(y, m, d) - _dt.date.today()).days
        except Exception:
            return
        if left >= 0:
            QMessageBox.warning(
                self, "API 키 안내",
                f"제공되던 공용 API 키가 {g}에 종료됩니다. (D-{left})\n\n"
                f"종료 이후에는 글쓰기가 되지 않으니, 그 전에 설정에서 본인 API 키\n"
                f"(딥시크 · 챗GPT · 제미나이 중 택1)를 발급받아 입력해주세요.\n\n"
                f"딥시크 발급: platform.deepseek.com")
        else:
            QMessageBox.critical(
                self, "API 키 만료",
                "제공되던 공용 API 키가 만료되었습니다.\n\n"
                "이제부터는 본인 API 키를 입력해야 글쓰기가 가능합니다.\n"
                "설정 → API 키에서 딥시크 · 챗GPT · 제미나이 중 하나를 발급받아 입력해주세요.\n\n"
                "딥시크 발급: platform.deepseek.com")

    def _account_slot(self, acct_idx: int) -> int:
        """아이디 인덱스 → 명의(요금제 슬롯 1/2/3).
        명의는 3개 고정, 계정은 3명의로 나눠 담는다(명의당 아이디수=ceil(전체/3)).
        예) 9개 → 3개씩(0~2=1, 3~5=2, 6~8=3),  3개 → 1개씩(0=1,1=2,2=3)."""
        import math
        try:
            from config import load_config
            n = len((load_config() or {}).get("accounts", []) or [])
        except Exception:
            n = 3
        per = max(1, math.ceil(n / 3)) if n else 1
        return min(3, int(acct_idx) // per + 1)

    def _is_user_expired(self) -> bool:
        """현재 활성 아이디가 속한 명의의 이용기간 만료 여부 (관리자는 항상 False)."""
        try:
            from users import is_account_expired
            from config import load_config
            slot = self._account_slot(int((load_config() or {}).get("active_account", 0) or 0))
            return is_account_expired(getattr(self, "current_user", {}) or {}, slot)
        except Exception:
            return False

    def _show_expired_bot1(self):
        """크롤링 현황 봇1 영역에 '이용기간 만료'를 현재 폰트보다 2pt 크게(빨강) 표시."""
        try:
            lbls = getattr(self, "crawl_worker_region_labels", None)
            if lbls:
                lbls[0].setText("이용기간 만료")
                # 봇 지역 라벨 기본 11px → +2pt = 13px
                lbls[0].setStyleSheet("font-size: 13px; color: #ef4444; font-weight: bold; padding-left: 4px;")
        except Exception:
            pass
        try:
            logs = getattr(self, "crawl_worker_logs", None)
            if logs:
                # 봇 로그 기본 11px → +2pt = 13px
                logs[0].setHtml("<div style='color:#ef4444;font-size:13px;font-weight:bold;'>이용기간 만료</div>")
        except Exception:
            pass

    def _clear_expired_bot1(self):
        """봇1의 '이용기간 만료' 표시 해제 (정상 명의로 전환 시)."""
        try:
            lbls = getattr(self, "crawl_worker_region_labels", None)
            if lbls:
                lbls[0].setText("")
                lbls[0].setStyleSheet("font-size: 11px; color: #334155; font-weight: bold; padding-left: 4px;")
        except Exception:
            pass
        try:
            logs = getattr(self, "crawl_worker_logs", None)
            if logs:
                logs[0].clear()
        except Exception:
            pass

    def _block_if_expired(self) -> bool:
        """만료면 봇1에 표시하고 안내 후 True(차단) 반환. 정상이면 False."""
        if self._is_user_expired():
            self._show_expired_bot1()
            try:
                QMessageBox.warning(self, "이용기간 만료", "이용기간 만료")
            except Exception:
                pass
            return True
        return False

    def _block_if_api_expired(self) -> bool:
        """활성 명의의 API(GPT 글쓰기) 구독이 만료면 안내 후 True(차단). 정상이면 False.
        단, 자기(본인) API 키를 쓰는 유저는 공용키 구독과 무관하므로 만료로 차단하지 않는다."""
        try:
            from users import is_api_expired as _is_api_exp, load_users as _lu
            from config import get_current_user as _gcu, load_config as _lc
            cfg = _lc() or {}
            uname = _gcu() or ""
            user = _lu().get(uname, {})
            # 자기 API 키 사용자 → 만료 차단 안 함 (본인이 API 결제 주체)
            def _own(src):
                return any(any(k for k in ((src or {}).get(f) or [])) for f in ("gpt_key_list", "gemini_key_list", "deepseek_key_list"))
            if cfg.get("_own_ai_key") or _own(user.get("api_keys")) or _own((cfg.get("api_keys_by_user") or {}).get(uname, {})):
                return False
            slot = self._account_slot(int(cfg.get("active_account", 0) or 0))
            if _is_api_exp(user, slot=slot):
                reply = QMessageBox.question(
                    self, "API 기간 만료",
                    "API 기간이 만료되었습니다.\n키를 수정하거나 결제 부탁드립니다.\n\n결제 창을 여시겠습니까?",
                    QMessageBox.Yes | QMessageBox.No)
                if reply == QMessageBox.Yes:
                    self._open_payment_dialog()
                return True
        except Exception:
            pass
        return False

    def _require_privacy_consent(self):
        """개인정보 수집·이용 동의가 없는 사용자에게 동의 창을 띄우고, 동의 시 Firebase 저장.
        동의 안 하면 프로그램 종료. (관리자·이미 동의한 사용자는 통과)"""
        try:
            user = getattr(self, "current_user", {}) or {}
            if user.get("role") == "admin" or user.get("privacy_consent"):
                return
            dlg = QDialog(self)
            dlg.setWindowTitle("개인정보 수집·이용 동의")
            dlg.setMinimumWidth(460)
            dlg.setModal(True)
            v = QVBoxLayout(dlg)
            v.setContentsMargins(22, 20, 22, 18)
            v.setSpacing(12)
            _title = QLabel("개인정보 수집·이용 동의 (필수)")
            _title.setStyleSheet("font-size:16px; font-weight:bold;")
            v.addWidget(_title)
            _body = QLabel(
                "서비스 제공 및 본인확인을 위해 아래 개인정보를 수집·이용합니다.\n\n"
                "• 수집 항목: 아이디, 이름, 생년월일, 연락처, 이메일\n"
                "• 이용 목적: 회원 식별·본인확인, 서비스 제공, 고객 지원\n"
                "• 보유 기간: 회원 탈퇴 시까지 (관련 법령에 따른 보존 제외)\n\n"
                "동의하셔야 서비스를 이용하실 수 있습니다."
            )
            _body.setWordWrap(True)
            _body.setStyleSheet("font-size:13px; color:#334155;")
            v.addWidget(_body)
            _chk = QCheckBox("위 개인정보 수집·이용에 동의합니다.")
            _chk.setStyleSheet("font-size:13px; font-weight:bold;")
            v.addWidget(_chk)
            _row = QHBoxLayout()
            _btn_ok = QPushButton("동의하고 시작")
            _btn_ok.setStyleSheet("background:#6366f1; color:white; border:none; border-radius:8px; padding:9px 18px; font-weight:bold;")
            _btn_no = QPushButton("동의 안 함(종료)")
            _btn_no.setStyleSheet("padding:9px 18px;")
            _row.addStretch(); _row.addWidget(_btn_no); _row.addWidget(_btn_ok)
            v.addLayout(_row)

            def _agree():
                if not _chk.isChecked():
                    QMessageBox.warning(dlg, "동의 필요", "동의 체크 후 진행해주세요.")
                    return
                import datetime as _dt
                _today = _dt.date.today().strftime("%Y-%m-%d")
                try:
                    from users import update_user
                    update_user(user.get("username", ""), privacy_consent=True, privacy_consent_date=_today)
                except Exception:
                    pass
                self.current_user["privacy_consent"] = True
                self.current_user["privacy_consent_date"] = _today
                dlg.accept()

            _btn_ok.clicked.connect(_agree)
            _btn_no.clicked.connect(dlg.reject)
            if dlg.exec() != QDialog.Accepted:
                QApplication.quit()
        except Exception:
            pass

    def _normalize_post_images(self, pix_key: str):
        """각 포스트의 실제 저장 이미지 갯수를 image_count에 맞춰 평준화
        - 부족: Pixabay 재다운로드로 채움
        - 초과: 앞 N장만 남기고 제거
        """
        import shutil as _shutil
        fixed = 0
        for post in self._generated_posts:
            if post.get("posted"):
                continue
            place = post.get("place", {}) or {}
            content = post.get("content", {}) or {}
            target = int(content.get("image_count", 3) or 3)
            if target < 1:
                continue
            pkey = self._place_key(place)
            safe = "".join(c for c in (pkey[0] + "_" + pkey[1]) if c.isalnum() or c in "-_")[:80] or "unknown"
            pdir = os.path.join(os.path.dirname(__file__), "saved_images",
                                self._account_key(), safe)
            if not os.path.isdir(pdir):
                continue
            imgs = sorted([f for f in os.listdir(pdir)
                          if f.lower().endswith((".jpg", ".jpeg", ".png")) and not f.startswith("._")])
            cur = len(imgs)
            if cur == target:
                continue
            if cur > target:
                for extra in imgs[target:]:
                    try: os.remove(os.path.join(pdir, extra))
                    except Exception: pass
                content["image_paths"] = [os.path.join(pdir, f) for f in imgs[:target]]
                self._emit_post_log(f"[평준화] {place.get('name','')}: {cur}→{target}장 (초과 제거)")
                fixed += 1
                continue
            # 부족 → 추가 다운로드
            need = target - cur
            biz = self._best_biz_term(place, "")
            # 업종별 Pixabay 검색어 후보
            _plist = []
            try:
                from app_paths import ensure_from_bundle as _efb
                _pp = _efb("prompts.json")
                with open(_pp, "r", encoding="utf-8") as _pf:
                    _prompts = json.load(_pf)
                _entry = _prompts.get(biz, {}) or {}
                _plist_raw = _entry.get("pixabay_list") or []
                if not _plist_raw and _entry.get("pixabay"):
                    _plist_raw = [_entry.get("pixabay", "")]
                _plist = [s.strip() for s in _plist_raw if s and s.strip()]
            except Exception:
                _plist = []
            import random as _rnd
            added = 0
            for _i in range(need):
                if _plist:
                    _picked = _rnd.choice(_plist)
                    _query = self._translate_ko_to_en(_picked)
                else:
                    _query = biz
                try:
                    _r = download_images(pix_key, _query, 1)
                except Exception:
                    _r = []
                if not _r:
                    continue
                dst = os.path.join(pdir, f"image_{cur + added + 1}.jpg")
                try:
                    _shutil.copyfile(_r[0], dst)
                    added += 1
                except Exception:
                    pass
            # image_paths 갱신
            imgs_new = sorted([f for f in os.listdir(pdir)
                              if f.lower().endswith((".jpg", ".jpeg", ".png")) and not f.startswith("._")])
            content["image_paths"] = [os.path.join(pdir, f) for f in imgs_new[:target]]
            new_cur = len(imgs_new)
            if new_cur != cur:
                self._emit_post_log(f"[평준화] {place.get('name','')}: {cur}→{new_cur}장 (목표 {target}장)")
                fixed += 1
        if fixed:
            self._emit_post_log(f"이미지 평준화 완료: {fixed}개 포스트 조정")

    def _refresh_excludes_button(self):
        total = sum(len(v) for v in self._exclude_keywords_by_biz.values())
        if total == 0:
            self.btn_excludes.setText("▼ 제외 키워드 (없음)")
        else:
            biz_count = len(self._exclude_keywords_by_biz)
            self.btn_excludes.setText(f"▼ 제외 키워드 ({biz_count}업종, 총 {total}개)")

    def _save_exclude_keywords(self):
        try:
            _c = load_config()
            _c["exclude_keywords_by_biz"] = {
                k: list(v) for k, v in self._exclude_keywords_by_biz.items() if v
            }
            save_config(_c)
            self.cfg = load_config()
        except Exception:
            pass

    def _current_biz_type(self) -> str:
        """키워드 마지막 단어 = 업종"""
        kw = (self.keyword_input.currentText() or "").strip()
        parts = kw.split()
        return parts[-1] if parts else ""

    def _open_excludes_dialog(self):
        """업종별 제외 키워드 — 업종 드롭다운(+추가 버튼), 키워드 입력 + Enter/+, 목록에서 - 삭제"""
        from PySide6.QtWidgets import (
            QDialog, QVBoxLayout, QHBoxLayout, QComboBox, QPushButton,
            QLabel, QLineEdit, QScrollArea, QWidget, QInputDialog
        )

        dlg = QDialog(self)
        dlg.setWindowTitle("제외 키워드 설정")
        dlg.resize(460, 560)
        layout = QVBoxLayout(dlg)

        info = QLabel("업체명/카테고리/주소에 제외 키워드가 하나라도 포함되면\n"
                      "크롤링 결과에서 제외됩니다. 업종별로 따로 저장됩니다.")
        info.setStyleSheet("color: #475569; font-size: 11px;")
        layout.addWidget(info)

        # 업종 선택 + 업종 추가 버튼
        row = QHBoxLayout()
        row.addWidget(QLabel("업종:"))
        biz_combo = QComboBox()
        biz_combo.setEditable(False)
        all_biz = set(self._exclude_keywords_by_biz.keys())
        cur_biz = self._current_biz_type()
        if cur_biz:
            all_biz.add(cur_biz)
        for b in sorted(all_biz):
            biz_combo.addItem(b)
        if cur_biz:
            idx = biz_combo.findText(cur_biz)
            if idx >= 0:
                biz_combo.setCurrentIndex(idx)
        row.addWidget(biz_combo, 1)
        btn_add_biz = QPushButton("업종 추가")
        btn_add_biz.setStyleSheet("padding: 4px 10px;")
        btn_add_biz.setAutoDefault(False)
        btn_add_biz.setDefault(False)
        row.addWidget(btn_add_biz)
        layout.addLayout(row)

        # 키워드 추가 줄 (입력 + 버튼)
        add_row = QHBoxLayout()
        kw_input = QLineEdit()
        kw_input.setPlaceholderText("제외할 키워드 입력 후 Enter 또는 +")
        add_row.addWidget(kw_input, 1)
        btn_plus = QPushButton("+")
        btn_plus.setStyleSheet(
            "background:#22c55e;color:white;font-weight:bold;font-size:16px;"
            "padding:4px 12px;border-radius:6px;min-width:36px;"
        )
        btn_plus.setAutoDefault(False)
        btn_plus.setDefault(False)
        add_row.addWidget(btn_plus)
        layout.addLayout(add_row)

        # 키워드 리스트 (각각 - 삭제 버튼)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        canvas = QWidget()
        list_layout = QVBoxLayout(canvas)
        list_layout.setContentsMargins(6, 6, 6, 6)
        list_layout.setSpacing(4)
        list_layout.addStretch()
        scroll.setWidget(canvas)
        layout.addWidget(scroll, 1)

        bottom = QHBoxLayout()
        bottom.addStretch()
        btn_close = QPushButton("닫기")
        btn_close.setStyleSheet("padding:6px 14px;")
        btn_close.setAutoDefault(False)
        btn_close.setDefault(False)
        btn_close.clicked.connect(dlg.accept)
        bottom.addWidget(btn_close)
        layout.addLayout(bottom)

        def _current_list() -> list:
            biz = biz_combo.currentText().strip()
            if not biz:
                return []
            return list(self._exclude_keywords_by_biz.get(biz, []))

        def _persist(items: list):
            biz = biz_combo.currentText().strip()
            if not biz:
                return
            cleaned = []
            for k in items:
                k = (k or "").strip()
                if k and k not in cleaned:
                    cleaned.append(k)
            if cleaned:
                self._exclude_keywords_by_biz[biz] = cleaned
            else:
                self._exclude_keywords_by_biz.pop(biz, None)
            self._save_exclude_keywords()
            self._refresh_excludes_button()

        def _render():
            # 기존 아이템 위젯 지우기 (마지막 stretch 유지)
            while list_layout.count() > 1:
                it = list_layout.takeAt(0)
                w = it.widget()
                if w:
                    w.setParent(None)
            items = _current_list()
            for idx, kw in enumerate(items):
                row_w = QWidget()
                row_l = QHBoxLayout(row_w)
                row_l.setContentsMargins(0, 0, 0, 0)
                lbl = QLabel(kw)
                lbl.setStyleSheet(
                    "background:#f1f5f9;border:1px solid #e2e8f0;"
                    "border-radius:6px;padding:6px 10px;"
                )
                row_l.addWidget(lbl, 1)
                btn_del = QPushButton("−")
                btn_del.setStyleSheet(
                    "background:#ef4444;color:white;font-weight:bold;"
                    "padding:4px 12px;border-radius:6px;min-width:36px;"
                )
                btn_del.setCursor(Qt.PointingHandCursor)
                def _mk_del(k=kw):
                    def _do():
                        cur = _current_list()
                        cur = [x for x in cur if x != k]
                        _persist(cur)
                        _render()
                    return _do
                btn_del.clicked.connect(_mk_del())
                row_l.addWidget(btn_del)
                list_layout.insertWidget(list_layout.count() - 1, row_w)

        def _add_keyword():
            kw = (kw_input.text() or "").strip()
            if not kw:
                return
            cur = _current_list()
            if kw not in cur:
                cur.append(kw)
                _persist(cur)
                _render()
            kw_input.clear()
            kw_input.setFocus()

        def _add_biz():
            text, ok = QInputDialog.getText(dlg, "업종 추가", "업종명:")
            if not ok:
                return
            text = (text or "").strip()
            if not text:
                return
            if biz_combo.findText(text) < 0:
                biz_combo.addItem(text)
            biz_combo.setCurrentText(text)
            # 빈 목록으로 초기화 (저장은 키워드 추가 시)
            _render()

        kw_input.returnPressed.connect(_add_keyword)
        btn_plus.clicked.connect(_add_keyword)
        btn_add_biz.clicked.connect(_add_biz)
        biz_combo.currentTextChanged.connect(lambda _: _render())

        _render()
        dlg.exec()

    def _refresh_regions_button(self):
        n = len(self._selected_regions)
        if n == 0:
            self.btn_regions.setText("▼ 지역설정 (선택 없음 — 키워드만 사용)")
        else:
            preview = ", ".join(self._selected_regions[:3])
            more = f" 외 {n-3}" if n > 3 else ""
            self.btn_regions.setText(f"▼ 지역설정 ({n}개: {preview}{more})")

    def _save_selected_regions(self):
        try:
            _c = load_config()
            _c["selected_regions"] = list(self._selected_regions)
            save_config(_c)
            self.cfg = load_config()
        except Exception:
            pass

    def _open_regions_dialog(self):
        """시/도(1차) → 구/군(2차) 트리 체크박스 다이얼로그"""
        from PySide6.QtWidgets import (
            QDialog, QVBoxLayout, QHBoxLayout, QTreeWidget, QTreeWidgetItem,
            QPushButton, QLabel
        )
        from places_crawler import SIDO_DISTRICTS

        dlg = QDialog(self)
        dlg.setWindowTitle("지역 설정")
        dlg.resize(420, 600)
        layout = QVBoxLayout(dlg)

        info = QLabel("크롤링할 시/도 및 구/군을 체크하세요.\n"
                      "선택 시 키워드 앞에 자동 부여되어 각 지역별로 크롤링됩니다.")
        info.setStyleSheet("color: #475569; font-size: 11px;")
        layout.addWidget(info)

        tree = QTreeWidget()
        tree.setHeaderHidden(True)
        layout.addWidget(tree, 1)

        selected_set = set(self._selected_regions)
        sido_items = {}
        for sido, dists in SIDO_DISTRICTS.items():
            parent = QTreeWidgetItem(tree, [sido])
            parent.setFlags(parent.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsAutoTristate)
            parent.setCheckState(0, Qt.Unchecked)
            sido_items[sido] = parent
            for d in dists:
                child = QTreeWidgetItem(parent, [d])
                child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
                # 시도명이 구 이름에 이미 포함되면 중복 제거 (세종/세종시 → 세종시, 제주/제주시 → 제주시)
                key = d if d.startswith(sido) else f"{sido} {d}"
                child.setCheckState(0, Qt.Checked if key in selected_set else Qt.Unchecked)
        # 시·도 항목들은 기본 접힘 — 사용자가 클릭해서 펼치도록
        tree.collapseAll()

        # 시/도 부모 클릭 → 그 아래 구/군 전체 토글 (서울 전체 등 일괄 선택 편의)
        _propagating = {"on": False}
        def _on_item_changed(item, col):
            if col != 0 or _propagating["on"]:
                return
            if item.parent() is None and item.childCount() > 0:
                st = item.checkState(0)
                if st == Qt.PartiallyChecked:
                    return
                _propagating["on"] = True
                try:
                    for i in range(item.childCount()):
                        item.child(i).setCheckState(0, st)
                finally:
                    _propagating["on"] = False
        tree.itemChanged.connect(_on_item_changed)

        btn_row = QHBoxLayout()
        btn_all = QPushButton("전체 선택")
        btn_none = QPushButton("전체 해제")
        btn_ok = QPushButton("확인")
        btn_cancel = QPushButton("취소")
        btn_ok.setStyleSheet("background:#3b82f6;color:white;padding:6px 14px;border-radius:6px;")
        btn_cancel.setStyleSheet("padding:6px 14px;")

        def _set_all(state):
            for p in sido_items.values():
                for i in range(p.childCount()):
                    p.child(i).setCheckState(0, state)
        btn_all.clicked.connect(lambda: _set_all(Qt.Checked))
        btn_none.clicked.connect(lambda: _set_all(Qt.Unchecked))
        btn_cancel.clicked.connect(dlg.reject)

        def _accept():
            picked = []
            for sido, parent in sido_items.items():
                for i in range(parent.childCount()):
                    ch = parent.child(i)
                    if ch.checkState(0) == Qt.Checked:
                        _ct = ch.text(0)
                        picked.append(_ct if _ct.startswith(sido) else f"{sido} {_ct}")
            self._selected_regions = picked
            self._save_selected_regions()
            self._refresh_regions_button()
            dlg.accept()
        btn_ok.clicked.connect(_accept)

        btn_row.addWidget(btn_all)
        btn_row.addWidget(btn_none)
        btn_row.addStretch()
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_ok)
        layout.addLayout(btn_row)

        dlg.exec()

    def _open_image_gallery(self):
        """saved_images/{user}/ 아래 포스트별 이미지를 Lazy Load 갤러리로 표시"""
        from PySide6.QtWidgets import (
            QDialog, QVBoxLayout, QHBoxLayout, QScrollArea, QLabel,
            QFrame, QComboBox, QPushButton, QSizePolicy
        )
        from PySide6.QtCore import QTimer, QThread, Signal
        from PySide6.QtGui import QPixmap, QFont

        THUMB = 160
        COLS = 6

        # safe_name → post 매핑 (pixabay_keywords 조회용)
        post_map = {}
        for gp in self._load_generated_posts():
            pl = gp.get("place", {})
            pkey = self._place_key(pl)
            safe = "".join(c for c in (pkey[0] + "_" + pkey[1]) if c.isalnum() or c in "-_")[:80] or "unknown"
            post_map[safe] = gp

        class LazyThumb(QLabel):
            def __init__(self, path):
                super().__init__()
                self.image_path = path
                self.loaded = False
                self.setFixedSize(THUMB, THUMB)
                self.setStyleSheet("background: #e5e7eb; border: 1px solid #d1d5db;")
                self.setAlignment(Qt.AlignCenter)
                self.setText("...")

            def load_image(self):
                if self.loaded:
                    return
                pm = QPixmap(self.image_path)
                if pm.isNull():
                    self.setText("!")
                    return
                self.setPixmap(pm.scaled(THUMB, THUMB, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                self.loaded = True

            def reload(self):
                self.loaded = False
                self.load_image()

        cfg_now = load_config()
        pix_api_keys = [k for k in (cfg_now.get("pixabay_key_list") or []) if k]

        dlg = QDialog(self)
        dlg.setWindowTitle("이미지 갤러리")
        dlg.resize(1200, 800)
        root = QVBoxLayout(dlg)

        top = QHBoxLayout()
        top.addWidget(QLabel("포스트 폴더:"))
        info_label = QLabel("")
        top.addStretch()
        top.addWidget(info_label)
        root.addLayout(top)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        root.addWidget(scroll, 1)
        canvas = QWidget()
        grid_layout = QVBoxLayout(canvas)
        grid_layout.setSpacing(16)
        scroll.setWidget(canvas)

        thumbs = []
        workers = []  # GC 방지

        def load_visible():
            viewport = scroll.viewport()
            vp_top = scroll.verticalScrollBar().value()
            vp_bottom = vp_top + viewport.height()
            preload = 200
            for t in thumbs:
                if t.loaded:
                    continue
                pos = t.mapTo(canvas, t.rect().topLeft())
                y_top = pos.y()
                y_bottom = y_top + t.height()
                if y_bottom + preload < vp_top:
                    continue
                if y_top - preload > vp_bottom:
                    continue
                t.load_image()

        scroll.verticalScrollBar().valueChanged.connect(lambda _: load_visible())

        base = os.path.join(os.path.dirname(__file__), "saved_images", self._account_key())
        if not os.path.isdir(base):
            info_label.setText("저장된 이미지 폴더 없음")
            dlg.exec()
            return

        post_folders = sorted([f for f in os.listdir(base)
                               if os.path.isdir(os.path.join(base, f))])
        total = 0
        for fname in post_folders:
            fpath = os.path.join(base, fname)
            imgs = sorted([f for f in os.listdir(fpath)
                          if f.lower().endswith((".jpg", ".jpeg", ".png")) and not f.startswith("._")])
            if not imgs:
                continue
            section = QFrame()
            section.setStyleSheet("background: #f9fafb; border-radius: 8px; padding: 8px;")
            sec_layout = QVBoxLayout(section)
            title = QLabel(f"{fname}  ({len(imgs)}장)")
            title.setFont(QFont("맑은 고딕", 11, QFont.Bold))
            sec_layout.addWidget(title)

            row_layout = QHBoxLayout()
            row_layout.setSpacing(8)
            sec_layout.addLayout(row_layout)
            col = 0
            for img in imgs:
                img_path = os.path.join(fpath, img)

                cell = QFrame()
                cell_layout = QVBoxLayout(cell)
                cell_layout.setContentsMargins(0, 0, 0, 0)
                cell_layout.setSpacing(2)

                thumb = LazyThumb(img_path)
                thumbs.append(thumb)
                total += 1
                cell_layout.addWidget(thumb)

                btn_change = QPushButton("이미지 변경")
                btn_change.setFixedWidth(THUMB)
                btn_change.setStyleSheet("font-size:11px; padding:2px;")
                btn_change.setAutoDefault(False)
                btn_change.setDefault(False)

                def _on_change(checked=False, _ipath=img_path, _fname=fname, _thumb=thumb, _btn=btn_change):
                    _btn.setEnabled(False)
                    _btn.setText("변경 중...")
                    gp = post_map.get(_fname)
                    place_ = gp.get("place", {}) if gp else {}
                    # search_keyword 마지막 토큰 우선 (강남구 치과 → 치과), 매핑 실패 시 GPT 번역으로 폴백
                    keyword = self._best_biz_term(place_, "") if place_ else ""
                    if not keyword:
                        keyword = "interior"
                    api_key = pix_api_keys[0] if pix_api_keys else ""
                    w = ImageReplaceWorker(_ipath, api_key, keyword, translator=self._translate_ko_to_en)
                    def _on_done(path, ok, msg, __btn=_btn, __thumb=_thumb):
                        if ok:
                            __thumb.reload()
                            __btn.setText("이미지 변경")
                        else:
                            __btn.setText(f"실패: {msg[:10]}")
                        __btn.setEnabled(True)
                    w.done.connect(_on_done)
                    workers.append(w)
                    w.start()

                btn_change.clicked.connect(_on_change)
                cell_layout.addWidget(btn_change)

                row_layout.addWidget(cell)
                col += 1
                if col >= COLS:
                    row_layout.addStretch()
                    col = 0
                    row_layout = QHBoxLayout()
                    row_layout.setSpacing(8)
                    sec_layout.addLayout(row_layout)
            if col > 0:
                row_layout.addStretch()
            grid_layout.addWidget(section)

        grid_layout.addStretch()
        info_label.setText(f"{len(post_folders)}개 포스트, {total}장 이미지")
        QTimer.singleShot(50, load_visible)
        dlg.exec()

    def _cleanup_orphan_image_folders(self):
        """saved_images/{user}/ 아래 폴더 중 현재 generated_posts에 없는 건 삭제 + macOS 메타파일 정리"""
        import shutil as _shutil
        base = os.path.join(os.path.dirname(__file__), "saved_images", self._account_key())
        if not os.path.isdir(base):
            return
        # 현재 유효한 place_key → 폴더명 매핑
        valid_names = set()
        for gp in self._load_generated_posts():
            pkey = self._place_key(gp.get("place", {}))
            safe = "".join(c for c in (pkey[0] + "_" + pkey[1]) if c.isalnum() or c in "-_")[:80] or "unknown"
            valid_names.add(safe)
        removed = 0
        meta_removed = 0
        for name in os.listdir(base):
            full = os.path.join(base, name)
            # macOS AppleDouble 메타파일/폴더는 즉시 삭제
            if name.startswith("._") or name == ".DS_Store":
                try:
                    if os.path.isdir(full):
                        _shutil.rmtree(full)
                    else:
                        os.remove(full)
                    meta_removed += 1
                except Exception:
                    pass
                continue
            if not os.path.isdir(full):
                continue
            if name in valid_names:
                # 유효한 폴더 안의 메타파일도 청소
                try:
                    for sub in os.listdir(full):
                        if sub.startswith("._") or sub == ".DS_Store":
                            try:
                                os.remove(os.path.join(full, sub))
                                meta_removed += 1
                            except Exception:
                                pass
                except Exception:
                    pass
                continue
            try:
                _shutil.rmtree(full)
                removed += 1
            except Exception:
                pass
        if removed:
            self._emit_log(f"이미지 폴더 자동 정리: {removed}개 고아 폴더 삭제")
        if meta_removed:
            self._emit_log(f"macOS 메타파일 정리: {meta_removed}개 삭제")

    def _scope_prompts_to_user(self, user: dict):
        """로그인한 사용자 기준으로 로컬 prompts.json 재설정 — '기본(캐논) + 이 사용자 프롬프트'만 남김.
        이전에 로그인했던 다른 사용자의 프롬프트는 화면/생성에서 사라진다(각 계정 클라우드 저장본은 유지)."""
        try:
            import json as _j
            from app_paths import data_file as _df
            from prompts import get_base_prompt as _gbp
            data = {"기본": _gbp()}
            user_prompts = (user or {}).get("prompts", {}) or {}
            for k, v in user_prompts.items():
                if k != "기본" and isinstance(v, dict):
                    data[k] = v
            with open(_df("prompts.json"), "w", encoding="utf-8") as f:
                _j.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _sync_admin_api_keys_from_firebase(self, user: dict):
        """admin 로그인 시 Firebase의 api_keys를 로컬 config.json의 api_keys_by_user에 덮어쓰기"""
        api_keys = user.get("api_keys") or {}
        if not api_keys:
            return
        username = user.get("username", "admin")
        try:
            import json as _json
            from config import CONFIG_FILE as _CF, load_config as _lc
            raw = {}
            if os.path.exists(_CF):
                with open(_CF, "r", encoding="utf-8") as f:
                    raw = _json.load(f)
            akbu = dict(raw.get("api_keys_by_user") or {})
            akbu[username] = {k: list(v or []) for k, v in api_keys.items()}
            raw["api_keys_by_user"] = akbu
            with open(_CF, "w", encoding="utf-8") as f:
                _json.dump(raw, f, ensure_ascii=False, indent=2)
            self.cfg = _lc()
        except Exception as e:
            self._emit_log(f"admin 키 동기화 실패: {e}")

    def _update_expires_label(self):
        if not hasattr(self, "dash_expires"):
            return
        user = getattr(self, "current_user", {}) or {}
        if user.get("role") == "admin":
            self.dash_expires.setText("[무제한]")
            self.dash_expires.setStyleSheet("font-size: 16px; font-weight: bold; color: #22c55e;")
            return
        import datetime as _dt
        today = _dt.date.today()

        def _fmt(exp):
            """만료일 문자열 → (표시텍스트, 색). 값 없으면 None."""
            exp = (exp or "").strip()
            if not exp:
                return None
            try:
                y, m, d = map(int, exp.split("-"))
                days = (_dt.date(y, m, d) - today).days
            except Exception:
                return ("?", "#94a3b8")
            if days > 3650:
                return ("무제한", "#22c55e")
            if days < 0:
                return ("만료", "#ef4444")
            c = "#22c55e" if days >= 31 else "#f59e0b" if days >= 8 else "#ef4444"
            return (f"{days}일", c)

        # 아이디 1개 = 명의 1개 구조 — 값이 있는 슬롯만 표시 (분할 전 구계정 호환용으로 2·3도 남김)
        slot_keys = [("expires", "api_expires"),
                     ("expires_2", "api_expires_2"),
                     ("expires_3", "api_expires_3")]
        blocks = []
        for i, (ek, ak) in enumerate(slot_keys):
            if i > 0 and not (user.get(ek) or user.get(ak)):
                continue  # 빈 명의2/3은 표시 안 함
            mv = _fmt(user.get(ek))
            if mv is None:
                blocks.append("<span style='color:#cbd5e1'>구독 없음</span>")
                continue
            mtxt, mc = mv
            av = _fmt(user.get(ak))
            if av is None:
                api_part = "<span style='color:#cbd5e1'>API-</span>"
            else:
                api_part = f"<span style='color:{av[1]}'>API {av[0]}</span>"
            label = "구독" if i == 0 else f"명의{i+1}"
            blocks.append(f"{label} <span style='color:{mc}'>{mtxt}</span>·{api_part}")
        html = " · ".join(blocks)
        self.dash_expires.setText(html)
        self.dash_expires.setTextFormat(Qt.RichText)
        self.dash_expires.setStyleSheet("font-size: 11px; font-weight: bold;")

    # ── 구독 결제 ──
    def _open_payment_dialog(self):
        # 결제는 웹페이지로 연결 (인앱 결제창 대신 n-jobs.kr 결제 페이지)
        import webbrowser as _wb
        _wb.open("https://n-jobs.kr/payment.html")
        return
        from PySide6.QtWidgets import QButtonGroup, QRadioButton, QCheckBox
        import datetime as _dt
        from config import get_current_user as _gcu
        from users import load_users as _lu

        username = _gcu()
        _user = _lu().get(username or "", {})

        def _days_left(exp_str):
            if not exp_str:
                return None
            try:
                exp = _dt.date.fromisoformat(exp_str)
                return max(0, (exp - _dt.date.today()).days)
            except Exception:
                return 0

        _d = [
            _days_left(_user.get("expires", "")),
            _days_left(_user.get("expires_2", "")),
            _days_left(_user.get("expires_3", "")),
        ]
        _ad = [
            _days_left(_user.get("api_expires", "")),
            _days_left(_user.get("api_expires_2", "")),
            _days_left(_user.get("api_expires_3", "")),
        ]

        def _status_txt(idx):
            days = _d[idx]
            if days is None:   return "❌ 사용불가"
            elif days > 3650:  return "♾️ 무제한"
            elif days > 0:     return f"✅ {days}일 남음"
            else:              return "❌ 만료"

        def _api_status_txt(idx):
            days = _ad[idx]
            if days is None:   return "❌ 미구독"
            elif days > 3650:  return "♾️ 무제한"
            elif days > 0:     return f"✅ {days}일 남음"
            else:              return "❌ 미구독"

        BASE_PRICES = [9900, 14900, 19900]
        API_PRICES  = [5500, 9900, 9900]
        FIELDS      = ["expires",    "expires_2",    "expires_3"]
        API_FIELDS  = ["api_expires","api_expires_2","api_expires_3"]

        dlg = QDialog(self)
        dlg.setWindowTitle("구독 연장")
        dlg.setMinimumWidth(520)
        layout = QVBoxLayout(dlg)
        layout.setSpacing(12)
        layout.setContentsMargins(22, 18, 22, 18)

        # ── 명의 수 선택 (가격) ──
        layout.addWidget(QLabel("<b>명의 수 선택</b>  (1명의 = 네이버 아이디 3개)"))
        myeong_bg = QButtonGroup(dlg)
        myeong_row = QHBoxLayout()
        myeong_row.setSpacing(8)
        _rm = []
        for _i, (_p, _s) in enumerate(zip(BASE_PRICES, ["1명의", "2명의", "3명의"])):
            _r = QRadioButton(f"{_s}  {_p:,}원")
            _r.setStyleSheet("padding: 4px;")
            myeong_bg.addButton(_r)
            myeong_row.addWidget(_r)
            _rm.append(_r)
        _rm[0].setChecked(True)
        layout.addLayout(myeong_row)

        # ── 적용할 명의(슬롯) 선택 ──
        layout.addWidget(QLabel("<b>어느 명의에 적용할지 선택</b>  (위에서 고른 개수만큼 체크)"))
        slot_row = QHBoxLayout()
        slot_row.setSpacing(8)
        _slot_chks = []
        for _i in range(3):
            _c = QCheckBox(f"명의{_i+1}  {_status_txt(_i)}")
            _c.setStyleSheet("padding: 4px;")
            slot_row.addWidget(_c)
            _slot_chks.append(_c)
        _slot_chks[0].setChecked(True)
        layout.addLayout(slot_row)

        def _selected_qty():
            return next((i for i, r in enumerate(_rm) if r.isChecked()), 0) + 1

        def _checked_slots():
            return [i for i, c in enumerate(_slot_chks) if c.isChecked()]

        # ── API 구독 추가 ──
        chk_api = QCheckBox()
        chk_api.setStyleSheet("padding: 2px;")
        layout.addWidget(chk_api)

        def _update_api_chk():
            _ap = API_PRICES[_selected_qty() - 1]
            chk_api.setText(f"🔑 글쓰기 API 구독  +{_ap:,}원/월  (선택한 명의에 적용)")

        _update_api_chk()

        # ── 기간 선택 ──
        layout.addWidget(QLabel("<b>기간 선택</b>"))
        period_bg = QButtonGroup(dlg)
        period_row = QHBoxLayout()
        period_row.setSpacing(4)
        rp1   = QRadioButton("1개월")
        rp3   = QRadioButton("3개월 -5%")
        rp6   = QRadioButton("6개월 -10%")
        rpInf = QRadioButton("무제한 770,000원/명의")
        rp1.setChecked(True)
        for _r in (rp1, rp3, rp6, rpInf):
            period_bg.addButton(_r)
            period_row.addWidget(_r)
        layout.addLayout(period_row)

        # ── 금액 표시 ──
        price_lbl = QLabel()
        price_lbl.setStyleSheet("font-size:15px; font-weight:bold; color:#6366f1; padding:6px 0;")
        price_lbl.setAlignment(Qt.AlignCenter)
        price_lbl.setWordWrap(True)
        layout.addWidget(price_lbl)

        def _enforce_slots():
            # 체크 개수가 명의 수를 넘으면 앞에서부터 명의 수만큼만 유지
            qty = _selected_qty()
            checked = _checked_slots()
            if len(checked) > qty:
                for i in checked[qty:]:
                    _slot_chks[i].blockSignals(True)
                    _slot_chks[i].setChecked(False)
                    _slot_chks[i].blockSignals(False)

        def _calc():
            _enforce_slots()
            qty = _selected_qty()
            _base = BASE_PRICES[qty - 1]
            _api_add = API_PRICES[qty - 1]
            chk = _checked_slots()
            _names = ", ".join(f"명의{i+1}" for i in chk) or "(미선택)"
            _warn = f"   ⚠ {qty}개 선택 필요 (현재 {len(chk)}개)" if len(chk) != qty else ""
            if rpInf.isChecked():
                _inf = 770000 * qty
                _ex = f"\n+ API {_api_add:,}원/월 별도" if chk_api.isChecked() else ""
                price_lbl.setText(f"무제한 770,000원 × {qty}명의 = {_inf:,}원\n적용: {_names}{_warn}{_ex}")
            else:
                if rp3.isChecked():   _m, _dc = 3, 0.05
                elif rp6.isChecked(): _m, _dc = 6, 0.10
                else:                 _m, _dc = 1, 0.0
                _sub = int(_base * _m * (1 - _dc))
                if chk_api.isChecked():
                    _sub += int(_api_add * _m * (1 - _dc))
                _d2 = f" ({int(_dc*100)}%할인)" if _dc else ""
                price_lbl.setText(f"{_base:,}원/월{_d2} × {_m}개월 = 총 {_sub:,}원\n적용: {_names}{_warn}")

        _calc()
        for _r in (*_rm, *_slot_chks, rp1, rp3, rp6, rpInf):
            _r.toggled.connect(lambda _: (_update_api_chk(), _calc()))
        chk_api.stateChanged.connect(lambda _: _calc())

        # ── 계좌이체 결제 안내 ──
        account_box = QLabel(
            "🏦  결제 안내  (현재 계좌이체)\n"
            "───────────────────────────\n"
            "①  위에 표시된 합계 금액을 확인합니다.\n"
            "②  아래 계좌로 입금합니다.\n"
            "        하나은행   806-910287-39407\n"
            "        예금주 : 김태현\n"
            "        ※ 입금자명은 '본인 이름'으로 입력해 주세요.\n"
            "③  아래 '입금 완료 알림' 버튼을 누릅니다.\n"
            "④  관리자 확인 후 자동으로 연장됩니다.\n"
            "───────────────────────────\n"
            "💳  카드 결제는 지원 준비중입니다. (곧 지원 예정)"
        )
        account_box.setStyleSheet(
            "background:#f8fafc; border:1px solid #e2e8f0; border-radius:8px; "
            "padding:14px; font-size:12px;"
        )
        account_box.setWordWrap(True)
        account_box.setAlignment(Qt.AlignLeft)
        layout.addWidget(account_box)

        # ── 결제 문의 → 카카오톡 오픈채팅 (QR + 바로입장 팝업) ──
        def _open_pay_inquiry():
            import webbrowser as _wb
            from PySide6.QtGui import QPixmap as _QPixmap
            _PAY_URL = "https://open.kakao.com/o/sqpEYNuf"
            try:
                from app_paths import get_bundle_dir as _gbd
                _qp = os.path.join(_gbd(), "kakao_qr_pay.png")
            except Exception:
                _qp = os.path.join(os.path.dirname(__file__), "kakao_qr_pay.png")
            d = QDialog(dlg)
            d.setWindowTitle("결제 문의 (카카오톡 오픈채팅)")
            _v = QVBoxLayout(d)
            _t = QLabel("💳 결제 문의 오픈채팅")
            _t.setAlignment(Qt.AlignCenter)
            _t.setStyleSheet("font-size:15px; font-weight:bold;")
            _v.addWidget(_t)
            if os.path.exists(_qp):
                _big = QLabel()
                _big.setPixmap(_QPixmap(_qp).scaled(280, 280, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                _big.setAlignment(Qt.AlignCenter)
                _v.addWidget(_big)
            _cap = QLabel("휴대폰 카메라로 QR을 찍거나 아래 버튼으로 입장하세요.\n\n"
                          "※ 이 채팅방은 결제에 대한 상담만 가능합니다.")
            _cap.setAlignment(Qt.AlignCenter)
            _cap.setStyleSheet("color:#475569;")
            _cap.setWordWrap(True)
            _v.addWidget(_cap)
            _enter = QPushButton("💬 바로입장")
            _enter.setCursor(Qt.PointingHandCursor)
            _enter.setStyleSheet("background:#FEE500; color:#3C1E1E; border:none; border-radius:8px; padding:10px 18px; font-weight:bold;")
            _enter.clicked.connect(lambda: _wb.open(_PAY_URL))
            _v.addWidget(_enter)
            d.exec()

        btn_inquiry = QPushButton("💬 결제 문의 (카카오톡 오픈채팅)")
        btn_inquiry.setCursor(Qt.PointingHandCursor)
        btn_inquiry.setStyleSheet(
            "background:#FEE500; color:#3C1E1E; border:none; border-radius:8px; "
            "padding:9px; font-size:13px; font-weight:bold;"
        )
        btn_inquiry.clicked.connect(_open_pay_inquiry)
        layout.addWidget(btn_inquiry)

        btn_pay = QPushButton("✅ 입금 완료 알림")
        btn_pay.setStyleSheet("background:#6366f1; color:white; border:none; border-radius:8px; padding:10px; font-size:14px; font-weight:bold;")
        btn_pay.setCursor(Qt.PointingHandCursor)

        def _send_notify():
            qty = _selected_qty()
            chk = _checked_slots()
            if len(chk) != qty:
                QMessageBox.warning(dlg, "선택 확인",
                    f"{qty}명의를 구매하셨으니 적용할 명의도 정확히 {qty}개 선택해주세요.\n(현재 {len(chk)}개 선택)")
                return
            _slot_codes = ",".join(str(i + 1) for i in chk)      # 예: "1,3"
            _slot_label = ", ".join(f"명의{i+1}" for i in chk)
            _plan = f"{qty}명의 ({_slot_label})"
            _base = BASE_PRICES[qty - 1]
            _api_add = API_PRICES[qty - 1]
            _with_api = chk_api.isChecked()
            _unlimited = rpInf.isChecked()
            if _unlimited:
                _months = 1
                _total = 770000 * qty + (_api_add if _with_api else 0)
            else:
                if rp3.isChecked():   _months, _dc = 3, 0.05
                elif rp6.isChecked(): _months, _dc = 6, 0.10
                else:                 _months, _dc = 1, 0.0
                _total = int(_base * _months * (1 - _dc))
                if _with_api: _total += int(_api_add * _months * (1 - _dc))
            try:
                import urllib.request as _ur, json as _jmod
                _payload = _jmod.dumps({
                    "username": username,
                    "plan": _plan,
                    "months": _months,
                    "amount": _total,
                    "field": _slot_codes,
                    "api_field": _slot_codes if _with_api else "",
                    "unlimited": _unlimited,
                    "api": _with_api,
                }).encode()
                _req = _ur.Request(
                    "https://automaster-kr-kingte0560-3092s-projects.vercel.app/api/notify",
                    data=_payload,
                    headers={"Content-Type": "application/json"},
                    method="POST"
                )
                _ur.urlopen(_req, timeout=10)
                QMessageBox.information(dlg, "완료", "입금 완료 알림을 전송했습니다.\n관리자 확인 후 자동으로 연장됩니다.")
                dlg.accept()
            except Exception as _e:
                QMessageBox.warning(dlg, "오류", f"알림 전송 실패: {_e}\n관리자에게 직접 문의해 주세요.")

        btn_pay.clicked.connect(_send_notify)
        layout.addWidget(btn_pay)

        dlg.exec()

    # ── 설정 ──
    def _open_settings(self):
        from settings_dialog import SettingsDialog
        cu = getattr(self, "current_user", {}) or {}
        is_admin = cu.get("role") == "admin"
        username = cu.get("username", "admin")
        dlg = SettingsDialog(self, is_admin=is_admin, app_user=username)
        dlg.exec()
        self.cfg = load_config()
        self._refresh_account_combo()

    def _logout(self):
        self._clear_own_session()
        self.hide()
        login = LoginDialog()
        if login.exec() == QDialog.Accepted:
            self.current_user = login.user
            # 신규 로그인 유저로 config 컨텍스트 전환 — 안 하면 이전(예: admin) 유저로 load_config()가 호출됨
            import config as _cfg
            _cfg.set_current_user(login.user.get("username", "admin"))
            self.show()
            self.apply_user_session(login.user)
        else:
            QApplication.quit()

    # ── 봇 개수 설정 (1/2/3) + RAM 자동 캡 ──
    @staticmethod
    def _get_system_ram_gb() -> float:
        """시스템 총 RAM(GB) 반환. psutil 없으면 -1.0."""
        try:
            import psutil
            return psutil.virtual_memory().total / (1024 ** 3)
        except Exception:
            return -1.0

    @staticmethod
    def _ram_bot_cap(ram_gb: float) -> int:
        """RAM 기반 봇 개수 상한: <4GB→1, 그 외→2 (최대 2봇 정책).
        throttle 회피를 위해 최대 봇 수를 2로 고정한다."""
        if ram_gb < 0:  # psutil 실패 → 보수적으로 2
            return 2
        if ram_gb < 4.0:
            return 1
        return 2

    def _save_bot_count(self):
        try:
            n = int(self.bot_count_group.checkedId() or 2)
            if n not in (1, 2):
                n = 2
            cfg = load_config()
            cfg["bot_count"] = n
            save_config(cfg)
            self.cfg["bot_count"] = n
            self._emit_log(f"크롤링 봇 개수 변경: {n}개")
        except Exception:
            pass

    def _get_bot_count(self) -> int:
        """사용자 설정 봇 개수, 단 RAM 기반 자동 상한 적용 (저사양 PC 보호)."""
        try:
            user_n = int(self.bot_count_group.checkedId() or 2)
            if user_n not in (1, 2):
                user_n = 2
        except Exception:
            user_n = 2
        ram = self._get_system_ram_gb()
        cap = self._ram_bot_cap(ram)
        if user_n > cap:
            self._emit_log(f"⚠️ RAM {ram:.1f}GB 감지 — 봇 {user_n}개 → {cap}개로 자동 제한")
            return cap
        return user_n

    # ── Pixabay 검색어 우선순위 도우미 ──
    def _best_biz_term(self, place: dict, fallback_keyword: str = "") -> str:
        """Pixabay 검색에 가장 적합한 업종 문자열을 골라 반환.
        우선순위: search_keyword 마지막 토큰 > category 마지막 토큰 > fallback_keyword > category 풀스트링.
        예: search_keyword='강남구 치과' → '치과', category='의원 > 치과의원' → '치과의원'."""
        # 사용자가 포스트 생성 시 직접 입력한 '사진 검색 키워드'가 있으면 그걸 최우선 사용
        _ov = (getattr(self, "_override_image_keyword", "") or "").strip()
        if _ov:
            return _ov
        from image_handler import _tokenize_biz, BIZ_TO_EN, _strip_region
        # category_2 (더 구체적) → search_keyword → category_1 → fallback 순
        # 매핑 키(ko)를 직접 반환 → 지역명 없는 '순수 개념'이 보장됨
        # (예: search_keyword='부산개인회생' → '개인회생', '강남구 치과' → '치과')
        for src in (place.get("category_2"), place.get("search_keyword"), place.get("category"), fallback_keyword):
            toks = _tokenize_biz(src or "")
            for tok in reversed(toks):
                tl = tok.lower()
                for ko in sorted(BIZ_TO_EN.keys(), key=len, reverse=True):
                    if ko in tl:
                        return ko
        # 매핑 실패 시: 지역명을 발라낸 뒤 마지막 토큰(가장 구체적) 사용
        for src in (place.get("category_2"), place.get("search_keyword"), place.get("category"), fallback_keyword):
            toks = _tokenize_biz(_strip_region(src or ""))
            if toks:
                return toks[-1]
        return (place.get("category_2") or place.get("category") or fallback_keyword or "").strip()

    # ── 한영 번역 캐시 (Pixabay 검색어용) ──
    def _get_translation_cache_file(self):
        return os.path.join(os.path.dirname(__file__), "translation_cache.json")

    def _load_translation_cache(self) -> dict:
        fp = self._get_translation_cache_file()
        if os.path.exists(fp):
            try:
                with open(fp, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return {}

    def _save_translation_cache(self, cache: dict):
        fp = self._get_translation_cache_file()
        try:
            with open(fp, "w", encoding="utf-8") as f:
                json.dump(cache, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _translate_ko_to_en(self, text: str) -> str:
        """한글 검색어를 영어로 번역 (AI 사용, 결과 캐싱)"""
        text = (text or "").strip()
        if not text:
            return ""
        # 이미 영어면 그대로
        if not any(0xAC00 <= ord(c) <= 0xD7A3 for c in text):
            return text
        cache = self._load_translation_cache()
        if text in cache:
            return cache[text]
        # AI로 번역
        try:
            cfg = self.cfg
            api_keys = [k for k in cfg.get("gpt_key_list", []) if k]
            if not api_keys:
                return text
            prompt = (
                f"Translate this Korean search term to a concise English keyword (max 3 words) suitable for Pixabay image search. "
                f"Output ONLY the English term, no explanation, no quotes.\n\n"
                f"Korean: {text}\nEnglish:"
            )
            from openai import OpenAI
            client = OpenAI(api_key=api_keys[0])
            resp = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=20,
            )
            en = (resp.choices[0].message.content or "").strip()
            # 정제: 따옴표/마침표 제거, 한 줄로
            en = en.split("\n")[0].strip().strip('"\'`.')
            if en:
                cache[text] = en
                self._save_translation_cache(cache)
                return en
        except Exception as _e:
            self._emit_log(f"번역 실패 ('{text}'): {_e}")
        return text

    # ── 업종별 프롬프트 자동 생성 (크롤 완료 후 호출) ──
    def _auto_generate_prompt_for_biz(self, biz_type: str):
        """키워드의 업종이 prompts.json에 없으면 GPT로 블로그/제목 프롬프트 생성 후 저장.
        백그라운드 스레드에서 실행, 이미 있으면 skip.
        ※ 비활성화됨 — 모든 업종을 단일 '기본' 프롬프트로 처리(기본이 업종/톤 동적 분기)."""
        return  # 업종별 자동생성 OFF — '기본' 프롬프트 하나로 통일
        biz_type = (biz_type or "").strip()
        if not biz_type or biz_type == "기본":
            return
        from app_paths import ensure_from_bundle as _efb
        prompts_path = _efb("prompts.json")
        try:
            with open(prompts_path, "r", encoding="utf-8") as f:
                prompts = json.load(f)
        except Exception:
            prompts = {}
        # 이미 있는 업종이면 skip (정확 매칭 + 부분 매칭)
        if biz_type in prompts:
            return
        for key in prompts:
            if key == "기본":
                continue
            if key in biz_type or biz_type in key:
                return  # 유사 업종 이미 있음
        cfg = load_config()
        api_keys = [k for k in cfg.get("gpt_key_list", []) if k]
        if not api_keys:
            return
        base = prompts.get("기본", {"blog": "", "title": ""})
        base_blog = base.get("blog", "")
        base_title = base.get("title", "")
        meta_prompt = (
            f"아래는 블로그 자동 포스팅에 사용하는 '기본' 시스템 프롬프트다.\n\n"
            f"[기본 블로그 본문 프롬프트]\n{base_blog}\n\n"
            f"[기본 제목 프롬프트]\n{base_title}\n\n"
            f"---\n\n"
            f"위 기본 프롬프트를 '{biz_type}' 업종 전용으로 변환해라.\n\n"
            f"[절대 규칙 — 반드시 지켜라]\n"
            f"- 기본 프롬프트의 모든 섹션(역할/동적 페르소나, 입력 데이터, 금지 단어, 수익형 3단 구조(서론/본론/결론), 네이버 최적화 태그 규칙)을 그대로 유지해라.\n"
            f"- 글자 수 조건(1,600~2,000자), 문장 수(40문장 이상), 서론/본론/결론 구조, [이미지] 마커 3~5개 규칙, 해시태그 7~10개 규칙, 주소·지역 필수 포함 — 모두 그대로 유지해라.\n"
            f"- 플레이스홀더 유지 필수: {{업체명}}, {{업종}}, {{키워드}}, {{시}}, {{구}}, {{동}}, {{주소}}, {{근처역}}, {{근처역상세}}\n"
            f"- 오직 '역할' 섹션의 업종 설명과 본론 항목만 '{biz_type}' 업종에 맞게 수정해라.\n"
            f"- 실제 블로그 글을 쓰지 말 것. 지시문(시스템 프롬프트)만 출력해라.\n"
            f"- 이모티콘 금지, 광고성 표현 금지.\n\n"
            f"[출력 형식 — 반드시 아래 구분자 그대로 단독 줄에 출력]\n"
            f"===BLOG===\n"
            f"('{biz_type}' 업종용 블로그 본문 시스템 프롬프트)\n"
            f"===TITLE===\n"
            f"('{biz_type}' 업종용 제목 시스템 프롬프트)\n"
        )

        def _worker():
            try:
                self._emit_log(f"업종 프롬프트 자동 생성 시작: '{biz_type}'")
                from openai import OpenAI
                client = OpenAI(api_key=api_keys[0])
                resp = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[{"role": "user", "content": meta_prompt}],
                )
                text = resp.choices[0].message.content or ""
                import re as _re
                m_blog = _re.search(r"===\s*BLOG\s*===\s*(.*?)\s*===\s*TITLE\s*===", text, _re.S)
                m_title = _re.search(r"===\s*TITLE\s*===\s*(.*)", text, _re.S)
                blog_p = (m_blog.group(1).strip() if m_blog else "")
                title_p = (m_title.group(1).strip() if m_title else "")
                if not blog_p or not title_p:
                    self._emit_log(f"  [경고] '{biz_type}' 프롬프트 파싱 실패 — AI 응답: {text[:200]!r}")
                    return
                try:
                    with open(prompts_path, "r", encoding="utf-8") as _f:
                        cur = json.load(_f)
                except Exception:
                    cur = {}
                if biz_type in cur:  # 레이스 조건 방지
                    return
                cur[biz_type] = {
                    "blog": blog_p,
                    "title": title_p,
                    "title_prefix": "dong",
                    "pixabay_list": ["", "", ""],
                }
                with open(prompts_path, "w", encoding="utf-8") as _f:
                    json.dump(cur, _f, ensure_ascii=False, indent=2)
                # Firebase에 유저별 프롬프트 누적 저장
                try:
                    from users import update_user as _upd_fb, load_users as _lu_fb
                    from config import get_current_user as _gcu_fb
                    _uname = _gcu_fb()
                    if _uname:
                        _fb_cur = (_lu_fb().get(_uname) or {}).get("prompts") or {}
                        _fb_cur[biz_type] = cur[biz_type]
                        _upd_fb(_uname, prompts=_fb_cur)
                except Exception:
                    pass
                self._emit_log(f"  '{biz_type}' 프롬프트 자동 생성 완료 (저장 완료)")
            except Exception as _e:
                self._emit_log(f"  '{biz_type}' 프롬프트 자동 생성 실패: {_e}")

        threading.Thread(target=_worker, daemon=True).start()

    # ── 프롬프트 편집 ──
    def _open_prompt_editor(self):
        from app_paths import ensure_from_bundle as _efb
        prompts_path = _efb("prompts.json")
        is_admin = (getattr(self, "current_user", {}) or {}).get("role") == "admin"

        prompts = {}
        if is_admin:
            # 관리자: 로컬 전체 로드
            try:
                with open(prompts_path, "r", encoding="utf-8") as f:
                    prompts = json.load(f)
            except Exception:
                prompts = {}
        else:
            # 일반 유저: 기본(공유) + Firebase 개인 프롬프트만
            try:
                with open(prompts_path, "r", encoding="utf-8") as f:
                    _local = json.load(f)
                if "기본" in _local:
                    prompts["기본"] = _local["기본"]
            except Exception:
                pass
            try:
                from users import load_users as _lu_p
                from config import get_current_user as _gcu
                _fb_prompts = _lu_p().get(_gcu(), {}).get("prompts", {}) or {}
                for _k, _v in _fb_prompts.items():
                    prompts[_k] = _v
            except Exception:
                pass
        if not prompts:
            prompts = {"기본": {"blog": "", "title": ""}}
        # '기본'은 항상 번들 캐논으로 고정 (수정·삭제 불가, 실수로 비운 경우 자동 복구)
        try:
            from prompts import get_base_prompt as _gbp
            prompts["기본"] = _gbp()
        except Exception:
            pass

        dlg = QDialog(self)
        dlg.setWindowTitle("프롬프트 편집")
        dlg.resize(900, 700)
        layout = QVBoxLayout(dlg)

        top = QHBoxLayout()
        top.addWidget(QLabel("업종 선택:"))
        type_combo = QComboBox()
        # '기본'은 목록에서 숨김 — 파생/직접추가한 업종만 보이게 (기본은 내부 폴백으로만 사용)
        type_combo.addItems([k for k in prompts.keys() if k != "기본"])
        top.addWidget(type_combo, 1)

        btn_add = QPushButton("+ 업종 추가")
        btn_add_user = QPushButton("+ 사용자 프롬프트")
        btn_add_user.setStyleSheet("background: #6366f1; color: white; border-radius: 6px; padding: 4px 10px;")
        btn_auto = QPushButton("AI 자동 생성")
        btn_auto.setStyleSheet("background: #22c55e; color: white; border-radius: 6px; padding: 4px 10px;")
        btn_del = QPushButton("삭제")
        btn_del.setStyleSheet("background: #ef4444; color: white; border-radius: 6px; padding: 4px 10px;")
        top.addWidget(btn_add)
        top.addWidget(btn_add_user)
        top.addWidget(btn_auto)
        top.addWidget(btn_del)
        layout.addLayout(top)

        _edit_style = "QTextEdit { border: 1px solid #cbd5e1; border-radius: 6px; padding: 6px; background: #ffffff; }"
        _lbl_blog = QLabel("블로그 본문 프롬프트:")
        layout.addWidget(_lbl_blog)
        blog_edit = QTextEdit()
        blog_edit.setAcceptRichText(False)
        blog_edit.setStyleSheet(_edit_style)
        layout.addWidget(blog_edit, 3)

        _lbl_title = QLabel("제목 프롬프트:")
        layout.addWidget(_lbl_title)
        title_edit = QTextEdit()
        title_edit.setAcceptRichText(False)
        title_edit.setStyleSheet(_edit_style)
        layout.addWidget(title_edit, 1)

        # 파생(기본 기반) 프롬프트 선택 시 일반 사용자에게는 내용 대신 보여줄 안내 (load_current에서 토글)
        _info = QLabel("이 업종 프롬프트는 기본을 기반으로 자동 생성됩니다.\n'AI 자동 생성'을 누르면 만들어져 저장되며, 내용은 비공개입니다.")
        _info.setStyleSheet("color:#64748b; padding:24px; font-size:13px;")
        _info.setWordWrap(True)
        _info.setVisible(False)
        layout.addWidget(_info, 1)

        def _set_content_visible(show):
            _lbl_blog.setVisible(show); blog_edit.setVisible(show)
            _lbl_title.setVisible(show); title_edit.setVisible(show)
            _info.setVisible(not show)

        def _kind_of(key):
            # '_kind': 'user'(직접작성=공개) / 'derived'(기본기반=비공개). 없으면 파생 취급.
            return (prompts.get(key, {}) or {}).get("_kind", "derived")

        # 제목 키워드 형식은 포스트 생성 다이얼로그에서 결정 (여기서 제거)

        current = {"key": type_combo.currentText()}

        def load_current():
            key = type_combo.currentText()
            data = prompts.get(key, {"blog": "", "title": "", "pixabay_list": []})
            kind = _kind_of(key)
            # 내용은 항상 박스에 로드(파생이라 숨겨도 저장 무결성 유지)
            blog_edit.blockSignals(True); title_edit.blockSignals(True)
            blog_edit.setPlainText(data.get("blog", ""))
            title_edit.setPlainText(data.get("title", ""))
            blog_edit.blockSignals(False); title_edit.blockSignals(False)
            current["key"] = key
            # 표시: 관리자는 항상 / 일반 사용자는 '직접작성(user)'만 보임, '파생'은 숨김(비공개)
            show = bool(is_admin or kind == "user")
            _set_content_visible(show)
            # 잠금: 기본은 항상 / 일반 사용자의 파생도 잠금(어차피 숨김)
            locked = (key == "기본") or (not is_admin and kind != "user")
            blog_edit.setReadOnly(locked)
            title_edit.setReadOnly(locked)
            btn_del.setEnabled(key != "기본")
            btn_auto.setEnabled(key != "기본")   # AI 자동생성은 파생/직접 모두 가능

        def save_current_to_memory():
            if current["key"] == "기본":
                return  # 기본은 캐논 고정 — 편집 내용 저장 안 함
            if current["key"] in prompts:
                _k = (prompts[current["key"]] or {}).get("_kind", "derived")
                prompts[current["key"]] = {
                    "blog": blog_edit.toPlainText(),
                    "title": title_edit.toPlainText(),
                    "pixabay_list": [],
                    "_kind": _k,
                }

        def on_type_changed():
            save_current_to_memory()
            load_current()

        type_combo.currentIndexChanged.connect(lambda _i: on_type_changed())

        def add_type():
            from PySide6.QtWidgets import QInputDialog
            name, ok = QInputDialog.getText(dlg, "업종 추가", "업종명:")
            name = (name or "").strip()
            if ok and name and name not in prompts:
                save_current_to_memory()
                # 기본 기반 '파생' 업종 — 내용 비공개(일반 사용자). AI 자동생성으로 완성
                _base = prompts.get("기본", {}) or {}
                prompts[name] = {
                    "blog": _base.get("blog", ""),
                    "title": _base.get("title", ""),
                    "pixabay_list": ["", "", ""],
                    "_kind": "derived",
                }
                type_combo.addItem(name)
                type_combo.setCurrentText(name)

        def add_user_prompt():
            from PySide6.QtWidgets import QInputDialog
            name, ok = QInputDialog.getText(dlg, "사용자 프롬프트 추가", "이름:")
            name = (name or "").strip()
            if ok and name and name not in prompts:
                save_current_to_memory()
                # 사용자가 '직접 작성'하는 프롬프트 — 내용 보이고 편집 가능(공개)
                prompts[name] = {"blog": "", "title": "", "pixabay_list": ["", "", ""], "_kind": "user"}
                type_combo.addItem(name)
                type_combo.setCurrentText(name)

        def del_type():
            key = type_combo.currentText()
            if not key:
                return
            if key == "기본":
                QMessageBox.warning(dlg, "안내", "기본 프롬프트는 삭제할 수 없습니다.")
                return
            if len(prompts) <= 1:
                QMessageBox.warning(dlg, "경고", "최소 1개는 남겨야 합니다.")
                return
            if QMessageBox.question(dlg, "확인", f"'{key}' 업종을 삭제할까요?") != QMessageBox.Yes:
                return
            prompts.pop(key, None)
            type_combo.removeItem(type_combo.currentIndex())
            load_current()

        # AI 응답 수신 시그널 (워커 스레드에서 GUI로 전달)
        class _AISignal(QObject):
            done = Signal(str, str)  # text, error
        ai_sig = _AISignal()

        def _on_ai_done(text, error):
            btn_auto.setEnabled(True)
            btn_auto.setText("AI 자동 생성")
            if error:
                QMessageBox.critical(dlg, "오류", f"AI 호출 실패: {error}")
                return
            import re
            m_blog = re.search(r"===\s*BLOG\s*===\s*(.*?)\s*===\s*TITLE\s*===", text, re.S)
            m_title = re.search(r"===\s*TITLE\s*===\s*(.*)", text, re.S)
            blog_p = (m_blog.group(1).strip() if m_blog else "")
            title_p = (m_title.group(1).strip() if m_title else "")
            if not blog_p and not title_p:
                QMessageBox.warning(dlg, "경고", f"응답 파싱 실패. 전체 응답:\n{text[:500]}")
                return
            if blog_p:
                blog_edit.setPlainText(blog_p)
            if title_p:
                title_edit.setPlainText(title_p)
            QMessageBox.information(dlg, "완료", "프롬프트 자동 생성 완료. 확인 후 '저장'을 눌러주세요.")

        ai_sig.done.connect(_on_ai_done)

        def auto_generate():
            key = type_combo.currentText()
            if not key:
                return
            cfg2 = load_config()
            provider = "GPT"  # GPT 전용
            api_keys = [k for k in cfg2.get("gpt_key_list", []) if k]
            if not api_keys:
                QMessageBox.critical(dlg, "오류", "API 키를 설정해주세요.")
                return
            base = prompts.get("기본", {"blog": "", "title": ""})
            ref_example = """# 역할
너는 부모님을 직접 모시지 못한다는 무거운 마음을 안고 수개월간 요양원을 찾아 헤맸던 **'현실적인 효자/효녀 블로거'**다. 단순한 정보 나열이 아니라, 같은 처지에 놓인 자녀들이 밤잠 설치며 고민하는 지점을 정확히 짚어주고, 그들의 죄책감을 덜어주면서도 **'전문적인 사람'**으로서의 선택임을 강조하는 따뜻한 조언자다
글은 광고 느낌이 아니라, 실제로 부모님을 맡길 요양원을 고르면서 고민했던 사람이 지인에게 조용히 털어놓는 이야기처럼 진심 있게 작성한다.
기본은 존댓말을 사용하고, **1인칭(저, 우리 가족)** 을 자연스럽게 사용해도 된다.
※ 이 프롬프트에서 가장 중요한 규칙은 **'분량'과 '문장 수'**이다. 다른 어떤 규칙보다 **글자 수·문장 수 조건을 우선해서** 지킨다. 글이 짧다고 느껴지면, 이미 쓴 내용을 더 구체적인 장면·감정·예시로 풀어 써서라도 분량을 반드시 채운다.

# 작업 지시
아래에 제공된 '입력 데이터'를 활용하여 **부모님 요양원 선택 과정과 실제 이용 후기**를 블로그 본문 형식으로 작성한다.
* 전체 글자 수는 한글 기준 **공백 포함 1500자 안팎**으로 맞춘다.
* **1450자 이상 1650자 이하** 범위에서 작성하고, **1400자 미만이 되지 않도록** 충분히 내용을 확장한다.
글의 전체적인 흐름은 다음 구조를 따른다.
1. **서론** — 방문 계기/동기, 처음의 감정(죄책감·불안 등)
2. **본론** — 검색·상담·방문 과정 / 비교 기준 / 최종 선택 요양원의 위치·접근성 / 시설·프로그램·돌봄 환경 / 부모님의 변화, 가족의 심리 변화
3. **결론** — 만족도·느낀 점 / "나도 이렇게 선택했다" 식의 부드러운 상담/방문 권유

# 입력 데이터
* 요양원명(업체명): {업체명}
* 주소: {주소}
* 근처역/교통: {근처역}
* 카테고리: {카테고리}
* 앞 키워드: {앞키워드}
* 태그: {태그}
(앞 키워드와 태그는 글의 주제와 분위기를 파악하기 위한 참고용이다.)

# 준수 사항
## 1. 형식 및 문장 구성
* **문장마다 줄바꿈**한다. 한 문장은 한 줄, 문장 끝마다 줄바꿈.
* 서론 → 본론 → 결론 흐름에 맞춰 배치한다.
* 전체 문장 수는 **최소 40문장 이상**.
* 각 문장은 너무 짧지 않게, **대략 25자 이상**이 되도록 상황·감정·설명을 충분히 풀어 쓴다.

## 2. 분량
* 공백 포함 1500자 안팎(1450~1650자). 1400자 미만 금지.
* 부족하면 구체적 예시·상황·대화·표정·몸짓 묘사로 늘린다.

## 3. 내용 구성 (각 파트 최소 문장 수)
1) 서론 — 처음 상황과 감정 (최소 8문장)
2) 검색·상담·방문 과정 (최소 10문장, 실제 상담 질문 3가지 이상 풀어쓰기)
3) 여러 곳 비교 + 최종 선택 기준 (최소 7문장)
4) 위치·접근성·주변 환경 (최소 5문장)
5) 시설·프로그램·돌봄 세부 묘사 (최소 8문장)
6) 부모님 변화·가족 심리 변화 (최소 6문장, 시간 흐름 단계적 서술)
7) 장점/아쉬운 점 + 결론 (최소 8문장)

## 4. 어조
* 일상 대화체 + 진심 담은 후기체.
* "솔직히 처음에는…", "막상 다녀와 보니까…" 같은 표현 적절히.
* 자극·과장 없이 담담하지만 진심 있는 톤.

## 5. 금지 사항
* 이모티콘 금지.
* '카테고리', '앞 키워드', '태그'라는 단어 본문 언급 금지.
* **광고·홍보·이벤트·할인·"무조건 여기로 오세요" 식 문구 절대 금지.**
* 가격·비용·금액 구체 언급 금지(간접 언급만 가능).

## 6. 정보 출처 표현
* '크롤링', '검색 엔진', '데이터 수집' 같은 단어 본문 사용 금지.
* 본인이 직접 상담·방문해서 느낀 후기처럼 자연스럽게 표현.

## 이미지 배치
* 본문 흐름상 사진이 들어가면 좋을 위치에 '[이미지]' 마커를 해당 줄 단독으로 3~5개 삽입.
"""
            ref_title = """# 역할
너는 대한민국의 블로그 포스팅 전문가다. 너의 임무는 주어진 키워드에 대해 사용자의 클릭을 유발하는 매력적인 블로그 제목의 '마무리 문구'를 생성하는 것이다.

# 작업 설명
1. '핵심 키워드'의 특성을 분석한다.
2. 분석한 내용을 바탕으로, 블로그 제목의 마지막 부분에 위치할 '마무리 문구'를 생성한다.
3. 생성되는 문구는 아래 '카테고리'를 참고하여 다양하게 구성한다.
4. 최종 결과물은 리스트 형태로, '마무리 문구'만 간결하게 제공한다.

# 블로그 제목 구조
[핵심 키워드] + [업체명] + [생성된 마무리 문구]

# 카테고리
- 정보성: 가격, 위치, 영업시간, 주차, 예약 방법 등 사실 정보
- 장점 및 특징: 시설, 전문성, 특별한 메뉴나 서비스, 분위기
- 경험 및 후기: 솔직 후기, 내돈내산, 방문 경험, 추천 이유
- 질문 및 호기심 유발: ~방법?, ~후기, ~꿀팁, ~총정리

# 예시
- 핵심 키워드: "강남역 파스타 맛집"
- 결과물 예시:
  데이트 코스로 강력 추천
  내돈내산 솔직 후기
  메뉴와 가격 총정리
  웨이팅 없이 즐기는 꿀팁
  분위기 좋은 곳 찾는다면


# 지시
아래 '핵심 키워드'에 대한 '마무리 문구' 10개를 생성해라.
마무리 문구는 20자 이내, 지역명/구/동/역/업종명/업체명은 절대 포함하지 않는다.
숫자나 번호 없이 한 줄씩 10개.

핵심 키워드: {키워드}
"""
            base_blog = base.get("blog", "")
            base_title = base.get("title", "")
            meta_prompt = (
                f"아래는 블로그 자동 포스팅에 사용하는 '기본' 시스템 프롬프트 예시다.\n\n"
                f"[기본 블로그 본문 프롬프트 예시]\n{base_blog}\n\n"
                f"[기본 제목 프롬프트 예시]\n{base_title}\n\n"
                f"---\n\n"
                f"위 예시를 참고해서 '{key}' 업종에 특화된 시스템 프롬프트를 새로 작성해라.\n\n"
                f"[반드시 지킬 규칙]\n"
                f"- 실제 블로그 글을 쓰지 말 것. GPT에게 블로그를 쓰도록 지시하는 '시스템 프롬프트(지시문)'를 작성해야 한다.\n"
                f"- 플레이스홀더 유지 필수: {{업체명}}, {{업종}}, {{키워드}}, {{시}}, {{구}}, {{동}}, {{주소}}, {{근처역}}, {{근처역상세}}\n"
                f"- '{key}' 업종을 방문하는 사람의 관점과 관심사에 맞게 역할·지시 내용을 조정\n"
                f"- 이모티콘 금지, 광고성 표현 금지\n\n"
                f"[출력 형식 — 반드시 아래 구분자 그대로 단독 줄에 출력]\n"
                f"===BLOG===\n"
                f"('{key}' 업종용 블로그 본문 시스템 프롬프트)\n"
                f"===TITLE===\n"
                f"('{key}' 업종용 제목 시스템 프롬프트)\n"
            )

            btn_auto.setEnabled(False)
            btn_auto.setText("생성 중...")

            def _worker():
                try:
                    from openai import OpenAI
                    client = OpenAI(api_key=api_keys[0])
                    resp = client.chat.completions.create(
                        model="gpt-4o-mini",
                        messages=[{"role": "user", "content": meta_prompt}],
                    )
                    text = resp.choices[0].message.content or ""
                    ai_sig.done.emit(text, "")
                except Exception as e:
                    ai_sig.done.emit("", str(e))

            threading.Thread(target=_worker, daemon=True).start()

        btn_add.clicked.connect(add_type)
        btn_add_user.clicked.connect(add_user_prompt)
        btn_auto.clicked.connect(auto_generate)
        btn_del.clicked.connect(del_type)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_save = QPushButton("저장")
        btn_save.setStyleSheet("background: #4a6cf7; color: white; border-radius: 8px; padding: 8px 20px; font-weight: bold;")
        btn_cancel = QPushButton("취소")
        btn_cancel.setStyleSheet("padding: 8px 20px;")
        btn_row.addWidget(btn_save)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

        def do_save():
            save_current_to_memory()
            try:
                # 비관리자: 파일 전체를 덮어쓰지 않고 변경된 키만 병합
                if not is_admin:
                    try:
                        with open(prompts_path, "r", encoding="utf-8") as _f:
                            full = json.load(_f)
                    except Exception:
                        full = {}
                    for k, v in prompts.items():
                        full[k] = v
                    save_data = full
                else:
                    save_data = prompts
                with open(prompts_path, "w", encoding="utf-8") as f:
                    json.dump(save_data, f, ensure_ascii=False, indent=2)
                # Firebase에 유저별 프롬프트 저장
                try:
                    from users import update_user as _upd_p
                    from config import get_current_user as _gcu
                    _upd_p(_gcu(), prompts=save_data)
                except Exception:
                    pass
                QMessageBox.information(dlg, "완료", "프롬프트를 저장했습니다.")
                dlg.accept()
            except Exception as e:
                QMessageBox.critical(dlg, "오류", f"저장 실패: {e}")

        btn_save.clicked.connect(do_save)
        btn_cancel.clicked.connect(dlg.reject)

        load_current()
        dlg.exec()

    # ── 크롤링 ──
    def _check_api_keys(self) -> bool:
        cfg = load_config()
        has_key = any(k for k in cfg.get("gpt_key_list", []) if k) or any(k for k in cfg.get("gemini_key_list", []) if k) or any(k for k in cfg.get("deepseek_key_list", []) if k)
        if not has_key:
            QMessageBox.warning(self, "API 키 필요", "API 키가 설정되지 않았습니다.\n설정에서 API 키를 입력하거나 관리자에게 문의하세요.")
            return False
        return True

    def _open_engage_dialog(self):
        """💬 댓글·이웃 관리 — 서로이웃 목록 / 댓글목록+원글삭제 / 공감·자동답글."""
        from PySide6.QtWidgets import (
            QDialog, QVBoxLayout, QHBoxLayout, QCheckBox, QPushButton, QLabel,
            QTextEdit, QTextBrowser, QTabWidget, QWidget, QTableWidget, QTableWidgetItem,
            QAbstractItemView, QSpinBox, QComboBox,
        )
        from PySide6.QtCore import QTimer, Qt
        import threading, queue
        import comment_buddy_manager as cbm

        cfg = load_config()
        accounts = cfg.get("accounts", []) or []
        accounts = [a for a in accounts if a.get("blog_id") or a.get("naver_id")]
        if not accounts:
            QMessageBox.critical(self, "오류", "네이버 계정 정보를 먼저 설정해주세요.")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("💬 댓글·이웃 관리")
        dlg.resize(1040, 820)
        dlg.setMinimumSize(820, 640)
        root = QVBoxLayout(dlg)

        import math as _math
        GROUP = max(1, _math.ceil(len(accounts) / 3)) if accounts else 1  # 한 명의당 아이디 수

        def _acct_disp(i):
            """네이버아이디/블로그아이디 형식."""
            a = accounts[i]
            nid = (a.get("naver_id", "") or "").strip() or "-"
            bid = (a.get("blog_id", "") or "").strip()
            return f"{nid}/{bid}" if bid else nid

        def _slot_expired(acct_idx):
            """계정(아이디)이 속한 명의의 요금제 만료/미결제 여부."""
            try:
                from users import is_account_expired
                return is_account_expired(getattr(self, "current_user", {}) or {}, self._account_slot(acct_idx))
            except Exception:
                return False

        acct_row = QHBoxLayout()
        acct_row.addWidget(QLabel("<b>명의(계정):</b>"))
        cmb_acct = QComboBox()
        # 명의별로 그룹 헤더 + 그 아래 아이디들
        for g in range(0, len(accounts), GROUP):
            myeongui = g // GROUP + 1
            cmb_acct.addItem(f"━━ {myeongui}명의 ━━")
            try:
                cmb_acct.model().item(cmb_acct.count() - 1).setEnabled(False)  # 헤더 선택 불가
            except Exception:
                pass
            for j in range(g, min(g + GROUP, len(accounts))):
                lab = f"    {j - g + 1}. {_acct_disp(j)}"
                if _slot_expired(j):
                    lab += "   🔒미결제"
                cmb_acct.addItem(lab)
                cmb_acct.setItemData(cmb_acct.count() - 1, j, Qt.UserRole)
        # 활성 계정 선택 (없으면 첫 아이디)
        _active = cfg.get("active_account", 0)
        _sel = 1
        for r in range(cmb_acct.count()):
            if cmb_acct.itemData(r, Qt.UserRole) == _active:
                _sel = r
                break
        cmb_acct.setCurrentIndex(_sel)
        acct_row.addWidget(cmb_acct, 1)
        b_login = QPushButton("로그인")
        b_login.setStyleSheet("background:#2563eb;color:white;border:none;border-radius:6px;padding:6px 18px;font-weight:bold;")
        acct_row.addWidget(b_login)
        b_sweep = QPushButton("🔁 전체 자동 관리")
        b_sweep.setToolTip("결제된 모든 아이디를 순서대로 자동 로그인하며\n서이추 수락 · 삭제요청 원글 삭제 · 공감/자동답글을 알아서 처리")
        b_sweep.setStyleSheet("background:#7c3aed;color:white;border:none;border-radius:6px;padding:6px 18px;font-weight:bold;")
        acct_row.addWidget(b_sweep)
        root.addLayout(acct_row)

        def _cur_acct_idx():
            """현재 선택된 아이디의 accounts 인덱스 (헤더면 None)."""
            return cmb_acct.currentData(Qt.UserRole)

        tabs = QTabWidget()
        root.addWidget(tabs, 1)

        state = {"stop": False, "busy": False, "poster": None, "mgr": None, "acct_idx": -1}
        q = queue.Queue()

        def _log(m):
            q.put(("log", str(m)))
            try:
                print(f"[engage] {m}", flush=True)  # 진단용 — _gui_run.log 에도 남김
            except Exception:
                pass

        log_view = QTextEdit()
        log_view.setReadOnly(True)
        log_view.setStyleSheet("background:#0f172a;color:#e2e8f0;font-family:Consolas;font-size:12px;")

        def _kill_profile_chrome(blog_id):
            """해당 계정 프로필을 쓰는 고아 크롬 프로세스 종료 (close 후에도 남는 것 정리)."""
            if not blog_id:
                return
            try:
                import subprocess
                ps = ("Get-CimInstance Win32_Process -Filter \"Name='chrome.exe'\" | "
                      "Where-Object { $_.CommandLine -like '*chrome_profile*" + blog_id + "*' } | "
                      "ForEach-Object { Stop-Process -Id $_.ProcessId -Force -ErrorAction SilentlyContinue }")
                subprocess.run(["powershell", "-NoProfile", "-Command", ps],
                               timeout=15, capture_output=True)
            except Exception:
                pass

        def _ensure_mgr(force_idx=None):
            idx = _cur_acct_idx() if force_idx is None else force_idx
            if idx is None:
                return None
            account = accounts[idx]
            # 계정이 바뀌었으면 기존 세션 닫기
            if state["mgr"] is not None and state["acct_idx"] != idx:
                old_idx = state["acct_idx"]
                old_blog = accounts[old_idx].get("blog_id", "") if 0 <= old_idx < len(accounts) else ""
                _log(f"계정 전환 → {idx//GROUP+1}명의 {idx%GROUP+1}. {_acct_disp(idx)} 재로그인")
                try:
                    state["poster"].close()
                except Exception:
                    pass
                state["mgr"] = None
                state["poster"] = None
                _kill_profile_chrome(old_blog)  # 고아 크롬 확실히 정리
                import time as _t_sw
                _t_sw.sleep(2.0)  # 이전 크롬 완전 종료 대기 (바로 새 크롬 띄우면 충돌)
            # 같은 계정이고 세션 살아있으면 재사용
            elif state["mgr"] is not None and state["poster"] is not None:
                try:
                    _ = state["poster"].driver.current_url
                    return state["mgr"]
                except Exception:
                    _log("브라우저가 닫혀 다시 로그인합니다...")
                    try:
                        state["poster"].close()
                    except Exception:
                        pass
                    state["mgr"] = None
                    state["poster"] = None
            if not account.get("naver_id") or not account.get("naver_pw"):
                _log(f"⚠ {account.get('blog_id','')}에 네이버 아이디/비번이 없습니다.")
                return None
            _log(f"로그인 중... 크롬 창에서 진행됩니다 ({idx//GROUP+1}명의 {idx%GROUP+1}. {_acct_disp(idx)})")
            _kill_profile_chrome(account.get("blog_id", ""))  # 시작 전 이 계정 고아 크롬 정리
            import time as _t_st
            poster = None
            for _attempt in range(3):
                try:
                    poster = NaverBlogPoster(
                        naver_id=account["naver_id"], naver_pw=account["naver_pw"],
                        blog_id=account["blog_id"], headless=False,
                        window_w=1180, window_h=820, stop_flag=lambda: state["stop"])
                    poster.start_browser()
                    break
                except Exception as e:
                    _log(f"브라우저 시작 실패 ({_attempt+1}/3): {str(e)[:55]}")
                    try:
                        if poster:
                            poster.close()
                    except Exception:
                        pass
                    poster = None
                    _t_st.sleep(2.5)
            if poster is None:
                _log("브라우저를 시작하지 못했습니다. 잠시 후 다시 시도해주세요.")
                return None
            try:
                self._active_posters.append(poster)
            except Exception:
                pass
            # 로그인 창을 화면 안에 적당한 크기로 (전체화면 X)
            try:
                poster.driver.set_window_rect(x=80, y=60, width=1180, height=860)
                poster.driver.execute_script("window.focus();")
            except Exception:
                pass
            if not poster.login():
                _log("로그인 실패! 캡차/기기인증이면 뜬 창에서 직접 풀고 다시 눌러주세요.")
                try:
                    poster.close()
                except Exception:
                    pass
                return None
            _log("로그인 성공 — 창을 숨깁니다")
            # 로그인 끝나면 창 숨김 (서로이웃 수락 시에만 잠깐 다시 띄움)
            try:
                poster.driver.set_window_position(-32000, -32000)
            except Exception:
                try:
                    poster.driver.minimize_window()
                except Exception:
                    pass
            state["poster"] = poster
            state["acct_idx"] = idx
            state["mgr"] = cbm.CommentBuddyManager(poster, log=_log, stop_flag=lambda: state["stop"])
            return state["mgr"]

        def _run(fn):
            idx = _cur_acct_idx()
            if idx is None:
                QMessageBox.information(dlg, "안내", "명의 아래의 아이디를 선택하세요."); return
            if _slot_expired(idx):
                myn = idx // GROUP + 1
                QMessageBox.warning(
                    dlg, "요금제 미결제",
                    f"{myn}명의 {idx % GROUP + 1}.  {_acct_disp(idx)}\n요금제가 만료되었거나\n"
                    "결제되지 않아 이 기능을 사용할 수 없습니다.\n\n구독 결제 후 이용해주세요.")
                _log(f"⚠ {myn}명의 요금제 미결제/만료 — 차단됨")
                return
            if state["busy"]:
                _log("이미 작업 중입니다. 잠시만요...")
                return
            state["busy"] = True
            state["stop"] = False

            def _w():
                try:
                    mgr = _ensure_mgr()
                    if mgr is not None:
                        fn(mgr)
                except Exception as e:
                    _log(f"오류: {e}")
                finally:
                    state["busy"] = False
                    q.put(("done",))
            threading.Thread(target=_w, daemon=True).start()

        def _checked_rows(tbl):
            return [r for r in range(tbl.rowCount())
                    if tbl.item(r, 0) is not None and tbl.item(r, 0).checkState() == Qt.Checked]

        def _set_all(tbl, checked):
            st = Qt.Checked if checked else Qt.Unchecked
            for r in range(tbl.rowCount()):
                if tbl.item(r, 0) is not None:
                    tbl.item(r, 0).setCheckState(st)

        def _mk_check(value):
            it = QTableWidgetItem()
            it.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            it.setCheckState(Qt.Unchecked)
            it.setData(Qt.UserRole, value)
            return it

        # ── 탭1: 서로이웃 ──
        t1 = QWidget(); l1 = QVBoxLayout(t1)
        l1.addWidget(QLabel("받은 서로이웃 신청 — 선택 후 수락/거절"))
        tbl_b = QTableWidget(0, 5)
        tbl_b.setHorizontalHeaderLabels(["선택", "신청자(닉네임)", "아이디", "메시지", "신청일"])
        tbl_b.setEditTriggers(QAbstractItemView.NoEditTriggers)
        tbl_b.setSelectionBehavior(QAbstractItemView.SelectRows)
        tbl_b.horizontalHeader().setStretchLastSection(True)
        l1.addWidget(tbl_b, 1)
        rb = QHBoxLayout()
        b_load = QPushButton("목록 불러오기")
        b_all = QPushButton("전체선택")
        b_clr = QPushButton("전체해제")
        b_acc = QPushButton("선택 수락"); b_acc.setStyleSheet("background:#10b981;color:white;padding:6px 12px;border:none;border-radius:6px;")
        b_rej = QPushButton("선택 거절"); b_rej.setStyleSheet("background:#ef4444;color:white;padding:6px 12px;border:none;border-radius:6px;")
        rb.addWidget(b_load); rb.addWidget(b_all); rb.addWidget(b_clr); rb.addStretch(); rb.addWidget(b_acc); rb.addWidget(b_rej)
        l1.addLayout(rb)
        b_all.clicked.connect(lambda: _set_all(tbl_b, True))
        b_clr.clicked.connect(lambda: _set_all(tbl_b, False))
        # '서로이웃 신청' 탭 제거 — 서이추 수락은 '서이추+자동댓글' 버튼으로 일괄 처리

        def _load_buddies():
            def fn(mgr):
                data = mgr.list_buddy_requests()
                q.put(("buddy", data))
                _log(f"서로이웃 신청 {len(data)}건")
            _run(fn)

        def _act_buddies(accept):
            rows = _checked_rows(tbl_b)
            if not rows:
                QMessageBox.information(dlg, "안내", "신청을 선택하세요."); return
            ids = [tbl_b.item(r, 0).data(Qt.UserRole) for r in rows]
            ids = [x for x in ids if x]
            def fn(mgr):
                mgr.act_on_buddies(ids, accept=accept)
                q.put(("buddy", mgr.list_buddy_requests()))
            _run(fn)
        b_load.clicked.connect(_load_buddies)
        b_acc.clicked.connect(lambda: _act_buddies(True))
        b_rej.clicked.connect(lambda: _act_buddies(False))

        # ── 탭2: 댓글 목록 + 원글삭제 ──
        t2 = QWidget(); l2 = QVBoxLayout(t2)
        l2.addWidget(QLabel("삭제/법적조치 등 키워드가 있는 '삭제요청 의심' 댓글만 표시됩니다 — 선택 후 원글 삭제"))
        tbl_c = QTableWidget(0, 4)
        tbl_c.setHorizontalHeaderLabels(["선택", "글 제목", "댓글 내용", "작성자"])
        tbl_c.setEditTriggers(QAbstractItemView.NoEditTriggers)
        tbl_c.setSelectionBehavior(QAbstractItemView.SelectRows)
        tbl_c.horizontalHeader().setStretchLastSection(True)
        tbl_c.verticalHeader().setDefaultSectionSize(30)
        tbl_c.setMinimumHeight(210)   # 최소 5행 보이게
        l2.addWidget(tbl_c, 3)

        l2.addWidget(QLabel("<b>선택한 댓글 상세</b> (행을 클릭 / 원글 주소 클릭하면 브라우저로 열림)"))
        cmt_detail = QTextBrowser()
        cmt_detail.setOpenExternalLinks(True)   # 링크 클릭 시 기본 브라우저로 열기
        cmt_detail.setMinimumHeight(150)
        cmt_detail.setStyleSheet("background:#f8fafc; font-size:16px; padding:10px;")
        l2.addWidget(cmt_detail, 2)

        def _show_cmt_detail():
            r = tbl_c.currentRow()
            if r < 0 or tbl_c.item(r, 0) is None:
                return
            logno = tbl_c.item(r, 0).data(Qt.UserRole) or ""
            title = tbl_c.item(r, 1).text() if tbl_c.item(r, 1) else ""
            content = tbl_c.item(r, 2).text() if tbl_c.item(r, 2) else ""
            writer = tbl_c.item(r, 3).text() if tbl_c.item(r, 3) else ""
            url = f"https://blog.naver.com/{accounts[(_cur_acct_idx() or 0)].get('blog_id','')}/{logno}" if logno else ""
            cmt_detail.setHtml(
                f"<b>글 제목:</b> {title}<br>"
                f"<b>작성자:</b> {writer}<br>"
                f"<b>원글 주소:</b> <a href='{url}'>{url}</a><br><hr>"
                f"<b>댓글 내용</b><br><div style='white-space:pre-wrap'>{content}</div>")
        tbl_c.itemSelectionChanged.connect(_show_cmt_detail)
        tbl_c.cellClicked.connect(lambda *_: _show_cmt_detail())

        rc = QHBoxLayout()
        c_load = QPushButton("댓글 불러오기")
        c_all = QPushButton("전체선택")
        c_clr = QPushButton("전체해제")
        c_open = QPushButton("🔗 원글 열기")
        c_del = QPushButton("선택한 원글 삭제"); c_del.setStyleSheet("background:#ef4444;color:white;padding:6px 12px;border:none;border-radius:6px;")
        rc.addWidget(c_load); rc.addWidget(c_all); rc.addWidget(c_clr); rc.addWidget(c_open); rc.addStretch(); rc.addWidget(c_del)
        l2.addLayout(rc)
        c_all.clicked.connect(lambda: _set_all(tbl_c, True))
        c_clr.clicked.connect(lambda: _set_all(tbl_c, False))

        def _open_post():
            r = tbl_c.currentRow()
            if r < 0 or tbl_c.item(r, 0) is None:
                QMessageBox.information(dlg, "안내", "목록에서 글을 먼저 클릭해 선택하세요."); return
            logno = tbl_c.item(r, 0).data(Qt.UserRole) or ""
            if not logno:
                QMessageBox.information(dlg, "안내", "이 항목의 글 주소를 못 찾았습니다."); return
            bid = accounts[(_cur_acct_idx() or 0)].get("blog_id", "")
            import webbrowser
            webbrowser.open(f"https://blog.naver.com/{bid}/{logno}")
        c_open.clicked.connect(_open_post)
        tabs.insertTab(0, t2, "🚨 삭제요청 댓글")   # 맨 앞으로

        def _load_cmts():
            def fn(mgr):
                # 내가 공감/답글 단 댓글 + 삭제된 글 댓글은 제외하고 불러오기
                data = mgr.list_comments(exclude_engaged=True, check_deleted=True)
                q.put(("cmts", data))
                _log(f"댓글 {len(data)}건")
            _run(fn)

        def _del_posts():
            rows = _checked_rows(tbl_c)
            if not rows:
                QMessageBox.information(dlg, "안내", "삭제할 댓글(원글)을 선택하세요."); return
            lognos = [tbl_c.item(r, 0).data(Qt.UserRole) for r in rows]
            lognos = [x for x in lognos if x]
            if not lognos:
                QMessageBox.warning(dlg, "안내", "선택한 항목에서 글 주소를 못 찾았습니다."); return
            if QMessageBox.question(dlg, "원글 삭제 확인",
                    f"선택한 {len(lognos)}개 글을 정말 삭제할까요?\n되돌릴 수 없습니다.") != QMessageBox.Yes:
                return
            def fn(mgr):
                n = mgr.delete_posts(lognos)
                _log(f"원글 {n}개 삭제 완료")
                q.put(("cmts", mgr.list_comments()))
            _run(fn)
        c_load.clicked.connect(_load_cmts)
        c_del.clicked.connect(_del_posts)

        # ── 탭3: 공감·자동답글 ──
        t3 = QWidget(); l3 = QVBoxLayout(t3)
        cb_like = QCheckBox("댓글 공감(좋아요)"); cb_like.setChecked(True)
        cb_reply = QCheckBox("자동답글 (reply_phrases.json 문구 사용)"); cb_reply.setChecked(True)
        l3.addWidget(cb_like); l3.addWidget(cb_reply)
        hb = QHBoxLayout(); hb.addWidget(QLabel("처리할 글 최대 개수:"))
        sp_max = QSpinBox(); sp_max.setRange(1, 1000); sp_max.setValue(1000)
        hb.addWidget(sp_max); hb.addStretch()
        l3.addLayout(hb)
        e_start = QPushButton("🚀 서이추 수락 + 자동댓글 (한번에)")
        e_start.setStyleSheet("background:#10b981;color:white;padding:10px 16px;border:none;border-radius:6px;font-weight:bold;font-size:13px;")
        l3.addWidget(e_start); l3.addStretch()
        tabs.addTab(t3, "서이추+자동댓글")

        def _start_engage():
            if not (cb_like.isChecked() or cb_reply.isChecked()):
                QMessageBox.information(dlg, "안내", "공감/답글 중 하나는 선택하세요."); return
            phrases = cbm.load_reply_phrases()
            do_l, do_r, mx = cb_like.isChecked(), cb_reply.isChecked(), sp_max.value()
            def fn(mgr):
                # 1) 서로이웃 신청 전체 수락
                try:
                    reqs = mgr.list_buddy_requests()
                    ids = [(r.get("id") or r.get("nick", "")) for r in reqs]
                    ids = [x for x in ids if x]
                    if ids:
                        mgr.act_on_buddies(ids, accept=True)
                        _log(f"서로이웃 {len(ids)}건 수락 완료")
                        try:
                            q.put(("buddy", mgr.list_buddy_requests()))
                        except Exception:
                            pass
                    else:
                        _log("대기 중인 서로이웃 신청 없음")
                except Exception as e:
                    _log(f"서로이웃 수락 실패: {e}")
                # 2) 자동댓글(공감+답글)
                _log("공감·답글 시작...")
                mgr.like_and_reply_comments(phrases, do_like=do_l, do_reply=do_r, max_posts=mx)
                _log("✅ 서이추 수락 + 자동댓글 완료")
            _run(fn)
        e_start.clicked.connect(_start_engage)

        # ── 전체 자동 관리(sweep): 모든 아이디를 순회하며 알아서 처리 ──
        def _run_sweep():
            if state["busy"]:
                _log("이미 작업 중입니다. 잠시만요..."); return
            targets = [i for i in range(len(accounts)) if not _slot_expired(i)]
            if not targets:
                QMessageBox.information(dlg, "안내", "관리할 수 있는(결제된) 아이디가 없습니다."); return
            if QMessageBox.question(
                    dlg, "전체 자동 관리",
                    f"결제된 아이디 {len(targets)}개를 순서대로 자동 로그인하며\n"
                    "  ① 서이추 자동 수락\n"
                    "  ② 삭제요청 의심 댓글의 원글 자동 삭제\n"
                    "  ③ 공감 · 자동답글\n"
                    "을 알아서 진행합니다.\n\n"
                    "⚠ ②에서 '삭제/법적/권리침해' 등 키워드가 달린 글은 자동 삭제되며\n"
                    "되돌릴 수 없습니다.\n\n계속할까요?") != QMessageBox.Yes:
                return
            phrases = cbm.load_reply_phrases()
            state["busy"] = True
            state["stop"] = False

            def _sweep():
                try:
                    _log(f"🔁 전체 자동 관리 시작 — 대상 아이디 {len(targets)}개")
                    for n, idx in enumerate(targets, 1):
                        if state["stop"]:
                            _log("중단됨 — 관리 종료"); break
                        _log(f"━━ [{n}/{len(targets)}] {self._account_slot(idx)}명의 "
                             f"{idx % GROUP + 1}. {_acct_disp(idx)} ━━")
                        mgr = _ensure_mgr(force_idx=idx)
                        if mgr is None:
                            _log("  ⚠ 로그인 실패 — 다음 아이디로 넘어감"); continue
                        # ① 서이추 자동 수락
                        try:
                            reqs = mgr.list_buddy_requests()
                            ids = [(r.get("id") or r.get("nick", "")) for r in reqs]
                            ids = [x for x in ids if x]
                            if ids:
                                mgr.act_on_buddies(ids, accept=True)
                                _log(f"  ✓ 서이추 {len(ids)}건 수락")
                            else:
                                _log("  · 서이추 신청 없음")
                        except Exception as e:
                            _log(f"  · 서이추 실패: {str(e)[:50]}")
                        if state["stop"]:
                            break
                        # ② 삭제요청 의심 댓글 → 원글 자동 삭제
                        try:
                            cmts = mgr.list_comments(check_deleted=True)
                            lognos = list(dict.fromkeys(
                                c.get("logno") for c in (cmts or [])
                                if c.get("logno") and _is_del_request(c.get("content", ""))))
                            if lognos:
                                _log(f"  🚨 삭제요청 의심 {len(lognos)}건 → 원글 자동 삭제")
                                nd = mgr.delete_posts(lognos)
                                _log(f"  ✓ 원글 {nd}건 삭제")
                            else:
                                _log("  · 삭제요청 의심 댓글 없음")
                        except Exception as e:
                            _log(f"  · 삭제요청 처리 실패: {str(e)[:50]}")
                        if state["stop"]:
                            break
                        # ③ 공감 + 자동답글
                        try:
                            _log("  · 공감 · 자동답글 시작")
                            mgr.like_and_reply_comments(
                                phrases, do_like=True, do_reply=True, max_posts=1000)
                        except Exception as e:
                            _log(f"  · 자동댓글 실패: {str(e)[:50]}")
                    _log("✅ 전체 자동 관리 완료")
                except Exception as e:
                    _log(f"관리 오류: {e}")
                finally:
                    state["busy"] = False
                    q.put(("done",))
            threading.Thread(target=_sweep, daemon=True).start()
        b_sweep.clicked.connect(_run_sweep)

        # ── 로그 + 중단/닫기 ──
        root.addWidget(QLabel("로그"))
        log_view.setMaximumHeight(90)
        root.addWidget(log_view, 0)
        bottom = QHBoxLayout()
        b_stop = QPushButton("중단")
        b_close = QPushButton("닫기")
        bottom.addStretch(); bottom.addWidget(b_stop); bottom.addWidget(b_close)
        root.addLayout(bottom)

        def _fill_buddy(data):
            tbl_b.setRowCount(0)
            for d in data:
                r = tbl_b.rowCount(); tbl_b.insertRow(r)
                tbl_b.setItem(r, 0, _mk_check(d.get("id") or d.get("nick", "")))
                tbl_b.setItem(r, 1, QTableWidgetItem(d.get("nick", "")))
                tbl_b.setItem(r, 2, QTableWidgetItem(d.get("id", "")))
                tbl_b.setItem(r, 3, QTableWidgetItem(d.get("msg", "")))
                tbl_b.setItem(r, 4, QTableWidgetItem(d.get("date", "")))
            tbl_b.resizeColumnsToContents()
            tbl_b.horizontalHeader().setStretchLastSection(True)
            if not data:
                _log("대기 중인 서로이웃 신청이 없습니다.")

        _DEL_KW = ("삭제", "내려", "지워", "비공개", "권리침해", "초상권", "저작권",
                   "법적", "고소", "수정부탁", "수정 부탁", "게시중단", "게시 중단",
                   "내려주", "삭제요청", "삭제 부탁", "삭제부탁", "삭제해", "지워주")

        def _is_del_request(text):
            t = text or ""
            return any(k in t for k in _DEL_KW)

        def _fill_cmts(data):
            from PySide6.QtGui import QColor
            warn = QColor(255, 210, 210)  # 연한 빨강
            tbl_c.setRowCount(0)
            # 삭제/법적조치 등 키워드가 있는 '삭제요청 의심' 댓글만 추려서 표시
            filtered = [d for d in (data or []) if _is_del_request(d.get("content", ""))]
            for d in filtered:
                r = tbl_c.rowCount(); tbl_c.insertRow(r)
                items = [_mk_check(d.get("logno", "")),
                         QTableWidgetItem(d.get("title", "")),
                         QTableWidgetItem(d.get("content", "")),
                         QTableWidgetItem(d.get("writer", ""))]
                for ci, it in enumerate(items):
                    tbl_c.setItem(r, ci, it)
                    it.setBackground(warn)
                    it.setToolTip("삭제요청 의심 댓글")
            tbl_c.resizeColumnsToContents()
            tbl_c.horizontalHeader().setStretchLastSection(True)
            if not filtered:
                _log(f"삭제요청 의심 댓글 없음 (받은 댓글 {len(data or [])}개 확인)")
            else:
                _log(f"⚠ 삭제요청 의심 댓글 {len(filtered)}개 추려냄")

        timer = QTimer(dlg)

        def _drain():
            while not q.empty():
                try:
                    item = q.get_nowait()
                except Exception:
                    break
                kind = item[0]
                if kind == "log":
                    log_view.append(item[1]); log_view.moveCursor(QTextCursor.End)
                elif kind == "buddy":
                    _fill_buddy(item[1])
                elif kind == "cmts":
                    _fill_cmts(item[1])
        timer.timeout.connect(_drain)
        timer.start(200)

        def _on_acct_changed(_i):
            # 명의 바꾸면 목록 비우기 (다음 작업 시 그 계정으로 재로그인)
            tbl_b.setRowCount(0)
            tbl_c.setRowCount(0)
            _log(f"명의 선택: {cmb_acct.currentText()} — 작업 누르면 이 계정으로 로그인합니다.")
        cmb_acct.currentIndexChanged.connect(_on_acct_changed)

        def _on_stop():
            state["stop"] = True
            _log("중단 요청됨...")

        def _on_close():
            state["stop"] = True
            timer.stop()
            try:
                if state["poster"]:
                    state["poster"].close()
            except Exception:
                pass
            dlg.reject()
        b_stop.clicked.connect(_on_stop)
        b_close.clicked.connect(_on_close)

        def _do_login():
            idx = _cur_acct_idx()
            if idx is None:
                QMessageBox.information(dlg, "안내", "명의 아래의 아이디를 선택하세요."); return
            myn = self._account_slot(idx)
            pos = idx % GROUP + 1
            disp = _acct_disp(idx)
            if _slot_expired(idx):
                QMessageBox.warning(dlg, "요금제 미결제",
                    f"{myn}명의 {pos}.  {disp}\n요금제가 만료되었거나 결제되지 않아 로그인할 수 없습니다."); return
            if QMessageBox.question(dlg, "로그인",
                    f"{myn}명의 {pos}.  {disp}\n이 아이디로 로그인할까요?") != QMessageBox.Yes:
                return
            # 로그인 후 서로이웃 신청·댓글을 '한 작업'으로 자동 불러오기 (순차 busy 충돌 방지)
            def _after_login(mgr):
                _log("✅ 로그인 완료 — 서로이웃 신청·댓글 자동 불러오는 중...")
                try:
                    b = mgr.list_buddy_requests()
                    q.put(("buddy", b)); _log(f"서로이웃 신청 {len(b)}건")
                except Exception as e:
                    _log(f"서로이웃 불러오기 실패: {e}")
                try:
                    c = mgr.list_comments(exclude_engaged=True, check_deleted=True)
                    q.put(("cmts", c)); _log(f"댓글 {len(c)}건")
                except Exception as e:
                    _log(f"댓글 불러오기 실패: {e}")
            _run(_after_login)
        b_login.clicked.connect(_do_login)

        dlg.exec()

    def _start_crawl(self):
        if self._block_if_expired():
            return
        if not self._check_api_keys():
            return
        _url_mode = self._radio_url.isChecked()
        keyword = self.keyword_input.currentText().strip()
        if not keyword:
            QMessageBox.warning(self, "경고", "키워드를 입력해주세요.")
            return
        _direct_url = None  # 키워드 기반도 동일한 URL 로딩, no_filter로 필터만 끔
        if self.is_crawling:
            return

        self._save_keyword_history(keyword)
        self.last_keyword = keyword
        # 설정에도 저장 (계정별, 재시작해도 유지)
        try:
            _c = load_config()
            _c["last_keyword"] = keyword  # 호환용
            lkm = _c.get("last_keyword_by_account", {}) or {}
            lkm[self._current_blog_id()] = keyword
            _c["last_keyword_by_account"] = lkm
            save_config(_c)
        except Exception:
            pass
        count = self.count_spin.value()
        self.is_crawling = True
        self.stop_flag = False
        self.crawled_data = []
        self._current_crawl_mode = "keyword" if self._radio_url.isChecked() else "region"

        log_dir = self._get_logs_dir()
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        self.result_file = os.path.join(log_dir, f"places_{timestamp}.json")

        existing_for_resume = []

        self._emit_status("크롤링 중...", "#f59e0b")
        if _url_mode:
            self._emit_log(f"크롤링 시작: URL 기반 (필터 없음)")
        else:
            self._emit_log(f"크롤링 시작: '{keyword}' (검색 지역 일치 항목만 수집)")
        # 크롤 시작과 동시에 업종 프롬프트 자동 생성 (없는 경우만)
        try:
            _biz_now = keyword.strip().split()[-1] if keyword.strip() else ""
            if _biz_now:
                self._auto_generate_prompt_for_biz(_biz_now)
        except Exception:
            pass

        def _worker():
            try:
                last_save_count = [len(existing_for_resume)]
                latest_results = [list(existing_for_resume)]

                def on_progress(current, scanned, name, results_ref=None):
                    if results_ref is not None:
                        latest_results[0] = results_ref
                    if self.stop_flag:
                        try:
                            self._save_crawled(latest_results[0], keyword)
                        except Exception:
                            pass
                        raise InterruptedError("중단됨")
                    self._emit_log(f"  {name} ({current}개)")
                    self._emit_status(f"수집 {current}개", "#f59e0b")
                    self._emit_crawl_count(f"{current}개")
                    # 메인창 캐시만 갱신 (logs/{kw}.json은 _on_item이 매번 저장)
                    if current - last_save_count[0] >= 1:
                        try:
                            self._save_crawled(latest_results[0], keyword)
                            last_save_count[0] = current
                        except Exception:
                            pass

                # 업체 1개 크롤 끝날 때마다 즉시 저장 (logs/{account}/{kw}.json)
                _crawl_start_times = {}  # kw → 첫 저장 시각
                def _on_item(place, items_so_far, kw):
                    try:
                        import re as _re
                        _safe = _re.sub(r"[^가-힣A-Za-z0-9_]", "_", str(kw or "unknown"))[:80]
                        _dir = self._get_logs_dir()
                        os.makedirs(_dir, exist_ok=True)
                        _fp = os.path.join(_dir, f"{_safe}.json")
                        if kw not in _crawl_start_times:
                            _crawl_start_times[kw] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        with open(_fp, "w", encoding="utf-8") as _f:
                            json.dump({"keyword": kw, "crawled_at": _crawl_start_times[kw], "crawl_mode": getattr(self, "_current_crawl_mode", ""), "items": list(items_so_far)}, _f, ensure_ascii=False, indent=2)
                        pass
                    except Exception as _e:
                        self._emit_log(f"업체별 저장 실패: {_e}")

                def _crawl_with_retry(kw, target_count, existing, excludes, max_retries=3, d_url=None, no_filter=False):
                    """timeout/webdriver 예외 발생 시 드라이버 재시작 + resume."""
                    acc = list(existing or [])
                    last_err = None
                    for attempt in range(max_retries):
                        try:
                            return crawl_places(
                                kw, target_count, on_progress,
                                existing_places=acc,
                                exclude_keywords=excludes,
                                on_item=_on_item,
                                direct_url=d_url,
                                no_filter=no_filter,
                                stop_flag=lambda: self.stop_flag,
                                emit_log=self._emit_log,
                            )
                        except InterruptedError:
                            raise
                        except Exception as e:
                            last_err = e
                            msg = str(e)
                            is_timeout = ("Read timed out" in msg) or ("Timeout" in msg) or ("WebDriverException" in type(e).__name__)
                            if is_timeout and attempt < max_retries - 1:
                                self._emit_log(f"  [재시도 {attempt+1}/{max_retries}] 드라이버 timeout → 재시작 후 resume ({len(acc)}개 유지)")
                                import time as _t
                                # 3초 대기 — 중단 즉시 반응 (0.2초 단위)
                                _end = _t.time() + 3
                                while _t.time() < _end:
                                    if self.stop_flag: raise InterruptedError("중단됨")
                                    _t.sleep(0.2)
                                continue
                            raise
                    if last_err:
                        raise last_err
                    return acc

                # 키워드 기반 모드: 업종/지역 필터 없이 전체 수집 (제외 키워드는 지역기반과 동일하게 적용)
                if _url_mode:
                    _kp = keyword.split()
                    _biz_kw = _kp[-1] if _kp else keyword
                    _ex_map_kw = getattr(self, "_exclude_keywords_by_biz", {}) or {}
                    _excludes_kw = list(_ex_map_kw.get(_biz_kw, []))
                    if _excludes_kw:
                        self._emit_log(f"제외 키워드 {len(_excludes_kw)}개: {', '.join(_excludes_kw)}")
                    # 기존 파일 삭제 후 처음부터 수집
                    import re as _re_kw
                    _kw_safe = _re_kw.sub(r"[^가-힣A-Za-z0-9_]", "_", str(keyword or "unknown"))[:80]
                    _kw_fp = os.path.join(self._get_logs_dir(), f"{_kw_safe}.json")
                    if os.path.exists(_kw_fp):
                        try:
                            os.remove(_kw_fp)
                            self._emit_log(f"기존 수집 데이터 초기화: {keyword}")
                        except Exception:
                            pass
                    self._emit_log(f"키워드 기반 크롤링: {keyword}")
                    results = _crawl_with_retry(keyword, count, [], _excludes_kw, no_filter=True)
                    self.crawled_data = results
                    self._save_crawled(results, keyword)
                    return

                regions = list(getattr(self, "_selected_regions", []) or [])
                _kparts = keyword.split()
                _biz = _kparts[-1] if _kparts else keyword
                if regions:
                    expanded = [f"{r} {_biz}" for r in regions]
                else:
                    expanded = [keyword]
                # 제외 키워드 (업종 기준)
                _ex_map = getattr(self, "_exclude_keywords_by_biz", {}) or {}
                _excludes = list(_ex_map.get(_biz, []))
                if _excludes:
                    self._emit_log(f"제외 키워드 {len(_excludes)}개: {', '.join(_excludes)}")
                if len(expanded) > 1:
                    # 기존 파일 삭제 후 처음부터 수집
                    import re as _re2
                    _log_dir = self._get_logs_dir()
                    _deleted = 0
                    for _ekw in expanded:
                        _safe = _re2.sub(r"[^가-힣A-Za-z0-9_]", "_", str(_ekw))[:80]
                        _fp = os.path.join(_log_dir, f"{_safe}.json")
                        if os.path.exists(_fp):
                            try:
                                os.remove(_fp)
                                _deleted += 1
                            except Exception:
                                pass
                    if _deleted:
                        self._emit_log(f"기존 수집 데이터 초기화: {_deleted}개 지역")

                    # 병렬 크롤 — 사용자가 설정한 봇 개수(1/2/3)와 키워드 개수 중 작은 값
                    _user_bots = self._get_bot_count()
                    _workers = min(_user_bots, max(1, len(expanded)))
                    self._emit_log(f"지역 {len(expanded)}개 × '{_biz}' 병렬 크롤 (봇 {_workers}개)")
                    _per = count  # 지역별 목표 개수 = 사용자 설정값
                    _existing_by_kw = {}

                    def _on_batch(items, kw=None, raw=None):
                        # 키워드별 단일 파일은 _on_item이 매번 update 중. 여기선 메인창 캐시만 갱신.
                        latest_results[0] = list(items)
                        try:
                            self._save_crawled(list(items), keyword)
                        except Exception:
                            pass
                    results = crawl_places_parallel(
                        expanded,
                        count_per=_per,
                        on_progress=on_progress,
                        existing_places=existing_for_resume,
                        existing_by_keyword=_existing_by_kw,
                        exclude_keywords=_excludes,
                        max_workers=_workers,
                        stop_flag=lambda: self.stop_flag,
                        emit_log=self._emit_log,
                        save_batch=_on_batch,
                        on_item=_on_item,
                        no_filter=False,  # 지역(구) 필터 적용 — 검색한 구만 수집(인접구 제거). 업종필터는 이미 없음
                    )
                    latest_results[0] = results
                    try:
                        self._save_crawled(results, keyword)
                        save_results(results, self.result_file, keyword)
                    except Exception:
                        pass
                else:
                    results = _crawl_with_retry(expanded[0], count, existing_for_resume, _excludes, no_filter=False)
                    for _p in results:
                        _p.setdefault("search_keyword", keyword)
                latest_results[0] = results
                self.crawled_data = results
                # app_user별 영속 저장 (메인창 캐시) — 키워드별 logs/{kw}.json은 _on_item이 처리
                self._save_crawled(results, keyword)

                self._emit_log(f"크롤링 완료! {len(results)}개 수집")
                self._emit_status("완료", "#22c55e")
                self._emit_crawl_count(f"{len(results)}개")
                self._update_history_result_count(keyword, len(results))

            except InterruptedError:
                self._emit_log(f"크롤링 중단완료 ({len(latest_results[0])}개 저장)")
                self._emit_status("중단완료", "#ef4444")
                try:
                    self._save_crawled(latest_results[0], keyword)
                except Exception:
                    pass
            except Exception as e:
                self._emit_log(f"크롤링 오류: {e}")
                self._emit_status("오류", "#ef4444")
                try:
                    self._save_crawled(latest_results[0], keyword)
                    self._emit_log(f"오류 전까지 {len(latest_results[0])}개 저장됨 → F7에서 이어서 선택 가능")
                except Exception:
                    pass
            finally:
                self.is_crawling = False

        threading.Thread(target=_worker, daemon=True).start()

    def _on_crawl_mode_changed(self, keyword_checked: bool):
        url_mode = not keyword_checked
        self._crawl_url_input.setVisible(False)  # URL 입력창 항상 숨김
        self.keyword_input.setVisible(True)       # 키워드 입력은 항상 표시
        self.btn_regions.setVisible(not url_mode)
        self.btn_excludes.setVisible(True)        # 제외 키워드는 두 모드 모두 표시

    def _force_quit_posters(self):
        """활성 NaverBlogPoster 브라우저를 백그라운드 스레드에서 강제 quit — selenium 블로킹 호출을 즉시 깨움."""
        if not getattr(self, "_active_posters", None):
            return
        import threading as _th
        def _kill():
            for p in list(self._active_posters):
                try:
                    if getattr(p, "driver", None):
                        p.driver.quit()
                except Exception:
                    pass
        _th.Thread(target=_kill, daemon=True).start()

    def _stop(self):
        if self.is_crawling or self.is_posting or getattr(self, 'is_generating', False):
            self.stop_flag = True
            self._emit_log("중단 요청 — 즉시 정지 중...")
            self._emit_status("중단 중...", "#f59e0b")
            self._force_quit_posters()

    def _stop_generating(self):
        if getattr(self, 'is_generating', False):
            self.stop_flag = True
            self._emit_post_log("글쓰기 중단 요청 — 즉시 정지 중...")

    def _stop_posting_only(self):
        if self.is_posting:
            self.stop_flag = True
            self._emit_post_log("포스팅 중단 요청 — 즉시 정지 중...")
            self._force_quit_posters()

    # ── 결과보기 ──
    def _show_results(self):
        if not self.crawled_data:
            log_dir = self._get_logs_dir()
            if os.path.exists(log_dir):
                files = sorted(
                    [f for f in os.listdir(log_dir) if f.endswith(".json") and not f.startswith("._")],
                    reverse=True
                )
                if files:
                    filepath = os.path.join(log_dir, files[0])
                    raw = load_results(filepath)
                    self.crawled_data = raw.get("items", []) if isinstance(raw, dict) else raw
                    self._emit_log(f"결과 불러옴: {files[0]}")

        if not self.crawled_data:
            QMessageBox.information(self, "결과", "크롤링 결과가 없습니다.")
            return

        from PySide6.QtWidgets import QTreeWidget, QTreeWidgetItem

        # 지역 추출: address에서 "시/도 + 구/군"
        def _region_of(p):
            import re as _re
            addr = p.get("address", "") or p.get("jibun_address", "")
            m = _re.search(
                r"((?:서울|부산|대구|인천|광주|대전|울산|세종|경기|강원|충북|충남|전북|전남|경북|경남|제주)\S*)\s+(\S+[시군구])",
                addr,
            )
            if m:
                return f"{m.group(1)} {m.group(2)}"
            m2 = _re.search(r"(\S+[시군구])", addr)
            if m2:
                return m2.group(1)
            return "지역 미상"

        groups = {}
        for p in self.crawled_data:
            r = _region_of(p)
            groups.setdefault(r, []).append(p)

        dlg = QDialog(self)
        dlg.setWindowTitle(f"크롤링 결과 — 총 {len(self.crawled_data)}건 ({len(groups)}개 지역)")
        dlg.resize(1350, 650)
        layout = QVBoxLayout(dlg)

        # 상단 컨트롤
        top = QHBoxLayout()
        btn_expand = QPushButton("전체 펼치기")
        btn_collapse = QPushButton("전체 접기")
        for b in (btn_expand, btn_collapse):
            b.setStyleSheet("padding: 6px 14px;")
            b.setCursor(Qt.PointingHandCursor)
        top.addWidget(btn_expand)
        top.addWidget(btn_collapse)
        top.addStretch()
        layout.addLayout(top)

        tree = QTreeWidget()
        tree.setAlternatingRowColors(True)
        headers = ["지역 / 업체명", "업체주소", "카테고리", "근처역", "앞 키워드", "태그", "픽사베이 키워드"]
        tree.setColumnCount(len(headers))
        tree.setHeaderLabels(headers)
        widths = [260, 280, 130, 150, 200, 160, 200]
        for i, w in enumerate(widths):
            tree.setColumnWidth(i, w)
        tree.setRootIsDecorated(True)
        tree.setUniformRowHeights(True)

        # 지역 가나다순
        for region in sorted(groups.keys()):
            items = groups[region]
            root = QTreeWidgetItem(tree, [f"▼ {region} ({len(items)}건)"])
            from PySide6.QtGui import QFont as _QF
            _f = _QF()
            _f.setBold(True)
            root.setFont(0, _f)
            for p in items:
                addr = p.get("address", "")
                jibun = p.get("jibun_address", "")
                full_addr = (jibun or addr).strip()  # jibun이 행정동+지번 완전주소라 중복 방지
                child = QTreeWidgetItem(root, [
                    p.get("name", ""),
                    full_addr,
                    p.get("category", ""),
                    p.get("nearby_station", ""),
                    p.get("front_keywords", ""),
                    p.get("tags", ""),
                    p.get("pixabay_keywords", ""),
                ])
                # 2차카테고리 — 있으면 접힌 상태로 하위 항목 추가
                cat2 = (p.get("category_2") or "").strip()
                if cat2:
                    sub = QTreeWidgetItem(child, ["", "", f"└ {cat2}", "", "", "", ""])
                    from PySide6.QtGui import QColor as _QC
                    sub.setForeground(2, _QC("#6b7280"))
                # child는 접힌 상태 유지 (setExpanded 호출 안 함)
            root.setExpanded(True)

        btn_expand.clicked.connect(lambda: tree.expandAll())
        btn_collapse.clicked.connect(lambda: tree.collapseAll())

        layout.addWidget(tree)

        def copy_to_clipboard():
            # 펼쳐진 자식 행만 복사 (지역 헤더 제외)
            lines = []
            for i in range(tree.topLevelItemCount()):
                region_item = tree.topLevelItem(i)
                for j in range(region_item.childCount()):
                    ch = region_item.child(j)
                    row = [ch.text(k) for k in range(tree.columnCount())]
                    lines.append("\t".join(row))
            QApplication.clipboard().setText("\n".join(lines))
            QMessageBox.information(dlg, "복사 완료", "엑셀에 붙여넣기(Ctrl+V) 하세요.")

        btn_copy = QPushButton("엑셀용 복사")
        btn_copy.setStyleSheet("background: #4a6cf7; color: white; border: none; border-radius: 8px; font-size: 13px; font-weight: bold; padding: 10px;")
        btn_copy.setCursor(Qt.PointingHandCursor)
        btn_copy.clicked.connect(copy_to_clipboard)
        layout.addWidget(btn_copy)

        dlg.exec()

    # ── 생성된 포스트 보기 (크롤 업체 목록 + 생성여부 표시) ──
    def _view_generated_posts(self):
        if self._block_if_expired():
            return
        posts = self._load_generated_posts()
        self._generated_posts = posts

        # (업체명+주소) → content 매핑
        post_map = {}
        for p in posts:
            pl = p.get("place", {})
            key = (pl.get("name", ""), pl.get("address", "") or pl.get("jibun_address", ""))
            post_map[key] = p

        # 크롤 결과 로드 — 지역기반/키워드기반 분리
        groups_region = self._load_all_crawl_results(mode="region")
        groups_keyword = self._load_all_crawl_results(mode="keyword")
        if self.crawled_data:
            cur_label = f"현재 크롤링 결과 ({len(self.crawled_data)}개)"
            cur_mode = getattr(self, "_current_crawl_mode", "region")
            if cur_mode == "keyword":
                groups_keyword[cur_label] = self.crawled_data
            else:
                groups_region[cur_label] = self.crawled_data

        if not groups_region and not groups_keyword and not posts:
            QMessageBox.information(self, "결과", "크롤링/생성된 포스트가 없습니다.")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("생성된 포스트 보기")
        dlg.resize(1000, 650)
        layout = QVBoxLayout(dlg)

        # 상단: 전체 선택/해제
        top_row = QHBoxLayout()
        btn_all_sel = QPushButton("전체 선택")
        btn_all_sel.setStyleSheet("padding: 6px 15px;")
        btn_all_none = QPushButton("전체 해제")
        btn_all_none.setStyleSheet("padding: 6px 15px;")
        top_row.addWidget(btn_all_sel)
        top_row.addWidget(btn_all_none)
        top_row.addStretch()
        layout.addLayout(top_row)

        info = QLabel("● 초록: 생성됨 (더블클릭해서 내용 확인 / 체크 후 삭제 가능)  [완] 파랑: 발행완료  ● 빨강: 미생성")
        info.setStyleSheet("font-size: 12px; color: #475569; padding: 5px;")
        layout.addWidget(info)

        # 탭 위젯 — 지역기반/키워드기반 분리
        from PySide6.QtWidgets import QTabWidget
        tab_widget = QTabWidget()

        # 더블클릭 시 포스트 내용 보기를 위한 매핑
        item_to_post = {}
        all_children = []  # (child_item, place) — 전체 선택/삭제용
        all_trees = []     # 탭별 QTreeWidget 목록

        def _build_tree(groups_dict):
            tw = QTreeWidget()
            tw.setHeaderLabels(["업체명", "업체주소", "카테고리", "근처역", "앞 키워드", "태그"])
            tw.setColumnWidth(0, 260)
            tw.setColumnWidth(1, 240)
            tw.setColumnWidth(2, 100)
            tw.setColumnWidth(3, 100)
            tw.setColumnWidth(4, 160)
            tw.setAlternatingRowColors(True)
            tab_count = 0
            for group_name, places in groups_dict.items():
                if not places:
                    continue
                parent = QTreeWidgetItem(tw)
                parent.setText(0, group_name)
                parent.setFlags(parent.flags() | Qt.ItemIsUserCheckable)
                parent.setCheckState(0, Qt.Unchecked)
                parent.setExpanded(False)

                gen_count = 0
                for num, p in enumerate(places, 1):
                    addr = p.get("address", "")
                    jibun = p.get("jibun_address", "")
                    full_addr = (jibun or addr).strip()  # jibun이 행정동+지번 완전주소라 중복 방지

                    key = (p.get("name", ""), p.get("address", "") or p.get("jibun_address", ""))
                    generated_post = post_map.get(key)

                    child = QTreeWidgetItem(parent)
                    child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
                    child.setCheckState(0, Qt.Unchecked)
                    is_posted = bool(generated_post and generated_post.get("posted", False))
                    if is_posted:
                        child.setText(0, f"[완] {num}. {p.get('name', '')}")
                        child.setForeground(0, QColor("#3b82f6"))
                        gen_count += 1
                        item_to_post[id(child)] = generated_post
                    elif generated_post:
                        child.setText(0, f"● {num}. {p.get('name', '')}")
                        child.setForeground(0, QColor("#22c55e"))
                        gen_count += 1
                        item_to_post[id(child)] = generated_post
                    else:
                        child.setText(0, f"● {num}. {p.get('name', '')}")
                        child.setForeground(0, QColor("#ef4444"))
                    child.setText(1, full_addr)
                    child.setText(2, p.get("category", ""))
                    child.setText(3, p.get("nearby_station", ""))
                    child.setText(4, p.get("front_keywords", ""))
                    child.setText(5, p.get("tags", ""))
                    all_children.append((child, p))
                    tab_count += 1

                # 폴더 색
                if len(places) > 0:
                    if gen_count == len(places):
                        parent.setForeground(0, QColor("#22c55e"))
                    elif gen_count == 0:
                        parent.setForeground(0, QColor("#ef4444"))
                    else:
                        parent.setForeground(0, QColor("#f59e0b"))
                    parent.setText(0, f"●  {group_name}  ({gen_count}/{len(places)} 생성)")
            return tw, tab_count

        tree_r, count_r = _build_tree(groups_region)
        tree_k, count_k = _build_tree(groups_keyword)
        all_trees.extend([tree_r, tree_k])

        tab_widget.addTab(tree_r, f"지역기반 ({count_r}개)")
        tab_widget.addTab(tree_k, f"키워드기반 ({count_k}개)")

        layout.addWidget(tab_widget)

        # 호환성을 위해 tree 참조는 현재 활성 탭의 트리로 유지
        tree = tree_r

        def on_double_click(item, col):
            post = item_to_post.get(id(item))
            if not post:
                return
            self._show_single_post_view(post)

        for _tw in all_trees:
            _tw.itemDoubleClicked.connect(on_double_click)

        # 부모 체크 시 하위 전체 토글
        def _make_on_changed(_tw):
            def _h(item, col):
                if col != 0:
                    return
                _tw.blockSignals(True)
                state = item.checkState(0)
                if item.childCount() > 0:
                    for i in range(item.childCount()):
                        item.child(i).setCheckState(0, state)
                _tw.blockSignals(False)
            return _h
        for _tw in all_trees:
            _tw.itemChanged.connect(_make_on_changed(_tw))

        def delete_selected():
            # F8 선택 삭제: 업체 + 생성된 포스트 모두 제거 (크롤링 목록에서도 사라짐)
            checked_places = [p for (c, p) in all_children if c.checkState(0) == Qt.Checked]
            if not checked_places:
                QMessageBox.information(dlg, "안내", "선택된 항목이 없습니다.")
                return
            reply = QMessageBox.question(dlg, "삭제 확인", f"{len(checked_places)}개 항목을 완전 삭제할까요?\n(업체 + 생성된 포스트 모두 제거, F7에서도 사라짐)")
            if reply != QMessageBox.Yes:
                return
            keys = {self._place_key(p) for p in checked_places}
            self._generated_posts = [gp for gp in self._generated_posts if self._place_key(gp.get("place", {})) not in keys]
            self._save_generated_posts()
            self._purge_places_from_logs(checked_places)
            QMessageBox.information(dlg, "완료", f"{len(checked_places)}개 완전 삭제됨.")
            dlg.accept()

        def delete_all():
            # F8 전체 삭제: 표시된 모든 항목(업체+포스트) 완전 제거
            self._del_log(f"=== F8 delete_all 클릭 (all_children={len(all_children)}개) ===")
            all_places = [p for (_c, p) in all_children]
            if not all_places:
                self._del_log("F8 delete_all: all_places 비어있음 → 종료")
                return
            reply = QMessageBox.question(dlg, "전체 삭제 확인", f"표시된 전체 {len(all_places)}개 항목을 완전 삭제할까요?\n(업체 + 생성된 포스트 모두 제거)")
            if reply != QMessageBox.Yes:
                self._del_log("F8 delete_all: 사용자 취소")
                return
            self._del_log(f"F8 delete_all: 사용자 확인 — {len(all_places)}개 삭제 진행")
            keys = {self._place_key(p) for p in all_places}
            self._generated_posts = [gp for gp in self._generated_posts if self._place_key(gp.get("place", {})) not in keys]
            self._save_generated_posts()
            removed = self._purge_places_from_logs(all_places)
            self._del_log(f"F8 delete_all: _purge가 {removed}개 제거 보고")
            QMessageBox.information(dlg, "완료", f"{len(all_places)}개 완전 삭제됨.")
            dlg.accept()

        def select_all_f8():
            # 현재 활성 탭만 전체 선택
            cur_tree = all_trees[tab_widget.currentIndex()] if all_trees else None
            if cur_tree is None: return
            cur_tree.blockSignals(True)
            for i in range(cur_tree.topLevelItemCount()):
                p = cur_tree.topLevelItem(i)
                p.setCheckState(0, Qt.Checked)
                for j in range(p.childCount()):
                    p.child(j).setCheckState(0, Qt.Checked)
            cur_tree.blockSignals(False)

        def select_none_f8():
            # 현재 활성 탭만 전체 해제
            cur_tree = all_trees[tab_widget.currentIndex()] if all_trees else None
            if cur_tree is None: return
            cur_tree.blockSignals(True)
            for i in range(cur_tree.topLevelItemCount()):
                p = cur_tree.topLevelItem(i)
                p.setCheckState(0, Qt.Unchecked)
                for j in range(p.childCount()):
                    p.child(j).setCheckState(0, Qt.Unchecked)
            cur_tree.blockSignals(False)

        btn_all_sel.clicked.connect(select_all_f8)
        btn_all_none.clicked.connect(select_none_f8)

        # 랜덤 포스팅 순서 토글 — (v1.6.6) "첫 포스팅 지금 바로 시작" 체크박스 제거. 자동 분기로 통합
        opt_row = QHBoxLayout()
        info_auto = QLabel("기존 예약 있으면 이어쓰고, 없으면 첫 글 즉시발행 + 나머지 예약")
        info_auto.setStyleSheet("font-size: 11px; color: #64748b;")
        opt_row.addWidget(info_auto)
        opt_row.addStretch()
        random_cb = ToggleSwitch()
        random_lbl = QLabel("랜덤포스팅")
        random_lbl.setStyleSheet("font-size: 12px; color: #334155; margin-left: 6px;")
        opt_row.addWidget(random_cb)
        opt_row.addWidget(random_lbl)
        layout.addLayout(opt_row)

        def post_selected():
            targets = []
            # 모든 탭의 체크된 항목 모음
            for _tw in all_trees:
                for i in range(_tw.topLevelItemCount()):
                    parent = _tw.topLevelItem(i)
                    for j in range(parent.childCount()):
                        child = parent.child(j)
                        if child.checkState(0) == Qt.Checked:
                            post = item_to_post.get(id(child))
                            if post:
                                targets.append(post)
            if not targets:
                QMessageBox.information(dlg, "안내", "포스팅할 생성된 포스트를 선택해주세요.")
                return
            if random_cb.isChecked():
                import random as _r
                _r.shuffle(targets)
            self.posting_targets = targets
            dlg.accept()
            self._post_generated()

        def delete_done():
            # 발행 완료([완]) 상태인 포스트 + 그 업체를 크롤링 목록에서도 일괄 제거
            done_pairs = []
            for (c, p) in all_children:
                gp = item_to_post.get(id(c))
                if gp and gp.get("posted"):
                    done_pairs.append(p)
            if not done_pairs:
                QMessageBox.information(dlg, "안내", "완료된 항목이 없습니다.")
                return
            reply = QMessageBox.question(dlg, "완료 목록 삭제", f"발행 완료된 {len(done_pairs)}개 항목을 완전 삭제할까요?\n(업체 + 생성된 포스트 모두 제거)")
            if reply != QMessageBox.Yes:
                return
            keys = {self._place_key(p) for p in done_pairs}
            self._generated_posts = [gp for gp in self._generated_posts if self._place_key(gp.get("place", {})) not in keys]
            self._save_generated_posts()
            self._purge_places_from_logs(done_pairs)
            QMessageBox.information(dlg, "완료", f"완료된 {len(done_pairs)}개 항목 삭제됨.")
            dlg.accept()

        btn_row = QHBoxLayout()
        btn_delete = QPushButton("선택 삭제")
        btn_delete.setStyleSheet("background: #ef4444; color: white; border: none; border-radius: 8px; padding: 8px 20px;")
        btn_delete.setCursor(Qt.PointingHandCursor)
        btn_delete.clicked.connect(delete_selected)
        btn_del_done = QPushButton("완료 목록 삭제")
        btn_del_done.setStyleSheet("background: #3b82f6; color: white; border: none; border-radius: 8px; padding: 8px 20px;")
        btn_del_done.setCursor(Qt.PointingHandCursor)
        btn_del_done.clicked.connect(delete_done)
        btn_del_all = QPushButton("전체 삭제")
        btn_del_all.setStyleSheet("background: #dc2626; color: white; border: none; border-radius: 8px; padding: 8px 20px;")
        btn_del_all.setCursor(Qt.PointingHandCursor)
        btn_del_all.clicked.connect(delete_all)
        btn_post = QPushButton("자동 포스팅")
        btn_post.setStyleSheet("background: #3b82f6; color: white; border: none; border-radius: 8px; font-weight: bold; padding: 8px 24px;")
        btn_post.setCursor(Qt.PointingHandCursor)
        btn_post.clicked.connect(post_selected)
        btn_close = QPushButton("닫기")
        btn_close.setStyleSheet("padding: 8px 20px;")
        btn_close.clicked.connect(dlg.reject)
        btn_row.addWidget(btn_delete)
        btn_row.addWidget(btn_del_done)
        btn_row.addWidget(btn_del_all)
        btn_row.addStretch()
        btn_row.addWidget(btn_post)
        btn_row.addWidget(btn_close)
        layout.addLayout(btn_row)

        dlg.exec()

    def _show_single_post_view(self, post: dict):
        """단일 포스트 내용 보기"""
        from PySide6.QtWidgets import QTextEdit, QLineEdit, QScrollArea
        from PySide6.QtGui import QPixmap
        from PySide6.QtCore import Qt
        place = post.get("place", {})
        content = post.get("content", {})

        dlg = QDialog(self)
        dlg.setWindowTitle(place.get("name", "포스트"))
        dlg.resize(1000, 750)
        layout = QVBoxLayout(dlg)

        title_edit = QLineEdit(content.get("title", ""))
        title_edit.setStyleSheet("font-weight: bold; font-size: 14px; padding: 8px;")
        layout.addWidget(title_edit)

        # ── 이미지 썸네일 패널 ──
        img_scroll = QScrollArea()
        img_scroll.setFixedHeight(160)
        img_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        img_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        img_scroll.setWidgetResizable(True)
        img_container = QWidget()
        img_hbox = QHBoxLayout(img_container)
        img_hbox.setContentsMargins(4, 4, 4, 4)
        img_hbox.setSpacing(6)
        img_hbox.addStretch()
        img_scroll.setWidget(img_container)
        layout.addWidget(img_scroll)

        loading_lbl = QLabel("이미지 불러오는 중...")
        loading_lbl.setStyleSheet("color: #94a3b8; font-size: 12px; padding: 4px;")
        layout.addWidget(loading_lbl)

        def _add_image_labels(paths):
            # stretch 제거 후 이미지 추가
            item = img_hbox.takeAt(img_hbox.count() - 1)
            if item:
                del item
            for path in paths:
                lbl = QLabel()
                lbl.setFixedHeight(140)
                lbl.setFixedWidth(200)
                lbl.setAlignment(Qt.AlignCenter)
                lbl.setStyleSheet("border: 1px solid #e2e8f0; border-radius: 6px; background: #f8fafc;")
                px = QPixmap(path)
                if not px.isNull():
                    lbl.setPixmap(px.scaled(200, 140, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                else:
                    lbl.setText("로드 실패")
                def _open_full(p=path):
                    fdlg = QDialog(dlg)
                    fdlg.setWindowTitle(os.path.basename(p))
                    fv = QVBoxLayout(fdlg)
                    fl = QLabel()
                    fpx = QPixmap(p)
                    fl.setPixmap(fpx.scaled(900, 700, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                    fv.addWidget(fl)
                    fdlg.exec()
                lbl.mousePressEvent = lambda e, p=path: _open_full(p)
                lbl.setCursor(Qt.PointingHandCursor)
                img_hbox.addWidget(lbl)
            img_hbox.addStretch()
            if paths:
                loading_lbl.hide()
            else:
                loading_lbl.setText("이미지 없음")

        class _ImgLoader(QThread):
            done = Signal(list)
            def __init__(self, place_, content_, cfg_, parent_=None):
                super().__init__()
                self.place_ = place_
                self.content_ = content_
                self.cfg_ = cfg_
                self.parent_ = parent_
            def run(self):
                paths = [p for p in (self.content_.get("image_paths") or []) if p and os.path.exists(p)]
                if not paths:
                    pix_keys = [k for k in self.cfg_.get("pixabay_key_list", []) if k]
                    if pix_keys:
                        try:
                            import image_handler as _ih
                            # search_keyword 마지막 토큰 → category 마지막 토큰 순으로 시도
                            biz = self.parent_._best_biz_term(self.place_, "") if self.parent_ else (self.place_.get("category") or "").strip()
                            name_ = self.place_.get("name", "")
                            img_count = int(self.content_.get("image_count", 3) or 3)
                            translator = self.parent_._translate_ko_to_en if self.parent_ else None
                            paths = _ih.download_images(pix_keys[0], biz, img_count, watermark_text=name_, translator=translator)
                        except Exception:
                            paths = []
                self.done.emit(paths)

        cfg_ = load_config()
        _loader = _ImgLoader(place, content, cfg_, parent_=self)
        _loader.done.connect(_add_image_labels)
        _loader.start()
        dlg._img_loader = _loader  # GC 방지

        body_edit = QTextEdit()
        body_edit.setPlainText(content.get("body", ""))
        layout.addWidget(body_edit)

        meta = QLabel(
            f"주소: {place.get('jibun_address','') or place.get('address','')}  |  "
            f"카테고리: {place.get('category','')}  |  "
            f"근처역: {place.get('nearby_station','')}  |  "
            f"태그: {','.join(content.get('tags', [])) if isinstance(content.get('tags'), list) else content.get('tags', '')}"
        )
        meta.setStyleSheet("color: #475569; padding: 5px; font-size: 12px;")
        meta.setWordWrap(True)
        layout.addWidget(meta)

        btn_row = QHBoxLayout()
        btn_save = QPushButton("저장")
        btn_save.setStyleSheet("background: #22c55e; color: white; border: none; border-radius: 6px; padding: 8px 20px;")
        def save():
            content["title"] = title_edit.text()
            content["body"] = body_edit.toPlainText()
            self._save_generated_posts()
            QMessageBox.information(dlg, "저장", "저장되었습니다.")
        btn_save.clicked.connect(save)
        btn_close = QPushButton("닫기")
        btn_close.setStyleSheet("padding: 8px 20px;")
        btn_close.clicked.connect(dlg.reject)
        btn_row.addWidget(btn_save)
        btn_row.addStretch()
        btn_row.addWidget(btn_close)
        layout.addLayout(btn_row)

        dlg.exec()

    # ── 포스트 생성 ──
    def _account_key(self) -> str:
        # 블로그마스터 로그인 사용자(app_user) 기준 — 같은 app_user면 모든 네이버 슬롯 데이터 공유
        try:
            from config import get_current_user as _gcu
            uid = _gcu() or "default"
        except Exception:
            uid = "default"
        return "".join(c for c in uid if c.isalnum() or c in "-_") or "default"

    def _refresh_status_panes(self):
        # 라벨에 숫자 표시 안 함 — 실시간 로그는 아래 crawl_log/post_log에서 보임
        pass

    def _get_crawled_file(self):
        app_user = self._account_key()
        try:
            idx = self.cfg.get("active_account", 0)
            bid = (self.cfg.get("accounts", [{}])[idx].get("blog_id", "") or "").strip()
            bid = "".join(c for c in bid if c.isalnum() or c in "-_")
            if not bid:
                bid = f"acc{idx}"
        except Exception:
            bid = "default"
        return os.path.join(os.path.dirname(__file__), f"crawled_{app_user}_{bid}.json")

    def _save_crawled(self, items: list, keyword: str = ""):
        mode = getattr(self, "_current_crawl_mode", "")
        try:
            with open(self._get_crawled_file(), "w", encoding="utf-8") as f:
                json.dump({"keyword": keyword, "crawl_mode": mode, "items": items}, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self._emit_log(f"crawled 저장 실패: {e}")

    def _load_crawled(self) -> list:
        fp = self._get_crawled_file()
        if not os.path.exists(fp):
            return []
        try:
            with open(fp, "r", encoding="utf-8") as f:
                d = json.load(f)
            return d.get("items", []) if isinstance(d, dict) else (d or [])
        except Exception:
            return []

    def _load_crawled_meta(self) -> dict:
        """저장된 크롤 데이터 + 키워드 반환. {'keyword': str, 'items': list}"""
        fp = self._get_crawled_file()
        if not os.path.exists(fp):
            return {"keyword": "", "items": []}
        try:
            with open(fp, "r", encoding="utf-8") as f:
                d = json.load(f)
            if isinstance(d, dict):
                return {"keyword": d.get("keyword", ""), "items": d.get("items", []) or []}
            return {"keyword": "", "items": d or []}
        except Exception:
            return {"keyword": "", "items": []}

    def _del_log(self, msg: str):
        """삭제 진단 로그 — delete_debug.log에 timestamp 포함으로 기록"""
        try:
            import datetime as _dt
            fp = os.path.join(os.path.dirname(__file__), "delete_debug.log")
            with open(fp, "a", encoding="utf-8") as f:
                f.write(f"[{_dt.datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        except Exception:
            pass

    # ── 삭제: deleted_keys 시스템 제거. logs 파일에서 즉시 영구 삭제 ──
    def _purge_places_from_logs(self, places_to_remove: list):
        """logs 폴더의 모든 places_*.json/{kw}.json 파일에서 (이름,주소) 매칭 항목 제거.
        파일이 비면 통째로 삭제."""
        self._del_log(f"_purge 호출 (places={len(places_to_remove or [])}개)")
        if not places_to_remove:
            self._del_log("_purge: places_to_remove 비어있음 → 0 반환")
            return 0
        keys = {self._place_key(p) for p in places_to_remove}
        self._del_log(f"_purge: {len(keys)}개 고유 key. 첫 key={list(keys)[0] if keys else None}")
        log_dir = self._get_logs_dir()
        self._del_log(f"_purge: log_dir={log_dir}")
        if not os.path.isdir(log_dir):
            self._del_log("_purge: log_dir 없음 → 0 반환")
            return 0
        removed_total = 0
        all_files = os.listdir(log_dir)
        valid_files = [f for f in all_files if f.endswith(".json") and not f.startswith("._")]
        self._del_log(f"_purge: 전체 {len(all_files)}개 / 유효 .json {len(valid_files)}개")
        for fname in valid_files:
            fp = os.path.join(log_dir, fname)
            try:
                with open(fp, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if not isinstance(data, dict):
                    self._del_log(f"  {fname}: dict 아님 → skip")
                    continue
                items = data.get("items", []) or []
                kept = [p for p in items if self._place_key(p) not in keys]
                if len(kept) == len(items):
                    self._del_log(f"  {fname}: items {len(items)}개 — 매칭 0 → skip")
                    continue
                removed_total += (len(items) - len(kept))
                if not kept:
                    try:
                        os.remove(fp)
                        self._del_log(f"  {fname}: items {len(items)}개 모두 삭제 → 파일 삭제")
                        pass
                    except Exception as e:
                        self._del_log(f"  {fname}: 파일 삭제 실패: {e}")
                else:
                    data["items"] = kept
                    with open(fp, "w", encoding="utf-8") as f:
                        json.dump(data, f, ensure_ascii=False, indent=2)
                    self._del_log(f"  {fname}: items {len(items)}→{len(kept)} 재기록")
                    pass
            except Exception as e:
                self._del_log(f"  {fname}: 예외 {type(e).__name__}: {e}")
                continue
        # 메모리 self.crawled_data도 동기화
        try:
            before_cnt = len(self.crawled_data or [])
            self.crawled_data = [p for p in (self.crawled_data or []) if self._place_key(p) not in keys]
            self._save_crawled(self.crawled_data, getattr(self, "last_keyword", ""))
            self._del_log(f"_purge: crawled_data {before_cnt}→{len(self.crawled_data)} 동기화")
        except Exception as e:
            self._del_log(f"_purge: crawled_data 동기화 실패 {type(e).__name__}: {e}")
        self._del_log(f"_purge 완료: 총 {removed_total}개 제거")
        return removed_total

    def _place_key(self, place: dict) -> tuple:
        return (place.get("name", ""), place.get("address", "") or place.get("jibun_address", ""))

    def _get_posts_file(self):
        # 계정별로 분리 저장
        per_acc = os.path.join(os.path.dirname(__file__), f"generated_posts_{self._account_key()}.json")
        legacy = os.path.join(os.path.dirname(__file__), "generated_posts.json")
        # 최초 전환 시 공용 파일 → 현재 계정 파일로 자동 승격 (1회)
        if not os.path.exists(per_acc) and os.path.exists(legacy):
            try:
                import shutil
                shutil.copyfile(legacy, per_acc)
            except Exception:
                pass
        return per_acc

    def _get_logs_dir(self) -> str:
        app_user = self._account_key()
        try:
            idx = self.cfg.get("active_account", 0)
            bid = (self.cfg.get("accounts", [{}])[idx].get("blog_id", "") or "").strip()
            bid = "".join(c for c in bid if c.isalnum() or c in "-_")
            if not bid:
                bid = f"acc{idx}"
        except Exception:
            bid = "default"
        d = os.path.join(os.path.dirname(__file__), "logs", f"{app_user}_{bid}")
        os.makedirs(d, exist_ok=True)
        return d

    def _migrate_legacy_logs(self):
        """logs/*.json 을 1번 계정 폴더로 1회 이동 (구버전 데이터 복구용)"""
        import shutil
        base = os.path.join(os.path.dirname(__file__), "logs")
        if not os.path.isdir(base):
            return
        legacy = [f for f in os.listdir(base) if f.endswith(".json") and not f.startswith("._") and os.path.isfile(os.path.join(base, f))]
        if not legacy:
            return
        accounts = self.cfg.get("accounts", [])
        target_idx = 0
        for i, acc in enumerate(accounts):
            if (acc.get("blog_id") or "").strip():
                target_idx = i
                break
        bid = (accounts[target_idx].get("blog_id") if target_idx < len(accounts) else "") or f"acc{target_idx}"
        safe = "".join(c for c in bid if c.isalnum() or c in "-_") or "default"
        target_dir = os.path.join(base, safe)
        os.makedirs(target_dir, exist_ok=True)
        moved = 0
        for f in legacy:
            src = os.path.join(base, f)
            dst = os.path.join(target_dir, f)
            if os.path.exists(dst):
                continue
            try:
                shutil.move(src, dst)
                moved += 1
            except Exception:
                pass
        if moved:
            self._emit_log(f"기존 크롤 결과 {moved}개를 '{bid}' 계정으로 이동했습니다")

    def _save_generated_posts(self):
        data = []
        for item in self._generated_posts:
            data.append({
                "place": item["place"],
                "content": item["content"],
                "posted": item.get("posted", False),
            })
        with open(self._get_posts_file(), "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        try:
            self._cleanup_orphan_image_folders()
        except Exception:
            pass

    # ── 주간 글쓰기 한도 (공유키 유저: 네이버 아이디당 주 200개, 관리자 제외) ──
    WEEKLY_GEN_LIMIT = 200

    def _weekly_limit_active(self) -> bool:
        """이 유저에게 주간 한도를 적용해야 하나 — 관리자 제외, 공용키(shared) 사용 시에만."""
        user = getattr(self, "current_user", {}) or {}
        if user.get("role") == "admin":
            return False
        try:
            return bool((self.cfg or {}).get("_has_shared_keys"))
        except Exception:
            return False

    def _week_key(self) -> str:
        """ISO 주(월요일 시작) 식별자 — 예: '2026-W27'."""
        import datetime as _dt
        y, w, _ = _dt.date.today().isocalendar()
        return f"{y}-W{w:02d}"

    def _weekly_count_path(self) -> str:
        return os.path.join(os.path.dirname(__file__), "weekly_gen_count.json")

    def _load_weekly_counts(self) -> dict:
        """{week_key: {blog_id: count}} — 이번 주 데이터만 유지."""
        try:
            from app_paths import safe_load_json as _slj
            data = _slj(self._weekly_count_path(), default={}, max_mb=5) or {}
        except Exception:
            data = {}
        wk = self._week_key()
        return {wk: data.get(wk, {})} if isinstance(data, dict) else {wk: {}}

    def _weekly_used(self, blog_id: str) -> int:
        return int(self._load_weekly_counts().get(self._week_key(), {}).get(blog_id or "", 0))

    def _weekly_remaining(self, blog_id: str) -> int:
        return max(0, self.WEEKLY_GEN_LIMIT - self._weekly_used(blog_id))

    def _weekly_add(self, blog_id: str, n: int):
        """이번 주 blog_id 카운트 n 증가 (원자적, 병렬 안전은 lock으로 감싸 호출)."""
        if not blog_id or n <= 0:
            return
        import json as _json
        wk = self._week_key()
        data = self._load_weekly_counts()
        bucket = data.get(wk, {})
        bucket[blog_id] = int(bucket.get(blog_id, 0)) + int(n)
        data[wk] = bucket
        try:
            with open(self._weekly_count_path(), "w", encoding="utf-8") as f:
                _json.dump(data, f, ensure_ascii=False)
        except Exception:
            pass

    def _load_generated_posts(self) -> list:
        filepath = self._get_posts_file()
        if not os.path.exists(filepath):
            return []
        # ★ 근본 대비: 파일이 비정상적으로 크거나(폭주) 깨졌으면 백업 후 초기화 → 앱이 멈추지 않음
        try:
            if os.path.getsize(filepath) > 60 * 1024 * 1024:  # 60MB 초과 = 비정상
                try:
                    os.replace(filepath, filepath + ".corrupt.bak")
                except Exception:
                    pass
                try:
                    self._emit_post_log("⚠ 생성글 파일이 비정상적으로 커서 초기화했습니다 (.corrupt.bak 백업).")
                except Exception:
                    pass
                return []
        except Exception:
            pass
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data if isinstance(data, list) else []
        except Exception:
            # 손상된 파일 → 백업 후 초기화 (다음 저장 때 새로 생성)
            try:
                os.replace(filepath, filepath + ".corrupt.bak")
            except Exception:
                pass
            return []

    def _generate_posts(self):
        if self._block_if_expired():
            return
        if self._block_if_api_expired():   # API 기간 만료면 글 생성 차단
            return
        if not self._check_api_keys():
            return
        # 크롤 결과 파일에서 스냅샷 로드 (크롤이 동시에 돌아도 안 꼬이게)
        snap = self._load_crawled()
        if snap:
            self.crawled_data = snap
        # 업체 선택
        selected = self._show_posting_selector()
        if not selected:
            return

        cfg = load_config()
        provider = "GPT"  # 하위호환 인자
        gpt_keys = [k for k in cfg.get("gpt_key_list", []) if k]
        ds_keys = [k for k in cfg.get("deepseek_key_list", []) if k]
        gm_keys = [k for k in cfg.get("gemini_key_list", []) if k]
        ai_engine = (cfg.get("ai_engine") or "deepseek").strip().lower()
        # 선택 엔진의 키가 있는지 확인 → 없으면 안내
        _engkey = {"deepseek": ds_keys, "gpt": gpt_keys, "gemini": gm_keys}.get(ai_engine, [])
        if not _engkey:
            _engname = {"deepseek": "딥시크", "gpt": "챗GPT", "gemini": "제미나이"}.get(ai_engine, ai_engine)
            QMessageBox.critical(self, "API 키 필요",
                f"선택하신 글쓰기 엔진({_engname})의 API 키가 없습니다.\n\n"
                f"설정 → API 키에서 {_engname} 키를 입력하거나, 다른 엔진을 선택해주세요.")
            return

        deepseek_key = ds_keys[0] if ds_keys else None
        gpt_key = gpt_keys[0] if gpt_keys else None
        gemini_key = gm_keys[0] if gm_keys else None
        api_key = (gpt_keys or ds_keys or gm_keys)[0]  # 검색어 추출기용
        # AI 검색어 추출기(gpt-4o-mini)에 키 주입 — 실패/타임아웃 시 하드코딩 폴백
        try:
            import image_handler as _ih
            _ih.configure_ai_extractor(api_key, os.path.join(os.path.dirname(__file__), "pixabay_query_cache.json"))
        except Exception:
            pass
        _lkm = self.cfg.get("last_keyword_by_account", {}) or {}
        keyword = self.keyword_input.currentText().strip() or getattr(self, 'last_keyword', '') or _lkm.get(self._current_blog_id(), "") or self.cfg.get("last_keyword", "")
        # 크롤된 데이터에서 키워드 자동 복원 (앞키워드 첫 항목이 "{지역}{업종}" 형식)
        if not keyword and selected:
            first = selected[0]
            fk = first.get("front_keywords", "")
            if fk:
                keyword = fk.split(",")[0].strip()
                self._emit_post_log(f"크롤 데이터에서 키워드 복원: '{keyword}'")
        # ── 주간 한도 적용 (공유키 유저: 네이버 아이디당 주 200개) ──
        if self._weekly_limit_active():
            _bid_lim = self._current_blog_id() or self._active_blog_id() or ""
            _remain = self._weekly_remaining(_bid_lim)
            if _remain <= 0:
                QMessageBox.warning(self, "주간 글쓰기 한도",
                    f"이 아이디는 이번 주 글쓰기 한도({self.WEEKLY_GEN_LIMIT}개)를 모두 사용했습니다.\n"
                    f"다음 주에 다시 이용하실 수 있습니다.")
                return
            if len(selected) > _remain:
                selected = selected[:_remain]
                QMessageBox.information(self, "주간 글쓰기 한도",
                    f"이번 주 남은 글쓰기 한도가 {_remain}개입니다.\n{_remain}개만 생성합니다. (아이디당 주 {self.WEEKLY_GEN_LIMIT}개)")

        total = len(selected)

        self._emit_status("포스트 생성 중...", "#8b5cf6")
        self._emit_post_log(f"포스트 생성 시작: {total}개 업체")
        if self._weekly_limit_active():
            _bidL = self._current_blog_id() or self._active_blog_id() or ""
            self._emit_post_log(f"(이번 주 사용 {self._weekly_used(_bidL)}/{self.WEEKLY_GEN_LIMIT} · 남음 {self._weekly_remaining(_bidL)})")

        self._generating_data = {"selected": selected, "keyword": keyword, "api_key": api_key, "provider": provider}
        self._generated_posts = self._load_generated_posts()

        self.is_generating = True
        self.stop_flag = False

        def _worker():
            from concurrent.futures import ThreadPoolExecutor, as_completed
            import threading
            lock = threading.Lock()
            done_count = [0]
            new_posts = []

            pix_keys = [k for k in cfg.get("pixabay_key_list", []) if k]

            def _gen_one(idx, place):
                if self.stop_flag:
                    return
                name = place.get("name", "")
                try:
                    content = generate_content(
                        provider=provider,
                        api_key=api_key,
                        place=place,
                        keyword=keyword,
                        prompt_override=getattr(self, "_override_prompt_name", None),
                        title_prefix=getattr(self, "_override_title_prefix", None),
                        deepseek_key=deepseek_key,
                        gpt_key=gpt_key,
                        gemini_key=gemini_key,
                        engine=ai_engine,
                    )
                    # 제미나이(무료 한도)면 다음 글 전 딜레이 — 분당 호출 제한 회피
                    if ai_engine == "gemini":
                        import time as _tgm
                        _tgm.sleep(5)
                    if self.stop_flag:
                        return

                    # 이미지: 네이버 실사 1장 + Pixabay 나머지 (중복 방지)
                    img_paths = []
                    if pix_keys:
                        try:
                            # 업종 우선순위: search_keyword 마지막 토큰 > 카테고리 마지막 토큰 > 검색어
                            biz = self._best_biz_term(place, keyword)
                            img_count = int(content.get("image_count", 3) or 3)
                            pkey = self._place_key(place)
                            safe_name = "".join(c for c in (pkey[0] + "_" + pkey[1]) if c.isalnum() or c in "-_")[:80] or "unknown"
                            persist_dir = os.path.join(
                                os.path.dirname(__file__), "saved_images",
                                self._account_key(), safe_name
                            )
                            os.makedirs(persist_dir, exist_ok=True)
                            for _f in os.listdir(persist_dir):
                                try: os.remove(os.path.join(persist_dir, _f))
                                except: pass

                            # 1) 실사 수집 OFF — 글마다 크롬 띄우는 비용 제거(속도), Pixabay만 사용
                            real_paths = []
                            pid = (place.get("place_id") or "").strip()
                            if False:  # 실사 수집 비활성화
                                try:
                                    import real_photos as _rp
                                    from selenium import webdriver as _wd
                                    from selenium.webdriver.chrome.options import Options as _Opt
                                    from selenium.webdriver.chrome.service import Service as _Svc
                                    from webdriver_manager.chrome import ChromeDriverManager as _CM
                                    _opt = _Opt()
                                    _opt.add_argument("--headless=new")
                                    _opt.add_argument("--disable-blink-features=AutomationControlled")
                                    _opt.add_argument("--no-sandbox")
                                    _opt.add_argument("--disable-gpu")
                                    _opt.add_argument("--window-size=1920,1080")
                                    _opt.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36")
                                    _opt.add_experimental_option("excludeSwitches", ["enable-automation"])
                                    _drv = _wd.Chrome(service=_Svc(_CM().install()), options=_opt)
                                    _drv.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
                                        "source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
                                    })
                                    try:
                                        real_paths = _rp.pick_real_photos(
                                            _drv, pid, place.get("name", ""),
                                            target_real=1, out_dir=persist_dir,
                                            emit_log=self._emit_post_log,
                                        )
                                    finally:
                                        _drv.quit()
                                except Exception as _e:
                                    self._emit_post_log(f"실사 수집 실패 ({name}): {_e}")
                            img_paths.extend(real_paths)

                            # 2) Pixabay 폴백 — 중복 방지 ID 파일 계정별로
                            need = max(0, img_count - len(img_paths))
                            if need > 0:
                                try:
                                    import image_handler as _ih
                                    _used_file = os.path.join(
                                        os.path.dirname(__file__), "saved_images",
                                        self._account_key(), "used_pixabay_ids.json",
                                    )
                                    _ih.configure_used_ids_file(_used_file)
                                    self._emit_post_log(f"  Pixabay 보충 {need}장 (업종: {biz})")
                                    _tmp = _ih.download_images(pix_keys[0], biz, need,
                                                                watermark_text=name,
                                                                translator=self._translate_ko_to_en)
                                    import shutil as _sh
                                    for _i, _p in enumerate(_tmp):
                                        _dst = os.path.join(persist_dir, f"pix_{_i+1}.jpg")
                                        try:
                                            _sh.copyfile(_p, _dst)
                                            img_paths.append(_dst)
                                        except Exception:
                                            pass
                                except Exception as _e:
                                    self._emit_post_log(f"Pixabay 실패 ({name}): {_e}")
                        except Exception as _ie:
                            self._emit_post_log(f"이미지 저장 실패 ({name}): {_ie}")

                    content["image_paths"] = img_paths

                    with lock:
                        new_post = {"place": place, "content": content, "posted": False}
                        new_posts.append(new_post)
                        done_count[0] += 1
                        # 주간 한도 카운트 증가 (공유키 유저만)
                        if self._weekly_limit_active():
                            try:
                                self._weekly_add(self._current_blog_id() or self._active_blog_id() or "", 1)
                            except Exception:
                                pass
                        # 1개 생성 완료마다 즉시 디스크 저장 — F8에서 실시간 확인 가능
                        try:
                            _new_key = (place.get("name", ""), place.get("address", "") or place.get("jibun_address", ""))
                            _existing = [gp for gp in (self._generated_posts or [])
                                         if (gp.get("place", {}).get("name", ""),
                                             gp.get("place", {}).get("address", "") or gp.get("place", {}).get("jibun_address", "")) != _new_key]
                            self._generated_posts = _existing + [new_post]
                            self._save_generated_posts()
                        except Exception as _se:
                            self._emit_post_log(f"즉시 저장 실패 ({name}): {_se}")
                        self._emit_post_log(f"[{done_count[0]}/{total}] '{name}' 생성 완료 (이미지 {len(img_paths)}장)")
                        self._emit_status(f"생성 {done_count[0]}/{total}", "#8b5cf6")
                        self._emit_post_count(f"생성 {done_count[0]}/{total}")
                except Exception as e:
                    _emsg = str(e)
                    with lock:
                        done_count[0] += 1
                        self._emit_post_log(f"[{done_count[0]}/{total}] '{name}' 생성 실패: {e}")
                        # API 토큰(잔액) 부족 → 팝업 1회 + 생성 중단 (같은 오류 스팸 방지, 자기키/공용키 공통)
                        _low = _emsg.lower()
                        if (("토큰(잔액)이 부족" in _emsg) or "insufficient_quota" in _low
                                or "insufficient balance" in _low or ("quota" in _low and "exceed" in _low)) \
                                and not getattr(self, "_balance_err_shown", False):
                            self._balance_err_shown = True
                            self.stop_flag = True
                            try:
                                from PySide6.QtCore import QMetaObject as _QM
                                _QM.invokeMethod(self, "_show_api_balance_error", Qt.QueuedConnection)
                            except Exception:
                                pass

            self._balance_err_shown = False  # 이번 생성 세션 잔액오류 팝업 1회 제한 초기화
            # 제미나이(무료 한도)는 병렬 3 + 글당 딜레이로 분당 호출 제한 회피, 딥시크/GPT는 5
            max_workers = 3 if ai_engine == "gemini" else 5
            with ThreadPoolExecutor(max_workers=max_workers) as ex:
                futures = [ex.submit(_gen_one, i, p) for i, p in enumerate(selected, 1)]
                for f in as_completed(futures):
                    if self.stop_flag:
                        for fu in futures:
                            fu.cancel()
                        break

            # 스마트 병합: 새로 생성된 키만 교체, 나머지 기존 포스트 보존
            def _k(p):
                pl = p.get("place", {})
                return (pl.get("name", ""), pl.get("address", "") or pl.get("jibun_address", ""))
            def _kname(p):
                return p.get("place", {}).get("name", "")
            new_keys = {_k(p) for p in new_posts}
            # 기존 포스트 중 이번에 새로 생성된 키와 겹치지 않는 것 보존
            remaining = [gp for gp in (self._generated_posts or []) if _k(gp) not in new_keys]
            merged = remaining + new_posts
            # 업체명 기준 중복 제거 (새로 생성된 것 우선)
            name_seen = {}
            deduped = []
            for p in reversed(merged):
                n = _kname(p)
                if n not in name_seen:
                    name_seen[n] = True
                    deduped.append(p)
            deduped.reverse()
            self._generated_posts = deduped
            # 미리보기는 이번에 생성한 글만 표시 (디스크에는 기존+신규 누적 저장)
            self._last_generated = [p for p in deduped if _k(p) in new_keys]
            self.is_generating = False

            if self.stop_flag:
                self._emit_post_log(f"포스트 생성 중단완료 ({len(new_posts)}/{total}개 완료)")
                self._emit_status("중단완료", "#ef4444")
            else:
                self._emit_status("생성 완료", "#22c55e")
                self._emit_post_log(f"포스트 생성 완료: {len(new_posts)}/{total}개")
            # 이미지 갯수 평준화 비활성화 (사용자 요청)
            self._save_generated_posts()

            # 메인 스레드에서 미리보기 열기
            from PySide6.QtCore import QMetaObject, Q_ARG
            QMetaObject.invokeMethod(self, "_show_post_preview", Qt.QueuedConnection)

        threading.Thread(target=_worker, daemon=True).start()

    @Slot()
    def _show_api_balance_error(self):
        """API 토큰(잔액) 부족 안내 — 메인스레드에서 팝업."""
        try:
            QMessageBox.critical(
                self, "API 잔액 부족",
                "API 토큰(잔액)이 부족합니다.\n\n사용 중인 API 키에 결제·충전을 하신 뒤 다시 시도해주세요.")
        except Exception:
            pass

    @Slot()
    def _show_post_preview(self):
        """생성된 포스트 미리보기 — 이번에 생성한 글만 표시."""
        posts = getattr(self, "_last_generated", None) or self._generated_posts
        if not posts:
            QMessageBox.information(self, "결과", "생성된 포스트가 없습니다.")
            return

        from PySide6.QtWidgets import QTextEdit, QLineEdit

        dlg = QDialog(self)
        dlg.setWindowTitle("생성된 포스트 미리보기")
        dlg.resize(1000, 700)

        layout = QVBoxLayout(dlg)

        # 상단
        top = QHBoxLayout()
        select_all_cb = QCheckBox("전체 선택")
        count_label = QLabel(f"0 / {len(posts)}개 선택됨")
        count_label.setStyleSheet("font-weight: bold; font-size: 13px;")
        top.addWidget(select_all_cb)
        top.addStretch()
        top.addWidget(count_label)

        btn_delete = QPushButton("선택 삭제")
        btn_delete.setStyleSheet("background: #ef4444; color: white; border: none; border-radius: 8px; font-size: 13px; padding: 8px 15px;")
        btn_delete.setCursor(Qt.PointingHandCursor)
        btn_post_selected = QPushButton("선택 포스트 생성")
        btn_post_selected.setStyleSheet("background: #4a6cf7; color: white; border: none; border-radius: 8px; font-size: 13px; font-weight: bold; padding: 8px 20px;")
        btn_post_selected.setCursor(Qt.PointingHandCursor)
        btn_close = QPushButton("닫기")
        btn_close.setStyleSheet("padding: 8px 20px;")
        btn_close.clicked.connect(dlg.reject)
        top.addWidget(btn_delete)
        top.addWidget(btn_post_selected)
        top.addWidget(btn_close)
        layout.addLayout(top)

        # 스크롤 영역
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        post_checkboxes = []
        post_editors = []
        title_editors = []

        for idx, item in enumerate(posts):
            place = item["place"]
            content = item["content"]

            # 카드
            card = QFrame()
            card.setStyleSheet("background: white; border: 1px solid #e2e8f0; border-radius: 8px; padding: 10px; margin-bottom: 5px;")
            card_layout = QVBoxLayout(card)
            card_layout.setSpacing(5)

            # 헤더: 체크박스 + 업체명 + 저장 버튼
            hdr = QHBoxLayout()
            cb = QCheckBox(place.get("name", ""))
            cb.setStyleSheet("font-weight: bold; font-size: 13px;")
            post_checkboxes.append(cb)
            hdr.addWidget(cb)
            hdr.addStretch()

            btn_save = QPushButton("저장")
            btn_save.setStyleSheet("background: #22c55e; color: white; border: none; border-radius: 6px; padding: 5px 15px;")
            btn_save.setCursor(Qt.PointingHandCursor)
            hdr.addWidget(btn_save)
            card_layout.addLayout(hdr)

            # 제목 (편집 가능)
            title_edit = QLineEdit()
            title_edit.setText(content.get("title", ""))
            title_edit.setStyleSheet("color: #4a6cf7; font-size: 12px; padding: 5px; border: 1px solid #e2e8f0; border-radius: 4px;")
            title_editors.append(title_edit)
            card_layout.addWidget(title_edit)

            # 본문 (편집 가능)
            editor = QTextEdit()
            editor.setPlainText(content.get("body", ""))
            editor.setMaximumHeight(200)
            editor.setStyleSheet("border: 1px solid #e2e8f0; border-radius: 4px; padding: 5px; font-size: 11px;")
            post_editors.append(editor)
            card_layout.addWidget(editor)

            # 수집 데이터 표시
            addr = place.get("address", "")
            jibun = place.get("jibun_address", "")
            full_addr = (jibun or addr).strip()  # jibun이 행정동+지번 완전주소라 중복 방지
            info_text = f"주소: {full_addr}  |  카테고리: {place.get('category', '')}  |  근처역: {place.get('nearby_station', '')}  |  태그: {', '.join(content.get('tags', []))}"
            info_label = QLabel(info_text)
            info_label.setStyleSheet("color: #64748b; font-size: 10px; padding: 3px 0; background: #f8fafc; border-radius: 4px; padding: 5px;")
            info_label.setWordWrap(True)
            card_layout.addWidget(info_label)

            # 저장 버튼 기능
            def make_save(i, te, ed):
                def save():
                    posts[i]["content"]["title"] = te.text()
                    posts[i]["content"]["body"] = ed.toPlainText()
                    self._save_generated_posts()
                    QMessageBox.information(dlg, "저장", "저장되었습니다.")
                return save

            btn_save.clicked.connect(make_save(idx, title_edit, editor))

            scroll_layout.addWidget(card)

        scroll_layout.addStretch()
        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)

        # 선택 카운트 업데이트
        def update_count():
            cnt = sum(1 for cb in post_checkboxes if cb.isChecked())
            count_label.setText(f"{cnt} / {len(posts)}개 선택됨")

        for cb in post_checkboxes:
            cb.stateChanged.connect(lambda: update_count())

        # 선택 삭제
        def delete_selected():
            reply = QMessageBox.question(dlg, "삭제 확인", "선택된 포스트를 삭제하시겠습니까?",
                                          QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply != QMessageBox.Yes:
                return
            # 역순으로 삭제
            cards = [scroll_layout.itemAt(i).widget() for i in range(scroll_layout.count()) if scroll_layout.itemAt(i).widget()]
            for i in range(len(post_checkboxes) - 1, -1, -1):
                if post_checkboxes[i].isChecked():
                    _gp = posts.pop(i)
                    try:
                        self._generated_posts.remove(_gp)
                    except ValueError:
                        pass
                    post_checkboxes.pop(i)
                    post_editors.pop(i)
                    title_editors.pop(i)
                    widget = cards[i]
                    scroll_layout.removeWidget(widget)
                    widget.deleteLater()
            self._save_generated_posts()
            update_count()
            count_label.setText(f"0 / {len(posts)}개 선택됨")

        btn_delete.clicked.connect(delete_selected)

        # 전체 선택
        def toggle_all(state):
            checked = select_all_cb.isChecked()
            for cb in post_checkboxes:
                cb.setChecked(checked)
            update_count()
        select_all_cb.stateChanged.connect(toggle_all)

        # 선택 포스팅
        def post_selected():
            self.posting_targets = []
            for i, cb in enumerate(post_checkboxes):
                if cb.isChecked():
                    posts[i]["content"]["title"] = title_editors[i].text()
                    posts[i]["content"]["body"] = post_editors[i].toPlainText()
                    self.posting_targets.append(posts[i])
            if not self.posting_targets:
                QMessageBox.warning(dlg, "경고", "선택된 포스트가 없습니다.")
                return
            dlg.accept()
            self._post_generated()

        btn_post_selected.clicked.connect(post_selected)
        dlg.exec()

    def _post_generated(self):
        """미리보기에서 선택된 포스트를 블로그에 포스팅"""
        if not self.posting_targets:
            return

        # 이미 발행 완료된 포스트는 미리 제외 — i==1 즉시발행 분기가 미발행 첫 항목으로 정확히 가도록
        original_count = len(self.posting_targets)
        self.posting_targets = [t for t in self.posting_targets if not t.get("posted")]
        skipped_done = original_count - len(self.posting_targets)
        if not self.posting_targets:
            QMessageBox.information(self, "안내",
                f"선택된 {original_count}개 모두 이미 발행 완료된 포스트입니다.\n포스팅할 항목이 없습니다.")
            return

        cfg = load_config()
        from config import get_active_account
        account = get_active_account(cfg)
        if not account.get("naver_id") or not account.get("naver_pw"):
            QMessageBox.critical(self, "오류", "네이버 계정 정보를 설정해주세요.")
            return

        total = len(self.posting_targets)
        interval_sec = self._get_interval_seconds()
        h = int(self.interval_hour.currentText())
        m = int(self.interval_min.currentText())
        rand_suffix = " (±20분 랜덤)" if (getattr(self, "interval_random", None) and self.interval_random.isChecked()) else ""
        interval_label = f"{h}시간 {m}분{rand_suffix}"
        acc_idx = cfg.get("active_account", 0) + 1

        skipped_msg = f"\n(이미 발행 완료 {skipped_done}개 제외됨)" if skipped_done > 0 else ""
        reply = QMessageBox.question(self, "포스팅 확인",
            f"[아이디 {acc_idx}] {account['naver_id']}\n{total}개 포스팅{skipped_msg}\n간격: {interval_label}\n\n시작할까요?")
        if reply != QMessageBox.Yes:
            return

        keyword = self.keyword_input.currentText().strip()
        pix_keys = [k for k in cfg.get("pixabay_key_list", []) if k]
        self.is_posting = True
        self.stop_flag = False

        def _worker():
            _pl = self._emit_post_log  # 포스팅 로그 → 포스트 현황 패널
            poster = None
            publish_done = [0]
            # ★ 근본방어: 앱/로그 폴더에 비정상적으로 큰(80MB+) json이 있으면 백업 후 제거 (JSON 파싱 폭주 크래시 원천 차단)
            try:
                import os as _os3
                _dirs = [_os3.path.dirname(__file__)]
                try: _dirs.append(self._get_logs_dir())
                except Exception: pass
                _cleaned = 0
                for _dd in _dirs:
                    if not _dd or not _os3.path.isdir(_dd):
                        continue
                    for _fn in _os3.listdir(_dd):
                        if not _fn.endswith(".json"):
                            continue
                        _fp = _os3.path.join(_dd, _fn)
                        try:
                            if _os3.path.getsize(_fp) > 80 * 1024 * 1024:
                                _os3.replace(_fp, _fp + ".corrupt.bak")
                                _cleaned += 1
                        except Exception:
                            pass
                if _cleaned:
                    _pl(f"⚠ 비정상적으로 큰 데이터 파일 {_cleaned}개 자동 정리(.corrupt.bak 백업).")
            except Exception:
                pass
            try:
                self._emit_status("포스팅 중...", "#8b5cf6")
                poster = NaverBlogPoster(
                    naver_id=account["naver_id"],
                    naver_pw=account["naver_pw"],
                    blog_id=account["blog_id"],
                    window_x=0,
                    window_y=0,
                    window_w=1280,
                    window_h=900,
                    stop_flag=lambda: self.stop_flag,
                )
                self._active_posters.append(poster)
                _pl("브라우저 시작...")
                poster.start_browser()
                # 크롬 창을 앞으로 (적당한 크기, 전체화면 X)
                try:
                    poster.driver.execute_script("window.focus();")
                    poster.driver.set_window_rect(x=80, y=60, width=1180, height=860)
                except Exception:
                    pass
                _pl("네이버 로그인 중...")
                if not poster.login():
                    _pl("로그인 실패!")
                    return
                _pl("로그인 성공")

                import datetime as _dt
                base_time = _dt.datetime.now()
                existing_slots = set()   # 기존 예약이 찜한 10분 슬롯 (충돌 회피용)
                # 예약은 '지금부터' 간격대로 잡되, 이미 찬 시간대만 건너뜀 (맨 뒤로 안 밀림)
                reserved_count = 0
                first_immediate = True
                try:
                    _pl("기존 예약 발행 확인 중...")
                    existing = poster.peek_reservations()
                    _resv_total = getattr(poster, "reservation_total", -1)
                    if existing:
                        reserved_count = len(existing)
                        latest = max(existing)
                        _pl(f"기존 예약 {reserved_count}개 / 가장 늦은 시간: {latest.strftime('%Y-%m-%d %H:%M')}")
                        # 기존 예약이 찜한 10분 슬롯 기록 → 맨 뒤로 밀지 않고 '지금부터' 잡되 이미 찬 슬롯만 건너뜀
                        for _e in existing:
                            _es = _e.replace(minute=(_e.minute // 10) * 10, second=0, microsecond=0)
                            existing_slots.add(_es.strftime("%Y-%m-%d %H:%M"))
                        first_immediate = False
                        _pl("→ 지금부터 예약 (이미 찬 시간대는 자동으로 건너뜀)")
                    else:
                        _pl("기존 예약 없음 — 첫 글 즉시발행 + 나머지 예약")
                    # 예약 목록 패널은 가상 스크롤이라 날짜 개수가 실제보다 적게 읽힘.
                    # 네이버 '예약 발행 N건' 라벨이 정확한 총수 → 99 한도 판정은 이 값으로 보정.
                    if isinstance(_resv_total, int) and _resv_total > reserved_count:
                        reserved_count = _resv_total
                        _pl(f"네이버 예약 카운트 보정: 실제 예약 {reserved_count}건 (99 한도 기준)")
                        if reserved_count > 0:
                            first_immediate = False
                    # 예약 현황 캐시 저장 (계정 옆 라벨에 자동 표시)
                    try:
                        _ls = latest.strftime("%Y-%m-%d %H:%M") if existing else ""
                        self._save_resv_status(self._active_blog_id(), reserved_count, _ls)
                    except Exception:
                        pass
                except Exception as _pe:
                    _pl(f"기존 예약 조회 실패: {_pe} — 첫 글 즉시발행으로 진행")
                    first_immediate = True

                running_dt = base_time
                used_slots = set()       # 이번 배치에서 잡은 슬롯 (중복 방지)
                last_reserved_str = ""   # 이번 포스팅으로 잡힌 가장 늦은 예약 시간
                resv_fail_streak = 0     # 예약 발행 연속 실패 수 — 네이버 99 한도 도달 감지용
                for i, item in enumerate(self.posting_targets, 1):
                    if self.stop_flag:
                        _pl(f"포스팅 중단완료 ({i-1}/{len(self.posting_targets)}개 완료)")
                        self._emit_status("중단완료", "#ef4444")
                        break

                    # 가드: 사전 필터링됐어야 하지만, 만약 사이에 상태가 바뀐 경우 skip
                    if item.get("posted"):
                        continue

                    # 예약발행 99개 한도 체크
                    if not (i == 1 and first_immediate) and reserved_count >= 99:
                        _pl(f"예약 발행 99개 한도 도달 ({reserved_count}개) — 포스팅 중단")
                        break

                    # 첫 포스트: 즉시 발행 (옵션에 따라) / 나머지: 예약 발행
                    if i == 1 and first_immediate:
                        schedule_time = None  # 즉시 발행
                    else:
                        # 매 포스트마다 _get_interval_seconds() 호출 → 랜덤 모드면 매번 다른 간격
                        running_dt = running_dt + _dt.timedelta(seconds=self._get_interval_seconds())
                        # 네이버 예약은 10분 단위만 허용 → 항상 10분 단위 정렬 (누적 드리프트/패턴 방지)
                        minute = (running_dt.minute // 10) * 10
                        running_dt = running_dt.replace(minute=minute, second=0, microsecond=0)
                        sched_dt = running_dt
                        if sched_dt <= _dt.datetime.now():
                            sched_dt += _dt.timedelta(minutes=10)
                        # 이미 예약된 시간 / 이번에 잡은 시간이면 10분씩 밀어 '빈 시간'을 찾음 (겹침 방지)
                        _st = sched_dt.strftime("%Y-%m-%d %H:%M")
                        while _st in existing_slots or _st in used_slots:
                            sched_dt += _dt.timedelta(minutes=10)
                            _st = sched_dt.strftime("%Y-%m-%d %H:%M")
                        running_dt = sched_dt
                        used_slots.add(_st)
                        schedule_time = _st
                        _pl(f"[{i}/{total}] 예약 시간: {schedule_time}")

                    place = item["place"]
                    content = item["content"]
                    name = place.get("name", "")
                    _pl(f"[{i}/{total}] '{name}' 포스팅 중...")
                    self._emit_status(f"포스팅 {i}/{total}", "#8b5cf6")
                    self._emit_post_count(f"포스팅 {i}/{total}")

                    # 포스트 생성 시 저장된 이미지 우선 사용 (재다운로드 방지)
                    img_paths = []
                    saved_paths = content.get("image_paths", []) or []
                    saved_paths = [p for p in saved_paths if p and os.path.exists(p)
                                   and not os.path.basename(p).startswith("._")]
                    if saved_paths:
                        img_paths = saved_paths
                        _pl(f"저장된 이미지 {len(img_paths)}장 사용")
                    elif pix_keys:
                        try:
                            # 완전 자동: download_images 내부 체인(AI→테이블→번역)
                            biz = self._best_biz_term(place, keyword)
                            _pl(f"이미지 재검색 (업종 '{biz}', 저장본 없음)")
                            img_paths = download_images(
                                pix_keys[0], biz,
                                3,  # 무조건 3장 고정
                                watermark_text=name,
                                translator=self._translate_ko_to_en,
                            )
                            _pl(f"이미지 {len(img_paths)}장 다운로드")
                        except Exception as e:
                            _pl(f"이미지 다운로드 실패: {e}")
                    else:
                        _pl("Pixabay API 키 미설정 - 이미지 없이 포스팅")

                    body = content.get("body", "")

                    # 제목의 잘못된 지역명 교체 (검색 키워드의 구 ≠ 실제 주소의 구인 경우)
                    actual_address = place.get("address", "") or place.get("jibun_address", "")
                    import re
                    actual_gu_match = re.search(r"([가-힣]+구)", actual_address)
                    if actual_gu_match:
                        actual_gu = actual_gu_match.group(1)
                        title = content.get("title", "")
                        wrong_gus = re.findall(r"([가-힣]+구)", title)
                        for wg in wrong_gus:
                            if wg != actual_gu:
                                title = title.replace(wg, actual_gu)
                                self._emit_post_log(f"제목 지역 교체: {wg} → {actual_gu}")
                        # 동 이름도 확인
                        actual_dong_match = re.search(r"([가-힣]+동)", actual_address)
                        if actual_dong_match:
                            actual_dong = actual_dong_match.group(1)
                            wrong_dongs = re.findall(r"([가-힣]+동)", title)
                            for wd in wrong_dongs:
                                if wd != actual_dong and wd != actual_gu:
                                    title = title.replace(wd, actual_dong)
                        content["title"] = title

                    # 이미지가 있는데 본문에 사진 마커가 없으면 균등 분포로 자동 삽입
                    _has_marker = re.search(r'\[사진\d+\]', body) or '[이미지]' in body
                    if img_paths and not _has_marker:
                        paragraphs = [p for p in body.split("\n\n") if p.strip()]
                        if len(paragraphs) >= 2:
                            step = max(1, len(paragraphs) // (len(img_paths) + 1))
                            new_parts = []
                            img_used = 0
                            for idx, p in enumerate(paragraphs):
                                new_parts.append(p)
                                if img_used < len(img_paths) and (idx + 1) % step == 0 and idx < len(paragraphs) - 1:
                                    img_used += 1
                                    new_parts.append(f"[사진{img_used}]")
                            body = "\n\n".join(new_parts)
                            content["body"] = body
                            _pl(f"본문에 사진 마커 {img_used}개 자동 삽입")

                    try:
                        success = poster.write_post(
                            title=content["title"],
                            body=content["body"],
                            tags=content["tags"],
                            image_paths=img_paths,
                            category=account.get("blog_category", ""),
                            place_name=name,
                            place_address=place.get("address", "") or place.get("jibun_address", ""),
                            schedule_time=schedule_time,
                        )
                        msg = "완료!" if success else "실패"
                        _pl(f"[{i}/{total}] '{name}' {msg}")
                        # 예약 발행 연속 실패 감지 — 네이버 99개 한도에 걸리면 예약이 계속 실패하는데
                        # 성공 카운트(reserved_count)가 안 올라가 무한 시도됨. 연속 실패면 한도 도달로 보고 종료.
                        if schedule_time is not None:
                            if success:
                                resv_fail_streak = 0
                            else:
                                resv_fail_streak += 1
                                if resv_fail_streak >= 3:
                                    _pl(f"예약 발행 연속 {resv_fail_streak}회 실패 — 네이버 예약 한도(99개) 도달로 판단하여 포스팅 종료")
                                    self._emit_status("예약 한도 도달", "#f59e0b")
                                    break
                        if success:
                            publish_done[0] += 1
                            self._emit_publish_count(f"{publish_done[0]}개")
                            if schedule_time is not None:
                                reserved_count += 1
                                last_reserved_str = schedule_time   # 가장 늦은 예약 시간 갱신
                        # 브라우저 세션 확인 — 죽었으면 즉시 중단
                        try:
                            _ = poster.driver.current_url
                        except Exception as _se:
                            _pl(f"!!! 브라우저 세션이 종료됨: {str(_se)[:120]}. 포스팅 중단")
                            self.stop_flag = True
                            break
                        if success:
                            # 날짜별 발행 일정표에 자동 기록 (로컬, 조회용)
                            try:
                                self._append_schedule_record(
                                    account.get("blog_id", ""), schedule_time,
                                    name, keyword, content.get("title", ""))
                            except Exception:
                                pass
                            # 포스팅 히스토리 Firestore 저장 (백그라운드)
                            try:
                                import threading as _th_h
                                from users import add_posting_history as _aph
                                _app_user = getattr(self, "current_user", {}).get("username", "")
                                _th_h.Thread(
                                    target=_aph,
                                    args=(_app_user, account.get("blog_id", ""),
                                          name, place.get("address", "") or place.get("jibun_address", ""),
                                          keyword, content.get("title", ""), schedule_time or ""),
                                    daemon=True,
                                ).start()
                            except Exception:
                                pass
                            # 포스팅 완료 시: posted=True 마킹 (F8에서 [완] 표시로 확인 가능)
                            _k = self._place_key(place)
                            for _gp in self._generated_posts:
                                if self._place_key(_gp.get("place", {})) == _k:
                                    _gp["posted"] = True
                                    break
                            self._save_generated_posts()
                            _pl(f"'{place.get('name','')}' 포스팅 완료 → F8에서 [완] 표시로 확인 가능")
                    except Exception as e:
                        _pl(f"'{name}' 오류: {e}")
                        # 예약 발행이 예외로 실패한 경우도 연속 실패로 집계 — 한도 도달 시 무한 시도 방지
                        if schedule_time is not None:
                            resv_fail_streak += 1
                            if resv_fail_streak >= 3:
                                _pl(f"예약 발행 연속 {resv_fail_streak}회 실패/오류 — 네이버 예약 한도(99개) 도달로 판단하여 포스팅 종료")
                                self._emit_status("예약 한도 도달", "#f59e0b")
                                break

                if not self.stop_flag:
                    _pl(f"전체 포스팅 완료! ({total}개)")
                    self._emit_post_count(f"{total}/{total}")
                self._emit_status("완료", "#22c55e")
                # 포스팅 결과로 예약 현황 자동 저장 → 조회 안 해도 라벨에 '며칠까지' 표시
                try:
                    if last_reserved_str:
                        self._save_resv_status(self._active_blog_id(), reserved_count, last_reserved_str)
                        self.resv_signal.emit()
                except Exception:
                    pass

            except Exception as e:
                _pl(f"오류: {e}")
                # 정확한 발생 위치(파일·줄) 기록 — 원인 추적용
                try:
                    import traceback as _tb, os as _os2
                    _tbs = _tb.extract_tb(e.__traceback__)
                    # 우리 코드(블로그마스터 .py) 프레임을 우선 표시 — 라이브러리(decoder 등) 말고 진짜 호출 지점
                    _mine = [fr for fr in _tbs if fr.filename.endswith((".py",)) and
                             _os2.path.basename(fr.filename) in (
                                 "main.py","naver_poster.py","image_handler.py","content_generator.py",
                                 "local_image_store.py","places_crawler.py","comment_buddy_manager.py",
                                 "naver_crawler.py","config.py","places_api_crawler.py")]
                    _show = _mine[-1] if _mine else (_tbs[-1] if _tbs else None)
                    if _show:
                        _pl(f"  ↳ 위치: {_os2.path.basename(_show.filename)}:{_show.lineno} ({_show.name})")
                    _crash = _os2.path.join(_os2.path.dirname(__file__), "posting_error.log")
                    with open(_crash, "a", encoding="utf-8") as _cf:
                        _cf.write(f"\n--- {e} ---\n{_tb.format_exc()}\n")
                except Exception:
                    pass
                self._emit_status("오류", "#ef4444")
            finally:
                self.is_posting = False
                if poster:
                    try: self._active_posters.remove(poster)
                    except Exception: pass
                    poster.close()

        threading.Thread(target=_worker, daemon=True).start()

    # ── 포스팅 업체 선택 ──
    @staticmethod
    def _gu_of_place(p: dict) -> str:
        """지역 그룹 라벨 — 광역시(서울/부산 등)는 '구' 단위, 도(경기 등)는 '시/군' 단위로 묶는다.
        (효자구가 어느 시인지 몰라도 도 지역은 시로 묶여 찾기 쉽게)"""
        import re as _re
        addr = ((p.get("address", "") or "") + " " + (p.get("jibun_address", "") or "")).strip()
        first = addr.split()[0] if addr.split() else ""
        is_metro = first.endswith(("특별시", "광역시")) or first == "세종특별자치시"
        m_si = _re.search(r"([가-힣]+시)", addr)
        m_gu = _re.search(r"([가-힣]+[구군])", addr)
        if is_metro:
            # 광역시: 구 단위 (예: 강남구)
            if m_gu:
                return m_gu.group(1)
            return m_si.group(1) if m_si else "지역 미상"
        # 도 지역: 시 단위 우선 (예: 수원시), 없으면 군, 없으면 구
        if m_si:
            return m_si.group(1)
        if m_gu:
            return m_gu.group(1)
        return "지역 미상"

    def _load_all_crawl_results(self, mode: str = "") -> dict:
        """현재 계정의 logs 폴더의 모든 크롤링 결과를 (구 × 업종)별로 집계.
        mode: 'region' 또는 'keyword' 지정 시 해당 모드만 로드. 빈 문자열이면 전체.
        여러 크롤 파일에 걸쳐 같은 (구, 업종)은 한 그룹으로 합치고 (업체명, 주소) 기준 중복 제거.
        라벨에 가장 최근 크롤 날짜 표시."""
        from places_crawler import load_results
        log_dir = self._get_logs_dir()
        # (gu, keyword) → {"items": list, "seen": set, "latest": str}
        buckets: dict = {}
        if not os.path.exists(log_dir):
            return {}
        files = sorted(
            [f for f in os.listdir(log_dir) if f.endswith(".json") and not f.startswith("._")],
            reverse=True
        )
        for f in files:
            try:
                raw = load_results(os.path.join(log_dir, f))
                keyword = (raw.get("keyword") or "").strip()
                items = raw.get("items", [])
                if not items:
                    continue
                # mode 필터 (crawl_mode 없는 기존 데이터는 region으로 취급)
                if mode:
                    file_mode = (raw.get("crawl_mode") or "region").strip()
                    if file_mode != mode:
                        continue
                # crawled_at 필드 우선, 없으면 파일명에서 추출
                date_part = (raw.get("crawled_at") or "").strip()
                if not date_part:
                    date_raw = f.replace("places_", "").replace(".json", "")
                    try:
                        if len(date_raw) >= 13 and date_raw[:4].isdigit():
                            date_part = f"{date_raw[:4]}-{date_raw[4:6]}-{date_raw[6:8]} {date_raw[9:11]}:{date_raw[11:13]}"
                        else:
                            date_part = ""
                    except Exception:
                        date_part = ""
                # 업종명 정규화 — 지역 토큰 제거 (앞쪽)
                biz = keyword
                _parts = keyword.split()
                if len(_parts) >= 2:
                    biz = _parts[-1]
                # 이 파일의 수집 날짜 (YYYY-MM-DD)
                _crawl_day = (date_part[:10] if date_part else "").strip()
                for p in items:
                    gu = self._gu_of_place(p)
                    key = (gu, biz)
                    b = buckets.get(key)
                    if b is None:
                        b = {"items": [], "seen": set(), "latest": date_part}
                        buckets[key] = b
                    n = (p.get("name") or "").strip()
                    a = (p.get("address") or p.get("jibun_address") or "").strip()
                    sig = (n, a)
                    if n and sig not in b["seen"]:
                        b["seen"].add(sig)
                        # 첫 등장(최신) 파일의 수집 날짜를 부여 — F7 날짜별 일괄선택용
                        if _crawl_day and not p.get("_crawl_day"):
                            p["_crawl_day"] = _crawl_day
                        b["items"].append(p)
                    # 파일은 최신→과거 순으로 처리 중이므로 첫 등장이 최신
            except Exception:
                continue
        # 라벨 생성 — 업종별 최신 시간 내림차순, 같은 업종 내엔 구 이름 오름차순
        biz_latest = {}
        for (gu, biz), b in buckets.items():
            t = b.get("latest", "")
            if t > biz_latest.get(biz, ""):
                biz_latest[biz] = t
        groups = {}
        for (gu, biz), b in sorted(
            buckets.items(),
            key=lambda x: (biz_latest.get(x[0][1], ""), x[0][0]),
            reverse=True
        ):
            label = f"[{gu}] {biz} ({len(b['items'])}개) · {b['latest']}"
            groups[label] = b["items"]
        return groups

    def _show_posting_selector(self) -> list:
        # 새로고침 시 보존된 체크/편집 상태 (한 번 적용 후 폐기)
        _preserved_checks = getattr(self, "_selector_preserved_checks", None) or set()
        _preserved_edits = getattr(self, "_selector_preserved_edits", None) or {}
        self._selector_preserved_checks = None
        self._selector_preserved_edits = None

        groups_region = self._load_all_crawl_results(mode="region")
        groups_keyword = self._load_all_crawl_results(mode="keyword")

        if not groups_region and not groups_keyword:
            QMessageBox.warning(self, "경고", "크롤링 결과가 없습니다.")
            return []

        # 생성된 포스트 로드해서 (업체명+주소) → post 매핑
        _gen_posts = self._load_generated_posts()
        _post_map = {}
        for _gp in _gen_posts:
            _pl = _gp.get("place", {})
            _key = (_pl.get("name", ""), _pl.get("address", "") or _pl.get("jibun_address", ""))
            _post_map[_key] = _gp

        dlg = QDialog(self)
        dlg.setWindowTitle("포스팅할 업체 선택")
        dlg.resize(950, 620)

        layout = QVBoxLayout(dlg)

        # 상단
        top = QHBoxLayout()
        btn_all = QPushButton("전체 선택")
        btn_all.setStyleSheet("padding: 6px 15px;")
        btn_none = QPushButton("전체 해제")
        btn_none.setStyleSheet("padding: 6px 15px;")
        btn_cat_filter = QPushButton("2차 필터")
        btn_cat_filter.setStyleSheet("padding: 6px 15px; background: #f59e0b; color: white; border: none; border-radius: 6px; font-weight: bold;")
        btn_cat_filter.setToolTip("카테고리별로 일괄 선택/해제 — 무관한 업체 빠르게 거르기")
        btn_remove_done = QPushButton("포스팅완료 제거")
        btn_remove_done.setStyleSheet("padding: 6px 15px; background: #6366f1; color: white; border: none; border-radius: 6px; font-weight: bold;")
        btn_remove_done.setToolTip("포스팅 완료된 업체([완])를 목록에서 영구 삭제")
        count_label = QLabel("0개 선택됨")
        count_label.setStyleSheet("font-weight: bold; font-size: 13px;")
        top.addWidget(btn_all)
        top.addWidget(btn_none)
        top.addWidget(btn_cat_filter)
        top.addWidget(btn_remove_done)
        top.addStretch()
        top.addWidget(count_label)
        layout.addLayout(top)

        # 탭 위젯
        from PySide6.QtWidgets import QTabWidget
        tab_widget = QTabWidget()

        all_items = []  # (QTreeWidgetItem, place_data)
        all_trees = []  # 탭별 QTreeWidget 목록

        def _build_tree(groups_dict):
            from PySide6.QtGui import QFont
            container = QWidget()
            container_layout = QVBoxLayout(container)
            container_layout.setContentsMargins(0, 0, 0, 0)
            container_layout.setSpacing(0)

            tw = QTreeWidget()
            tw.setHeaderLabels(["업체명", "업체주소", "카테고리", "근처역", "앞 키워드", "태그", "픽사베이 키워드"])
            tw.setColumnWidth(0, 320)
            tw.setColumnWidth(1, 220)
            tw.setColumnWidth(2, 120)
            tw.setColumnWidth(3, 120)
            tw.setColumnWidth(4, 180)
            tw.setColumnWidth(5, 140)
            tw.setColumnWidth(6, 180)
            tw.setAlternatingRowColors(True)
            tw.setStyleSheet("""
                QTreeWidget { font-size: 12px; }
                QTreeWidget::item { padding: 3px 0; }
                alternate-background-color: #f5f5f5;
            """)
            container_layout.addWidget(tw)

            # 날짜별 그룹 묶기 — 그룹 라벨 " · YYYY-MM-DD ..." 에서 날짜 추출
            date_map = {}  # date_str -> [(group_name, places)]
            for group_name, places in groups_dict.items():
                if not places:
                    continue
                _date_key = "기타"
                if " · " in group_name:
                    _ts = group_name.split(" · ", 1)[1].strip()
                    _date_key = _ts[:10] if len(_ts) >= 10 else _ts
                date_map.setdefault(_date_key, []).append((group_name, places))

            # 최신 날짜 먼저, "기타" 맨 뒤
            sorted_dates = sorted(date_map.keys(), key=lambda d: (d == "기타", d), reverse=True)

            date_font = QFont()
            date_font.setBold(True)
            date_font.setPointSize(10)

            tab_items = []
            for date_str in sorted_dates:
                groups_in_date = date_map[date_str]
                total_places = sum(len(pl) for _, pl in groups_in_date)

                # 날짜 헤더 (1단계, 접힘)
                date_display = date_str[5:] if len(date_str) >= 10 else date_str
                date_item = QTreeWidgetItem(tw)
                date_item.setText(0, f"📅 {date_display}  ({total_places}개)")
                date_item.setFlags(date_item.flags() | Qt.ItemIsUserCheckable)
                date_item.setCheckState(0, Qt.Unchecked)
                date_item.setExpanded(False)
                date_item.setFont(0, date_font)
                date_item.setForeground(0, QColor("#1e40af"))
                date_item.setBackground(0, QColor("#eff6ff"))

                for group_name, places in groups_in_date:
                    # " · 날짜" 제거한 깨끗한 이름
                    clean_name = group_name.split(" · ")[0].strip() if " · " in group_name else group_name

                    # 그룹 아이템 (2단계, 접힘)
                    parent = QTreeWidgetItem(date_item)
                    parent.setFlags(parent.flags() | Qt.ItemIsUserCheckable)
                    parent.setCheckState(0, Qt.Unchecked)
                    parent.setExpanded(False)

                    posted_count = 0
                    child_count = 0

                    for p in places:
                        addr = p.get("address", "")
                        jibun = p.get("jibun_address", "")
                        full_addr = (jibun or addr).strip()  # jibun이 행정동+지번 완전주소라 중복 방지

                        child = QTreeWidgetItem(parent)
                        child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
                        child.setCheckState(0, Qt.Unchecked)

                        _k = (p.get("name", ""), p.get("address", "") or p.get("jibun_address", ""))
                        _gp = _post_map.get(_k)
                        is_done = bool(_gp and _gp.get("posted", False))
                        if is_done:
                            child.setText(0, "[완] " + p.get("name", ""))
                            child.setForeground(0, QColor("#3b82f6"))
                            posted_count += 1
                        elif _gp is not None:
                            child.setText(0, "●  " + p.get("name", ""))
                            child.setForeground(0, QColor("#22c55e"))
                            posted_count += 1
                        else:
                            child.setText(0, "●  " + p.get("name", ""))
                            child.setForeground(0, QColor("#ef4444"))

                        child.setText(1, full_addr)
                        child.setText(2, p.get("category", ""))
                        child.setText(3, p.get("nearby_station", ""))
                        child.setText(4, p.get("front_keywords", ""))
                        child.setText(5, p.get("tags", ""))
                        child.setText(6, p.get("pixabay_keywords", ""))
                        child.setFlags(child.flags() | Qt.ItemIsEditable | Qt.ItemIsUserCheckable)

                        _pkey = (p.get("name", ""), p.get("address", "") or p.get("jibun_address", ""))
                        if _pkey in _preserved_checks:
                            child.setCheckState(0, Qt.Checked)
                        _e = _preserved_edits.get(_pkey)
                        if _e:
                            if "front_keywords" in _e: child.setText(4, _e["front_keywords"])
                            if "tags" in _e: child.setText(5, _e["tags"])
                            if "pixabay_keywords" in _e: child.setText(6, _e["pixabay_keywords"])

                        tab_items.append((child, p))
                        child_count += 1

                    if child_count > 0:
                        if posted_count == child_count:
                            parent.setForeground(0, QColor("#22c55e"))
                            parent.setText(0, "●  " + clean_name)
                        elif posted_count == 0:
                            parent.setForeground(0, QColor("#ef4444"))
                            parent.setText(0, "●  " + clean_name)
                        else:
                            parent.setForeground(0, QColor("#f59e0b"))
                            parent.setText(0, "●  " + clean_name)

            return container, tw, tab_items

        groups_region = self._load_all_crawl_results(mode="region")
        groups_keyword = self._load_all_crawl_results(mode="keyword")

        container_r, tree_r, items_r = _build_tree(groups_region)
        container_k, tree_k, items_k = _build_tree(groups_keyword)

        all_items.extend(items_r)
        all_items.extend(items_k)
        all_trees.extend([tree_r, tree_k])

        tab_widget.addTab(container_r, f"지역기반 ({len(items_r)}개)")
        tab_widget.addTab(container_k, f"키워드기반 ({len(items_k)}개)")

        def update_count():
            cnt = sum(1 for item, _ in all_items if item.checkState(0) == Qt.Checked)
            count_label.setText(f"{cnt}개 선택됨")

        def on_item_changed(item, col):
            if col != 0:
                return
            if item.childCount() > 0:
                state = item.checkState(0)
                for i in range(item.childCount()):
                    item.child(i).setCheckState(0, state)
            update_count()

        def _make_sort_handler(tw):
            _st = {"col": -1, "order": Qt.AscendingOrder}
            def _on_section(col):
                if _st["col"] == col:
                    _st["order"] = Qt.DescendingOrder if _st["order"] == Qt.AscendingOrder else Qt.AscendingOrder
                else:
                    _st["col"] = col
                    _st["order"] = Qt.AscendingOrder
                tw.sortItems(col, _st["order"])
            return _on_section

        for tw in all_trees:
            tw.itemChanged.connect(on_item_changed)
            # 컬럼 헤더(카테고리/업체명/주소 등) 클릭 시 정렬 — 클릭마다 오름/내림 토글
            tw.header().setSectionsClickable(True)
            tw.header().sectionClicked.connect(_make_sort_handler(tw))

        _tree_map = {id(container_r): tree_r, id(container_k): tree_k}

        def current_tree():
            w = tab_widget.currentWidget()
            return _tree_map.get(id(w), w)

        def select_all():
            tw = current_tree()
            tw.blockSignals(True)
            for i in range(tw.topLevelItemCount()):
                date_item = tw.topLevelItem(i)
                date_item.setCheckState(0, Qt.Checked)
                for j in range(date_item.childCount()):
                    grp = date_item.child(j)
                    grp.setCheckState(0, Qt.Checked)
                    for k in range(grp.childCount()):
                        grp.child(k).setCheckState(0, Qt.Checked)
            tw.blockSignals(False)
            update_count()

        def select_none():
            tw = current_tree()
            tw.blockSignals(True)
            for i in range(tw.topLevelItemCount()):
                date_item = tw.topLevelItem(i)
                date_item.setCheckState(0, Qt.Unchecked)
                for j in range(date_item.childCount()):
                    grp = date_item.child(j)
                    grp.setCheckState(0, Qt.Unchecked)
                    for k in range(grp.childCount()):
                        grp.child(k).setCheckState(0, Qt.Unchecked)
            tw.blockSignals(False)
            update_count()

        btn_all.clicked.connect(select_all)
        btn_none.clicked.connect(select_none)

        # ── 카테고리 2차 필터 다이얼로그 ──
        def open_category_filter():
            """현재 표시된 모든 업체의 카테고리 목록 → 체크박스 → 선택된 카테고리만 체크 유지."""
            # 카테고리별 업체 매핑 — 체크된(선택된) 업체만 대상 (예: 강북구만 체크 → 강북구 카테고리만).
            # 아무것도 체크 안 됐으면 전체 대상.
            cat_map = {}  # cat_str -> [(item, place), ...]
            _checked_src = [(it, p) for it, p in all_items if it.checkState(0) == Qt.Checked]
            _src = _checked_src if _checked_src else all_items
            for _item, _p in _src:
                _c = (_p.get("category") or "").strip() or "(카테고리 없음)"
                cat_map.setdefault(_c, []).append((_item, _p))
            if not cat_map:
                QMessageBox.information(dlg, "안내", "표시된 업체가 없습니다.")
                return

            fdlg = QDialog(dlg)
            fdlg.setWindowTitle("2차 필터 — 카테고리")
            fdlg.resize(480, 560)
            flay = QVBoxLayout(fdlg)

            info = QLabel("✅ 체크된 카테고리의 업체만 선택 유지됩니다.\n"
                          "해제된 카테고리의 업체는 자동으로 체크 해제됩니다.")
            info.setStyleSheet("color: #475569; font-size: 12px; padding: 4px;")
            flay.addWidget(info)

            # 위/아래 일괄 버튼
            qtop = QHBoxLayout()
            btn_fall = QPushButton("전체 선택")
            btn_fall.setStyleSheet("padding: 4px 12px;")
            btn_fnone = QPushButton("전체 해제")
            btn_fnone.setStyleSheet("padding: 4px 12px;")
            qtop.addWidget(btn_fall)
            qtop.addWidget(btn_fnone)
            qtop.addStretch()
            flay.addLayout(qtop)

            # 스크롤 영역에 카테고리 체크박스
            from PySide6.QtWidgets import QScrollArea
            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            scroll_w = QWidget()
            scroll_lay = QVBoxLayout(scroll_w)
            scroll_lay.setSpacing(2)

            cat_checks = []  # (QCheckBox, cat_str, [(item,place)...])
            # 카테고리 개수 많은 순으로 정렬
            for _cat in sorted(cat_map.keys(), key=lambda c: -len(cat_map[c])):
                _places = cat_map[_cat]
                # 현재 어느 하나라도 체크되어 있으면 기본 체크
                any_checked = any(it.checkState(0) == Qt.Checked for it, _ in _places)
                cb = QCheckBox(f"{_cat}  ({len(_places)}개)")
                cb.setStyleSheet("font-size: 12px; padding: 3px;")
                cb.setChecked(any_checked)
                scroll_lay.addWidget(cb)
                cat_checks.append((cb, _cat, _places))
            scroll_lay.addStretch()
            scroll.setWidget(scroll_w)
            flay.addWidget(scroll, 1)

            def _f_all():
                for cb, _, _ in cat_checks:
                    cb.setChecked(True)
            def _f_none():
                for cb, _, _ in cat_checks:
                    cb.setChecked(False)
            btn_fall.clicked.connect(_f_all)
            btn_fnone.clicked.connect(_f_none)

            # 적용/취소/삭제
            qbot = QHBoxLayout()
            btn_cancel = QPushButton("취소")
            btn_cancel.setStyleSheet("padding: 6px 18px;")
            btn_delete_filter = QPushButton("선택 삭제")
            btn_delete_filter.setStyleSheet("background: #ef4444; color: white; padding: 6px 18px; border: none; border-radius: 6px; font-weight: bold;")
            btn_delete_filter.setToolTip("체크된 카테고리의 업체를 데이터에서 영구 삭제")
            btn_apply = QPushButton("적용")
            btn_apply.setStyleSheet("background: #4a6cf7; color: white; padding: 6px 18px; border: none; border-radius: 6px; font-weight: bold;")
            qbot.addStretch()
            qbot.addWidget(btn_cancel)
            qbot.addWidget(btn_delete_filter)
            qbot.addWidget(btn_apply)
            flay.addLayout(qbot)

            btn_cancel.clicked.connect(fdlg.reject)

            def _sync_parents(tw):
                """place 체크 변경 후 group/date 부모 체크 상태 동기화."""
                tw.blockSignals(True)
                try:
                    for i in range(tw.topLevelItemCount()):
                        date_item = tw.topLevelItem(i)
                        all_date_checked = True
                        any_date_checked = False
                        for j in range(date_item.childCount()):
                            grp = date_item.child(j)
                            n = grp.childCount()
                            n_checked = sum(1 for k in range(n) if grp.child(k).checkState(0) == Qt.Checked)
                            if n > 0 and n_checked == n:
                                grp.setCheckState(0, Qt.Checked)
                                any_date_checked = True
                            else:
                                grp.setCheckState(0, Qt.Unchecked)
                                all_date_checked = False
                                if n_checked > 0:
                                    any_date_checked = True
                        n_grps = date_item.childCount()
                        if n_grps > 0 and all_date_checked:
                            date_item.setCheckState(0, Qt.Checked)
                        elif any_date_checked:
                            date_item.setCheckState(0, Qt.Unchecked)
                        else:
                            date_item.setCheckState(0, Qt.Unchecked)
                finally:
                    tw.blockSignals(False)

            def _apply_filter():
                # 체크된 카테고리 → 그 업체들 체크, 해제된 → 체크 해제
                for _tw in all_trees:
                    _tw.blockSignals(True)
                try:
                    for cb, _cat_str, _places in cat_checks:
                        state = Qt.Checked if cb.isChecked() else Qt.Unchecked
                        for _it, _ in _places:
                            _it.setCheckState(0, state)
                finally:
                    for _tw in all_trees:
                        _tw.blockSignals(False)
                for _tw in all_trees:
                    _sync_parents(_tw)
                # 선택(체크)된 업체만 목록에 표시, 나머지 + 빈 그룹/날짜는 숨김
                for _tw in all_trees:
                    for i in range(_tw.topLevelItemCount()):
                        date_item = _tw.topLevelItem(i)
                        date_vis = False
                        for j in range(date_item.childCount()):
                            grp = date_item.child(j)
                            grp_vis = False
                            for k in range(grp.childCount()):
                                leaf = grp.child(k)
                                show = leaf.checkState(0) == Qt.Checked
                                leaf.setHidden(not show)
                                grp_vis = grp_vis or show
                            grp.setHidden(not grp_vis)
                            date_vis = date_vis or grp_vis
                        date_item.setHidden(not date_vis)
                update_count()
                fdlg.accept()
            btn_apply.clicked.connect(_apply_filter)

            def _delete_filter():
                # 체크된(선택된) 카테고리의 업체 수집
                to_delete_places = []
                to_delete_items = []
                for cb, _cat_str, _places in cat_checks:
                    if cb.isChecked():
                        for _it, _p in _places:
                            to_delete_places.append(_p)
                            to_delete_items.append(_it)
                if not to_delete_places:
                    QMessageBox.information(fdlg, "안내", "삭제할 항목이 없습니다.\n체크된 카테고리가 없어요.")
                    return
                reply = QMessageBox.question(
                    fdlg, "삭제 확인",
                    f"체크된 카테고리의 업체 {len(to_delete_places)}개를\n데이터에서 영구 삭제할까요?",
                    QMessageBox.Yes | QMessageBox.No
                )
                if reply != QMessageBox.Yes:
                    return
                # 트리에서 제거
                del_ids = set(id(_it) for _it in to_delete_items)
                for _tw in all_trees:
                    _tw.blockSignals(True)
                try:
                    for _it in to_delete_items:
                        _par = _it.parent()
                        if _par:
                            _par.removeChild(_it)
                    all_items[:] = [(_it, _p) for _it, _p in all_items if id(_it) not in del_ids]
                    # 빈 그룹/날짜 정리
                    for _tw in all_trees:
                        for i in range(_tw.topLevelItemCount() - 1, -1, -1):
                            date_item = _tw.topLevelItem(i)
                            for j in range(date_item.childCount() - 1, -1, -1):
                                grp = date_item.child(j)
                                if grp.childCount() == 0:
                                    date_item.removeChild(grp)
                            if date_item.childCount() == 0:
                                _tw.takeTopLevelItem(i)
                finally:
                    for _tw in all_trees:
                        _tw.blockSignals(False)
                # 영구 삭제 (logs + generated_posts 모두)
                _persist_delete(to_delete_places)
                update_count()
                fdlg.accept()
            btn_delete_filter.clicked.connect(_delete_filter)

            fdlg.exec()

        btn_cat_filter.clicked.connect(open_category_filter)

        layout.addWidget(tab_widget)

        # 하단
        bottom = QHBoxLayout()

        # 삭제 버튼들 — 영구 저장 (deleted_keys_*.json) + generated_posts 동기화
        def _persist_delete(places_to_delete: list):
            if not places_to_delete:
                return
            keys = {self._place_key(p) for p in places_to_delete}
            # logs 파일에서 즉시 영구 삭제
            self._purge_places_from_logs(places_to_delete)
            # generated_posts에도 있으면 제거
            changed = False
            new_gp = []
            current = self._load_generated_posts()
            for gp in current:
                if self._place_key(gp.get("place", {})) in keys:
                    changed = True
                    continue
                new_gp.append(gp)
            if changed:
                self._generated_posts = new_gp
                self._save_generated_posts()

        def remove_posted():
            done_places = []
            done_item_ids = set()
            for it, p in all_items:
                _k = (p.get("name", ""), p.get("address", "") or p.get("jibun_address", ""))
                _gp = _post_map.get(_k)
                if _gp and _gp.get("posted", False):
                    done_places.append(p)
                    done_item_ids.add(id(it))
            if not done_places:
                QMessageBox.information(dlg, "안내", "포스팅 완료된 업체가 없습니다.")
                return
            reply = QMessageBox.question(
                dlg, "삭제 확인",
                f"포스팅 완료된 업체 {len(done_places)}개를\n데이터에서 영구 삭제할까요?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return
            _persist_delete(done_places)
            for tw in all_trees:
                tw.blockSignals(True)
                try:
                    for i in range(tw.topLevelItemCount() - 1, -1, -1):
                        date_item = tw.topLevelItem(i)
                        for j in range(date_item.childCount() - 1, -1, -1):
                            grp = date_item.child(j)
                            for k in range(grp.childCount() - 1, -1, -1):
                                place_item = grp.child(k)
                                if id(place_item) in done_item_ids:
                                    grp.removeChild(place_item)
                            if grp.childCount() == 0:
                                date_item.removeChild(grp)
                        if date_item.childCount() == 0:
                            tw.takeTopLevelItem(i)
                finally:
                    tw.blockSignals(False)
            all_items[:] = [(it, p) for it, p in all_items if id(it) not in done_item_ids]
            update_count()
            self._emit_log(f"포스팅 완료 업체 {len(done_places)}개 영구 삭제")
        btn_remove_done.clicked.connect(remove_posted)

        def delete_selected():
            # 체크된 places 수집
            checked_places = []
            for it, p in all_items:
                if it.checkState(0) == Qt.Checked:
                    checked_places.append(p)
            if not checked_places:
                QMessageBox.information(dlg, "안내", "선택된 항목이 없습니다.")
                return
            reply = QMessageBox.question(dlg, "삭제 확인", f"{len(checked_places)}개 항목을 삭제하시겠습니까?\n(F8에서도 같이 제거됩니다)",
                                          QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply != QMessageBox.Yes:
                return
            _persist_delete(checked_places)
            # UI에서 제거 (3단계 트리: 날짜→그룹→업체)
            checked_ids = set(id(it) for it, _ in all_items if it.checkState(0) == Qt.Checked)
            for tw in all_trees:
                tw.blockSignals(True)
                try:
                    for i in range(tw.topLevelItemCount() - 1, -1, -1):
                        date_item = tw.topLevelItem(i)
                        for j in range(date_item.childCount() - 1, -1, -1):
                            grp = date_item.child(j)
                            for k in range(grp.childCount() - 1, -1, -1):
                                place_item = grp.child(k)
                                if id(place_item) in checked_ids:
                                    grp.removeChild(place_item)
                            if grp.childCount() == 0:
                                date_item.removeChild(grp)
                        if date_item.childCount() == 0:
                            tw.takeTopLevelItem(i)
                finally:
                    tw.blockSignals(False)
            all_items[:] = [(it, p) for it, p in all_items if id(it) not in checked_ids]
            update_count()
            self._emit_log(f"{len(checked_places)}개 항목 삭제 (영구 반영)")

        def delete_all():
            self._del_log(f"=== F7-Selector delete_all 클릭 (all_items={len(all_items)}개) ===")
            all_places = [p for _, p in all_items]
            if not all_places:
                self._del_log("F7-Selector delete_all: all_places 비어있음 → 종료")
                return
            reply = QMessageBox.question(dlg, "전체 삭제 확인", f"현재 표시된 전체 {len(all_places)}개 항목을 삭제하시겠습니까?\n(F8에서도 같이 제거됩니다)",
                                          QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply != QMessageBox.Yes:
                self._del_log("F7-Selector delete_all: 사용자 취소")
                return
            self._del_log(f"F7-Selector delete_all: {len(all_places)}개 삭제 진행")
            _persist_delete(all_places)
            for tw in all_trees:
                tw.blockSignals(True)
                tw.clear()
                tw.blockSignals(False)
            all_items.clear()
            update_count()
            self._emit_log(f"{len(all_places)}개 전체 삭제 (영구 반영)")

        btn_del_sel = QPushButton("선택 삭제")
        btn_del_sel.setStyleSheet("background: #ef4444; color: white; border: none; border-radius: 8px; padding: 8px 15px;")
        btn_del_sel.setCursor(Qt.PointingHandCursor)
        btn_del_sel.clicked.connect(delete_selected)

        btn_del_all = QPushButton("전체 삭제")
        btn_del_all.setStyleSheet("background: #dc2626; color: white; border: none; border-radius: 8px; padding: 8px 15px;")
        btn_del_all.setCursor(Qt.PointingHandCursor)
        btn_del_all.clicked.connect(delete_all)

        btn_refresh = QPushButton("새로고침")
        btn_refresh.setStyleSheet("background: #64748b; color: white; border: none; border-radius: 8px; padding: 8px 15px;")
        btn_refresh.setCursor(Qt.PointingHandCursor)
        def do_refresh():
            # 현재 체크/편집 상태 보존 → 재오픈 시 복원
            _ck = set()
            _ed = {}
            for it, p in all_items:
                _k = (p.get("name", ""), p.get("address", "") or p.get("jibun_address", ""))
                if it.checkState(0) == Qt.Checked:
                    _ck.add(_k)
                _e = {}
                if it.text(4) != (p.get("front_keywords") or ""): _e["front_keywords"] = it.text(4)
                if it.text(5) != (p.get("tags") or ""): _e["tags"] = it.text(5)
                if it.text(6) != (p.get("pixabay_keywords") or ""): _e["pixabay_keywords"] = it.text(6)
                if _e: _ed[_k] = _e
            self._selector_preserved_checks = _ck
            self._selector_preserved_edits = _ed
            dlg.done(2)  # 특수 코드: 새로고침
        btn_refresh.clicked.connect(do_refresh)

        bottom.addWidget(btn_del_sel)
        bottom.addWidget(btn_del_all)
        bottom.addWidget(btn_refresh)
        bottom.addStretch()

        btn_cancel = QPushButton("취소")
        btn_cancel.setStyleSheet("padding: 8px 20px;")
        btn_cancel.clicked.connect(dlg.reject)

        # 프롬프트 선택 드롭다운 (취소 옆)
        try:
            import json as _pj
            from app_paths import ensure_from_bundle as _efb
            _prompts_path = _efb("prompts.json")
            with open(_prompts_path, "r", encoding="utf-8") as _pf:
                _prompts_dict = _pj.load(_pf)
            _prompt_keys = list(_prompts_dict.keys())
        except Exception:
            _prompt_keys = []
        prompt_combo = QComboBox()
        prompt_combo.addItem("(자동 선택)")
        for k in _prompt_keys:
            if k == "기본":   # 기본은 숨김 — '(자동 선택)'이 폴백 처리
                continue
            prompt_combo.addItem(k)
        prompt_combo.setStyleSheet("padding: 6px 10px; font-size: 12px;")
        prompt_lbl = QLabel("프롬프트:")
        prompt_lbl.setStyleSheet("font-size: 12px; color: #475569;")

        # 제목 형식 드롭다운 (동/역/구/자유) — 기본: 동
        prefix_combo = QComboBox()
        prefix_combo.addItem("동", "dong")
        prefix_combo.addItem("역", "station")
        prefix_combo.addItem("구", "gu")
        prefix_combo.addItem("자유", "free")
        prefix_combo.setStyleSheet("padding: 6px 10px; font-size: 12px;")
        prefix_lbl = QLabel("형식:")
        prefix_lbl.setStyleSheet("font-size: 12px; color: #475569;")

        free_prefix_input = QLineEdit()
        free_prefix_input.setPlaceholderText("앞부분 직접 입력")
        free_prefix_input.setStyleSheet("padding: 6px 8px; font-size: 12px;")
        free_prefix_input.setFixedWidth(140)
        free_prefix_input.setVisible(False)

        def _on_prefix_changed(idx):
            free_prefix_input.setVisible(prefix_combo.currentData() == "free")
        prefix_combo.currentIndexChanged.connect(_on_prefix_changed)

        # 사진 검색 키워드 (직접 입력 — 비우면 업종 자동 인식)
        img_kw_lbl = QLabel("사진 검색어:")
        img_kw_lbl.setStyleSheet("font-size: 12px; color: #475569;")
        img_kw_input = QLineEdit()
        img_kw_input.setPlaceholderText("예: 헬스장 (비우면 자동)")
        img_kw_input.setStyleSheet("padding: 6px 8px; font-size: 12px;")
        img_kw_input.setFixedWidth(160)
        img_kw_input.setText((getattr(self, "_override_image_keyword", "") or ""))

        btn_ok = QPushButton("포스트 생성")
        btn_ok.setStyleSheet("background: #8b5cf6; color: white; border: none; border-radius: 8px; font-size: 13px; font-weight: bold; padding: 10px 25px;")
        btn_ok.setCursor(Qt.PointingHandCursor)

        def _on_ok():
            self._override_image_keyword = img_kw_input.text().strip()
            sel = prompt_combo.currentText()
            self._override_prompt_name = sel if sel and sel != "(자동 선택)" else None
            pdata = prefix_combo.currentData()
            if pdata == "free":
                self._override_title_prefix = "free:" + free_prefix_input.text().strip()
            else:
                self._override_title_prefix = pdata
            dlg.accept()
        btn_ok.clicked.connect(_on_ok)

        # 맨 좌측에 사진 검색어 → 그다음 기존 위젯들
        bottom.addWidget(img_kw_lbl)
        bottom.addWidget(img_kw_input)
        bottom.addWidget(btn_cancel)
        bottom.addWidget(prompt_lbl)
        bottom.addWidget(prompt_combo)
        bottom.addWidget(prefix_lbl)
        bottom.addWidget(prefix_combo)
        bottom.addWidget(free_prefix_input)
        bottom.addWidget(btn_ok)
        layout.addLayout(bottom)

        _rc = dlg.exec()
        if _rc == 2:
            # 새로고침: 다이얼로그 재오픈 (mode 유지)
            return self._show_posting_selector()
        if _rc == QDialog.Accepted:
            selected = []
            for item, p in all_items:
                if item.checkState(0) == Qt.Checked:
                    # 수정된 값 반영
                    p["front_keywords"] = item.text(4)
                    p["tags"] = item.text(5)
                    p["pixabay_keywords"] = item.text(6)
                    selected.append(p)
            return selected
        return []

    # ── 블로그 포스팅 ──
    def _start_blog_posting(self):
        if self._block_if_expired():
            return
        # 크롤링 결과 불러오기 (현재 없으면 logs에서 로드)
        if not self.crawled_data:
            log_dir = self._get_logs_dir()
            if os.path.exists(log_dir):
                files = sorted(
                    [f for f in os.listdir(log_dir) if f.endswith(".json") and not f.startswith("._")],
                    reverse=True
                )
                if files:
                    from places_crawler import load_results
                    filepath = os.path.join(log_dir, files[0])
                    raw = load_results(filepath)
                    self.crawled_data = raw.get("items", []) if isinstance(raw, dict) else raw

        if not self.crawled_data:
            QMessageBox.warning(self, "경고", "먼저 크롤링을 실행해주세요.")
            return

        # 업체 선택 다이얼로그
        selected = self._show_posting_selector()
        if not selected:
            return

        cfg = load_config()
        provider = "GPT"  # 하위호환 인자
        gpt_keys = [k for k in cfg.get("gpt_key_list", []) if k]
        ds_keys = [k for k in cfg.get("deepseek_key_list", []) if k]
        gm_keys = [k for k in cfg.get("gemini_key_list", []) if k]
        ai_engine = (cfg.get("ai_engine") or "deepseek").strip().lower()
        _engkey = {"deepseek": ds_keys, "gpt": gpt_keys, "gemini": gm_keys}.get(ai_engine, [])
        if not _engkey:
            _engname = {"deepseek": "딥시크", "gpt": "챗GPT", "gemini": "제미나이"}.get(ai_engine, ai_engine)
            QMessageBox.critical(self, "API 키 필요",
                f"선택하신 글쓰기 엔진({_engname})의 API 키가 없습니다.\n\n"
                f"설정 → API 키에서 {_engname} 키를 입력하거나, 다른 엔진을 선택해주세요.")
            return
        deepseek_key = ds_keys[0] if ds_keys else None
        gpt_key = gpt_keys[0] if gpt_keys else None
        gemini_key = gm_keys[0] if gm_keys else None
        api_key = (gpt_keys or ds_keys or gm_keys)[0]  # 검색어 추출기용
        # AI 검색어 추출기(gpt-4o-mini)에 키 주입 — 실패/타임아웃 시 하드코딩 폴백
        try:
            import image_handler as _ih
            _ih.configure_ai_extractor(api_key, os.path.join(os.path.dirname(__file__), "pixabay_query_cache.json"))
        except Exception:
            pass
        from config import get_active_account
        account = get_active_account(cfg)
        if not account.get("naver_id") or not account.get("naver_pw"):
            QMessageBox.critical(self, "오류", "네이버 계정 정보를 설정해주세요.")
            return
        if self.is_posting:
            return

        # API 구독 만료 체크 (명의 슬롯별) — 만료면 "API 기간이 만료되었습니다..." 안내 후 차단
        acc_idx = cfg.get("active_account", 0) + 1
        if self._block_if_api_expired():
            return

        total = len(selected)
        interval_sec = self._get_interval_seconds()
        h = int(self.interval_hour.currentText())
        m = int(self.interval_min.currentText())
        rand_suffix = " (±20분 랜덤)" if (getattr(self, "interval_random", None) and self.interval_random.isChecked()) else ""
        interval_label = f"{h}시간 {m}분{rand_suffix}"

        reply = QMessageBox.question(self, "포스팅 확인",
            f"[아이디 {acc_idx}] {account['naver_id']}\n선택된 {total}개 업체 포스팅\n간격: {interval_label}\n\n시작할까요?")
        if reply != QMessageBox.Yes:
            return

        keyword = self.keyword_input.currentText().strip()
        self.is_posting = True
        self.stop_flag = False
        self.posting_targets = selected

        def _worker():
            poster = None
            publish_done = [0]
            try:
                self._emit_status("포스팅 중...", "#8b5cf6")
                poster = NaverBlogPoster(
                    naver_id=account["naver_id"],
                    naver_pw=account["naver_pw"],
                    blog_id=account["blog_id"],
                    window_x=0,
                    window_y=0,
                    window_w=1280,
                    window_h=900,
                    stop_flag=lambda: self.stop_flag,
                )
                self._active_posters.append(poster)
                self._emit_post_log("브라우저 시작...")
                poster.start_browser()
                try:
                    poster.driver.execute_script("window.focus();")
                    poster.driver.set_window_rect(x=80, y=60, width=1180, height=860)
                except Exception:
                    pass

                self._emit_post_log("네이버 로그인 중...")
                if not poster.login():
                    self._emit_post_log("로그인 실패!")
                    return

                self._emit_post_log("로그인 성공")

                for i, place in enumerate(self.posting_targets, 1):
                    if self.stop_flag:
                        self._emit_post_log(f"포스팅 중단완료 ({i-1}/{total}개 완료)")
                        self._emit_status("중단완료", "#ef4444")
                        break

                    name = place.get("name", "")
                    self._emit_post_log(f"[{i}/{total}] '{name}' 글 생성 중...")
                    self._emit_status(f"포스팅 {i}/{total}", "#8b5cf6")
                    self._emit_post_count(f"포스팅 {i}/{total}")

                    try:
                        content = generate_content(
                            provider=provider,
                            api_key=api_key,
                            place=place, keyword=keyword,
                            deepseek_key=deepseek_key, gpt_key=gpt_key,
                            gemini_key=gemini_key, engine=ai_engine
                        )
                        if ai_engine == "gemini":
                            import time as _tgm2; _tgm2.sleep(5)
                    except Exception as e:
                        self._emit_post_log(f"'{name}' 글 생성 실패: {e}")
                        continue

                    img_paths = []
                    pix_keys = [k for k in cfg.get("pixabay_key_list", []) if k]
                    if pix_keys:
                        try:
                            # 완전 자동: download_images 내부 체인(로컬→픽사베이)
                            biz = self._best_biz_term(place, keyword)
                            img_paths = download_images(
                                pix_keys[0], biz,
                                3,  # 무조건 3장 고정
                                watermark_text=name,
                                translator=self._translate_ko_to_en,
                            )
                        except Exception as e:
                            self._emit_post_log(f"'{name}' 이미지 다운로드 실패(이미지 없이 진행): {e}")

                    self._emit_post_log(f"[{i}/{total}] '{name}' 포스팅 중...")

                    try:
                        success = poster.write_post(
                            title=content["title"], body=content["body"],
                            tags=content["tags"], image_paths=img_paths,
                            category=account.get("blog_category", ""),
                        )
                        msg = "완료!" if success else "실패"
                        self._emit_post_log(f"[{i}/{total}] '{name}' {msg}")
                        if success:
                            publish_done[0] += 1
                            self._emit_publish_count(f"{publish_done[0]}개")
                    except Exception as e:
                        self._emit_post_log(f"'{name}' 오류: {e}")

                    if i < total and not self.stop_flag:
                        # 매 포스트마다 새 random 간격 (체크된 경우)
                        this_interval = self._get_interval_seconds()
                        if this_interval <= 0:
                            continue
                        remaining = this_interval
                        while remaining > 0 and not self.stop_flag:
                            m_left, s_left = divmod(remaining, 60)
                            self._emit_status(f"대기 {m_left}분 {s_left}초", "#94a3b8")
                            import time as _time
                            # 1초 단위로 sleep — 중단 즉시 반응
                            _time.sleep(1)
                            remaining -= 1

                if not self.stop_flag:
                    self._emit_post_log(f"전체 포스팅 완료! ({total}개)")
                    self._emit_post_count(f"{total}/{total}")

                self._emit_status("완료", "#22c55e")

            except Exception as e:
                self._emit_post_log(f"오류: {e}")
                self._emit_status("오류", "#ef4444")
            finally:
                self.is_posting = False
                if poster:
                    try: self._active_posters.remove(poster)
                    except Exception: pass
                    poster.close()

        threading.Thread(target=_worker, daemon=True).start()


LOGIN_CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "login_cache.json")


def _load_login_cache():
    try:
        with open(LOGIN_CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_login_cache(data):
    try:
        with open(LOGIN_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


class LoginDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("로그인")
        self.setFixedSize(340, 290)
        self.user = None  # 로그인 성공 시 유저 dict

        layout = QVBoxLayout(self)
        layout.setContentsMargins(25, 20, 25, 20)

        title = QLabel("블로그마스터")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #1e293b;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        layout.addSpacing(12)

        self.id_edit = QLineEdit()
        self.id_edit.setPlaceholderText("아이디")
        self.id_edit.setStyleSheet("padding: 8px; font-size: 13px;")
        self.id_edit.returnPressed.connect(self._try_login)
        layout.addWidget(self.id_edit)

        self.pw_edit = QLineEdit()
        self.pw_edit.setEchoMode(QLineEdit.Password)
        self.pw_edit.setPlaceholderText("비밀번호")
        self.pw_edit.setStyleSheet("padding: 8px; font-size: 13px;")
        self.pw_edit.returnPressed.connect(self._try_login)
        layout.addWidget(self.pw_edit)

        chk_row = QHBoxLayout()
        self.chk_save_id = QCheckBox("아이디 저장")
        self.chk_save_pw = QCheckBox("비밀번호 저장")
        self.chk_save_id.setStyleSheet("font-size: 11px; color: #64748b;")
        self.chk_save_pw.setStyleSheet("font-size: 11px; color: #64748b;")
        chk_row.addWidget(self.chk_save_id)
        chk_row.addWidget(self.chk_save_pw)
        chk_row.addStretch()
        layout.addLayout(chk_row)

        self.msg = QLabel("")
        self.msg.setStyleSheet("color: #ef4444; font-size: 11px;")
        layout.addWidget(self.msg)

        btn_row = QHBoxLayout()
        btn_login = QPushButton("로그인")
        btn_login.setStyleSheet("background: #4a6cf7; color: white; border: none; border-radius: 6px; padding: 8px 20px; font-weight: bold;")
        btn_login.setCursor(Qt.PointingHandCursor)
        btn_login.setDefault(True)
        btn_login.clicked.connect(self._try_login)
        btn_cancel = QPushButton("종료")
        btn_cancel.setStyleSheet("padding: 8px 20px;")
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(btn_login)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

        layout.addSpacing(4)
        signup_row = QHBoxLayout()
        signup_row.addStretch()
        self.find_pw_link = QLabel('<a href="#" style="color:#64748b; text-decoration:none;">비밀번호 찾기</a>')
        self.find_pw_link.setStyleSheet("font-size: 11px; color: #64748b;")
        self.find_pw_link.setCursor(Qt.PointingHandCursor)
        self.find_pw_link.setTextInteractionFlags(Qt.TextBrowserInteraction)
        self.find_pw_link.setOpenExternalLinks(False)
        self.find_pw_link.linkActivated.connect(lambda: self._open_find_pw())
        signup_row.addWidget(self.find_pw_link)
        sep = QLabel("|")
        sep.setStyleSheet("font-size: 11px; color: #cbd5e1; margin: 0 4px;")
        signup_row.addWidget(sep)
        self.signup_link = QLabel('<a href="#" style="color:#64748b; text-decoration:none;">회원가입</a>')
        self.signup_link.setStyleSheet("font-size: 11px; color: #64748b;")
        self.signup_link.setCursor(Qt.PointingHandCursor)
        self.signup_link.linkActivated.connect(
            lambda: __import__('webbrowser').open("https://n-jobs.kr/login.html?tab=signup&trial=blog"))
        signup_row.addWidget(self.signup_link)
        signup_row.addStretch()
        layout.addLayout(signup_row)

        self._load_cached()

    def _load_cached(self):
        cache = _load_login_cache()
        if cache.get("save_id"):
            self.chk_save_id.setChecked(True)
            self.id_edit.setText(cache.get("id", ""))
        if cache.get("save_pw"):
            self.chk_save_pw.setChecked(True)
            self.pw_edit.setText(cache.get("pw", ""))

    def _open_signup(self):
        dlg = SignupDialog(self)
        if dlg.exec() == QDialog.Accepted and dlg.username:
            self.id_edit.setText(dlg.username)
            self.pw_edit.setFocus()
            self.msg.setStyleSheet("color: #22c55e; font-size: 11px;")
            self.msg.setText("가입 완료. 로그인해주세요.")

    def _open_find_pw(self):
        dlg = FindPasswordDialog(self)
        dlg.exec()

    def _open_register(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("회원가입")
        dlg.setFixedSize(400, 680)
        layout = QVBoxLayout(dlg)

        title = QLabel("회원가입")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #1e293b;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        layout.addSpacing(8)

        dup_checked = {"phone": False, "email": False}
        fields = {}

        for label, key, placeholder, is_pw in [
            ("아이디 *", "uid", "영문/숫자", False),
            ("비밀번호 *", "pw", "비밀번호", True),
            ("비밀번호 확인 *", "pw2", "비밀번호 재입력", True),
            ("이름 *", "name", "홍길동", False),
            ("생년월일 *", "birth", "1990-01-01", False),
            ("연락처 *", "phone", "010-1234-5678", False),
            ("이메일 *", "email", "example@email.com", False),
            ("추천인", "referrer", "추천인 아이디 (선택)", False),
        ]:
            row = QHBoxLayout()
            row.addWidget(QLabel(label))
            row.addStretch()
            # 중복확인 버튼 (연락처/이메일)
            if key in ("phone", "email"):
                btn_dup = QPushButton("중복확인")
                btn_dup.setStyleSheet("background: #64748b; color: white; border: none; border-radius: 4px; padding: 3px 8px; font-size: 11px;")
                btn_dup.setCursor(Qt.PointingHandCursor)
                btn_dup.setProperty("field_key", key)
                row.addWidget(btn_dup)
            layout.addLayout(row)
            entry = QLineEdit()
            entry.setPlaceholderText(placeholder)
            entry.setStyleSheet("padding: 6px;")
            if is_pw:
                entry.setEchoMode(QLineEdit.Password)
            if key == "birth":
                entry.setMaxLength(10)
                def _dash_date(t, e=entry):
                    digits = t.replace("-", "")
                    if not digits.isdigit(): return
                    e.blockSignals(True)
                    digits = digits[:8]
                    if len(digits) >= 6: r = f"{digits[:4]}-{digits[4:6]}-{digits[6:]}"
                    elif len(digits) >= 4: r = f"{digits[:4]}-{digits[4:]}"
                    else: r = digits
                    e.setText(r); e.setCursorPosition(len(r))
                    e.blockSignals(False)
                entry.textChanged.connect(_dash_date)
            if key == "phone":
                entry.setMaxLength(13)
                def _dash_phone(t, e=entry):
                    digits = t.replace("-", "")
                    if not digits.isdigit(): return
                    e.blockSignals(True)
                    digits = digits[:11]
                    if len(digits) >= 8: r = f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
                    elif len(digits) >= 3: r = f"{digits[:3]}-{digits[3:]}"
                    else: r = digits
                    e.setText(r); e.setCursorPosition(len(r))
                    e.blockSignals(False)
                entry.textChanged.connect(_dash_phone)
                entry.textChanged.connect(lambda: dup_checked.update({"phone": False}))
            if key == "email":
                entry.textChanged.connect(lambda: dup_checked.update({"email": False}))
            layout.addWidget(entry)
            fields[key] = entry
            # 중복확인 연결
            if key in ("phone", "email"):
                def _check_dup(checked, k=key):
                    val = fields[k].text().strip()
                    if not val:
                        QMessageBox.warning(dlg, "경고", "값을 입력해주세요.")
                        return
                    from users import load_users
                    users = load_users()
                    taken = any(u.get(k, "") == val for u in users.values())
                    if taken:
                        QMessageBox.warning(dlg, "중복", f"이미 등록된 {('연락처' if k=='phone' else '이메일')}입니다.")
                        dup_checked[k] = False
                    else:
                        QMessageBox.information(dlg, "확인", "사용 가능합니다.")
                        dup_checked[k] = True
                btn_dup.clicked.connect(_check_dup)

        msg = QLabel("")
        msg.setStyleSheet("color: #ef4444; font-size: 11px;")
        layout.addWidget(msg)

        btn_row = QHBoxLayout()
        btn_ok = QPushButton("가입하기")
        btn_ok.setStyleSheet("background: #22c55e; color: white; border: none; border-radius: 6px; padding: 8px 20px; font-weight: bold;")
        btn_cancel = QPushButton("취소")
        btn_cancel.setStyleSheet("padding: 8px 20px;")
        btn_cancel.clicked.connect(dlg.reject)
        btn_row.addWidget(btn_ok)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

        def do_register():
            uid = fields["uid"].text().strip()
            pw = fields["pw"].text()
            name = fields["name"].text().strip()
            birth = fields["birth"].text().strip()
            phone = fields["phone"].text().strip()
            email = fields["email"].text().strip()
            referrer = fields["referrer"].text().strip()
            pw2 = fields["pw2"].text()
            if not uid or not pw or not name or not birth or not phone or not email:
                msg.setText("필수 항목을 모두 입력해주세요.")
                return
            if pw != pw2:
                msg.setText("비밀번호가 일치하지 않습니다.")
                return
            if not dup_checked["phone"]:
                msg.setText("연락처 중복확인을 해주세요.")
                return
            if not dup_checked["email"]:
                msg.setText("이메일 중복확인을 해주세요.")
                return
            import datetime as _dt
            expire_date = (_dt.date.today() + _dt.timedelta(days=30)).strftime("%Y-%m-%d")
            from users import create_user
            if not create_user(uid, pw, role="user", expires=expire_date, max_accounts=3,
                             name=name, birth=birth, phone=phone, referrer=referrer, email=email):
                msg.setText("이미 존재하는 아이디입니다.")
                return
            self.id_edit.setText(uid)
            self.pw_edit.setFocus()
            self.msg.setStyleSheet("color: #22c55e; font-size: 11px;")
            self.msg.setText("가입 완료! 로그인해주세요.")
            QMessageBox.information(dlg, "가입 완료",
                f"가입을 축하드립니다!\n\n아이디: {uid}\n무료 체험: 30일 ({expire_date}까지)\n네이버 계정: 3개 사용 가능\n\nAPI 키는 설정에서 직접 발급받아 입력해주세요.")
            dlg.accept()

        btn_ok.clicked.connect(do_register)
        dlg.exec()

    def _try_login(self):
        from users import verify, is_expired, check_session_conflict, set_session
        import uuid as _uuid
        uid = self.id_edit.text().strip()
        pw = self.pw_edit.text()
        user = verify(uid, pw)
        if not user:
            self.msg.setStyleSheet("color: #ef4444; font-size: 11px;")
            self.msg.setText("아이디 또는 비밀번호가 틀렸습니다.")
            return
        # 만료 사용자도 로그인은 허용 — 메인 화면 진입 후 모든 기능을 '이용기간 만료'로 차단한다.
        # (관리자는 is_expired에서 항상 False라 영향 없음)
        if user.get("role") != "admin" and check_session_conflict(uid):
            self.msg.setStyleSheet("color: #ef4444; font-size: 11px;")
            self.msg.setText("이미 다른 기기에서 로그인 중입니다.")
            return
        session_id = str(_uuid.uuid4())
        set_session(uid, session_id)
        cache = {
            "save_id": self.chk_save_id.isChecked(),
            "save_pw": self.chk_save_pw.isChecked(),
            "id": uid if self.chk_save_id.isChecked() else "",
            "pw": pw if self.chk_save_pw.isChecked() else "",
        }
        _save_login_cache(cache)
        self.user = {**user, "_session_id": session_id}
        self.accept()


class FindPasswordDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("비밀번호 찾기")
        self.setFixedSize(360, 300)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 20, 28, 20)
        layout.setSpacing(8)

        title = QLabel("비밀번호 찾기")
        title.setStyleSheet("font-size: 15px; font-weight: bold; color: #1e293b;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        layout.addSpacing(6)

        info = QLabel("가입 시 입력한 이름, 휴대폰번호, 이메일이 일치하면\n임시 비밀번호를 이메일로 발송합니다.")
        info.setStyleSheet("font-size: 11px; color: #64748b;")
        info.setWordWrap(True)
        layout.addWidget(info)
        layout.addSpacing(4)

        for placeholder, attr in [("이름", "name_edit"), ("휴대폰번호 (예: 01012345678)", "phone_edit"), ("이메일", "email_edit")]:
            e = QLineEdit()
            e.setPlaceholderText(placeholder)
            e.setStyleSheet("padding: 8px; font-size: 13px;")
            layout.addWidget(e)
            setattr(self, attr, e)

        self.msg = QLabel("")
        self.msg.setStyleSheet("font-size: 11px;")
        self.msg.setWordWrap(True)
        layout.addWidget(self.msg)

        btn_row = QHBoxLayout()
        btn_send = QPushButton("임시 비밀번호 발송")
        btn_send.setStyleSheet("background: #4a6cf7; color: white; border: none; border-radius: 6px; padding: 8px 16px; font-weight: bold;")
        btn_send.setCursor(Qt.PointingHandCursor)
        btn_send.clicked.connect(self._send)
        btn_cancel = QPushButton("닫기")
        btn_cancel.setStyleSheet("padding: 8px 16px;")
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(btn_send)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

    def _send(self):
        name = self.name_edit.text().strip()
        phone = self.phone_edit.text().strip()
        email = self.email_edit.text().strip()
        if not name or not phone or not email:
            self.msg.setStyleSheet("color: #ef4444; font-size: 11px;")
            self.msg.setText("이름, 휴대폰번호, 이메일을 모두 입력하세요.")
            return
        from users import find_user_by_identity, update_user
        uid = find_user_by_identity(name, phone, email)
        if not uid:
            self.msg.setStyleSheet("color: #ef4444; font-size: 11px;")
            self.msg.setText("일치하는 정보가 없습니다.")
            return
        import random, string, smtplib
        from email.mime.text import MIMEText
        from config import load_config
        cfg = load_config()
        smtp_email = cfg.get("smtp_email", "").strip()
        smtp_password = cfg.get("smtp_password", "").strip()
        if not smtp_email or not smtp_password:
            self.msg.setStyleSheet("color: #ef4444; font-size: 11px;")
            self.msg.setText("발송 이메일이 설정되지 않았습니다.\n관리자에게 문의하세요.")
            return
        new_pw = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
        try:
            msg = MIMEText(f"임시 비밀번호: {new_pw}\n\n로그인 후 반드시 비밀번호를 변경해주세요.", "plain", "utf-8")
            msg["Subject"] = "[블로그마스터] 임시 비밀번호 안내"
            msg["From"] = smtp_email
            msg["To"] = email
            with smtplib.SMTP("smtp.gmail.com", 587) as server:
                server.starttls()
                server.login(smtp_email, smtp_password)
                server.sendmail(smtp_email, email, msg.as_string())
        except Exception as e:
            self.msg.setStyleSheet("color: #ef4444; font-size: 11px;")
            self.msg.setText(f"이메일 발송 실패: {e}")
            return
        update_user(uid, password=new_pw)
        self.msg.setStyleSheet("color: #22c55e; font-size: 11px;")
        self.msg.setText(f"{email}으로 임시 비밀번호를 발송했습니다.")


class SignupDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("회원가입")
        self.setFixedSize(340, 300)
        self.username = None

        layout = QVBoxLayout(self)
        layout.setContentsMargins(25, 20, 25, 20)

        title = QLabel("회원가입")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #1e293b;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        layout.addSpacing(10)

        self.id_edit = QLineEdit()
        self.id_edit.setPlaceholderText("아이디 (3자 이상)")
        self.id_edit.setStyleSheet("padding: 8px; font-size: 13px;")
        layout.addWidget(self.id_edit)

        self.pw_edit = QLineEdit()
        self.pw_edit.setEchoMode(QLineEdit.Password)
        self.pw_edit.setPlaceholderText("비밀번호 (4자 이상)")
        self.pw_edit.setStyleSheet("padding: 8px; font-size: 13px;")
        layout.addWidget(self.pw_edit)

        self.pw2_edit = QLineEdit()
        self.pw2_edit.setEchoMode(QLineEdit.Password)
        self.pw2_edit.setPlaceholderText("비밀번호 확인")
        self.pw2_edit.setStyleSheet("padding: 8px; font-size: 13px;")
        self.pw2_edit.returnPressed.connect(self._try_signup)
        layout.addWidget(self.pw2_edit)

        self.msg = QLabel("")
        self.msg.setStyleSheet("color: #ef4444; font-size: 11px;")
        self.msg.setWordWrap(True)
        layout.addWidget(self.msg)

        btn_row = QHBoxLayout()
        btn_ok = QPushButton("가입")
        btn_ok.setStyleSheet("background: #22c55e; color: white; border: none; border-radius: 6px; padding: 8px 20px; font-weight: bold;")
        btn_ok.setCursor(Qt.PointingHandCursor)
        btn_ok.clicked.connect(self._try_signup)
        btn_cancel = QPushButton("취소")
        btn_cancel.setStyleSheet("padding: 8px 20px;")
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(btn_ok)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

    def _try_signup(self):
        from users import create_user, load_users
        uid = self.id_edit.text().strip()
        pw = self.pw_edit.text()
        pw2 = self.pw2_edit.text()
        if len(uid) < 3:
            self.msg.setText("아이디는 3자 이상이어야 합니다.")
            return
        if len(pw) < 4:
            self.msg.setText("비밀번호는 4자 이상이어야 합니다.")
            return
        if pw != pw2:
            self.msg.setText("비밀번호가 일치하지 않습니다.")
            return
        if uid in load_users():
            self.msg.setText("이미 존재하는 아이디입니다.")
            return
        expires = (datetime.date.today() + datetime.timedelta(days=30)).strftime("%Y-%m-%d")
        if not create_user(uid, pw, role="user", expires=expires):
            self.msg.setText("가입 실패. 다시 시도해주세요.")
            return
        self.username = uid
        self.accept()


class ExpiresEditor(QWidget):
    """만료일 편집기 — 달력 팝업 + +30/+90/+180/+365 + 무제한"""
    def __init__(self, expires_str: str = "", parent=None):
        super().__init__(parent)
        from PySide6.QtWidgets import QSizePolicy
        layout = QHBoxLayout(self)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(2)

        self.line = QLineEdit()
        self.line.setPlaceholderText("무제한")
        self.line.setText(expires_str or "")
        self.line.setMinimumWidth(110)
        self.line.setStyleSheet("padding: 2px 4px; font-size: 12px;")
        layout.addWidget(self.line)

        btn_cal = QPushButton("📅")
        btn_cal.setFixedWidth(28)
        btn_cal.setStyleSheet("padding: 2px;")
        btn_cal.setCursor(Qt.PointingHandCursor)
        btn_cal.clicked.connect(self._open_calendar)
        layout.addWidget(btn_cal)

        for days in (30, 90, 180, 365):
            b = QPushButton(f"+{days}")
            b.setFixedWidth(38)
            b.setStyleSheet("padding: 2px; font-size: 10px;")
            b.setCursor(Qt.PointingHandCursor)
            b.clicked.connect(lambda _, d=days: self._add_days(d))
            layout.addWidget(b)

        btn_inf = QPushButton("무제한")
        btn_inf.setFixedWidth(54)
        btn_inf.setStyleSheet("padding: 2px; font-size: 10px;")
        btn_inf.setCursor(Qt.PointingHandCursor)
        btn_inf.clicked.connect(lambda: self.line.setText("2099-12-31"))
        layout.addWidget(btn_inf)

        # 사용불가 — 칸을 비워 빈값(=차단) 상태로. 빈칸이면 placeholder에 '사용불가' 표시됨
        btn_block = QPushButton("사용불가")
        btn_block.setFixedWidth(60)
        btn_block.setStyleSheet("padding: 2px; font-size: 10px; color: #ef4444;")
        btn_block.setCursor(Qt.PointingHandCursor)
        btn_block.clicked.connect(lambda: self.line.setText(""))
        layout.addWidget(btn_block)
        layout.addStretch()

    def _open_calendar(self):
        from PySide6.QtWidgets import QCalendarWidget, QDialogButtonBox
        from PySide6.QtCore import QDate
        dlg = QDialog(self)
        dlg.setWindowTitle("만료일 선택")
        v = QVBoxLayout(dlg)
        cal = QCalendarWidget()
        txt = self.line.text().strip()
        if txt:
            d = QDate.fromString(txt, "yyyy-MM-dd")
            if d.isValid():
                cal.setSelectedDate(d)
        v.addWidget(cal)
        bbox = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bbox.accepted.connect(dlg.accept)
        bbox.rejected.connect(dlg.reject)
        v.addWidget(bbox)
        if dlg.exec() == QDialog.Accepted:
            self.line.setText(cal.selectedDate().toString("yyyy-MM-dd"))

    def _add_days(self, days: int):
        from PySide6.QtCore import QDate
        txt = self.line.text().strip()
        base = QDate.currentDate()
        if txt:
            d = QDate.fromString(txt, "yyyy-MM-dd")
            if d.isValid():
                base = d
        self.line.setText(base.addDays(days).toString("yyyy-MM-dd"))

    def get_value(self) -> str:
        return self.line.text().strip()


class MyeongExpWidget(QWidget):
    """관리자 — 3개 명의 만료일을 탭 버튼으로 전환 편집"""
    def __init__(self, exp1="", exp2=None, exp3=None, parent=None):
        super().__init__(parent)
        # None = DB에 없는 필드 → 저장 시 건드리지 않음
        self._orig = [exp1, exp2, exp3]
        self._values = [exp1 or "", exp2 or "" if exp2 is not None else "", exp3 or "" if exp3 is not None else ""]
        self._active = 0
        lay = QVBoxLayout(self)
        lay.setContentsMargins(4, 2, 4, 2)
        lay.setSpacing(2)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(3)
        self._btns = []
        for i, label in enumerate(["1명의", "2명의", "3명의"]):
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.setFixedHeight(20)
            btn.clicked.connect(lambda _, idx=i: self._select(idx))
            self._btns.append(btn)
            btn_row.addWidget(btn)
        lay.addLayout(btn_row)

        self._editor = ExpiresEditor(exp1 or "")
        lay.addWidget(self._editor)

        self._select(0)

    def _select(self, idx):
        self._values[self._active] = self._editor.get_value()
        self._active = idx
        self._editor.line.setText(self._values[idx])
        # 원래 없던 필드는 플레이스홀더를 "미구독"으로 표시
        if self._orig[idx] is None:
            self._editor.line.setPlaceholderText("미구독")
        else:
            # 빈값 = 사용불가 (무제한은 2099-12-31 날짜로 부여)
            self._editor.line.setPlaceholderText("사용불가")
        for i, btn in enumerate(self._btns):
            btn.setChecked(i == idx)
            if i == idx:
                btn.setStyleSheet("padding: 2px 6px; font-size: 10px; font-weight: bold; background: #4a6cf7; color: white; border-radius: 3px;")
            else:
                btn.setStyleSheet("padding: 2px 6px; font-size: 10px; font-weight: bold; border-radius: 3px;")

    def get_values(self):
        self._values[self._active] = self._editor.get_value()
        result = []
        for i, v in enumerate(self._values):
            # 원래 없던 필드(None)이고 여전히 비어있으면 None 반환 → update_user가 건드리지 않음
            if self._orig[i] is None and v == "":
                result.append(None)
            else:
                result.append(v)
        return result


class YesNoRadio(QWidget):
    """예/아니오 라디오 그룹 (셀 위젯용)"""
    def __init__(self, checked: bool = False, parent=None):
        super().__init__(parent)
        from PySide6.QtWidgets import QRadioButton, QButtonGroup
        h = QHBoxLayout(self)
        h.setContentsMargins(6, 0, 6, 0)
        h.setSpacing(8)
        self.yes = QRadioButton("예")
        self.no = QRadioButton("아니오")
        self.group = QButtonGroup(self)
        self.group.setExclusive(True)
        self.group.addButton(self.yes)
        self.group.addButton(self.no)
        (self.yes if checked else self.no).setChecked(True)
        h.addWidget(self.yes)
        h.addWidget(self.no)
        h.addStretch()

    def is_yes(self) -> bool:
        return self.yes.isChecked()


class AdminDialog(QDialog):
    """관리자 전용 — 사용자 계정 + 기간 관리"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("관리자 메뉴")
        self.resize(1100, 500)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("사용자 관리 (아이디 / 역할 / 명의별 만료일, 빈값=무제한)"))

        # 검색 (아이디 · 이름 · 전화번호 · 이메일)
        search_row = QHBoxLayout()
        search_row.addWidget(QLabel("🔍 검색:"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("아이디 · 이름 · 전화번호 · 이메일로 검색")
        self.search_edit.setStyleSheet("padding:6px;")
        self.search_edit.textChanged.connect(self._filter_users)
        search_row.addWidget(self.search_edit, 1)
        self.search_count = QLabel("")
        self.search_count.setStyleSheet("color:#64748b;")
        search_row.addWidget(self.search_count)
        btn_clear = QPushButton("전체보기")
        btn_clear.clicked.connect(lambda: self.search_edit.clear())
        search_row.addWidget(btn_clear)
        layout.addLayout(search_row)

        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["아이디", "역할", "명의별 만료일", "API키 부여", "키워드마스터", "네이버ID 초기화", "비밀번호 재설정"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setColumnWidth(0, 110)
        self.table.setColumnWidth(1, 80)
        self.table.setColumnWidth(2, 380)
        self.table.setColumnWidth(3, 100)
        self.table.setColumnWidth(4, 100)
        self.table.setColumnWidth(5, 210)
        self.table.verticalHeader().setDefaultSectionSize(62)
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #e2e8f0;
                selection-background-color: #e2e8f0;
                selection-color: #000;
                background: white;
            }
            QTableWidget::item:selected {
                background-color: #e2e8f0;
                color: #000;
            }
        """)
        layout.addWidget(self.table)
        self.table.cellDoubleClicked.connect(self._show_user_info)

        self._reload()

        bottom = QHBoxLayout()
        btn_add = QPushButton("+ 사용자 추가")
        btn_add.setStyleSheet("background: #22c55e; color: white; border: none; border-radius: 6px; padding: 8px 14px;")
        btn_add.clicked.connect(self._add_user)
        btn_del = QPushButton("선택 삭제")
        btn_del.setStyleSheet("background: #ef4444; color: white; border: none; border-radius: 6px; padding: 8px 14px;")
        btn_del.clicked.connect(self._del_user)
        btn_save = QPushButton("저장")
        btn_save.setStyleSheet("background: #4a6cf7; color: white; border: none; border-radius: 6px; padding: 8px 20px; font-weight: bold;")
        btn_save.clicked.connect(self._save_all)
        btn_close = QPushButton("닫기")
        btn_close.setStyleSheet("padding: 8px 20px;")
        btn_close.clicked.connect(self.reject)
        bottom.addWidget(btn_add)
        bottom.addWidget(btn_del)
        bottom.addStretch()
        bottom.addWidget(btn_save)
        bottom.addWidget(btn_close)
        layout.addLayout(bottom)

    def _filter_users(self, _text=None):
        """검색어로 사용자 행 필터 (아이디/이름/전화/이메일/추천인). 더블클릭하면 상세."""
        from users import load_users
        q = (self.search_edit.text() or "").strip().lower()
        users = load_users()
        qd = q.replace("-", "").replace(" ", "")
        shown = 0
        for r in range(self.table.rowCount()):
            it = self.table.item(r, 0)
            uid = it.text().strip() if it else ""
            if not q:
                self.table.setRowHidden(r, False)
                shown += 1
                continue
            u = users.get(uid, {}) or {}
            hay = " ".join([
                uid, str(u.get("name", "")), str(u.get("phone", "")),
                str(u.get("email", "")), str(u.get("referrer", "")),
            ]).lower()
            haed = hay.replace("-", "").replace(" ", "")
            match = (q in hay) or (qd and qd in haed)
            self.table.setRowHidden(r, not match)
            if match:
                shown += 1
        try:
            self.search_count.setText("" if not q else f"{shown}명")
        except Exception:
            pass

    def _show_user_info(self, row, _col):
        from users import load_users
        uid_item = self.table.item(row, 0)
        if not uid_item:
            return
        uid = uid_item.text().strip()
        users = load_users()
        u = users.get(uid, {})

        dlg = QDialog(self)
        dlg.setWindowTitle(f"가입 정보 — {uid}")
        dlg.setMinimumWidth(460)
        lay = QVBoxLayout(dlg)
        lay.setSpacing(10)

        read_only = [
            ("아이디",   uid),
            ("이름",     u.get("name", "") or "-"),
            ("생년월일", u.get("birth", "") or "-"),
            ("연락처",   u.get("phone", "") or "-"),
            ("추천인",   u.get("referrer", "") or "-"),
            ("가입일",   u.get("created_at", "") or "-"),
        ]
        for label, value in read_only:
            row_w = QWidget()
            row_lay = QHBoxLayout(row_w)
            row_lay.setContentsMargins(0, 0, 0, 0)
            lbl = QLabel(f"<b>{label}</b>")
            lbl.setFixedWidth(70)
            val = QLabel(value)
            val.setTextInteractionFlags(Qt.TextSelectableByMouse)
            row_lay.addWidget(lbl)
            row_lay.addWidget(val, 1)
            lay.addWidget(row_w)

        # 이메일 편집
        email_row = QWidget()
        email_lay = QHBoxLayout(email_row)
        email_lay.setContentsMargins(0, 0, 0, 0)
        email_lbl = QLabel("<b>이메일</b>")
        email_lbl.setFixedWidth(70)
        email_edit = QLineEdit(u.get("email", "") or "")
        email_edit.setPlaceholderText("example@email.com")
        email_edit.setStyleSheet("padding: 4px;")
        email_lay.addWidget(email_lbl)
        email_lay.addWidget(email_edit, 1)
        lay.addWidget(email_row)

        # ── 명의별 네이버 아이디 + 명의별 초기화 ──
        from PySide6.QtWidgets import QFrame
        line = QFrame(); line.setFrameShape(QFrame.HLine); line.setStyleSheet("color:#e2e8f0;")
        lay.addWidget(line)
        lay.addWidget(QLabel("<b>네이버 아이디 (명의별)</b>"))

        myeongui_vals = {}

        def _load_user_naver_ids():
            # 등록 계정은 users[uid].accounts (Firebase) 에 있음. 없으면 로컬 config 폴백.
            from users import load_users as _lu
            accs = list((_lu().get(uid) or {}).get("accounts") or [])
            if not any((a or {}).get("blog_id") or (a or {}).get("naver_id") for a in accs):
                try:
                    import json as _json
                    from config import CONFIG_FILE as _CF
                    if os.path.exists(_CF):
                        raw = _json.load(open(_CF, encoding="utf-8"))
                        accs = list((raw.get("accounts_by_user") or {}).get(uid) or []) or accs
                except Exception:
                    pass
            return accs

        def _refresh_myeongui():
            uaccs = _load_user_naver_ids()
            for m in range(3):
                ids = []
                for i in range(m * 3, m * 3 + 3):
                    if i < len(uaccs) and isinstance(uaccs[i], dict):
                        bid = (uaccs[i].get("blog_id") or "").strip()
                        nid = (uaccs[i].get("naver_id") or "").strip()
                        if bid or nid:
                            ids.append(bid or nid)
                myeongui_vals[m].setText("   •   ".join(ids) if ids else "(없음)")

        for m in range(3):
            box = QWidget(); h = QHBoxLayout(box); h.setContentsMargins(0, 2, 0, 2)
            lbl = QLabel(f"<b>{m+1}명의</b>"); lbl.setFixedWidth(56)
            val = QLabel("(없음)"); val.setWordWrap(True)
            val.setTextInteractionFlags(Qt.TextSelectableByMouse)
            myeongui_vals[m] = val
            rbtn = QPushButton("초기화")
            rbtn.setStyleSheet("background:#ef4444;color:white;border:none;border-radius:5px;padding:4px 10px;")
            rbtn.clicked.connect(
                lambda _, _m=m: (self._reset_naver_myeongui(uid, _m), _refresh_myeongui()))
            h.addWidget(lbl); h.addWidget(val, 1); h.addWidget(rbtn)
            lay.addWidget(box)
        _refresh_myeongui()

        msg_lbl = QLabel("")
        msg_lbl.setStyleSheet("font-size: 11px;")
        lay.addWidget(msg_lbl)

        btn_row = QHBoxLayout()
        btn_save = QPushButton("이메일 저장")
        btn_save.setStyleSheet("background: #4a6cf7; color: white; border: none; border-radius: 6px; padding: 6px 14px;")
        def _save_email():
            from users import update_user
            new_email = email_edit.text().strip()
            if update_user(uid, email=new_email):
                msg_lbl.setStyleSheet("color: #22c55e; font-size: 11px;")
                msg_lbl.setText("저장 완료")
            else:
                msg_lbl.setStyleSheet("color: #ef4444; font-size: 11px;")
                msg_lbl.setText("저장 실패")
        btn_save.clicked.connect(_save_email)
        btn_ok = QPushButton("닫기")
        btn_ok.setStyleSheet("padding: 6px 20px;")
        btn_ok.clicked.connect(dlg.accept)
        btn_row.addWidget(btn_save)
        btn_row.addStretch()
        btn_row.addWidget(btn_ok)
        lay.addLayout(btn_row)
        dlg.exec()

    def _reload(self):
        from users import load_users
        users = load_users()
        self.table.setRowCount(len(users))
        for row, (uid, u) in enumerate(sorted(users.items())):
            it_id = QTableWidgetItem(uid)
            if uid == "admin":
                it_id.setFlags(it_id.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 0, it_id)

            # 역할 드롭다운
            role_combo = QComboBox()
            role_combo.addItems(["user", "admin"])
            role_combo.setCurrentText(u.get("role", "user") if u.get("role") in ("user", "admin") else "user")
            role_combo.setStyleSheet("padding: 2px 6px;")
            self.table.setCellWidget(row, 1, role_combo)

            # 명의별 만료일 (compact 탭 위젯) — 어드민은 전 명의 무제한
            if u.get("role") == "admin":
                exp_widget = MyeongExpWidget("", "", "")
            else:
                exp_widget = MyeongExpWidget(u.get("expires", ""), u.get("expires_2"), u.get("expires_3"))
            self.table.setCellWidget(row, 2, exp_widget)

            # API키 부여 라디오 (예/아니오)
            api_toggle = YesNoRadio(checked=bool(u.get("shared_api_keys")))
            if uid == "admin":
                api_toggle.setEnabled(False)
            self.table.setCellWidget(row, 3, api_toggle)

            # 키워드마스터 권한 라디오 (예/아니오)
            km_toggle = YesNoRadio(checked=bool(u.get("keywordmaster_enabled")))
            if uid == "admin":
                km_toggle.setEnabled(False)
            self.table.setCellWidget(row, 4, km_toggle)

            # 네이버ID 초기화: 슬롯 선택 콤보 + 초기화 버튼
            reset_w = QWidget()
            reset_lay = QHBoxLayout(reset_w)
            reset_lay.setContentsMargins(4, 0, 4, 0)
            reset_lay.setSpacing(4)
            slot_combo = QComboBox()
            slot_combo.addItems([f"슬롯 {i+1}" for i in range(9)])
            slot_combo.setStyleSheet("padding: 2px 4px; font-size: 11px;")
            reset_btn = QPushButton("초기화")
            reset_btn.setStyleSheet("background: #ef4444; color: white; border: none; border-radius: 4px; padding: 4px 10px; font-size: 11px;")
            reset_btn.setCursor(Qt.PointingHandCursor)
            reset_btn.clicked.connect(lambda _, _uid=uid, _cb=slot_combo: self._reset_naver_slot(_uid, _cb.currentIndex()))
            reset_all_btn = QPushButton("전체 초기화")
            reset_all_btn.setStyleSheet("background: #b91c1c; color: white; border: none; border-radius: 4px; padding: 4px 10px; font-size: 11px; font-weight: bold;")
            reset_all_btn.setCursor(Qt.PointingHandCursor)
            reset_all_btn.clicked.connect(lambda _, _uid=uid: self._reset_naver_all(_uid))
            reset_lay.addWidget(slot_combo)
            reset_lay.addWidget(reset_btn)
            reset_lay.addWidget(reset_all_btn)
            self.table.setCellWidget(row, 5, reset_w)

            self.table.setItem(row, 6, QTableWidgetItem(""))  # 비번 재설정 입력칸

    def _reset_naver_all(self, uid: str):
        """관리자: 특정 사용자의 모든 슬롯 네이버 ID 데이터 + 잠금 전체 초기화"""
        reply = QMessageBox.question(
            self, "네이버 ID 전체 초기화",
            f"'{uid}'의 모든 슬롯(1~9) 네이버 계정 데이터를 초기화하고 잠금을 전부 해제할까요?"
        )
        if reply != QMessageBox.Yes:
            return
        try:
            import json as _json
            from config import CONFIG_FILE as _CF
            from users import update_user as _upd
            raw = {}
            if os.path.exists(_CF):
                raw = _json.load(open(_CF, encoding="utf-8"))
            abu = dict(raw.get("accounts_by_user") or {})
            empty_accs = [{"blog_id": "", "naver_id": "", "naver_pw": "", "blog_category": ""} for _ in range(9)]
            abu[uid] = empty_accs
            raw["accounts_by_user"] = abu
            with open(_CF, "w", encoding="utf-8") as f:
                _json.dump(raw, f, ensure_ascii=False, indent=2)
            _upd(uid, accounts=empty_accs, locked_naver_ids=[])
            QMessageBox.information(self, "완료", f"'{uid}' 전체 슬롯 초기화됨 (9개 슬롯 + 잠금 해제).")
        except Exception as e:
            QMessageBox.critical(self, "실패", f"전체 초기화 실패: {e}")

    def _reset_naver_myeongui(self, uid: str, myeongui: int):
        """관리자: 특정 사용자의 한 명의(아이디 3개) 전체 초기화 + 잠금 해제.
        myeongui: 0=1명의(슬롯0~2), 1=2명의(슬롯3~5), 2=3명의(슬롯6~8)."""
        reply = QMessageBox.question(
            self, f"{myeongui+1}명의 초기화",
            f"'{uid}'의 {myeongui+1}명의(아이디 3개)를 모두 초기화할까요?\n(잠금도 해제되어 다시 등록 가능)")
        if reply != QMessageBox.Yes:
            return
        try:
            import json as _json
            from config import CONFIG_FILE as _CF
            from users import load_users as _lu, update_user as _upd
            # 출처: users[uid].accounts (Firebase). 빈 dict 도 보존하며 9슬롯 보장
            ue = _lu().get(uid, {}) or {}
            accs = list(ue.get("accounts") or [])
            while len(accs) < 9:
                accs.append({"blog_id": "", "naver_id": "", "naver_pw": "", "blog_category": ""})
            removed = []
            for slot_idx in range(myeongui * 3, myeongui * 3 + 3):
                cur = accs[slot_idx] if isinstance(accs[slot_idx], dict) else {}
                nid = (cur.get("naver_id") or "").strip().lower()
                if nid:
                    removed.append(nid)
                accs[slot_idx] = {"blog_id": "", "naver_id": "", "naver_pw": "", "blog_category": ""}
            # 1) Firebase users.accounts + 잠금 해제
            locked = set((s or "").strip().lower() for s in ue.get("locked_naver_ids", []))
            for nid in removed:
                locked.discard(nid)
            _upd(uid, accounts=accs, locked_naver_ids=sorted(locked))
            # 2) 로컬 config 에도 해당 사용자 항목 있으면 동기화
            try:
                raw = {}
                if os.path.exists(_CF):
                    raw = _json.load(open(_CF, encoding="utf-8"))
                abu = dict(raw.get("accounts_by_user") or {})
                if uid in abu:
                    abu[uid] = accs
                    raw["accounts_by_user"] = abu
                    with open(_CF, "w", encoding="utf-8") as f:
                        _json.dump(raw, f, ensure_ascii=False, indent=2)
            except Exception:
                pass
            QMessageBox.information(self, "완료", f"'{uid}' {myeongui+1}명의 초기화 완료.")
        except Exception as e:
            QMessageBox.critical(self, "실패", f"초기화 실패: {e}")

    def _reset_naver_slot(self, uid: str, slot_idx: int):
        """관리자: 특정 사용자의 특정 슬롯 네이버 ID 데이터 초기화 + 잠금 해제"""
        reply = QMessageBox.question(
            self, "네이버 ID 초기화",
            f"'{uid}'의 슬롯 {slot_idx+1}번 네이버 계정 데이터를 초기화할까요?\n(잠금도 해제되어 다시 등록 가능)"
        )
        if reply != QMessageBox.Yes:
            return
        try:
            import json as _json
            from config import CONFIG_FILE as _CF
            from users import load_users as _lu, update_user as _upd
            # 1) 로컬 config의 그 슬롯 비우기
            raw = {}
            if os.path.exists(_CF):
                raw = _json.load(open(_CF, encoding="utf-8"))
            abu = dict(raw.get("accounts_by_user") or {})
            accs = list(abu.get(uid) or [])
            removed_nid = ""
            if 0 <= slot_idx < len(accs):
                removed_nid = (accs[slot_idx].get("naver_id") or "").strip().lower()
                accs[slot_idx] = {"blog_id": "", "naver_id": "", "naver_pw": "", "blog_category": ""}
                abu[uid] = accs
                raw["accounts_by_user"] = abu
                with open(_CF, "w", encoding="utf-8") as f:
                    _json.dump(raw, f, ensure_ascii=False, indent=2)
            # 2) Firebase users.accounts + locked_naver_ids 갱신
            ue = _lu().get(uid, {})
            locked = set((s or "").strip().lower() for s in ue.get("locked_naver_ids", []))
            if removed_nid in locked:
                locked.discard(removed_nid)
            _upd(uid, accounts=accs, locked_naver_ids=sorted(locked))
            QMessageBox.information(self, "완료", f"'{uid}' 슬롯 {slot_idx+1}번 초기화됨.")
        except Exception as e:
            QMessageBox.critical(self, "실패", f"초기화 실패: {e}")

    def _add_user(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("회원가입")
        dlg.setFixedSize(420, 560)
        layout = QVBoxLayout(dlg)

        fields = {}
        for label, key, placeholder, required in [
            ("아이디 *", "uid", "영문/숫자", True),
            ("비밀번호 *", "pw", "비밀번호", True),
            ("이름 *", "name", "홍길동", True),
            ("생년월일 *", "birth", "1990-01-01", True),
            ("연락처 *", "phone", "010-1234-5678", True),
            ("추천인", "referrer", "추천인 아이디 (없으면 비워두세요)", False),
        ]:
            layout.addWidget(QLabel(label))
            entry = QLineEdit()
            entry.setPlaceholderText(placeholder)
            entry.setStyleSheet("padding: 6px;")
            if key == "pw":
                entry.setEchoMode(QLineEdit.Password)
            if key == "birth":
                entry.setMaxLength(10)
                entry.textChanged.connect(lambda t, e=entry: self._auto_dash_date(e, t))
            if key == "phone":
                entry.setMaxLength(13)
                entry.textChanged.connect(lambda t, e=entry: self._auto_dash_phone(e, t))
            layout.addWidget(entry)
            fields[key] = entry

        btn_row = QHBoxLayout()
        btn_ok = QPushButton("가입")
        btn_ok.setMinimumWidth(100)
        btn_ok.setStyleSheet("background: #4a6cf7; color: white; border: none; border-radius: 6px; padding: 8px 20px; font-weight: bold; font-size: 13px;")
        btn_cancel = QPushButton("취소")
        btn_cancel.setMinimumWidth(100)
        btn_cancel.setStyleSheet("padding: 8px 20px; font-size: 13px;")
        btn_cancel.clicked.connect(dlg.reject)
        btn_row.addWidget(btn_ok)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

        def do_register():
            uid = fields["uid"].text().strip()
            pw = fields["pw"].text()
            name = fields["name"].text().strip()
            birth = fields["birth"].text().strip()
            phone = fields["phone"].text().strip()
            referrer = fields["referrer"].text().strip()
            if not uid or not pw or not name or not birth or not phone:
                QMessageBox.warning(dlg, "경고", "필수 항목을 모두 입력해주세요.")
                return
            from users import create_user
            if not create_user(uid, pw, role="user", expires="",
                             name=name, birth=birth, phone=phone, referrer=referrer):
                QMessageBox.warning(dlg, "경고", "이미 존재하는 아이디입니다.")
                return
            QMessageBox.information(dlg, "완료", f"'{uid}' 가입 완료!")
            dlg.accept()
            self._reload()

        btn_ok.clicked.connect(do_register)
        dlg.exec()

    def _auto_dash_date(self, entry, text):
        """생년월일 자동 하이픈: 19900101 → 1990-01-01"""
        digits = text.replace("-", "")
        if not digits.isdigit():
            return
        entry.blockSignals(True)
        if len(digits) > 8:
            digits = digits[:8]
        if len(digits) >= 6:
            result = f"{digits[:4]}-{digits[4:6]}-{digits[6:]}"
        elif len(digits) >= 4:
            result = f"{digits[:4]}-{digits[4:]}"
        else:
            result = digits
        entry.setText(result)
        entry.setCursorPosition(len(result))
        entry.blockSignals(False)

    def _auto_dash_phone(self, entry, text):
        """연락처 자동 하이픈: 01012345678 → 010-1234-5678"""
        digits = text.replace("-", "")
        if not digits.isdigit():
            return
        entry.blockSignals(True)
        if len(digits) > 11:
            digits = digits[:11]
        if len(digits) >= 8:
            result = f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
        elif len(digits) >= 3:
            result = f"{digits[:3]}-{digits[3:]}"
        else:
            result = digits
        entry.setText(result)
        entry.setCursorPosition(len(result))
        entry.blockSignals(False)

    def _del_user(self):
        row = self.table.currentRow()
        if row < 0:
            return
        uid = self.table.item(row, 0).text()
        if uid == "admin":
            QMessageBox.warning(self, "경고", "admin 계정은 삭제할 수 없습니다.")
            return
        if QMessageBox.question(self, "확인", f"'{uid}' 삭제?") != QMessageBox.Yes:
            return
        from users import delete_user
        delete_user(uid)
        self._reload()

    def _save_all(self):
        from users import update_user, load_users
        from config import load_config
        admin_cfg = load_config()
        admin_keys = {
            k: [v for v in admin_cfg.get(k, []) if v]
            for k in ("gpt_key_list", "gemini_key_list", "pixabay_key_list", "deepseek_key_list")
        }
        existing_users = load_users()
        for row in range(self.table.rowCount()):
            uid = self.table.item(row, 0).text().strip()
            role_widget = self.table.cellWidget(row, 1)
            role = role_widget.currentText() if role_widget else "user"
            myeong_widget = self.table.cellWidget(row, 2)
            exp_vals = myeong_widget.get_values() if myeong_widget else ["", "", ""]
            expires, expires_2, expires_3 = exp_vals[0], exp_vals[1], exp_vals[2]
            api_widget = self.table.cellWidget(row, 3)
            api_on = api_widget.is_yes() if api_widget else False
            km_widget = self.table.cellWidget(row, 4)
            km_on = km_widget.is_yes() if km_widget else False
            new_pw_item = self.table.item(row, 6)
            new_pw = (new_pw_item.text() if new_pw_item else "").strip()
            if role not in ("user", "admin"):
                role = "user"
            if uid == "admin":
                shared = None
                granted = None
            elif api_on:
                shared = {k: v for k, v in admin_keys.items() if v}
                granted = True   # 관리자가 직접 부여 → 앱이 공용키를 반영하도록 플래그 ON
            else:
                shared = {}
                granted = False
            update_user(
                uid,
                password=(new_pw if new_pw else None),
                role=role,
                expires=expires,
                expires_2=expires_2,
                expires_3=expires_3,
                shared_api_keys=shared,
                shared_api_keys_admin_granted=granted,
                keywordmaster_enabled=(True if uid != "admin" and km_on else None),
            )
        QMessageBox.information(self, "완료", "저장되었습니다.")
        self._reload()



def _ver_tuple(v: str):
    """버전 문자열을 튜플로 (예: '1.0.10' -> (1,0,10))"""
    out = []
    for x in (v or "").split("."):
        try:
            out.append(int(x))
        except ValueError:
            out.append(0)
    return tuple(out)


def _check_previous_update_result():
    """직전 인앱 패치 결과 로그 확인 — 실패면 사용자에게 알림"""
    import tempfile
    log = os.path.join(tempfile.gettempdir(), "blogmaster_update", "update_result.log")
    if not os.path.exists(log):
        return
    try:
        with open(log, "r", encoding="cp949", errors="ignore") as f:
            content = f.read().strip()
        os.remove(log)
        if content and not content.startswith("OK"):
            QMessageBox.warning(None, "업데이트 결과", f"직전 자동업데이트 실패:\n{content}")
    except Exception:
        pass


def _check_and_offer_update_pre_login(parent=None):
    """로그인 전 자동업데이트 — urllib + 부모 위젯 필수(QMessageBox None parent segfault 회피)"""
    _check_previous_update_result()
    def _log(msg):
        try:
            from app_paths import data_file
            with open(data_file("update_check.log"), "a", encoding="utf-8") as _f:
                import datetime as _dt
                _f.write(f"[{_dt.datetime.now()}] {msg}\n")
        except Exception:
            pass
    _log(f"check start, current={APP_VERSION}")
    try:
        import urllib.request as _urlreq
        import json as _json
        import ssl as _ssl
        # frozen exe에서 SSL 인증서 찾기 실패 대응 → certifi CA 번들 명시 + requests 폴백
        data = None
        try:
            import certifi as _certifi
            ctx = _ssl.create_default_context(cafile=_certifi.where())
        except Exception:
            ctx = _ssl.create_default_context()
        req = _urlreq.Request(
            "https://api.github.com/repos/kingth0506/BlogMaster/releases/latest",
            headers={"User-Agent": "BlogMaster", "Accept": "application/json"}
        )
        try:
            with _urlreq.urlopen(req, timeout=5, context=ctx) as resp:
                data = _json.loads(resp.read().decode("utf-8"))
        except Exception as e:
            _log(f"urllib fail, trying requests: {e}")
            # 2차 시도: requests (certifi 자동 사용)
            try:
                import requests as _req
                r = _req.get(
                    "https://api.github.com/repos/kingth0506/BlogMaster/releases/latest",
                    headers={"User-Agent": "BlogMaster", "Accept": "application/json"},
                    timeout=5,
                )
                r.raise_for_status()
                data = r.json()
            except Exception as e2:
                _log(f"requests fail too: {e2}")
                # 네트워크 완전 실패 시 수동 다운로드 안내 (한 번만)
                try:
                    reply = QMessageBox.question(
                        parent, "업데이트 확인 실패",
                        "인터넷 연결 또는 인증서 문제로 업데이트를 확인하지 못했습니다.\n"
                        "브라우저에서 최신 버전을 직접 다운로드하시겠습니까?",
                        QMessageBox.Yes | QMessageBox.No
                    )
                    if reply == QMessageBox.Yes:
                        import webbrowser
                        webbrowser.open("https://github.com/kingth0506/BlogMaster/releases/latest")
                except Exception:
                    pass
                return
        latest = (data.get("tag_name") or "").lstrip("v")
        _log(f"latest={latest}, current={APP_VERSION}")
        if not latest or _ver_tuple(latest) <= _ver_tuple(APP_VERSION):
            _log("up to date, no dialog")
            return
        # ── 무한루프 방지: 직전에 이 버전으로 자동패치를 시도했는데도 아직 구버전이면
        #    (= 이 PC에서 자동 업데이트 실패) 자동 재프롬프트를 멈추고 수동설치만 1회 안내 ──
        try:
            _cfg = load_config()
        except Exception:
            _cfg = {}
        if _cfg.get("update_attempted_version") == latest:
            _log("auto-update previously failed for this version -> manual guide, stop loop")
            try:
                reply = QMessageBox.question(
                    parent, "업데이트 안내",
                    f"새 버전(v{latest})이 있지만, 이 PC에서 자동 업데이트가 완료되지 않았습니다.\n"
                    "(백신 차단 · 파일 잠김 · 권한 등)\n\n"
                    "브라우저에서 최신 버전을 직접 설치하시겠습니까?",
                    QMessageBox.Yes | QMessageBox.No
                )
                if reply == QMessageBox.Yes:
                    import webbrowser
                    webbrowser.open("https://github.com/kingth0506/BlogMaster/releases/latest")
            except Exception:
                pass
            return
        try:
            reply = QMessageBox.question(
                parent, "업데이트",
                f"새 버전(v{latest})이 있습니다.\n지금 업데이트하시겠습니까?",
                QMessageBox.Yes | QMessageBox.No
            )
        except Exception as e:
            _log(f"QMessageBox fail: {e}")
            return
        if reply != QMessageBox.Yes:
            _log("user declined")
            return
        _log("user accepted, starting update")
        # 이 버전으로 자동패치 '시도'했음을 기록 → 실패로 재실행돼도 위 가드가 루프를 끊음
        try:
            _cfg["update_attempted_version"] = latest
            save_config(_cfg)
        except Exception:
            pass
        _do_in_app_update()
    except Exception as e:
        _log(f"outer exception: {e}")


def _do_in_app_update():
    """변경 파일만 다운(zip) → 임시 폴더 압축 해제 → updater.bat이 메인 종료 대기 후 교체 + 재시작"""
    import requests as _req
    import tempfile, zipfile, shutil, subprocess
    from PySide6.QtWidgets import QProgressDialog

    if getattr(sys, 'frozen', False):
        install_dir = os.path.dirname(os.path.abspath(sys.executable))
        exe_name = os.path.basename(sys.executable)
    else:
        install_dir = os.path.dirname(os.path.abspath(__file__))
        exe_name = "BlogMaster.exe"

    work_dir = os.path.join(tempfile.gettempdir(), "blogmaster_update")
    if os.path.exists(work_dir):
        shutil.rmtree(work_dir, ignore_errors=True)
    os.makedirs(work_dir, exist_ok=True)
    staging = os.path.join(work_dir, "staging")
    zip_path = os.path.join(work_dir, "BlogMaster_files.zip")

    url = "https://github.com/kingth0506/BlogMaster/releases/latest/download/BlogMaster_files.zip"

    progress = UpdateProgressDialog()
    progress.show()
    QApplication.processEvents()

    try:
        with _req.get(url, stream=True, timeout=600) as rr:
            rr.raise_for_status()
            total = int(rr.headers.get('Content-Length', 0))
            downloaded = 0
            with open(zip_path, "wb") as f:
                for chunk in rr.iter_content(chunk_size=65536):
                    if progress.wasCanceled():
                        progress.close()
                        return
                    f.write(chunk)
                    downloaded += len(chunk)
                    if total:
                        pct = int(downloaded * 100 / total)
                        mb_d = downloaded / (1024 * 1024)
                        mb_t = total / (1024 * 1024)
                        progress.setProgress(pct, f"{mb_d:.1f} / {mb_t:.1f} MB")
                    QApplication.processEvents()
        progress.setLabelText("압축 해제 중...")
        QApplication.processEvents()
        with zipfile.ZipFile(zip_path, 'r') as z:
            z.extractall(staging)
        progress.setProgress(100, "완료")
    except Exception as e:
        progress.close()
        QMessageBox.warning(None, "업데이트 실패", f"다운로드/해제 실패:\n{e}")
        return

    progress.close()

    bat_path = os.path.join(work_dir, "updater.bat")
    install_dir_w = install_dir.replace("/", "\\")
    staging_w = staging.replace("/", "\\")
    log_path = os.path.join(work_dir, "update_result.log").replace("/", "\\")
    tl_tmp = os.path.join(work_dir, "tl.txt").replace("/", "\\")
    current_pid = os.getpid()
    bat = (
        "@echo off\r\n"
        "setlocal\r\n"
        f"set LOG=\"{log_path}\"\r\n"
        f"set TARGET_PID={current_pid}\r\n"
        ":wait\r\n"
        "timeout /t 1 /nobreak >nul 2>nul\r\n"
        f"tasklist /FI \"PID eq %TARGET_PID%\" /NH > \"{tl_tmp}\" 2>nul\r\n"
        f"findstr /I /C:\"{exe_name}\" \"{tl_tmp}\" >nul 2>nul\r\n"
        "if %errorlevel%==0 goto wait\r\n"
        f"del /Q \"{tl_tmp}\" 2>nul\r\n"
        f"robocopy \"{staging_w}\" \"{install_dir_w}\" /E /R:5 /W:2 /NFL /NDL /NJH /NJS /NC /NS /NP >nul 2>nul\r\n"
        "if not errorlevel 8 goto copyok\r\n"
        "timeout /t 2 /nobreak >nul 2>nul\r\n"
        f"robocopy \"{staging_w}\" \"{install_dir_w}\" /E /R:5 /W:2 /NFL /NDL /NJH /NJS /NC /NS /NP >nul 2>nul\r\n"
        "if not errorlevel 8 goto copyok\r\n"
        "echo ROBOCOPY_FAIL errorlevel %errorlevel% > %LOG%\r\n"
        f"start \"\" \"{install_dir_w}\\{exe_name}\"\r\n"
        "del /Q \"%~f0\"\r\n"
        "exit /b 8\r\n"
        ":copyok\r\n"
        f"if not exist \"{install_dir_w}\\{exe_name}\" (\r\n"
        f"  echo MISSING_EXE after_robocopy > %LOG%\r\n"
        "  exit /b 2\r\n"
        ")\r\n"
        f"del /Q \"{install_dir_w}\\NaverBlogAuto.exe\" 2>nul\r\n"
        "del /Q \"%USERPROFILE%\\Desktop\\NaverBlogAuto.lnk\" 2>nul\r\n"
        "del /Q \"%USERPROFILE%\\OneDrive\\Desktop\\NaverBlogAuto.lnk\" 2>nul\r\n"
        "del /Q \"%APPDATA%\\Microsoft\\Windows\\Start Menu\\Programs\\NaverBlogAuto\\NaverBlogAuto.lnk\" 2>nul\r\n"
        "echo OK > %LOG%\r\n"
        f"start \"\" \"{install_dir_w}\\{exe_name}\"\r\n"
        "del /Q \"%~f0\"\r\n"
    )
    with open(bat_path, "w", encoding="cp949") as f:
        f.write(bat)

    # 설치 폴더 쓰기 권한 확인 — 없으면(관리자로 Program Files 설치 등) UAC 승격으로 bat 실행
    def _writable(p):
        try:
            _t = os.path.join(p, ".upd_wtest")
            with open(_t, "w") as _wf:
                _wf.write("x")
            os.remove(_t)
            return True
        except Exception:
            return False

    if not _writable(install_dir):
        try:
            import ctypes as _ct
            _ct.windll.shell32.ShellExecuteW(None, "runas", "cmd.exe", f'/c ""{bat_path}""', None, 0)
            sys.exit(0)
        except Exception:
            pass  # 승격 실패 시 아래 일반(숨김) 경로로 폴백

    # cmd 완전 숨김: VBS 런처로 bat을 hidden(WindowStyle=0) 실행 → 어떤 cmd 창도 안 뜸
    vbs_path = os.path.join(work_dir, "run_hidden.vbs").replace("/", "\\")
    # VBS는 큰따옴표를 두 번 써서 escape: "" → 실제 큰따옴표 1개
    vbs_content = (
        'Set WshShell = CreateObject("WScript.Shell")\r\n'
        f'WshShell.Run "cmd /c ""{bat_path}""", 0, False\r\n'
    )
    with open(vbs_path, "w", encoding="cp949") as f:
        f.write(vbs_content)
    CREATE_NO_WINDOW = 0x08000000
    si = subprocess.STARTUPINFO()
    si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    si.wShowWindow = 0
    subprocess.Popen(
        ["wscript.exe", vbs_path],
        creationflags=CREATE_NO_WINDOW,
        startupinfo=si,
        close_fds=True,
    )
    sys.exit(0)


if __name__ == "__main__":
    # 단일 인스턴스 보장 — 이미 실행 중이면 즉시 종료
    import ctypes as _ctypes
    _mutex = _ctypes.windll.kernel32.CreateMutexW(None, False, "BlogMasterSingleInstance_v1")
    if _ctypes.windll.kernel32.GetLastError() == 183:  # ERROR_ALREADY_EXISTS
        sys.exit(0)

    app = QApplication(sys.argv)
    # 작업표시줄/시작메뉴 아이콘 — AppUserModelID(독립 그룹) + 앱 아이콘 지정
    try:
        import ctypes as _ct_icon
        _ct_icon.windll.shell32.SetCurrentProcessExplicitAppUserModelID("BlogMaster.App.v1")
    except Exception:
        pass
    try:
        from PySide6.QtGui import QIcon as _QIcon_app
        # frozen(onedir)에서는 icon.ico가 _internal(_MEIPASS)에 번들되므로 exe 폴더만 보면 못 찾음 → 여러 경로 탐색
        _icon_dirs = []
        if getattr(sys, "frozen", False):
            _icon_dirs.append(os.path.dirname(sys.executable))
            _icon_dirs.append(getattr(sys, "_MEIPASS", ""))
        _icon_dirs.append(os.path.dirname(os.path.abspath(__file__)))
        for _d in _icon_dirs:
            if not _d:
                continue
            _app_icon_path = os.path.join(_d, "icon.ico")
            if os.path.exists(_app_icon_path):
                app.setWindowIcon(_QIcon_app(_app_icon_path))
                break
    except Exception:
        pass
    # Windows 다크모드 무시하고 전역 라이트 팔레트 강제
    try:
        from PySide6.QtGui import QPalette, QColor
        app.setStyle("Fusion")
        _pal = QPalette()
        _pal.setColor(QPalette.Window, QColor("#ffffff"))
        _pal.setColor(QPalette.WindowText, QColor("#1e293b"))
        _pal.setColor(QPalette.Base, QColor("#ffffff"))
        _pal.setColor(QPalette.AlternateBase, QColor("#f1f5f9"))
        _pal.setColor(QPalette.ToolTipBase, QColor("#ffffff"))
        _pal.setColor(QPalette.ToolTipText, QColor("#1e293b"))
        _pal.setColor(QPalette.Text, QColor("#1e293b"))
        _pal.setColor(QPalette.Button, QColor("#f1f5f9"))
        _pal.setColor(QPalette.ButtonText, QColor("#1e293b"))
        _pal.setColor(QPalette.BrightText, QColor("#ef4444"))
        _pal.setColor(QPalette.Link, QColor("#3b82f6"))
        _pal.setColor(QPalette.Highlight, QColor("#3b82f6"))
        _pal.setColor(QPalette.HighlightedText, QColor("#ffffff"))
        _pal.setColor(QPalette.PlaceholderText, QColor("#94a3b8"))
        app.setPalette(_pal)
    except Exception:
        pass
    app.setStyleSheet(STYLE)
    app.setFont(QFont("맑은 고딕", 10))
    # 트레이로 숨길 때 앱이 종료되지 않도록
    app.setQuitOnLastWindowClosed(False)

    # 로그인 다이얼로그 먼저 생성 → 자동업데이트의 parent로 사용 (None parent segfault 회피)
    login = LoginDialog()
    from PySide6.QtCore import QTimer
    QTimer.singleShot(500, lambda: _check_and_offer_update_pre_login(parent=login))
    if login.exec() != QDialog.Accepted or not login.user:
        sys.exit(0)

    # 로그인 유저 컨텍스트 설정 (MainWindow 생성 전에 필수 — load_config가 이걸로 유저별 accounts 분기)
    import config as _cfg
    _cfg.set_current_user(login.user.get("username", "admin"))

    window = MainWindow()
    window.apply_user_session(login.user)
    window.show()
    window.raise_()
    window.activateWindow()

    # Windows: focus 도용 방지 우회 — 강제 foreground
    def _force_foreground():
        try:
            import ctypes
            u32 = ctypes.windll.user32
            k32 = ctypes.windll.kernel32
            hwnd = int(window.winId())
            fg = u32.GetForegroundWindow()
            fg_thread = u32.GetWindowThreadProcessId(fg, None)
            my_thread = k32.GetCurrentThreadId()
            u32.AttachThreadInput(fg_thread, my_thread, True)
            u32.ShowWindow(hwnd, 9)  # SW_RESTORE
            u32.SetForegroundWindow(hwnd)
            u32.BringWindowToTop(hwnd)
            u32.AttachThreadInput(fg_thread, my_thread, False)
        except Exception:
            pass
        window.raise_()
        window.activateWindow()

    _force_foreground()
    from PySide6.QtCore import QTimer
    QTimer.singleShot(300, _force_foreground)

    sys.exit(app.exec())
