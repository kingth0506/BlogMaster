# -*- coding: utf-8 -*-
"""네이버 플레이스 블로그 자동 포스팅 — PySide6 GUI"""
APP_VERSION = "1.4.5"

import os
import sys
import faulthandler
import traceback
from datetime import datetime

# ── 크래시 로그 (frozen GUI segfault/예외 진단용) ──
_app_data = os.environ.get('LOCALAPPDATA') or os.environ.get('APPDATA') or os.path.expanduser("~")
_log_dir = os.path.join(_app_data, 'BlogMaster')
try:
    os.makedirs(_log_dir, exist_ok=True)
    _crash_log = os.path.join(_log_dir, 'crash_log.txt')
    _crash_file = open(_crash_log, 'a', encoding='utf-8')
    _crash_file.write(f"\n--- App Started v{APP_VERSION} at {datetime.now()} ---\n")
    _crash_file.flush()
    faulthandler.enable(file=_crash_file)

    def _handle_exception(exc_type, exc_value, exc_tb):
        if issubclass(exc_type, KeyboardInterrupt):
            sys.__excepthook__(exc_type, exc_value, exc_tb)
            return
        with open(_crash_log, 'a', encoding='utf-8') as _f:
            _f.write(f"\n--- Uncaught Exception at {datetime.now()} ---\n")
            traceback.print_exception(exc_type, exc_value, exc_tb, file=_f)
    sys.excepthook = _handle_exception
except Exception:
    pass

if sys.platform == "win32":
    os.environ["PYTHONUTF8"] = "1"

# frozen exe에선 cwd 변경 위험 → 개발 모드에서만 변경
if not getattr(sys, "frozen", False):
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QSpinBox, QPushButton, QComboBox, QTextEdit,
    QFrame, QMessageBox, QDialog, QScrollArea, QTableWidget,
    QTableWidgetItem, QHeaderView, QSplitter, QCheckBox,
    QTreeWidget, QTreeWidgetItem, QAbstractButton
)
from PySide6.QtCore import Qt, QThread, Signal, Slot, QObject
from PySide6.QtGui import QFont, QColor, QTextCursor, QShortcut, QKeySequence

import threading
import datetime
import json

from config import load_config, save_config
from content_generator import generate_content
from image_handler import download_images
from places_crawler import crawl_places, crawl_places_parallel, save_results, load_results, expand_keyword_to_districts
from naver_poster import NaverBlogPoster
from usage_tracker import get_remaining, add_usage, add_post


STYLE = """
QMainWindow { background: #f0f2f5; }
QLabel { color: #1e293b; }

#header { background: #4a6cf7; }
#header QLabel { color: white; font-size: 18px; font-weight: bold; }
#header QPushButton {
    background: #3b5de7; color: white; border: none; border-radius: 6px;
    padding: 6px 18px; font-weight: bold; font-size: 12px;
}
#header QPushButton:hover { background: #2d4ed8; }

#leftPanel {
    background: white; border: 1px solid #e2e8f0; border-radius: 10px;
}

#sectionLabel { font-size: 13px; font-weight: bold; color: #1e293b; }
#subLabel { font-size: 11px; color: #64748b; }

QLineEdit, QSpinBox {
    background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px;
    padding: 8px 12px; font-size: 13px; color: #1e293b;
}
QLineEdit:focus, QSpinBox:focus {
    border: 1px solid #94a3b8;
}

QComboBox {
    background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px;
    padding: 6px 10px; font-size: 12px;
}

#btnCrawl {
    background: #4a6cf7; color: white; border: none; border-radius: 8px;
    font-size: 13px; font-weight: bold; padding: 10px;
}
#btnCrawl:hover { background: #3b5de7; }

#btnStop {
    background: #94a3b8; color: white; border: none; border-radius: 8px;
    font-size: 12px; padding: 8px;
}
#btnStop:hover { background: #7c8da0; }

#btnPost {
    background: #8b5cf6; color: white; border: none; border-radius: 8px;
    font-size: 14px; font-weight: bold; padding: 12px;
}
#btnPost:hover { background: #7c3aed; }

#btnResult {
    background: transparent; color: #333; border: 1px solid #ccc; border-radius: 8px;
    font-size: 12px; padding: 8px;
}
#btnResult:hover { background: #e8e8e8; }

#logPanel {
    background: white; border: 1px solid #e2e8f0; border-radius: 10px;
}

#logHeader { font-size: 13px; font-weight: bold; }

QTextEdit {
    background: white; border: none; font-size: 12px; color: #334155;
    padding: 10px;
}
"""


class UpdateProgressDialog(QDialog):
    """프리미엄 모던 디자인 업데이트 진행 다이얼로그 (frameless + 그라디언트 + 라운드)"""
    def __init__(self, parent=None):
        super().__init__(parent)
        from PySide6.QtWidgets import QProgressBar, QGraphicsDropShadowEffect
        from PySide6.QtGui import QColor
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setFixedSize(520, 280)
        self._cancelled = False

        # 컨테이너 (그림자 + 둥근 모서리)
        container = QFrame(self)
        container.setObjectName("container")
        container.setGeometry(10, 10, 500, 260)
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(30)
        shadow.setColor(QColor(0, 0, 0, 80))
        shadow.setOffset(0, 4)
        container.setGraphicsEffect(shadow)

        container.setStyleSheet("""
            QFrame#container {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ffffff, stop:1 #f8fafc);
                border: 1px solid #e2e8f0;
                border-radius: 16px;
            }
            QLabel#title { font-size: 18px; font-weight: 800; color: #0f172a; }
            QLabel#sub { font-size: 12px; color: #64748b; }
            QLabel#pct {
                font-size: 56px; font-weight: 800;
                color: #4a6cf7;
                letter-spacing: -2px;
            }
            QLabel#status { font-size: 13px; color: #475569; }
            QProgressBar {
                border: none;
                border-radius: 6px;
                background: #e2e8f0;
                min-height: 12px;
                max-height: 12px;
            }
            QProgressBar::chunk {
                border-radius: 6px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #6366f1, stop:0.5 #4a6cf7, stop:1 #22c55e);
            }
        """)

        layout = QVBoxLayout(container)
        layout.setContentsMargins(32, 24, 32, 24)
        layout.setSpacing(8)

        title_row = QHBoxLayout()
        title_row.setSpacing(10)
        icon_lbl = QLabel("⬇")
        icon_lbl.setStyleSheet("font-size: 22px; color: #4a6cf7;")
        title_row.addWidget(icon_lbl)
        title = QLabel("업데이트 다운로드")
        title.setObjectName("title")
        title_row.addWidget(title)
        title_row.addStretch()
        layout.addLayout(title_row)

        sub = QLabel("새 버전을 받아오는 중입니다. 잠시만 기다려주세요.")
        sub.setObjectName("sub")
        layout.addWidget(sub)
        layout.addSpacing(12)

        self.pct_lbl = QLabel("0%")
        self.pct_lbl.setObjectName("pct")
        self.pct_lbl.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.pct_lbl)

        self.bar = QProgressBar()
        self.bar.setRange(0, 100)
        self.bar.setTextVisible(False)
        layout.addWidget(self.bar)
        layout.addSpacing(8)

        self.status_lbl = QLabel("준비 중...")
        self.status_lbl.setObjectName("status")
        self.status_lbl.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.status_lbl)

    def setProgress(self, pct: int, label: str = ""):
        self.pct_lbl.setText(f"{pct}%")
        self.bar.setValue(pct)
        if label:
            self.status_lbl.setText(label)

    def setLabelText(self, label: str):
        self.status_lbl.setText(label)

    def wasCanceled(self):
        return self._cancelled


class ToggleSwitch(QAbstractButton):
    """슬라이드 ON/OFF 토글 (녹색 ON / 회색 OFF, 하얀 thumb, 텍스트 표시)"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setCheckable(True)
        self.setFixedSize(64, 28)
        self.setCursor(Qt.PointingHandCursor)

    def paintEvent(self, _):
        from PySide6.QtGui import QPainter, QColor, QBrush, QPen, QFont
        from PySide6.QtCore import QRectF
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        h = self.height(); w = self.width()
        track = QColor("#22c55e") if self.isChecked() else QColor("#9ca3af")
        p.setBrush(QBrush(track))
        p.setPen(Qt.NoPen)
        p.drawRoundedRect(QRectF(0, 0, w, h), h/2, h/2)
        # 텍스트
        p.setPen(QPen(QColor("#ffffff")))
        f = QFont(); f.setBold(True); f.setPointSize(8)
        p.setFont(f)
        if self.isChecked():
            p.drawText(QRectF(0, 0, w - h, h), Qt.AlignCenter, "ON")
        else:
            p.drawText(QRectF(h, 0, w - h, h), Qt.AlignCenter, "OFF")
        # thumb
        d = h - 4
        x = w - d - 2 if self.isChecked() else 2
        p.setBrush(QBrush(QColor("#ffffff")))
        p.drawEllipse(QRectF(x, 2, d, d))


class MainWindow(QMainWindow):
    log_signal = Signal(str)
    status_signal = Signal(str, str)

    def __init__(self):
        super().__init__()
        self.setWindowTitle("블로그마스터")
        self.setMinimumSize(900, 600)
        self.resize(1200, 780)

        self.cfg = load_config()
        self.crawled_data = []
        self.is_crawling = False
        self.is_posting = False
        self.stop_flag = False
        self.result_file = ""
        self.last_keyword = self.cfg.get("last_keyword", "")
        self._active_posters = []  # 백그라운드 시 함께 최소화할 포스터 추적

        self.log_signal.connect(self._append_log)
        self.status_signal.connect(self._update_status)

        self._build_ui()
        self._migrate_legacy_logs()

        # 단축키
        QShortcut(QKeySequence("F5"), self).activated.connect(self._start_crawl)
        QShortcut(QKeySequence("F6"), self).activated.connect(self._stop)
        QShortcut(QKeySequence("F7"), self).activated.connect(self._generate_posts)
        QShortcut(QKeySequence("F8"), self).activated.connect(self._view_generated_posts)

        self._emit_log("프로그램 시작됨")
        self._emit_log("설정에서 API키와 계정 정보를 먼저 입력해주세요")

        # 시스템 트레이 아이콘 구성
        self._force_quit = False
        self._init_tray_icon()

    def _init_tray_icon(self):
        self.tray = None
        try:
            from PySide6.QtWidgets import QSystemTrayIcon, QMenu, QStyle
            from PySide6.QtGui import QIcon, QAction
            if not QSystemTrayIcon.isSystemTrayAvailable():
                return
            base_dirs = []
            if getattr(sys, 'frozen', False):
                base_dirs.append(os.path.dirname(sys.executable))
                base_dirs.append(getattr(sys, '_MEIPASS', ''))
            base_dirs.append(os.path.dirname(os.path.abspath(__file__)))
            icon = QIcon()
            for d in base_dirs:
                if not d:
                    continue
                p = os.path.join(d, "icon.ico")
                if os.path.exists(p):
                    icon = QIcon(p)
                    break
            if icon.isNull():
                try:
                    icon = self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon)
                except Exception:
                    icon = QIcon()
            self.tray = QSystemTrayIcon(icon, self)
            self.tray.setToolTip("블로그마스터")
            menu = QMenu()
            act_show = QAction("창 열기", self)
            act_show.triggered.connect(self._tray_show)
            act_quit = QAction("종료", self)
            act_quit.triggered.connect(self._tray_quit)
            menu.addAction(act_show)
            menu.addSeparator()
            menu.addAction(act_quit)
            self.tray.setContextMenu(menu)
            self.tray.activated.connect(self._on_tray_activated)
            self.tray.show()
        except Exception:
            self.tray = None

    def _on_tray_activated(self, reason):
        from PySide6.QtWidgets import QSystemTrayIcon
        if reason in (QSystemTrayIcon.Trigger, QSystemTrayIcon.DoubleClick):
            self._tray_show()

    def _tray_show(self):
        self.showNormal()
        self.raise_()
        self.activateWindow()
        # 최소화해둔 포스팅 브라우저도 복원
        for p in list(getattr(self, "_active_posters", [])):
            try:
                if getattr(p, "driver", None):
                    p.driver.execute_script("")  # ping
                    p.driver.set_window_rect(x=100, y=100, width=900, height=700)
            except Exception:
                pass

    def _tray_quit(self):
        self._force_quit = True
        QApplication.quit()

    def closeEvent(self, event):
        if getattr(self, "_force_quit", False):
            event.accept()
            return
        # X 버튼 → 3지선다 (종료/백그라운드/취소)
        if getattr(self, "tray", None) is not None:
            box = QMessageBox(self)
            box.setWindowTitle("종료 확인")
            box.setText("프로그램을 어떻게 하시겠습니까?")
            box.setIcon(QMessageBox.Question)
            btn_quit = box.addButton("종료", QMessageBox.DestructiveRole)
            btn_bg = box.addButton("백그라운드 실행", QMessageBox.AcceptRole)
            btn_cancel = box.addButton("취소", QMessageBox.RejectRole)
            box.setDefaultButton(btn_bg)
            box.exec()
            clicked = box.clickedButton()
            if clicked is btn_quit:
                self._force_quit = True
                event.accept()
                try:
                    QApplication.quit()
                except Exception:
                    pass
            elif clicked is btn_bg:
                event.ignore()
                self.hide()
                # 실행 중인 포스팅 브라우저도 함께 최소화
                for p in list(getattr(self, "_active_posters", [])):
                    try:
                        if getattr(p, "driver", None):
                            p.driver.minimize_window()
                    except Exception:
                        pass
                try:
                    self.tray.showMessage(
                        "블로그마스터",
                        "백그라운드에서 실행 중입니다. 트레이 아이콘을 우클릭하면 종료할 수 있습니다.",
                        msecs=3000,
                    )
                except Exception:
                    pass
            else:
                # 취소
                event.ignore()
        else:
            # 트레이 지원 안 되는 환경은 단순 확인
            reply = QMessageBox.question(self, "종료 확인", "프로그램을 종료하시겠습니까?",
                                          QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                event.accept()
            else:
                event.ignore()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root_layout = QVBoxLayout(central)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        # ── 헤더 ──
        header = QFrame()
        header.setObjectName("header")
        header.setFixedHeight(55)
        h_layout = QHBoxLayout(header)
        h_layout.setContentsMargins(25, 0, 20, 0)

        title = QLabel("블로그마스터")
        h_layout.addWidget(title)
        h_layout.addStretch()

        self.dash_expires = QLabel("")
        self.dash_expires.setStyleSheet("font-size: 20px; font-weight: bold; color: #000;")
        h_layout.addWidget(self.dash_expires)

        btn_prompts = QPushButton("프롬프트")
        btn_prompts.setCursor(Qt.PointingHandCursor)
        btn_prompts.clicked.connect(self._open_prompt_editor)
        h_layout.addWidget(btn_prompts)

        self.btn_admin = QPushButton("관리자")
        self.btn_admin.setStyleSheet("background: #f59e0b; color: white; border: none; border-radius: 6px; padding: 6px 14px;")
        self.btn_admin.setCursor(Qt.PointingHandCursor)
        self.btn_admin.clicked.connect(self._open_admin_dialog)
        self.btn_admin.setVisible(False)  # 기본 숨김, admin 로그인 시 노출
        h_layout.addWidget(self.btn_admin)

        btn_settings = QPushButton("설정")
        btn_settings.setCursor(Qt.PointingHandCursor)
        btn_settings.clicked.connect(self._open_settings)
        h_layout.addWidget(btn_settings)

        root_layout.addWidget(header)

        # ── 메인 ──
        main_widget = QWidget()
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(12)

        # ── 좌측 패널 ──
        left = QFrame()
        left.setObjectName("leftPanel")
        left.setFixedWidth(310)
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(18, 18, 18, 18)
        left_layout.setSpacing(5)

        # 키워드 (히스토리 드롭다운)
        lbl = QLabel("키워드 입력")
        lbl.setObjectName("sectionLabel")
        left_layout.addWidget(lbl)

        self.keyword_input = QComboBox()
        self.keyword_input.setEditable(True)
        self.keyword_input.setInsertPolicy(QComboBox.NoInsert)
        self.keyword_input.lineEdit().setPlaceholderText("키워드를 입력하세요")
        self.keyword_input.lineEdit().returnPressed.connect(self._start_crawl)
        self._load_keyword_history()
        left_layout.addWidget(self.keyword_input)

        # 지역설정 버튼 (드롭다운 트리 다이얼로그)
        self._selected_regions = list(self.cfg.get("selected_regions", []) or [])
        self.btn_regions = QPushButton()
        self.btn_regions.setCursor(Qt.PointingHandCursor)
        self.btn_regions.setStyleSheet(
            "text-align: left; padding: 6px 10px; background: #f1f5f9; "
            "border: 1px solid #cbd5e1; border-radius: 6px; font-size: 12px;"
        )
        self._refresh_regions_button()
        self.btn_regions.clicked.connect(self._open_regions_dialog)
        left_layout.addWidget(self.btn_regions)

        # 제외 키워드 버튼
        self._exclude_keywords_by_biz = dict(self.cfg.get("exclude_keywords_by_biz", {}) or {})
        self.btn_excludes = QPushButton()
        self.btn_excludes.setCursor(Qt.PointingHandCursor)
        self.btn_excludes.setStyleSheet(
            "text-align: left; padding: 6px 10px; background: #f1f5f9; "
            "border: 1px solid #cbd5e1; border-radius: 6px; font-size: 12px;"
        )
        self._refresh_excludes_button()
        self.btn_excludes.clicked.connect(self._open_excludes_dialog)
        left_layout.addWidget(self.btn_excludes)

        # 크롤링 결과보기 (지역설정/제외키워드 밑으로 이동)
        btn_result = QPushButton("크롤링 결과보기")
        btn_result.setObjectName("btnResult")
        btn_result.setCursor(Qt.PointingHandCursor)
        btn_result.clicked.connect(self._show_results)
        left_layout.addWidget(btn_result)

        left_layout.addSpacing(5)
        self._add_divider(left_layout)

        # 크롤링 설정
        lbl = QLabel("크롤링 설정")
        lbl.setObjectName("sectionLabel")
        left_layout.addWidget(lbl)

        sub = QLabel("수집 개수")
        sub.setObjectName("subLabel")
        left_layout.addWidget(sub)

        self.count_spin = QSpinBox()
        self.count_spin.setRange(1, 10000)
        self.count_spin.setValue(250)
        left_layout.addWidget(self.count_spin)

        btn_start = QPushButton("크롤링 시작 (F5)")
        btn_start.setObjectName("btnCrawl")
        btn_start.setCursor(Qt.PointingHandCursor)
        btn_start.clicked.connect(self._start_crawl)
        left_layout.addWidget(btn_start)

        btn_stop = QPushButton("중단 (F6)")
        btn_stop.setObjectName("btnStop")
        btn_stop.setCursor(Qt.PointingHandCursor)
        btn_stop.clicked.connect(self._stop)
        left_layout.addWidget(btn_stop)

        left_layout.addSpacing(5)
        self._add_divider(left_layout)

        # 포스팅 설정
        lbl = QLabel("포스팅 설정")
        lbl.setObjectName("sectionLabel")
        left_layout.addWidget(lbl)

        sub = QLabel("발행 간격")
        sub.setObjectName("subLabel")
        left_layout.addWidget(sub)

        iv_widget = QWidget()
        iv_layout = QHBoxLayout(iv_widget)
        iv_layout.setContentsMargins(0, 0, 0, 0)

        self.interval_hour = QComboBox()
        self.interval_hour.addItems([str(h) for h in range(25)])
        self.interval_hour.setCurrentText("2")
        iv_layout.addWidget(self.interval_hour)
        iv_layout.addWidget(QLabel("시간"))

        self.interval_min = QComboBox()
        self.interval_min.addItems([str(m) for m in range(0, 60, 3)])
        self.interval_min.setCurrentText("0")
        iv_layout.addWidget(self.interval_min)
        iv_layout.addWidget(QLabel("분"))
        iv_layout.addStretch()

        left_layout.addWidget(iv_widget)

        # 랜덤 발행간격 옵션 (체크 시 위 드롭다운 비활성)
        rand_widget = QWidget()
        rand_layout = QHBoxLayout(rand_widget)
        rand_layout.setContentsMargins(0, 0, 0, 0)
        self.random_interval_2h = QCheckBox("2시간 랜덤")
        self.random_interval_3h = QCheckBox("3시간 랜덤")
        self.random_interval_2h.setStyleSheet("font-size: 12px; color: #334155;")
        self.random_interval_3h.setStyleSheet("font-size: 12px; color: #334155;")
        rand_layout.addWidget(self.random_interval_2h)
        rand_layout.addWidget(self.random_interval_3h)
        rand_layout.addStretch()
        left_layout.addWidget(rand_widget)

        def _on_2h_toggle(state):
            if state == Qt.Checked.value if hasattr(Qt.Checked, 'value') else (state == 2):
                self.random_interval_3h.blockSignals(True)
                self.random_interval_3h.setChecked(False)
                self.random_interval_3h.blockSignals(False)
            randomized = self.random_interval_2h.isChecked() or self.random_interval_3h.isChecked()
            self.interval_hour.setEnabled(not randomized)
            self.interval_min.setEnabled(not randomized)

        def _on_3h_toggle(state):
            if state == Qt.Checked.value if hasattr(Qt.Checked, 'value') else (state == 2):
                self.random_interval_2h.blockSignals(True)
                self.random_interval_2h.setChecked(False)
                self.random_interval_2h.blockSignals(False)
            randomized = self.random_interval_2h.isChecked() or self.random_interval_3h.isChecked()
            self.interval_hour.setEnabled(not randomized)
            self.interval_min.setEnabled(not randomized)

        self.random_interval_2h.stateChanged.connect(_on_2h_toggle)
        self.random_interval_3h.stateChanged.connect(_on_3h_toggle)

        left_layout.addSpacing(8)

        btn_generate = QPushButton("포스트 글쓰기 (F7)")
        btn_generate.setObjectName("btnPost")
        btn_generate.setCursor(Qt.PointingHandCursor)
        btn_generate.clicked.connect(self._generate_posts)
        left_layout.addWidget(btn_generate)

        btn_view_posts = QPushButton("생성된 포스트 보기 및 업로드 (F8)")
        btn_view_posts.setStyleSheet(
            "background: #3b82f6; color: white; border: none; border-radius: 8px; "
            "font-size: 14px; font-weight: bold; padding: 10px; "
        )
        btn_view_posts.setCursor(Qt.PointingHandCursor)
        btn_view_posts.clicked.connect(self._view_generated_posts)
        left_layout.addWidget(btn_view_posts)

        btn_gallery = QPushButton("이미지 갤러리 보기")
        btn_gallery.setObjectName("btnResult")
        btn_gallery.setCursor(Qt.PointingHandCursor)
        btn_gallery.clicked.connect(self._open_image_gallery)
        left_layout.addWidget(btn_gallery)

        left_layout.addStretch()
        main_layout.addWidget(left)

        # ── 우측 패널 ──
        right = QFrame()
        right.setObjectName("logPanel")
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(15, 12, 15, 15)
        right_layout.setSpacing(5)

        # 대시보드
        dash = QFrame()
        dash.setStyleSheet("background: transparent; border: none;")
        dash_layout = QHBoxLayout(dash)
        dash_layout.setContentsMargins(4, 4, 4, 4)

        from PySide6.QtWidgets import QSizePolicy
        from PySide6.QtGui import QPixmap, QPainter, QPolygon
        from PySide6.QtCore import QPoint
        # 드롭다운 화살표 이미지 생성 (1회)
        _arrow_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_down_arrow.png").replace("\\", "/")
        if not os.path.exists(_arrow_path):
            pm = QPixmap(10, 6)
            pm.fill(Qt.transparent)
            p = QPainter(pm)
            p.setBrush(Qt.black)
            p.setPen(Qt.NoPen)
            p.drawPolygon(QPolygon([QPoint(0, 0), QPoint(10, 0), QPoint(5, 6)]))
            p.end()
            pm.save(_arrow_path)

        self.account_combo = QComboBox()
        self.account_combo.setCursor(Qt.PointingHandCursor)
        self.account_combo.setSizeAdjustPolicy(QComboBox.AdjustToMinimumContentsLengthWithIcon)
        self.account_combo.setMinimumContentsLength(6)
        self.account_combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.account_combo.setMaximumWidth(260)
        self.account_combo.setStyleSheet(f"""
            QComboBox {{
                font-size: 12px;
                color: #000;
                background-color: #ffffff;
                padding: 1px 4px;
                border: 1px solid #7f9db9;
                border-radius: 0px;
                min-height: 20px;
            }}
            QComboBox:hover {{
                background-color: #f0f8ff;
            }}
            QComboBox:focus {{
                border: 1px solid #316ac5;
            }}
            QComboBox::drop-down {{
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 18px;
                border-left: 1px solid #7f9db9;
                background-color: #ece9d8;
            }}
            QComboBox::down-arrow {{
                image: url({_arrow_path});
                width: 10px;
                height: 6px;
            }}
            QComboBox QAbstractItemView {{
                border: 1px solid #7f9db9;
                background-color: #ffffff;
                color: #000;
                selection-background-color: #316ac5;
                selection-color: #ffffff;
                font-size: 12px;
                padding: 0px;
                outline: none;
            }}
            QComboBox QAbstractItemView::item {{
                min-height: 20px;
                padding: 2px 6px;
                color: #000;
                background-color: #ffffff;
            }}
            QComboBox QAbstractItemView::item:hover {{
                background-color: #316ac5;
                color: #ffffff;
            }}
            QComboBox QAbstractItemView::item:selected {{
                background-color: #ffffff;
                color: #000;
            }}
            QComboBox QAbstractItemView::item:selected:hover {{
                background-color: #316ac5;
                color: #ffffff;
            }}
        """)
        self._refresh_account_combo()
        self.account_combo.currentIndexChanged.connect(self._on_account_changed)

        self.dash_ai = QLabel("AI: 0/1500")
        self.dash_ai.setStyleSheet("font-size: 11px; color: #64748b;")
        self.dash_pix = QLabel("Pixabay: 0/2400")
        self.dash_pix.setStyleSheet("font-size: 11px; color: #64748b;")

        dash_layout.addWidget(QLabel("계정:"))
        dash_layout.addWidget(self.account_combo)
        dash_layout.addSpacing(10)
        dash_layout.addStretch()
        dash_layout.addWidget(self.dash_ai)
        dash_layout.addWidget(self.dash_pix)
        right_layout.addWidget(dash)

        self._update_dashboard()
        self._refresh_status_panes()
        from PySide6.QtCore import QTimer as _QT
        self._status_pane_timer = _QT(self)
        self._status_pane_timer.setInterval(2000)
        self._status_pane_timer.timeout.connect(self._refresh_status_panes)
        self._status_pane_timer.start()

        # 로그 헤더
        log_hdr = QWidget()
        log_hdr_layout = QHBoxLayout(log_hdr)
        log_hdr_layout.setContentsMargins(0, 0, 0, 0)

        self.indicator = QLabel("●")
        self.indicator.setStyleSheet("color: #22c55e; font-size: 14px;")
        log_hdr_layout.addWidget(self.indicator)

        lbl = QLabel("실행 로그")
        lbl.setObjectName("logHeader")
        log_hdr_layout.addWidget(lbl)
        log_hdr_layout.addStretch()

        self.status_label = QLabel("준비")
        self.status_label.setStyleSheet("color: #22c55e; font-weight: bold; font-size: 12px;")
        log_hdr_layout.addWidget(self.status_label)

        right_layout.addWidget(log_hdr)

        # 크롤링 현황 / 포스트 현황 (위아래 분할)
        from PySide6.QtWidgets import QSplitter
        status_split = QSplitter(Qt.Vertical)

        crawl_box = QFrame()
        crawl_box.setStyleSheet("background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 6px;")
        crawl_lay = QVBoxLayout(crawl_box)
        crawl_lay.setContentsMargins(8, 6, 8, 6)
        self.crawl_status_label = QLabel("크롤링 현황")
        self.crawl_status_label.setStyleSheet("font-weight: bold; font-size: 12px; color: #1e293b;")
        crawl_lay.addWidget(self.crawl_status_label)

        # 워커별 로그 3개 (병렬 크롤 시각화)
        from PySide6.QtWidgets import QSplitter as _QS
        _worker_split = _QS(Qt.Horizontal)
        _worker_split.setChildrenCollapsible(False)
        self.crawl_worker_logs = []
        self.crawl_worker_region_labels = []
        for i in range(3):
            col = QFrame()
            col.setStyleSheet("background: #ffffff; border: 1px solid #cbd5e1; border-radius: 4px;")
            col.setMinimumWidth(120)
            cl = QVBoxLayout(col)
            cl.setContentsMargins(4, 2, 4, 4)
            cl.setSpacing(2)
            hdr_row = QHBoxLayout()
            hdr_row.setContentsMargins(0, 0, 0, 0)
            hdr = QLabel(f"{i+1}")
            hdr.setStyleSheet("font-weight: bold; font-size: 11px; color: #3b82f6;")
            hdr_row.addWidget(hdr)
            region_lbl = QLabel("")
            region_lbl.setStyleSheet("font-size: 11px; color: #64748b;")
            hdr_row.addWidget(region_lbl, 1)
            cl.addLayout(hdr_row)
            te = QTextEdit()
            te.setReadOnly(True)
            te.setStyleSheet("font-size: 11px; border: none;")
            te.setMinimumHeight(120)
            cl.addWidget(te, 1)
            _worker_split.addWidget(col)
            self.crawl_worker_logs.append(te)
            self.crawl_worker_region_labels.append(region_lbl)
        _worker_split.setSizes([300, 300, 300])
        crawl_lay.addWidget(_worker_split, 1)

        # 공용 로그 — 숨겨진 백업 (레이아웃 차지 안 함)
        self.crawl_log = QTextEdit()
        self.crawl_log.setReadOnly(True)
        self.crawl_log.hide()

        status_split.addWidget(crawl_box)

        post_box = QFrame()
        post_box.setStyleSheet("background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 6px;")
        post_lay = QVBoxLayout(post_box)
        post_lay.setContentsMargins(8, 6, 8, 6)
        self.post_status_label = QLabel("포스트 현황")
        self.post_status_label.setStyleSheet("font-weight: bold; font-size: 12px; color: #1e293b;")
        post_lay.addWidget(self.post_status_label)
        self.post_log = QTextEdit()
        self.post_log.setReadOnly(True)
        post_lay.addWidget(self.post_log)
        status_split.addWidget(post_box)

        status_split.setSizes([400, 400])
        right_layout.addWidget(status_split, stretch=1)

        # 통합 로그 (보이지 않는 백업 — 기존 _append_log 호환용)
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.hide()
        right_layout.addWidget(self.log_text)

        main_layout.addWidget(right, stretch=1)
        root_layout.addWidget(main_widget, stretch=1)

    def _add_divider(self, layout):
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setStyleSheet("color: #e2e8f0;")
        layout.addWidget(line)
        layout.addSpacing(5)

    # ── 대시보드 ──
    def _update_dashboard(self):
        cfg = load_config()
        provider = cfg.get("ai_provider", "gpt").lower()
        info = get_remaining(provider)
        self.dash_ai.setText(f"AI: {info['ai_used']}/{info['ai_limit']}")
        self.dash_pix.setText(f"Pixabay: {info['pix_used']}/{info['pix_limit']}")

    # ── 로그 ──
    def _emit_log(self, msg):
        self.log_signal.emit(msg)

    def _append_log(self, msg):
        now = datetime.datetime.now().strftime("%H:%M:%S")
        line = f"<span style='color:#9ca3af'>[{now}]</span>  {msg}"
        self.log_text.append(line)
        self.log_text.moveCursor(QTextCursor.End)
        # 키워드 기반 분배: 크롤링 → 위, 포스트 → 아래
        crawl_kw = ("수집", "스캔", "크롤", "크롤링", "봇", "워커")
        post_kw = ("포스트", "포스팅", "포스트 생성", "포스트생성", "업로드", "글 생성", "본문 생성", "제목 생성")
        try:
            # 봇별 태그 [봇N] 감지 → 해당 패널로 라우팅
            import re as _re
            m = _re.search(r"\[봇(\d+)\]", msg)
            if m and hasattr(self, "crawl_worker_logs"):
                idx = int(m.group(1)) - 1
                if 0 <= idx < len(self.crawl_worker_logs):
                    self.crawl_worker_logs[idx].append(line)
                    self.crawl_worker_logs[idx].moveCursor(QTextCursor.End)
                    # "크롤 시작" 메시지면 지역 라벨 업데이트
                    m2 = _re.search(r"\[봇\d+\]\s*(.+?)\s*크롤 시작", msg)
                    if m2 and hasattr(self, "crawl_worker_region_labels") and idx < len(self.crawl_worker_region_labels):
                        self.crawl_worker_region_labels[idx].setText(m2.group(1).strip())
                    return
            if any(k in msg for k in crawl_kw) and hasattr(self, "crawl_log"):
                self.crawl_log.append(line)
                self.crawl_log.moveCursor(QTextCursor.End)
            if any(k in msg for k in post_kw) and hasattr(self, "post_log"):
                self.post_log.append(line)
                self.post_log.moveCursor(QTextCursor.End)
        except Exception:
            pass

    def _emit_status(self, text, color="#22c55e"):
        self.status_signal.emit(text, color)

    def _update_status(self, text, color):
        self.status_label.setText(text)
        self.status_label.setStyleSheet(f"color: {color}; font-weight: bold; font-size: 12px;")
        self.indicator.setStyleSheet(f"color: {color}; font-size: 14px;")

    def _get_interval_seconds(self):
        # 2시간 랜덤: 109~133분, 3시간 랜덤: 169~193분, 둘 다 미체크: 드롭다운 값
        try:
            if getattr(self, "random_interval_2h", None) and self.random_interval_2h.isChecked():
                import random as _r
                return _r.randint(109, 133) * 60
            if getattr(self, "random_interval_3h", None) and self.random_interval_3h.isChecked():
                import random as _r
                return _r.randint(169, 193) * 60
        except Exception:
            pass
        return int(self.interval_hour.currentText()) * 3600 + int(self.interval_min.currentText()) * 60

    # ── 계정 선택 ──
    def _refresh_account_combo(self):
        accounts = self.cfg.get("accounts", [])
        active = self.cfg.get("active_account", 0)
        self.account_combo.blockSignals(True)
        self.account_combo.clear()
        for i in range(9):
            acc = accounts[i] if i < len(accounts) else {}
            nid = (acc.get("naver_id") or "").strip()
            label = f"{i+1}. {nid}" if nid else f"{i+1}. (미설정)"
            self.account_combo.addItem(label)
        if 0 <= active < self.account_combo.count():
            self.account_combo.setCurrentIndex(active)
        self.account_combo.blockSignals(False)

    def _on_account_changed(self, idx):
        if idx < 0:
            return
        cfg = load_config()
        cfg["active_account"] = idx
        save_config(cfg)
        self.cfg = cfg
        # 계정 전환 시 메모리 상태 초기화 (크롤/생성 데이터는 계정별 파일에서 다시 로드됨)
        self.crawled_data = []
        self._generated_posts = []
        self.result_file = ""
        self.keyword_input.clear()
        self._load_keyword_history()
        bid = cfg.get('accounts', [{}])[idx].get('blog_id', '') or '(미설정)'
        self._emit_log(f"계정 전환: {idx+1}번 ({bid})")

    def _current_blog_id(self) -> str:
        try:
            idx = self.cfg.get("active_account", 0)
            return self.cfg.get("accounts", [])[idx].get("blog_id", "") or f"acc{idx}"
        except Exception:
            return "default"

    # ── 키워드 히스토리 (엑셀 저장, 계정별) ──
    def _get_history_file(self):
        # app_user 기준 단일 파일 — 같은 로그인 사용자의 모든 네이버 슬롯 공유
        return os.path.join(os.path.dirname(__file__), f"search_history_{self._account_key()}.xlsx")

    def _load_keyword_history(self):
        import openpyxl
        filepath = self._get_history_file()
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    self.keyword_input.addItem(str(row[0]))
            wb.close()
        except Exception:
            pass
        # 입력칸은 비우고 드롭다운에만 히스토리 유지
        self.keyword_input.setCurrentText("")

    def _save_keyword_history(self, keyword: str):
        import openpyxl
        filepath = self._get_history_file()
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
        except Exception:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "검색 히스토리"
            ws.append(["키워드", "검색일시", "수집 결과수"])

        ws.append([keyword, now, ""])
        wb.save(filepath)
        wb.close()

        # 드롭다운 갱신 (중복 제거)
        existing = set()
        items = []
        for i in range(self.keyword_input.count()):
            items.append(self.keyword_input.itemText(i))
            existing.add(self.keyword_input.itemText(i))
        if keyword not in existing:
            self.keyword_input.insertItem(0, keyword)
        self.keyword_input.setCurrentText(keyword)

    def _update_history_result_count(self, keyword: str, count: int):
        import openpyxl
        filepath = self._get_history_file()
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            # 마지막 행에서 해당 키워드 찾아서 결과수 업데이트
            for row in reversed(list(ws.iter_rows(min_row=2))):
                if row[0].value == keyword and not row[2].value:
                    row[2].value = count
                    break
            wb.save(filepath)
            wb.close()
        except Exception:
            pass


    # ── 자동 업데이트 ──
    @Slot()
    def _show_update_dialog(self):
        reply = QMessageBox.question(self, "업데이트",
            "새 버전이 있습니다.\n지금 다운로드하고 설치하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return
        self._emit_log("업데이트 다운로드 중...")

        def _download():
            try:
                import requests as _req
                import tempfile
                url = "https://github.com/kingth0506/BlogMaster/releases/latest/download/BlogMaster_Install.exe"
                r = _req.get(url, stream=True, timeout=120)
                tmp = os.path.join(tempfile.gettempdir(), "NaverBlogAuto_Install.exe")
                with open(tmp, "wb") as f:
                    for chunk in r.iter_content(chunk_size=8192):
                        f.write(chunk)
                self._emit_log("다운로드 완료. 백그라운드 설치 중...")
                import subprocess
                subprocess.Popen(
                    [tmp, "/VERYSILENT", "/SUPPRESSMSGBOXES", "/NORESTART", "/SP-"],
                    shell=False
                )
                # 현재 프로그램 종료 (설치 후 자동 재시작됨)
                from PySide6.QtWidgets import QApplication
                QApplication.quit()
            except Exception as e:
                self._emit_log(f"업데이트 실패: {e}")

        import threading
        threading.Thread(target=_download, daemon=True).start()

    # ── 관리자 메뉴 ──
    def _open_admin_dialog(self):
        dlg = AdminDialog(self)
        dlg.exec()

    def apply_user_session(self, user: dict):
        """로그인 성공 후 호출 — 역할에 따라 관리자 버튼 노출 + 계정 슬롯 제한"""
        self.current_user = user
        try:
            self.btn_admin.setVisible(user.get("role") == "admin")
        except Exception:
            pass
        # 계정 슬롯 제한 (max_accounts: 3/6/9), 관리자는 9개 모두 허용
        if user.get("role") == "admin":
            max_acc = 9
        else:
            max_acc = int(user.get("max_accounts", 3))
        try:
            # 콤보박스 항목 수 제한
            while self.account_combo.count() > max_acc:
                self.account_combo.removeItem(self.account_combo.count() - 1)
        except Exception:
            pass
        # max_accounts 초과 슬롯 데이터 자동 삭제 (downgrade 시 7~9 슬롯 비우기)
        try:
            cfg = load_config()
            accounts = list(cfg.get("accounts", []))
            trimmed = False
            for i in range(max_acc, min(len(accounts), 9)):
                if any((accounts[i].get(k) or "").strip() for k in ("blog_id", "naver_id", "naver_pw", "blog_category")):
                    accounts[i] = {"blog_id": "", "naver_id": "", "naver_pw": "", "blog_category": ""}
                    trimmed = True
            if trimmed:
                cfg["accounts"] = accounts
                save_config(cfg)
                self.cfg = load_config()
                self._refresh_account_combo()
                self._emit_log(f"max_accounts={max_acc} 초과 슬롯 자동 삭제됨")
        except Exception:
            pass
        # 관리자: Firebase users.api_keys → 로컬 config 동기화 (다중PC 자동 적용)
        if user.get("role") == "admin":
            self._sync_admin_api_keys_from_firebase(user)
        self._update_expires_label()
        # 고아 이미지 폴더 자동 정리 (현재 생성 포스트 없는 폴더 제거)
        try:
            self._cleanup_orphan_image_folders()
        except Exception:
            pass

    def _normalize_post_images(self, pix_key: str):
        """각 포스트의 실제 저장 이미지 갯수를 image_count에 맞춰 평준화
        - 부족: Pixabay 재다운로드로 채움
        - 초과: 앞 N장만 남기고 제거
        """
        import shutil as _shutil
        fixed = 0
        for post in self._generated_posts:
            if post.get("posted"):
                continue
            place = post.get("place", {}) or {}
            content = post.get("content", {}) or {}
            target = int(content.get("image_count", 3) or 3)
            if target < 1:
                continue
            pkey = self._place_key(place)
            safe = "".join(c for c in (pkey[0] + "_" + pkey[1]) if c.isalnum() or c in "-_")[:80] or "unknown"
            pdir = os.path.join(os.path.dirname(__file__), "saved_images",
                                self._account_key(), safe)
            if not os.path.isdir(pdir):
                continue
            imgs = sorted([f for f in os.listdir(pdir)
                          if f.lower().endswith((".jpg", ".jpeg", ".png"))])
            cur = len(imgs)
            if cur == target:
                continue
            if cur > target:
                for extra in imgs[target:]:
                    try: os.remove(os.path.join(pdir, extra))
                    except Exception: pass
                content["image_paths"] = [os.path.join(pdir, f) for f in imgs[:target]]
                self._emit_log(f"[평준화] {place.get('name','')}: {cur}→{target}장 (초과 제거)")
                fixed += 1
                continue
            # 부족 → 추가 다운로드
            need = target - cur
            biz = (place.get("category") or "").strip() or "place"
            # 업종별 Pixabay 검색어 후보
            _plist = []
            try:
                _pp = os.path.join(os.path.dirname(__file__), "prompts.json")
                with open(_pp, "r", encoding="utf-8") as _pf:
                    _prompts = json.load(_pf)
                _entry = _prompts.get(biz, {}) or {}
                _plist_raw = _entry.get("pixabay_list") or []
                if not _plist_raw and _entry.get("pixabay"):
                    _plist_raw = [_entry.get("pixabay", "")]
                _plist = [s.strip() for s in _plist_raw if s and s.strip()]
            except Exception:
                _plist = []
            import random as _rnd
            added = 0
            for _i in range(need):
                if _plist:
                    _picked = _rnd.choice(_plist)
                    _query = self._translate_ko_to_en(_picked)
                else:
                    _query = biz
                try:
                    _r = download_images(pix_key, _query, 1)
                except Exception:
                    _r = []
                if not _r:
                    continue
                dst = os.path.join(pdir, f"image_{cur + added + 1}.jpg")
                try:
                    _shutil.copyfile(_r[0], dst)
                    added += 1
                except Exception:
                    pass
            # image_paths 갱신
            imgs_new = sorted([f for f in os.listdir(pdir)
                              if f.lower().endswith((".jpg", ".jpeg", ".png"))])
            content["image_paths"] = [os.path.join(pdir, f) for f in imgs_new[:target]]
            new_cur = len(imgs_new)
            if new_cur != cur:
                self._emit_log(f"[평준화] {place.get('name','')}: {cur}→{new_cur}장 (목표 {target}장)")
                fixed += 1
        if fixed:
            self._emit_log(f"이미지 평준화 완료: {fixed}개 포스트 조정")

    def _refresh_excludes_button(self):
        total = sum(len(v) for v in self._exclude_keywords_by_biz.values())
        if total == 0:
            self.btn_excludes.setText("▼ 제외 키워드 (없음)")
        else:
            biz_count = len(self._exclude_keywords_by_biz)
            self.btn_excludes.setText(f"▼ 제외 키워드 ({biz_count}업종, 총 {total}개)")

    def _save_exclude_keywords(self):
        try:
            _c = load_config()
            _c["exclude_keywords_by_biz"] = {
                k: list(v) for k, v in self._exclude_keywords_by_biz.items() if v
            }
            save_config(_c)
            self.cfg = load_config()
        except Exception:
            pass

    def _current_biz_type(self) -> str:
        """키워드 마지막 단어 = 업종"""
        kw = (self.keyword_input.currentText() or "").strip()
        parts = kw.split()
        return parts[-1] if parts else ""

    def _open_excludes_dialog(self):
        """업종별 제외 키워드 — 업종 드롭다운(+추가 버튼), 키워드 입력 + Enter/+, 목록에서 - 삭제"""
        from PySide6.QtWidgets import (
            QDialog, QVBoxLayout, QHBoxLayout, QComboBox, QPushButton,
            QLabel, QLineEdit, QScrollArea, QWidget, QInputDialog
        )

        dlg = QDialog(self)
        dlg.setWindowTitle("제외 키워드 설정")
        dlg.resize(460, 560)
        layout = QVBoxLayout(dlg)

        info = QLabel("업체명/카테고리/주소에 제외 키워드가 하나라도 포함되면\n"
                      "크롤링 결과에서 제외됩니다. 업종별로 따로 저장됩니다.")
        info.setStyleSheet("color: #475569; font-size: 11px;")
        layout.addWidget(info)

        # 업종 선택 + 업종 추가 버튼
        row = QHBoxLayout()
        row.addWidget(QLabel("업종:"))
        biz_combo = QComboBox()
        biz_combo.setEditable(False)
        all_biz = set(self._exclude_keywords_by_biz.keys())
        cur_biz = self._current_biz_type()
        if cur_biz:
            all_biz.add(cur_biz)
        for b in sorted(all_biz):
            biz_combo.addItem(b)
        if cur_biz:
            idx = biz_combo.findText(cur_biz)
            if idx >= 0:
                biz_combo.setCurrentIndex(idx)
        row.addWidget(biz_combo, 1)
        btn_add_biz = QPushButton("업종 추가")
        btn_add_biz.setStyleSheet("padding: 4px 10px;")
        btn_add_biz.setAutoDefault(False)
        btn_add_biz.setDefault(False)
        row.addWidget(btn_add_biz)
        layout.addLayout(row)

        # 키워드 추가 줄 (입력 + 버튼)
        add_row = QHBoxLayout()
        kw_input = QLineEdit()
        kw_input.setPlaceholderText("제외할 키워드 입력 후 Enter 또는 +")
        add_row.addWidget(kw_input, 1)
        btn_plus = QPushButton("+")
        btn_plus.setStyleSheet(
            "background:#22c55e;color:white;font-weight:bold;font-size:16px;"
            "padding:4px 12px;border-radius:6px;min-width:36px;"
        )
        btn_plus.setAutoDefault(False)
        btn_plus.setDefault(False)
        add_row.addWidget(btn_plus)
        layout.addLayout(add_row)

        # 키워드 리스트 (각각 - 삭제 버튼)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        canvas = QWidget()
        list_layout = QVBoxLayout(canvas)
        list_layout.setContentsMargins(6, 6, 6, 6)
        list_layout.setSpacing(4)
        list_layout.addStretch()
        scroll.setWidget(canvas)
        layout.addWidget(scroll, 1)

        bottom = QHBoxLayout()
        bottom.addStretch()
        btn_close = QPushButton("닫기")
        btn_close.setStyleSheet("padding:6px 14px;")
        btn_close.setAutoDefault(False)
        btn_close.setDefault(False)
        btn_close.clicked.connect(dlg.accept)
        bottom.addWidget(btn_close)
        layout.addLayout(bottom)

        def _current_list() -> list:
            biz = biz_combo.currentText().strip()
            if not biz:
                return []
            return list(self._exclude_keywords_by_biz.get(biz, []))

        def _persist(items: list):
            biz = biz_combo.currentText().strip()
            if not biz:
                return
            cleaned = []
            for k in items:
                k = (k or "").strip()
                if k and k not in cleaned:
                    cleaned.append(k)
            if cleaned:
                self._exclude_keywords_by_biz[biz] = cleaned
            else:
                self._exclude_keywords_by_biz.pop(biz, None)
            self._save_exclude_keywords()
            self._refresh_excludes_button()

        def _render():
            # 기존 아이템 위젯 지우기 (마지막 stretch 유지)
            while list_layout.count() > 1:
                it = list_layout.takeAt(0)
                w = it.widget()
                if w:
                    w.setParent(None)
            items = _current_list()
            for idx, kw in enumerate(items):
                row_w = QWidget()
                row_l = QHBoxLayout(row_w)
                row_l.setContentsMargins(0, 0, 0, 0)
                lbl = QLabel(kw)
                lbl.setStyleSheet(
                    "background:#f1f5f9;border:1px solid #e2e8f0;"
                    "border-radius:6px;padding:6px 10px;"
                )
                row_l.addWidget(lbl, 1)
                btn_del = QPushButton("−")
                btn_del.setStyleSheet(
                    "background:#ef4444;color:white;font-weight:bold;"
                    "padding:4px 12px;border-radius:6px;min-width:36px;"
                )
                btn_del.setCursor(Qt.PointingHandCursor)
                def _mk_del(k=kw):
                    def _do():
                        cur = _current_list()
                        cur = [x for x in cur if x != k]
                        _persist(cur)
                        _render()
                    return _do
                btn_del.clicked.connect(_mk_del())
                row_l.addWidget(btn_del)
                list_layout.insertWidget(list_layout.count() - 1, row_w)

        def _add_keyword():
            kw = (kw_input.text() or "").strip()
            if not kw:
                return
            cur = _current_list()
            if kw not in cur:
                cur.append(kw)
                _persist(cur)
                _render()
            kw_input.clear()
            kw_input.setFocus()

        def _add_biz():
            text, ok = QInputDialog.getText(dlg, "업종 추가", "업종명:")
            if not ok:
                return
            text = (text or "").strip()
            if not text:
                return
            if biz_combo.findText(text) < 0:
                biz_combo.addItem(text)
            biz_combo.setCurrentText(text)
            # 빈 목록으로 초기화 (저장은 키워드 추가 시)
            _render()

        kw_input.returnPressed.connect(_add_keyword)
        btn_plus.clicked.connect(_add_keyword)
        btn_add_biz.clicked.connect(_add_biz)
        biz_combo.currentTextChanged.connect(lambda _: _render())

        _render()
        dlg.exec()

    def _refresh_regions_button(self):
        n = len(self._selected_regions)
        if n == 0:
            self.btn_regions.setText("▼ 지역설정 (선택 없음 — 키워드만 사용)")
        else:
            preview = ", ".join(self._selected_regions[:3])
            more = f" 외 {n-3}" if n > 3 else ""
            self.btn_regions.setText(f"▼ 지역설정 ({n}개: {preview}{more})")

    def _save_selected_regions(self):
        try:
            _c = load_config()
            _c["selected_regions"] = list(self._selected_regions)
            save_config(_c)
            self.cfg = load_config()
        except Exception:
            pass

    def _open_regions_dialog(self):
        """시/도(1차) → 구/군(2차) 트리 체크박스 다이얼로그"""
        from PySide6.QtWidgets import (
            QDialog, QVBoxLayout, QHBoxLayout, QTreeWidget, QTreeWidgetItem,
            QPushButton, QLabel
        )
        from places_crawler import SIDO_DISTRICTS

        dlg = QDialog(self)
        dlg.setWindowTitle("지역 설정")
        dlg.resize(420, 600)
        layout = QVBoxLayout(dlg)

        info = QLabel("크롤링할 시/도 및 구/군을 체크하세요.\n"
                      "선택 시 키워드 앞에 자동 부여되어 각 지역별로 크롤링됩니다.")
        info.setStyleSheet("color: #475569; font-size: 11px;")
        layout.addWidget(info)

        tree = QTreeWidget()
        tree.setHeaderHidden(True)
        layout.addWidget(tree, 1)

        selected_set = set(self._selected_regions)
        sido_items = {}
        for sido, dists in SIDO_DISTRICTS.items():
            parent = QTreeWidgetItem(tree, [sido])
            parent.setFlags(parent.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsAutoTristate)
            parent.setCheckState(0, Qt.Unchecked)
            sido_items[sido] = parent
            for d in dists:
                child = QTreeWidgetItem(parent, [d])
                child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
                key = f"{sido} {d}"
                child.setCheckState(0, Qt.Checked if key in selected_set else Qt.Unchecked)
        tree.expandAll()

        # 시/도 부모 클릭 → 그 아래 구/군 전체 토글 (서울 전체 등 일괄 선택 편의)
        _propagating = {"on": False}
        def _on_item_changed(item, col):
            if col != 0 or _propagating["on"]:
                return
            if item.parent() is None and item.childCount() > 0:
                st = item.checkState(0)
                if st == Qt.PartiallyChecked:
                    return
                _propagating["on"] = True
                try:
                    for i in range(item.childCount()):
                        item.child(i).setCheckState(0, st)
                finally:
                    _propagating["on"] = False
        tree.itemChanged.connect(_on_item_changed)

        btn_row = QHBoxLayout()
        btn_all = QPushButton("전체 선택")
        btn_none = QPushButton("전체 해제")
        btn_ok = QPushButton("확인")
        btn_cancel = QPushButton("취소")
        btn_ok.setStyleSheet("background:#3b82f6;color:white;padding:6px 14px;border-radius:6px;")
        btn_cancel.setStyleSheet("padding:6px 14px;")

        def _set_all(state):
            for p in sido_items.values():
                for i in range(p.childCount()):
                    p.child(i).setCheckState(0, state)
        btn_all.clicked.connect(lambda: _set_all(Qt.Checked))
        btn_none.clicked.connect(lambda: _set_all(Qt.Unchecked))
        btn_cancel.clicked.connect(dlg.reject)

        def _accept():
            picked = []
            for sido, parent in sido_items.items():
                for i in range(parent.childCount()):
                    ch = parent.child(i)
                    if ch.checkState(0) == Qt.Checked:
                        picked.append(f"{sido} {ch.text(0)}")
            self._selected_regions = picked
            self._save_selected_regions()
            self._refresh_regions_button()
            dlg.accept()
        btn_ok.clicked.connect(_accept)

        btn_row.addWidget(btn_all)
        btn_row.addWidget(btn_none)
        btn_row.addStretch()
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_ok)
        layout.addLayout(btn_row)

        dlg.exec()

    def _open_image_gallery(self):
        """saved_images/{user}/ 아래 포스트별 이미지를 Lazy Load 갤러리로 표시"""
        from PySide6.QtWidgets import (
            QDialog, QVBoxLayout, QHBoxLayout, QScrollArea, QLabel,
            QFrame, QComboBox, QPushButton
        )
        from PySide6.QtCore import QTimer
        from PySide6.QtGui import QPixmap, QFont

        THUMB = 160
        COLS = 6

        class LazyThumb(QLabel):
            def __init__(self, path):
                super().__init__()
                self.image_path = path
                self.loaded = False
                self.setFixedSize(THUMB, THUMB)
                self.setStyleSheet("background: #e5e7eb; border: 1px solid #d1d5db;")
                self.setAlignment(Qt.AlignCenter)
                self.setText("...")

            def load_image(self):
                if self.loaded:
                    return
                pm = QPixmap(self.image_path)
                if pm.isNull():
                    self.setText("!")
                    return
                self.setPixmap(pm.scaled(THUMB, THUMB, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                self.loaded = True

        dlg = QDialog(self)
        dlg.setWindowTitle("이미지 갤러리")
        dlg.resize(1200, 800)
        root = QVBoxLayout(dlg)

        top = QHBoxLayout()
        top.addWidget(QLabel("포스트 폴더:"))
        info_label = QLabel("")
        top.addStretch()
        top.addWidget(info_label)
        root.addLayout(top)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        root.addWidget(scroll, 1)
        canvas = QWidget()
        grid_layout = QVBoxLayout(canvas)
        grid_layout.setSpacing(16)
        scroll.setWidget(canvas)

        thumbs = []

        def load_visible():
            viewport = scroll.viewport()
            vp_top = scroll.verticalScrollBar().value()
            vp_bottom = vp_top + viewport.height()
            preload = 200
            for t in thumbs:
                if t.loaded:
                    continue
                pos = t.mapTo(canvas, t.rect().topLeft())
                y_top = pos.y()
                y_bottom = y_top + t.height()
                if y_bottom + preload < vp_top:
                    continue
                if y_top - preload > vp_bottom:
                    continue
                t.load_image()

        scroll.verticalScrollBar().valueChanged.connect(lambda _: load_visible())

        base = os.path.join(os.path.dirname(__file__), "saved_images", self._account_key())
        if not os.path.isdir(base):
            info_label.setText("저장된 이미지 폴더 없음")
            dlg.exec()
            return

        post_folders = sorted([f for f in os.listdir(base)
                               if os.path.isdir(os.path.join(base, f))])
        total = 0
        for fname in post_folders:
            fpath = os.path.join(base, fname)
            imgs = sorted([f for f in os.listdir(fpath)
                          if f.lower().endswith((".jpg", ".jpeg", ".png"))])
            if not imgs:
                continue
            section = QFrame()
            section.setStyleSheet("background: #f9fafb; border-radius: 8px; padding: 8px;")
            sec_layout = QVBoxLayout(section)
            title = QLabel(f"{fname}  ({len(imgs)}장)")
            title.setFont(QFont("맑은 고딕", 11, QFont.Bold))
            sec_layout.addWidget(title)

            row_layout = QHBoxLayout()
            row_layout.setSpacing(8)
            sec_layout.addLayout(row_layout)
            col = 0
            for img in imgs:
                img_path = os.path.join(fpath, img)
                thumb = LazyThumb(img_path)
                thumbs.append(thumb)
                total += 1
                row_layout.addWidget(thumb)
                col += 1
                if col >= COLS:
                    row_layout.addStretch()
                    col = 0
                    row_layout = QHBoxLayout()
                    row_layout.setSpacing(8)
                    sec_layout.addLayout(row_layout)
            if col > 0:
                row_layout.addStretch()
            grid_layout.addWidget(section)

        grid_layout.addStretch()
        info_label.setText(f"{len(post_folders)}개 포스트, {total}장 이미지")
        QTimer.singleShot(50, load_visible)
        dlg.exec()

    def _cleanup_orphan_image_folders(self):
        """saved_images/{user}/ 아래 폴더 중 현재 generated_posts에 없는 건 삭제"""
        import shutil as _shutil
        base = os.path.join(os.path.dirname(__file__), "saved_images", self._account_key())
        if not os.path.isdir(base):
            return
        # 현재 유효한 place_key → 폴더명 매핑
        valid_names = set()
        for gp in self._load_generated_posts():
            pkey = self._place_key(gp.get("place", {}))
            safe = "".join(c for c in (pkey[0] + "_" + pkey[1]) if c.isalnum() or c in "-_")[:80] or "unknown"
            valid_names.add(safe)
        removed = 0
        for name in os.listdir(base):
            folder = os.path.join(base, name)
            if not os.path.isdir(folder):
                continue
            if name in valid_names:
                continue
            try:
                _shutil.rmtree(folder)
                removed += 1
            except Exception:
                pass
        if removed:
            self._emit_log(f"이미지 폴더 자동 정리: {removed}개 고아 폴더 삭제")

    def _sync_admin_api_keys_from_firebase(self, user: dict):
        """admin 로그인 시 Firebase의 api_keys를 로컬 config.json의 api_keys_by_user에 덮어쓰기"""
        api_keys = user.get("api_keys") or {}
        if not api_keys:
            return
        username = user.get("username", "admin")
        try:
            import json as _json
            from config import CONFIG_FILE as _CF, load_config as _lc
            raw = {}
            if os.path.exists(_CF):
                with open(_CF, "r", encoding="utf-8") as f:
                    raw = _json.load(f)
            akbu = dict(raw.get("api_keys_by_user") or {})
            akbu[username] = {k: list(v or []) for k, v in api_keys.items()}
            raw["api_keys_by_user"] = akbu
            with open(_CF, "w", encoding="utf-8") as f:
                _json.dump(raw, f, ensure_ascii=False, indent=2)
            self.cfg = _lc()
        except Exception as e:
            self._emit_log(f"admin 키 동기화 실패: {e}")

    def _update_expires_label(self):
        if not hasattr(self, "dash_expires"):
            return
        user = getattr(self, "current_user", {}) or {}
        expires = (user.get("expires") or "").strip()
        if not expires:
            self.dash_expires.setText("[무제한]")
            self.dash_expires.setStyleSheet("font-size: 20px; font-weight: bold; color: #000;")
            return
        try:
            import datetime as _dt
            y, m, d = map(int, expires.split("-"))
            days = (_dt.date(y, m, d) - _dt.date.today()).days
        except Exception:
            return
        if days < 0:
            text = "[만료]"
            color = "#ef4444"
        else:
            text = f"[잔여 {days}일]"
            color = "#22c55e" if days >= 31 else "#f59e0b" if days >= 8 else "#ef4444"
        self.dash_expires.setText(text)
        self.dash_expires.setStyleSheet(f"font-size: 20px; font-weight: bold; color: {color};")

    # ── 설정 ──
    def _open_settings(self):
        from settings_dialog import SettingsDialog
        cu = getattr(self, "current_user", {}) or {}
        is_admin = cu.get("role") == "admin"
        username = cu.get("username", "admin")
        dlg = SettingsDialog(self, is_admin=is_admin, app_user=username)
        dlg.exec()
        self.cfg = load_config()
        self._refresh_account_combo()

    # ── 한영 번역 캐시 (Pixabay 검색어용) ──
    def _get_translation_cache_file(self):
        return os.path.join(os.path.dirname(__file__), "translation_cache.json")

    def _load_translation_cache(self) -> dict:
        fp = self._get_translation_cache_file()
        if os.path.exists(fp):
            try:
                with open(fp, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return {}

    def _save_translation_cache(self, cache: dict):
        fp = self._get_translation_cache_file()
        try:
            with open(fp, "w", encoding="utf-8") as f:
                json.dump(cache, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _translate_ko_to_en(self, text: str) -> str:
        """한글 검색어를 영어로 번역 (AI 사용, 결과 캐싱)"""
        text = (text or "").strip()
        if not text:
            return ""
        # 이미 영어면 그대로
        if not any(0xAC00 <= ord(c) <= 0xD7A3 for c in text):
            return text
        cache = self._load_translation_cache()
        if text in cache:
            return cache[text]
        # AI로 번역
        try:
            cfg = self.cfg
            api_keys = [k for k in cfg.get("gpt_key_list", []) if k]
            if not api_keys:
                return text
            prompt = (
                f"Translate this Korean search term to a concise English keyword (max 3 words) suitable for Pixabay image search. "
                f"Output ONLY the English term, no explanation, no quotes.\n\n"
                f"Korean: {text}\nEnglish:"
            )
            from openai import OpenAI
            client = OpenAI(api_key=api_keys[0])
            resp = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=20,
            )
            en = (resp.choices[0].message.content or "").strip()
            # 정제: 따옴표/마침표 제거, 한 줄로
            en = en.split("\n")[0].strip().strip('"\'`.')
            if en:
                cache[text] = en
                self._save_translation_cache(cache)
                return en
        except Exception as _e:
            self._emit_log(f"번역 실패 ('{text}'): {_e}")
        return text

    # ── 업종별 프롬프트 자동 생성 (크롤 완료 후 호출) ──
    def _auto_generate_prompt_for_biz(self, biz_type: str):
        """키워드의 업종이 prompts.json에 없으면 GPT로 블로그/제목 프롬프트 생성 후 저장.
        백그라운드 스레드에서 실행, 이미 있으면 skip."""
        biz_type = (biz_type or "").strip()
        if not biz_type or biz_type == "기본":
            return
        prompts_path = os.path.join(os.path.dirname(__file__), "prompts.json")
        try:
            with open(prompts_path, "r", encoding="utf-8") as f:
                prompts = json.load(f)
        except Exception:
            prompts = {}
        # 이미 있는 업종이면 skip (정확 매칭 + 부분 매칭)
        if biz_type in prompts:
            return
        for key in prompts:
            if key == "기본":
                continue
            if key in biz_type or biz_type in key:
                return  # 유사 업종 이미 있음
        cfg = load_config()
        api_keys = [k for k in cfg.get("gpt_key_list", []) if k]
        if not api_keys:
            return
        base = prompts.get("기본", {"blog": "", "title": ""})
        meta_prompt = (
            f"아래는 블로그 방문 후기 자동 생성용 프롬프트 템플릿을 만드는 작업이다.\n"
            f"'{biz_type}' 업종에 맞는 **블로그 본문 프롬프트**와 **제목 프롬프트**를 생성한다.\n\n"
            f"★★ 절대 규칙 ★★\n"
            f"1) '{{업종}}', '{{업체명}}', '{{주소}}', '{{근처역}}', '{{카테고리}}', '{{앞키워드}}', '{{태그}}', '{{키워드}}' 같은 "
            f"중괄호 플레이스홀더는 반드시 그대로 유지한다.\n"
            f"2) **그 시설을 검색해서 방문할 만한 사람의 관점·페르소나**로 후기를 쓰도록 프롬프트를 구성한다.\n"
            f"3) 기본 템플릿이 가진 제약(사실 기반 작성 / 지어내기 금지 / 이모티콘 금지 등)은 유지한다.\n"
            f"4) 분량 기준(공백 포함 1500자 안팎, 1400자 미만 금지, 최소 40문장)은 그대로 유지한다.\n\n"
            f"=== 기존 기본 템플릿 (제약·형식 참고용) ===\n[BLOG]\n{base.get('blog','')}\n\n[TITLE]\n{base.get('title','')}\n\n"
            f"=== 출력 형식 (반드시 이 구분자 사용) ===\n"
            f"===BLOG===\n(여기에 '{biz_type}' 업종용 블로그 본문 프롬프트 전문)\n"
            f"===TITLE===\n(여기에 '{biz_type}' 업종용 제목 프롬프트 전문)\n"
        )

        def _worker():
            try:
                self._emit_log(f"업종 프롬프트 자동 생성 시작: '{biz_type}'")
                from openai import OpenAI
                client = OpenAI(api_key=api_keys[0])
                resp = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[{"role": "user", "content": meta_prompt}],
                )
                text = resp.choices[0].message.content or ""
                import re as _re
                m_blog = _re.search(r"===\s*BLOG\s*===\s*(.*?)\s*===\s*TITLE\s*===", text, _re.S)
                m_title = _re.search(r"===\s*TITLE\s*===\s*(.*)", text, _re.S)
                blog_p = (m_blog.group(1).strip() if m_blog else "")
                title_p = (m_title.group(1).strip() if m_title else "")
                if not blog_p or not title_p:
                    self._emit_log(f"  [경고] '{biz_type}' 프롬프트 응답 파싱 실패 — skip")
                    return
                try:
                    with open(prompts_path, "r", encoding="utf-8") as _f:
                        cur = json.load(_f)
                except Exception:
                    cur = {}
                if biz_type in cur:  # 레이스 조건 방지
                    return
                cur[biz_type] = {
                    "blog": blog_p,
                    "title": title_p,
                    "title_prefix": "dong",
                    "pixabay_list": ["", "", ""],
                }
                with open(prompts_path, "w", encoding="utf-8") as _f:
                    json.dump(cur, _f, ensure_ascii=False, indent=2)
                self._emit_log(f"  '{biz_type}' 프롬프트 자동 생성 완료 (prompts.json 저장)")
            except Exception as _e:
                self._emit_log(f"  '{biz_type}' 프롬프트 자동 생성 실패: {_e}")

        threading.Thread(target=_worker, daemon=True).start()

    # ── 프롬프트 편집 ──
    def _open_prompt_editor(self):
        prompts_path = os.path.join(os.path.dirname(__file__), "prompts.json")
        try:
            with open(prompts_path, "r", encoding="utf-8") as f:
                prompts = json.load(f)
        except Exception:
            prompts = {}
        if not prompts:
            prompts = {"기본": {"blog": "", "title": ""}}

        dlg = QDialog(self)
        dlg.setWindowTitle("프롬프트 편집")
        dlg.resize(900, 700)
        layout = QVBoxLayout(dlg)

        top = QHBoxLayout()
        top.addWidget(QLabel("업종 선택:"))
        type_combo = QComboBox()
        type_combo.addItems(list(prompts.keys()))
        top.addWidget(type_combo, 1)

        btn_add = QPushButton("+ 업종 추가")
        btn_auto = QPushButton("AI 자동 생성")
        btn_auto.setStyleSheet("background: #22c55e; color: white; border-radius: 6px; padding: 4px 10px;")
        btn_del = QPushButton("삭제")
        btn_del.setStyleSheet("background: #ef4444; color: white; border-radius: 6px; padding: 4px 10px;")
        top.addWidget(btn_add)
        top.addWidget(btn_auto)
        top.addWidget(btn_del)
        layout.addLayout(top)

        layout.addWidget(QLabel("블로그 본문 프롬프트:"))
        blog_edit = QTextEdit()
        blog_edit.setAcceptRichText(False)
        layout.addWidget(blog_edit, 3)

        layout.addWidget(QLabel("제목 프롬프트:"))
        title_edit = QTextEdit()
        title_edit.setAcceptRichText(False)
        layout.addWidget(title_edit, 1)

        # 제목 키워드 형식 선택 (업종별)
        prefix_row = QHBoxLayout()
        prefix_row.addWidget(QLabel("제목 키워드 형식:"))
        prefix_combo = QComboBox()
        prefix_combo.addItems(["XX동 + 업종", "XX역 + 업종", "XX구 + 업종"])
        prefix_row.addWidget(prefix_combo, 1)
        layout.addLayout(prefix_row)

        # Pixabay 검색어 입력 (슬롯 3개, 채운 것 중 랜덤 선택 / 전부 비면 키워드 기반 자동)
        pix_outer = QVBoxLayout()
        pix_outer.addWidget(QLabel("Pixabay 검색어 (빈 칸은 키워드로 자동):"))
        pix_row = QHBoxLayout()
        pix_edit = QLineEdit()
        pix_edit.setPlaceholderText("1번 사진 (예: fitness)")
        pix_edit2 = QLineEdit()
        pix_edit2.setPlaceholderText("2번 사진 (예: workout)")
        pix_edit3 = QLineEdit()
        pix_edit3.setPlaceholderText("3번 사진 (예: dumbbell)")
        pix_row.addWidget(pix_edit, 1)
        pix_row.addWidget(pix_edit2, 1)
        pix_row.addWidget(pix_edit3, 1)
        pix_outer.addLayout(pix_row)
        layout.addLayout(pix_outer)

        current = {"key": type_combo.currentText()}

        def load_current():
            key = type_combo.currentText()
            data = prompts.get(key, {"blog": "", "title": "", "pixabay_list": []})
            blog_edit.blockSignals(True); title_edit.blockSignals(True)
            pix_edit.blockSignals(True); pix_edit2.blockSignals(True); pix_edit3.blockSignals(True)
            blog_edit.setPlainText(data.get("blog", ""))
            title_edit.setPlainText(data.get("title", ""))
            # 제목 키워드 형식
            tp = (data.get("title_prefix") or "dong").strip().lower()
            if tp == "station":
                prefix_combo.setCurrentIndex(1)
            elif tp == "gu":
                prefix_combo.setCurrentIndex(2)
            else:
                prefix_combo.setCurrentIndex(0)
            # 구버전 호환: pixabay (단일) + pixabay_list (3개) 둘 다 지원
            pix_list = data.get("pixabay_list") or []
            if not pix_list and data.get("pixabay"):
                pix_list = [data.get("pixabay", "")]
            pix_list = (pix_list + ["", "", ""])[:3]
            pix_edit.setText(pix_list[0])
            pix_edit2.setText(pix_list[1])
            pix_edit3.setText(pix_list[2])
            blog_edit.blockSignals(False); title_edit.blockSignals(False)
            pix_edit.blockSignals(False); pix_edit2.blockSignals(False); pix_edit3.blockSignals(False)
            current["key"] = key

        def save_current_to_memory():
            if current["key"] in prompts:
                _tp_map = {0: "dong", 1: "station", 2: "gu"}
                prompts[current["key"]] = {
                    "blog": blog_edit.toPlainText(),
                    "title": title_edit.toPlainText(),
                    "title_prefix": _tp_map.get(prefix_combo.currentIndex(), "dong"),
                    "pixabay_list": [pix_edit.text().strip(), pix_edit2.text().strip(), pix_edit3.text().strip()],
                }

        def on_type_changed():
            save_current_to_memory()
            load_current()

        type_combo.currentIndexChanged.connect(lambda _i: on_type_changed())

        def add_type():
            from PySide6.QtWidgets import QInputDialog
            name, ok = QInputDialog.getText(dlg, "업종 추가", "업종명:")
            name = (name or "").strip()
            if ok and name and name not in prompts:
                save_current_to_memory()
                prompts[name] = {"blog": "", "title": "", "pixabay_list": ["", "", ""]}
                type_combo.addItem(name)
                type_combo.setCurrentText(name)

        def del_type():
            key = type_combo.currentText()
            if not key:
                return
            if len(prompts) <= 1:
                QMessageBox.warning(dlg, "경고", "최소 1개는 남겨야 합니다.")
                return
            if QMessageBox.question(dlg, "확인", f"'{key}' 업종을 삭제할까요?") != QMessageBox.Yes:
                return
            prompts.pop(key, None)
            type_combo.removeItem(type_combo.currentIndex())
            load_current()

        # AI 응답 수신 시그널 (워커 스레드에서 GUI로 전달)
        class _AISignal(QObject):
            done = Signal(str, str)  # text, error
        ai_sig = _AISignal()

        def _on_ai_done(text, error):
            btn_auto.setEnabled(True)
            btn_auto.setText("AI 자동 생성")
            if error:
                QMessageBox.critical(dlg, "오류", f"AI 호출 실패: {error}")
                return
            import re
            m_blog = re.search(r"===\s*BLOG\s*===\s*(.*?)\s*===\s*TITLE\s*===", text, re.S)
            m_title = re.search(r"===\s*TITLE\s*===\s*(.*)", text, re.S)
            blog_p = (m_blog.group(1).strip() if m_blog else "")
            title_p = (m_title.group(1).strip() if m_title else "")
            if not blog_p and not title_p:
                QMessageBox.warning(dlg, "경고", f"응답 파싱 실패. 전체 응답:\n{text[:500]}")
                return
            if blog_p:
                blog_edit.setPlainText(blog_p)
            if title_p:
                title_edit.setPlainText(title_p)
            QMessageBox.information(dlg, "완료", "프롬프트 자동 생성 완료. 확인 후 '저장'을 눌러주세요.")

        ai_sig.done.connect(_on_ai_done)

        def auto_generate():
            key = type_combo.currentText()
            if not key:
                return
            cfg2 = load_config()
            provider = "GPT"  # GPT 전용
            api_keys = [k for k in cfg2.get("gpt_key_list", []) if k]
            if not api_keys:
                QMessageBox.critical(dlg, "오류", "GPT API 키를 설정해주세요.")
                return
            base = prompts.get("기본", {"blog": "", "title": ""})
            ref_example = """# 역할
너는 부모님을 직접 모시지 못한다는 무거운 마음을 안고 수개월간 요양원을 찾아 헤맸던 **'현실적인 효자/효녀 블로거'**다. 단순한 정보 나열이 아니라, 같은 처지에 놓인 자녀들이 밤잠 설치며 고민하는 지점을 정확히 짚어주고, 그들의 죄책감을 덜어주면서도 **'전문적인 사람'**으로서의 선택임을 강조하는 따뜻한 조언자다
글은 광고 느낌이 아니라, 실제로 부모님을 맡길 요양원을 고르면서 고민했던 사람이 지인에게 조용히 털어놓는 이야기처럼 진심 있게 작성한다.
기본은 존댓말을 사용하고, **1인칭(저, 우리 가족)** 을 자연스럽게 사용해도 된다.
※ 이 프롬프트에서 가장 중요한 규칙은 **'분량'과 '문장 수'**이다. 다른 어떤 규칙보다 **글자 수·문장 수 조건을 우선해서** 지킨다. 글이 짧다고 느껴지면, 이미 쓴 내용을 더 구체적인 장면·감정·예시로 풀어 써서라도 분량을 반드시 채운다.

# 작업 지시
아래에 제공된 '입력 데이터'를 활용하여 **부모님 요양원 선택 과정과 실제 이용 후기**를 블로그 본문 형식으로 작성한다.
* 전체 글자 수는 한글 기준 **공백 포함 1500자 안팎**으로 맞춘다.
* **1450자 이상 1650자 이하** 범위에서 작성하고, **1400자 미만이 되지 않도록** 충분히 내용을 확장한다.
글의 전체적인 흐름은 다음 구조를 따른다.
1. **서론** — 방문 계기/동기, 처음의 감정(죄책감·불안 등)
2. **본론** — 검색·상담·방문 과정 / 비교 기준 / 최종 선택 요양원의 위치·접근성 / 시설·프로그램·돌봄 환경 / 부모님의 변화, 가족의 심리 변화
3. **결론** — 만족도·느낀 점 / "나도 이렇게 선택했다" 식의 부드러운 상담/방문 권유

# 입력 데이터
* 요양원명(업체명): {업체명}
* 주소: {주소}
* 근처역/교통: {근처역}
* 카테고리: {카테고리}
* 앞 키워드: {앞키워드}
* 태그: {태그}
(앞 키워드와 태그는 글의 주제와 분위기를 파악하기 위한 참고용이다.)

# 준수 사항
## 1. 형식 및 문장 구성
* **문장마다 줄바꿈**한다. 한 문장은 한 줄, 문장 끝마다 줄바꿈.
* 서론 → 본론 → 결론 흐름에 맞춰 배치한다.
* 전체 문장 수는 **최소 40문장 이상**.
* 각 문장은 너무 짧지 않게, **대략 25자 이상**이 되도록 상황·감정·설명을 충분히 풀어 쓴다.

## 2. 분량
* 공백 포함 1500자 안팎(1450~1650자). 1400자 미만 금지.
* 부족하면 구체적 예시·상황·대화·표정·몸짓 묘사로 늘린다.

## 3. 내용 구성 (각 파트 최소 문장 수)
1) 서론 — 처음 상황과 감정 (최소 8문장)
2) 검색·상담·방문 과정 (최소 10문장, 실제 상담 질문 3가지 이상 풀어쓰기)
3) 여러 곳 비교 + 최종 선택 기준 (최소 7문장)
4) 위치·접근성·주변 환경 (최소 5문장)
5) 시설·프로그램·돌봄 세부 묘사 (최소 8문장)
6) 부모님 변화·가족 심리 변화 (최소 6문장, 시간 흐름 단계적 서술)
7) 장점/아쉬운 점 + 결론 (최소 8문장)

## 4. 어조
* 일상 대화체 + 진심 담은 후기체.
* "솔직히 처음에는…", "막상 다녀와 보니까…" 같은 표현 적절히.
* 자극·과장 없이 담담하지만 진심 있는 톤.

## 5. 금지 사항
* 이모티콘 금지.
* '카테고리', '앞 키워드', '태그'라는 단어 본문 언급 금지.
* **광고·홍보·이벤트·할인·"무조건 여기로 오세요" 식 문구 절대 금지.**
* 가격·비용·금액 구체 언급 금지(간접 언급만 가능).

## 6. 정보 출처 표현
* '크롤링', '검색 엔진', '데이터 수집' 같은 단어 본문 사용 금지.
* 본인이 직접 상담·방문해서 느낀 후기처럼 자연스럽게 표현.

## 이미지 배치
* 본문 흐름상 사진이 들어가면 좋을 위치에 '[이미지]' 마커를 해당 줄 단독으로 3~5개 삽입.
"""
            ref_title = """# 역할
너는 대한민국의 블로그 포스팅 전문가다. 너의 임무는 주어진 키워드에 대해 사용자의 클릭을 유발하는 매력적인 블로그 제목의 '마무리 문구'를 생성하는 것이다.

# 작업 설명
1. '핵심 키워드'의 특성을 분석한다.
2. 분석한 내용을 바탕으로, 블로그 제목의 마지막 부분에 위치할 '마무리 문구'를 생성한다.
3. 생성되는 문구는 아래 '카테고리'를 참고하여 다양하게 구성한다.
4. 최종 결과물은 리스트 형태로, '마무리 문구'만 간결하게 제공한다.

# 블로그 제목 구조
[핵심 키워드] + [업체명] + [생성된 마무리 문구]

# 카테고리
- 정보성: 가격, 위치, 영업시간, 주차, 예약 방법 등 사실 정보
- 장점 및 특징: 시설, 전문성, 특별한 메뉴나 서비스, 분위기
- 경험 및 후기: 솔직 후기, 내돈내산, 방문 경험, 추천 이유
- 질문 및 호기심 유발: ~방법?, ~후기, ~꿀팁, ~총정리

# 예시
- 핵심 키워드: "강남역 파스타 맛집"
- 결과물 예시:
  데이트 코스로 강력 추천
  내돈내산 솔직 후기
  메뉴와 가격 총정리
  웨이팅 없이 즐기는 꿀팁
  분위기 좋은 곳 찾는다면


# 지시
아래 '핵심 키워드'에 대한 '마무리 문구' 10개를 생성해라.
마무리 문구는 20자 이내, 지역명/구/동/역/업종명/업체명은 절대 포함하지 않는다.
숫자나 번호 없이 한 줄씩 10개.

핵심 키워드: {키워드}
"""
            meta_prompt = (
                f"아래는 블로그 방문 후기 자동 생성용 프롬프트 템플릿을 만드는 작업이다.\n"
                f"'{key}' 업종에 맞는 **블로그 본문 프롬프트**와 **제목 프롬프트**를 생성한다.\n\n"
                f"★★ 절대 규칙 ★★\n"
                f"1) '{{업종}}', '{{업체명}}', '{{주소}}', '{{근처역}}', '{{카테고리}}', '{{앞키워드}}', '{{태그}}', '{{키워드}}' 같은 "
                f"중괄호 플레이스홀더는 반드시 그대로 유지한다 (한글명이든 영문이든 기존 그대로).\n"
                f"2) **그 시설을 검색해서 방문할 만한 사람의 관점·페르소나**로 후기를 쓰도록 프롬프트를 구성한다.\n"
                f"   - 예: 헬스장 → 운동하러 가려는 본인 입장 (체력/근력/감량 목표, 첫 등록 고민, 운동 초보 관점 등)\n"
                f"   - 예: 요양원/요양센터/요양병원 → 부모님/가족을 모시려는 자녀 입장 (시설 안전, 직원 태도, 어르신 편의, 면회 환경)\n"
                f"   - 예: 학원 → 자녀 보낼 학부모 또는 본인 수강생 입장\n"
                f"   - 예: 병원/치과 → 본인이나 가족 진료 받으러 간 입장\n"
                f"   - 예: 카페/식당 → 방문해서 먹고 즐긴 본인 입장\n"
                f"   - 예: 미용실/네일 → 서비스 받으러 간 본인 입장\n"
                f"   - 예: 산후조리원 → 출산한 본인 또는 배우자 입장\n"
                f"3) 기본 템플릿이 가진 제약(사실 기반 작성 / 지어내기 금지 / 이모티콘 금지 등)은 유지한다.\n\n"
                f"★ 아래 '참고 스타일(요양원 예시)'의 구조·디테일 수준·섹션 구성·최소 문장 수 형식을 그대로 본떠라.\n"
                f"단, 요양원 고유의 고민(부모님 모시기 죄책감 등)은 '{key}' 업종의 고민/기대 포인트로 **전면 교체**한다.\n"
                f"분량 기준(공백 포함 1500자 안팎, 1400자 미만 금지, 최소 40문장)은 그대로 유지한다.\n\n"
                f"=== 참고 스타일 — 본문 (요양원 예시) ===\n{ref_example}\n\n"
                f"=== 참고 스타일 — 제목 (공통 템플릿) ===\n{ref_title}\n"
                f"★ 제목 프롬프트는 위 템플릿의 구조·섹션·카테고리 4종 설명·지시를 **그대로 유지**한다.\n"
                f"★ 단, '# 예시' 섹션의 '핵심 키워드'만 '{key}' 업종에 맞는 대표 키워드(예: 헬스장이면 '강서구 헬스장', 요양원이면 '은평구 요양원')로 바꾸고,\n"
                f"   '결과물 예시'의 5개 문구도 '{key}' 업종 검색자가 클릭할 만한 마무리 문구로 자연스럽게 교체한다.\n"
                f"★ 단, 예시로 쓰는 5개 문구에도 **지역명/구/동/역이름/업종명({key})/업체명은 절대 포함하지 않는다.** "
                f"예: 헬스장용 예시에 '헬스장'이라는 단어가 들어가면 안 됨. '운동', '다닐 곳', '다녀봤어요' 같은 간접 표현으로.\n"
                f"★ 출력은 '프롬프트 템플릿'이지 10개 문구 자체가 아니다. '핵심 키워드: {{키워드}}' 플레이스홀더는 그대로 둘 것.\n\n"
                f"=== 기존 기본 템플릿 (필요 제약 참고용) ===\n[BLOG]\n{base.get('blog','')}\n\n[TITLE]\n{base.get('title','')}\n\n"
                f"=== 출력 형식 (반드시 이 구분자 사용) ===\n"
                f"===BLOG===\n(여기에 '{key}' 업종용 블로그 본문 프롬프트 전문)\n"
                f"===TITLE===\n(여기에 '{key}' 업종용 제목 프롬프트 전문 — 위 제목 공통 골격 유지)\n"
            )

            btn_auto.setEnabled(False)
            btn_auto.setText("생성 중...")

            def _worker():
                try:
                    from openai import OpenAI
                    client = OpenAI(api_key=api_keys[0])
                    resp = client.chat.completions.create(
                        model="gpt-4o-mini",
                        messages=[{"role": "user", "content": meta_prompt}],
                    )
                    text = resp.choices[0].message.content or ""
                    ai_sig.done.emit(text, "")
                except Exception as e:
                    ai_sig.done.emit("", str(e))

            threading.Thread(target=_worker, daemon=True).start()

        btn_add.clicked.connect(add_type)
        btn_auto.clicked.connect(auto_generate)
        btn_del.clicked.connect(del_type)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_save = QPushButton("저장")
        btn_save.setStyleSheet("background: #4a6cf7; color: white; border-radius: 8px; padding: 8px 20px; font-weight: bold;")
        btn_cancel = QPushButton("취소")
        btn_cancel.setStyleSheet("padding: 8px 20px;")
        btn_row.addWidget(btn_save)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

        def do_save():
            save_current_to_memory()
            try:
                with open(prompts_path, "w", encoding="utf-8") as f:
                    json.dump(prompts, f, ensure_ascii=False, indent=2)
                QMessageBox.information(dlg, "완료", "프롬프트를 저장했습니다.")
                dlg.accept()
            except Exception as e:
                QMessageBox.critical(dlg, "오류", f"저장 실패: {e}")

        btn_save.clicked.connect(do_save)
        btn_cancel.clicked.connect(dlg.reject)

        load_current()
        dlg.exec()

    # ── 크롤링 ──
    def _check_api_keys(self) -> bool:
        cfg = load_config()
        has_key = any(k for k in cfg.get("gpt_key_list", []) if k) or any(k for k in cfg.get("gemini_key_list", []) if k)
        if not has_key:
            QMessageBox.warning(self, "API 키 필요", "API 키가 설정되지 않았습니다.\n설정에서 API 키를 입력하거나 관리자에게 문의하세요.")
            return False
        return True

    def _start_crawl(self):
        if not self._check_api_keys():
            return
        keyword = self.keyword_input.currentText().strip()
        if not keyword:
            QMessageBox.warning(self, "경고", "키워드를 입력해주세요.")
            return
        if self.is_crawling:
            return

        self._save_keyword_history(keyword)
        self.last_keyword = keyword
        # 설정에도 저장 (계정별, 재시작해도 유지)
        try:
            _c = load_config()
            _c["last_keyword"] = keyword  # 호환용
            lkm = _c.get("last_keyword_by_account", {}) or {}
            lkm[self._current_blog_id()] = keyword
            _c["last_keyword_by_account"] = lkm
            save_config(_c)
        except Exception:
            pass
        count = self.count_spin.value()
        self.is_crawling = True
        self.stop_flag = False
        self.crawled_data = []

        log_dir = self._get_logs_dir()
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        self.result_file = os.path.join(log_dir, f"places_{timestamp}.json")

        # 같은 키워드로 이전 크롤 데이터가 있으면 이어서 크롤
        meta = self._load_crawled_meta()
        existing_for_resume = []
        if meta.get("keyword") == keyword and meta.get("items"):
            existing_for_resume = meta["items"]
            self._emit_log(f"이전 크롤 {len(existing_for_resume)}개 발견 — 이어서 크롤링")

        self._emit_status("크롤링 중...", "#f59e0b")
        self._emit_log(f"크롤링 시작: '{keyword}' (검색 지역 일치 항목만 수집)")

        def _worker():
            try:
                last_save_count = [len(existing_for_resume)]
                latest_results = [list(existing_for_resume)]

                def on_progress(current, scanned, name, results_ref=None):
                    if results_ref is not None:
                        latest_results[0] = results_ref
                    if self.stop_flag:
                        try:
                            self._save_crawled(latest_results[0], keyword)
                        except Exception:
                            pass
                        raise InterruptedError("중단됨")
                    self._emit_log(f"  {name} ({current}개)")
                    self._emit_status(f"수집 {current}개", "#f59e0b")
                    # 메인창 캐시만 갱신 (logs/{kw}.json은 _on_item이 매번 저장)
                    if current - last_save_count[0] >= 1:
                        try:
                            self._save_crawled(latest_results[0], keyword)
                            last_save_count[0] = current
                        except Exception:
                            pass

                # 업체 1개 크롤 끝날 때마다 즉시 저장 (logs/{account}/{kw}.json)
                def _on_item(place, items_so_far, kw):
                    try:
                        import re as _re
                        _safe = _re.sub(r"[^가-힣A-Za-z0-9_]", "_", str(kw or "unknown"))[:80]
                        _dir = self._get_logs_dir()
                        os.makedirs(_dir, exist_ok=True)
                        _fp = os.path.join(_dir, f"{_safe}.json")
                        with open(_fp, "w", encoding="utf-8") as _f:
                            json.dump({"keyword": kw, "items": list(items_so_far)}, _f, ensure_ascii=False, indent=2)
                    except Exception as _e:
                        self._emit_log(f"업체별 저장 실패: {_e}")

                def _crawl_with_retry(kw, target_count, existing, excludes, max_retries=3):
                    """timeout/webdriver 예외 발생 시 드라이버 재시작 + resume."""
                    acc = list(existing or [])
                    last_err = None
                    for attempt in range(max_retries):
                        try:
                            return crawl_places(
                                kw, target_count, on_progress,
                                existing_places=acc,
                                exclude_keywords=excludes,
                                on_item=_on_item,
                            )
                        except InterruptedError:
                            raise
                        except Exception as e:
                            last_err = e
                            msg = str(e)
                            is_timeout = ("Read timed out" in msg) or ("Timeout" in msg) or ("WebDriverException" in type(e).__name__)
                            if is_timeout and attempt < max_retries - 1:
                                self._emit_log(f"  [재시도 {attempt+1}/{max_retries}] 드라이버 timeout → 재시작 후 resume ({len(acc)}개 유지)")
                                import time as _t; _t.sleep(3)
                                continue
                            raise
                    if last_err:
                        raise last_err
                    return acc

                regions = list(getattr(self, "_selected_regions", []) or [])
                _kparts = keyword.split()
                _biz = _kparts[-1] if _kparts else keyword
                if regions:
                    expanded = [f"{r} {_biz}" for r in regions]
                else:
                    expanded = [keyword]
                # 제외 키워드 (업종 기준)
                _ex_map = getattr(self, "_exclude_keywords_by_biz", {}) or {}
                _excludes = list(_ex_map.get(_biz, []))
                if _excludes:
                    self._emit_log(f"제외 키워드 {len(_excludes)}개: {', '.join(_excludes)}")
                if len(expanded) > 1:
                    # 병렬 크롤 — 지역을 3개 워커로 나눠서 동시 수집 (각자 독립 드라이버)
                    _workers = min(3, max(1, len(expanded)))
                    self._emit_log(f"지역 {len(expanded)}개 × '{_biz}' 병렬 크롤 ({_workers}개 봇)")
                    _per = max(50, count // len(expanded) * 2)  # 지역별 상한
                    def _on_batch(items, kw=None, raw=None):
                        # 키워드별 단일 파일은 _on_item이 매번 update 중. 여기선 메인창 캐시만 갱신.
                        latest_results[0] = list(items)
                        try:
                            self._save_crawled(list(items), keyword)
                        except Exception:
                            pass
                    results = crawl_places_parallel(
                        expanded,
                        count_per=_per,
                        on_progress=on_progress,
                        existing_places=existing_for_resume,
                        exclude_keywords=_excludes,
                        max_workers=_workers,
                        stop_flag=lambda: self.stop_flag,
                        emit_log=self._emit_log,
                        save_batch=_on_batch,
                        on_item=_on_item,
                    )
                    latest_results[0] = results
                    try:
                        self._save_crawled(results, keyword)
                        save_results(results, self.result_file, keyword)
                    except Exception:
                        pass
                else:
                    results = _crawl_with_retry(keyword, count, existing_for_resume, _excludes)
                    for _p in results:
                        _p.setdefault("search_keyword", keyword)
                latest_results[0] = results
                self.crawled_data = results
                # app_user별 영속 저장 (메인창 캐시) — 키워드별 logs/{kw}.json은 _on_item이 처리
                self._save_crawled(results, keyword)

                self._emit_log(f"크롤링 완료! {len(results)}개 수집")
                self._emit_status("완료", "#22c55e")
                self._update_history_result_count(keyword, len(results))
                # 업종별 프롬프트 자동 생성 (없는 경우만)
                try:
                    self._auto_generate_prompt_for_biz(_biz)
                except Exception as _pe:
                    self._emit_log(f"프롬프트 자동 생성 스킵: {_pe}")

            except InterruptedError:
                self._emit_log("크롤링 중단됨")
                self._emit_status("중단됨", "#ef4444")
                try:
                    self._save_crawled(latest_results[0], keyword)
                except Exception:
                    pass
            except Exception as e:
                self._emit_log(f"크롤링 오류: {e}")
                self._emit_status("오류", "#ef4444")
                try:
                    self._save_crawled(latest_results[0], keyword)
                    self._emit_log(f"오류 전까지 {len(latest_results[0])}개 저장됨 → F7에서 이어서 선택 가능")
                except Exception:
                    pass
            finally:
                self.is_crawling = False

        threading.Thread(target=_worker, daemon=True).start()

    def _stop(self):
        if self.is_crawling or self.is_posting or getattr(self, 'is_generating', False):
            self.stop_flag = True
            self._emit_log("중단 요청...")

    # ── 결과보기 ──
    def _show_results(self):
        if not self.crawled_data:
            log_dir = self._get_logs_dir()
            if os.path.exists(log_dir):
                files = sorted(
                    [f for f in os.listdir(log_dir) if f.endswith(".json")],
                    reverse=True
                )
                if files:
                    filepath = os.path.join(log_dir, files[0])
                    raw = load_results(filepath)
                    self.crawled_data = raw.get("items", []) if isinstance(raw, dict) else raw
                    self._emit_log(f"결과 불러옴: {files[0]}")

        if not self.crawled_data:
            QMessageBox.information(self, "결과", "크롤링 결과가 없습니다.")
            return

        from PySide6.QtWidgets import QTreeWidget, QTreeWidgetItem

        # 지역 추출: address에서 "시/도 + 구/군"
        def _region_of(p):
            import re as _re
            addr = p.get("address", "") or p.get("jibun_address", "")
            m = _re.search(
                r"((?:서울|부산|대구|인천|광주|대전|울산|세종|경기|강원|충북|충남|전북|전남|경북|경남|제주)\S*)\s+(\S+[시군구])",
                addr,
            )
            if m:
                return f"{m.group(1)} {m.group(2)}"
            m2 = _re.search(r"(\S+[시군구])", addr)
            if m2:
                return m2.group(1)
            return "지역 미상"

        groups = {}
        for p in self.crawled_data:
            r = _region_of(p)
            groups.setdefault(r, []).append(p)

        dlg = QDialog(self)
        dlg.setWindowTitle(f"크롤링 결과 — 총 {len(self.crawled_data)}건 ({len(groups)}개 지역)")
        dlg.resize(1350, 650)
        layout = QVBoxLayout(dlg)

        # 상단 컨트롤
        top = QHBoxLayout()
        btn_expand = QPushButton("전체 펼치기")
        btn_collapse = QPushButton("전체 접기")
        for b in (btn_expand, btn_collapse):
            b.setStyleSheet("padding: 6px 14px;")
            b.setCursor(Qt.PointingHandCursor)
        top.addWidget(btn_expand)
        top.addWidget(btn_collapse)
        top.addStretch()
        layout.addLayout(top)

        tree = QTreeWidget()
        tree.setAlternatingRowColors(True)
        headers = ["지역 / 업체명", "업체주소", "카테고리", "근처역", "앞 키워드", "태그", "픽사베이 키워드"]
        tree.setColumnCount(len(headers))
        tree.setHeaderLabels(headers)
        widths = [260, 280, 130, 150, 200, 160, 200]
        for i, w in enumerate(widths):
            tree.setColumnWidth(i, w)
        tree.setRootIsDecorated(True)
        tree.setUniformRowHeights(True)

        # 지역 가나다순
        for region in sorted(groups.keys()):
            items = groups[region]
            root = QTreeWidgetItem(tree, [f"▼ {region} ({len(items)}건)"])
            from PySide6.QtGui import QFont as _QF
            _f = _QF()
            _f.setBold(True)
            root.setFont(0, _f)
            for p in items:
                addr = p.get("address", "")
                jibun = p.get("jibun_address", "")
                full_addr = (f"{addr} {jibun}" if addr and jibun else (addr or jibun))
                child = QTreeWidgetItem(root, [
                    p.get("name", ""),
                    full_addr,
                    p.get("category", ""),
                    p.get("nearby_station", ""),
                    p.get("front_keywords", ""),
                    p.get("tags", ""),
                    p.get("pixabay_keywords", ""),
                ])
                # 자식은 기본 글씨
            root.setExpanded(True)

        btn_expand.clicked.connect(lambda: tree.expandAll())
        btn_collapse.clicked.connect(lambda: tree.collapseAll())

        layout.addWidget(tree)

        def copy_to_clipboard():
            # 펼쳐진 자식 행만 복사 (지역 헤더 제외)
            lines = []
            for i in range(tree.topLevelItemCount()):
                region_item = tree.topLevelItem(i)
                for j in range(region_item.childCount()):
                    ch = region_item.child(j)
                    row = [ch.text(k) for k in range(tree.columnCount())]
                    lines.append("\t".join(row))
            QApplication.clipboard().setText("\n".join(lines))
            QMessageBox.information(dlg, "복사 완료", "엑셀에 붙여넣기(Ctrl+V) 하세요.")

        btn_copy = QPushButton("엑셀용 복사")
        btn_copy.setStyleSheet("background: #4a6cf7; color: white; border: none; border-radius: 8px; font-size: 13px; font-weight: bold; padding: 10px;")
        btn_copy.setCursor(Qt.PointingHandCursor)
        btn_copy.clicked.connect(copy_to_clipboard)
        layout.addWidget(btn_copy)

        dlg.exec()

    # ── 생성된 포스트 보기 (크롤 업체 목록 + 생성여부 표시) ──
    def _view_generated_posts(self):
        posts = self._load_generated_posts()
        self._generated_posts = posts

        # (업체명+주소) → content 매핑
        post_map = {}
        for p in posts:
            pl = p.get("place", {})
            key = (pl.get("name", ""), pl.get("address", "") or pl.get("jibun_address", ""))
            post_map[key] = p

        # 크롤 결과 로드 (포스팅 업체 선택과 동일)
        all_groups = self._load_all_crawl_results()
        if self.crawled_data:
            all_groups["현재 크롤링 결과 (" + str(len(self.crawled_data)) + "개)"] = self.crawled_data

        if not all_groups and not posts:
            QMessageBox.information(self, "결과", "크롤링/생성된 포스트가 없습니다.")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("생성된 포스트 보기")
        dlg.resize(1000, 650)
        layout = QVBoxLayout(dlg)

        # 상단: 전체 선택/해제
        top_row = QHBoxLayout()
        btn_all_sel = QPushButton("전체 선택")
        btn_all_sel.setStyleSheet("padding: 6px 15px;")
        btn_all_none = QPushButton("전체 해제")
        btn_all_none.setStyleSheet("padding: 6px 15px;")
        top_row.addWidget(btn_all_sel)
        top_row.addWidget(btn_all_none)
        top_row.addStretch()
        layout.addLayout(top_row)

        info = QLabel("● 초록: 생성됨 (더블클릭해서 내용 확인 / 체크 후 삭제 가능)  ● 빨강: 미생성")
        info.setStyleSheet("font-size: 12px; color: #475569; padding: 5px;")
        layout.addWidget(info)

        tree = QTreeWidget()
        tree.setHeaderLabels(["업체명", "업체주소", "카테고리", "근처역", "앞 키워드", "태그"])
        tree.setColumnWidth(0, 260)
        tree.setColumnWidth(1, 240)
        tree.setColumnWidth(2, 100)
        tree.setColumnWidth(3, 100)
        tree.setColumnWidth(4, 160)
        tree.setAlternatingRowColors(True)

        # 더블클릭 시 포스트 내용 보기를 위한 매핑
        item_to_post = {}
        all_children = []  # (child_item, place) — 전체 선택/삭제용

        for group_name, places in all_groups.items():
            # 삭제된 키 필터
            places = [p for p in places if self._place_key(p) not in _deleted]
            if not places:
                continue
            parent = QTreeWidgetItem(tree)
            parent.setText(0, group_name)
            parent.setFlags(parent.flags() | Qt.ItemIsUserCheckable)
            parent.setCheckState(0, Qt.Unchecked)
            parent.setExpanded(False)

            gen_count = 0
            for p in places:
                addr = p.get("address", "")
                jibun = p.get("jibun_address", "")
                full_addr = f"{addr} {jibun}" if addr and jibun else (jibun or addr)

                key = (p.get("name", ""), p.get("address", "") or p.get("jibun_address", ""))
                generated_post = post_map.get(key)

                child = QTreeWidgetItem(parent)
                child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
                child.setCheckState(0, Qt.Unchecked)
                is_posted = bool(generated_post and generated_post.get("posted", False))
                if is_posted:
                    child.setText(0, "[완] " + p.get("name", ""))
                    child.setForeground(0, QColor("#3b82f6"))
                    gen_count += 1
                    item_to_post[id(child)] = generated_post
                elif generated_post:
                    child.setText(0, "●  " + p.get("name", ""))
                    child.setForeground(0, QColor("#22c55e"))
                    gen_count += 1
                    item_to_post[id(child)] = generated_post
                else:
                    child.setText(0, "●  " + p.get("name", ""))
                    child.setForeground(0, QColor("#ef4444"))
                child.setText(1, full_addr)
                child.setText(2, p.get("category", ""))
                child.setText(3, p.get("nearby_station", ""))
                child.setText(4, p.get("front_keywords", ""))
                child.setText(5, p.get("tags", ""))
                all_children.append((child, p))

            # 폴더 색
            if len(places) > 0:
                if gen_count == len(places):
                    parent.setForeground(0, QColor("#22c55e"))
                elif gen_count == 0:
                    parent.setForeground(0, QColor("#ef4444"))
                else:
                    parent.setForeground(0, QColor("#f59e0b"))
                parent.setText(0, f"●  {group_name}  ({gen_count}/{len(places)} 생성)")

        layout.addWidget(tree)

        def on_double_click(item, col):
            post = item_to_post.get(id(item))
            if not post:
                return
            self._show_single_post_view(post)

        tree.itemDoubleClicked.connect(on_double_click)

        # 부모 체크 시 하위 전체 토글
        def on_item_changed(item, col):
            if col != 0:
                return
            tree.blockSignals(True)
            state = item.checkState(0)
            if item.childCount() > 0:
                for i in range(item.childCount()):
                    item.child(i).setCheckState(0, state)
            tree.blockSignals(False)

        tree.itemChanged.connect(on_item_changed)

        def delete_selected():
            # F8 선택 삭제: 업체 + 생성된 포스트 모두 제거 (크롤링 목록에서도 사라짐)
            checked_places = [p for (c, p) in all_children if c.checkState(0) == Qt.Checked]
            if not checked_places:
                QMessageBox.information(dlg, "안내", "선택된 항목이 없습니다.")
                return
            reply = QMessageBox.question(dlg, "삭제 확인", f"{len(checked_places)}개 항목을 완전 삭제할까요?\n(업체 + 생성된 포스트 모두 제거, F7에서도 사라짐)")
            if reply != QMessageBox.Yes:
                return
            keys = {self._place_key(p) for p in checked_places}
            self._generated_posts = [gp for gp in self._generated_posts if self._place_key(gp.get("place", {})) not in keys]
            self._save_generated_posts()
            self._purge_places_from_logs(checked_places)
            QMessageBox.information(dlg, "완료", f"{len(checked_places)}개 완전 삭제됨.")
            dlg.accept()

        def delete_all():
            # F8 전체 삭제: 표시된 모든 항목(업체+포스트) 완전 제거
            all_places = [p for (_c, p) in all_children]
            if not all_places:
                return
            reply = QMessageBox.question(dlg, "전체 삭제 확인", f"표시된 전체 {len(all_places)}개 항목을 완전 삭제할까요?\n(업체 + 생성된 포스트 모두 제거)")
            if reply != QMessageBox.Yes:
                return
            keys = {self._place_key(p) for p in all_places}
            self._generated_posts = [gp for gp in self._generated_posts if self._place_key(gp.get("place", {})) not in keys]
            self._save_generated_posts()
            self._purge_places_from_logs(all_places)
            QMessageBox.information(dlg, "완료", f"{len(all_places)}개 완전 삭제됨.")
            dlg.accept()

        def select_all_f8():
            tree.blockSignals(True)
            for i in range(tree.topLevelItemCount()):
                p = tree.topLevelItem(i)
                p.setCheckState(0, Qt.Checked)
                for j in range(p.childCount()):
                    p.child(j).setCheckState(0, Qt.Checked)
            tree.blockSignals(False)

        def select_none_f8():
            tree.blockSignals(True)
            for i in range(tree.topLevelItemCount()):
                p = tree.topLevelItem(i)
                p.setCheckState(0, Qt.Unchecked)
                for j in range(p.childCount()):
                    p.child(j).setCheckState(0, Qt.Unchecked)
            tree.blockSignals(False)

        btn_all_sel.clicked.connect(select_all_f8)
        btn_all_none.clicked.connect(select_none_f8)

        # 첫 포스팅 즉시/대기 옵션 + 랜덤 포스팅 순서 토글
        opt_row = QHBoxLayout()
        first_immediate_cb = QCheckBox("첫 포스팅을 지금 바로 시작 (체크 해제 시 설정 간격만큼 대기 후 시작)")
        first_immediate_cb.setChecked(True)
        first_immediate_cb.setStyleSheet("font-size: 12px; color: #334155;")
        opt_row.addWidget(first_immediate_cb)
        opt_row.addStretch()
        random_cb = ToggleSwitch()
        random_lbl = QLabel("랜덤포스팅")
        random_lbl.setStyleSheet("font-size: 12px; color: #334155; margin-left: 6px;")
        opt_row.addWidget(random_cb)
        opt_row.addWidget(random_lbl)
        layout.addLayout(opt_row)

        def post_selected():
            targets = []
            for i in range(tree.topLevelItemCount()):
                parent = tree.topLevelItem(i)
                for j in range(parent.childCount()):
                    child = parent.child(j)
                    if child.checkState(0) == Qt.Checked:
                        post = item_to_post.get(id(child))
                        if post:
                            targets.append(post)
            if not targets:
                QMessageBox.information(dlg, "안내", "포스팅할 생성된 포스트를 선택해주세요.")
                return
            if random_cb.isChecked():
                import random as _r
                _r.shuffle(targets)
            self.posting_targets = targets
            self.first_post_immediate = first_immediate_cb.isChecked()
            dlg.accept()
            self._post_generated()

        btn_row = QHBoxLayout()
        btn_delete = QPushButton("선택 삭제")
        btn_delete.setStyleSheet("background: #ef4444; color: white; border: none; border-radius: 8px; padding: 8px 20px;")
        btn_delete.setCursor(Qt.PointingHandCursor)
        btn_delete.clicked.connect(delete_selected)
        btn_del_all = QPushButton("전체 삭제")
        btn_del_all.setStyleSheet("background: #dc2626; color: white; border: none; border-radius: 8px; padding: 8px 20px;")
        btn_del_all.setCursor(Qt.PointingHandCursor)
        btn_del_all.clicked.connect(delete_all)
        btn_post = QPushButton("자동 포스팅")
        btn_post.setStyleSheet("background: #3b82f6; color: white; border: none; border-radius: 8px; font-weight: bold; padding: 8px 24px;")
        btn_post.setCursor(Qt.PointingHandCursor)
        btn_post.clicked.connect(post_selected)
        btn_close = QPushButton("닫기")
        btn_close.setStyleSheet("padding: 8px 20px;")
        btn_close.clicked.connect(dlg.reject)
        btn_row.addWidget(btn_delete)
        btn_row.addWidget(btn_del_all)
        btn_row.addStretch()
        btn_row.addWidget(btn_post)
        btn_row.addWidget(btn_close)
        layout.addLayout(btn_row)

        dlg.exec()

    def _show_single_post_view(self, post: dict):
        """단일 포스트 내용 보기"""
        from PySide6.QtWidgets import QTextEdit, QLineEdit
        place = post.get("place", {})
        content = post.get("content", {})

        dlg = QDialog(self)
        dlg.setWindowTitle(place.get("name", "포스트"))
        dlg.resize(900, 700)
        layout = QVBoxLayout(dlg)

        title_edit = QLineEdit(content.get("title", ""))
        title_edit.setStyleSheet("font-weight: bold; font-size: 14px; padding: 8px;")
        layout.addWidget(title_edit)

        body_edit = QTextEdit()
        body_edit.setPlainText(content.get("body", ""))
        layout.addWidget(body_edit)

        meta = QLabel(
            f"주소: {place.get('address','')} {place.get('jibun_address','')}  |  "
            f"카테고리: {place.get('category','')}  |  "
            f"근처역: {place.get('nearby_station','')}  |  "
            f"태그: {','.join(content.get('tags', [])) if isinstance(content.get('tags'), list) else content.get('tags', '')}"
        )
        meta.setStyleSheet("color: #475569; padding: 5px; font-size: 12px;")
        meta.setWordWrap(True)
        layout.addWidget(meta)

        btn_row = QHBoxLayout()
        btn_save = QPushButton("저장")
        btn_save.setStyleSheet("background: #22c55e; color: white; border: none; border-radius: 6px; padding: 8px 20px;")
        def save():
            content["title"] = title_edit.text()
            content["body"] = body_edit.toPlainText()
            self._save_generated_posts()
            QMessageBox.information(dlg, "저장", "저장되었습니다.")
        btn_save.clicked.connect(save)
        btn_close = QPushButton("닫기")
        btn_close.setStyleSheet("padding: 8px 20px;")
        btn_close.clicked.connect(dlg.reject)
        btn_row.addWidget(btn_save)
        btn_row.addStretch()
        btn_row.addWidget(btn_close)
        layout.addLayout(btn_row)

        dlg.exec()

    # ── 포스트 생성 ──
    def _account_key(self) -> str:
        # 블로그마스터 로그인 사용자(app_user) 기준 — 같은 app_user면 모든 네이버 슬롯 데이터 공유
        try:
            from config import get_current_user as _gcu
            uid = _gcu() or "default"
        except Exception:
            uid = "default"
        return "".join(c for c in uid if c.isalnum() or c in "-_") or "default"

    def _refresh_status_panes(self):
        # 라벨에 숫자 표시 안 함 — 실시간 로그는 아래 crawl_log/post_log에서 보임
        pass

    def _get_crawled_file(self):
        return os.path.join(os.path.dirname(__file__), f"crawled_{self._account_key()}.json")

    def _save_crawled(self, items: list, keyword: str = ""):
        try:
            with open(self._get_crawled_file(), "w", encoding="utf-8") as f:
                json.dump({"keyword": keyword, "items": items}, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self._emit_log(f"crawled 저장 실패: {e}")

    def _load_crawled(self) -> list:
        fp = self._get_crawled_file()
        if not os.path.exists(fp):
            return []
        try:
            with open(fp, "r", encoding="utf-8") as f:
                d = json.load(f)
            return d.get("items", []) if isinstance(d, dict) else (d or [])
        except Exception:
            return []

    def _load_crawled_meta(self) -> dict:
        """저장된 크롤 데이터 + 키워드 반환. {'keyword': str, 'items': list}"""
        fp = self._get_crawled_file()
        if not os.path.exists(fp):
            return {"keyword": "", "items": []}
        try:
            with open(fp, "r", encoding="utf-8") as f:
                d = json.load(f)
            if isinstance(d, dict):
                return {"keyword": d.get("keyword", ""), "items": d.get("items", []) or []}
            return {"keyword": "", "items": d or []}
        except Exception:
            return {"keyword": "", "items": []}

    # ── 삭제: deleted_keys 시스템 제거. logs 파일에서 즉시 영구 삭제 ──
    def _purge_places_from_logs(self, places_to_remove: list):
        """logs 폴더의 모든 places_*.json/{kw}.json 파일에서 (이름,주소) 매칭 항목 제거.
        파일이 비면 통째로 삭제."""
        if not places_to_remove:
            return 0
        keys = {self._place_key(p) for p in places_to_remove}
        log_dir = self._get_logs_dir()
        if not os.path.isdir(log_dir):
            return 0
        removed_total = 0
        for fname in os.listdir(log_dir):
            if not fname.endswith(".json"):
                continue
            fp = os.path.join(log_dir, fname)
            try:
                with open(fp, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if not isinstance(data, dict): continue
                items = data.get("items", []) or []
                kept = [p for p in items if self._place_key(p) not in keys]
                if len(kept) == len(items):
                    continue
                removed_total += (len(items) - len(kept))
                if not kept:
                    try: os.remove(fp)
                    except Exception: pass
                else:
                    data["items"] = kept
                    with open(fp, "w", encoding="utf-8") as f:
                        json.dump(data, f, ensure_ascii=False, indent=2)
            except Exception:
                continue
        # 메모리 self.crawled_data도 동기화
        try:
            self.crawled_data = [p for p in (self.crawled_data or []) if self._place_key(p) not in keys]
            self._save_crawled(self.crawled_data, getattr(self, "last_keyword", ""))
        except Exception:
            pass
        return removed_total

    def _place_key(self, place: dict) -> tuple:
        return (place.get("name", ""), place.get("address", "") or place.get("jibun_address", ""))

    def _get_posts_file(self):
        # 계정별로 분리 저장
        per_acc = os.path.join(os.path.dirname(__file__), f"generated_posts_{self._account_key()}.json")
        legacy = os.path.join(os.path.dirname(__file__), "generated_posts.json")
        # 최초 전환 시 공용 파일 → 현재 계정 파일로 자동 승격 (1회)
        if not os.path.exists(per_acc) and os.path.exists(legacy):
            try:
                import shutil
                shutil.copyfile(legacy, per_acc)
            except Exception:
                pass
        return per_acc

    def _get_logs_dir(self) -> str:
        d = os.path.join(os.path.dirname(__file__), "logs", self._account_key())
        os.makedirs(d, exist_ok=True)
        return d

    def _migrate_legacy_logs(self):
        """logs/*.json 을 1번 계정 폴더로 1회 이동 (구버전 데이터 복구용)"""
        import shutil
        base = os.path.join(os.path.dirname(__file__), "logs")
        if not os.path.isdir(base):
            return
        legacy = [f for f in os.listdir(base) if f.endswith(".json") and os.path.isfile(os.path.join(base, f))]
        if not legacy:
            return
        accounts = self.cfg.get("accounts", [])
        target_idx = 0
        for i, acc in enumerate(accounts):
            if (acc.get("blog_id") or "").strip():
                target_idx = i
                break
        bid = (accounts[target_idx].get("blog_id") if target_idx < len(accounts) else "") or f"acc{target_idx}"
        safe = "".join(c for c in bid if c.isalnum() or c in "-_") or "default"
        target_dir = os.path.join(base, safe)
        os.makedirs(target_dir, exist_ok=True)
        moved = 0
        for f in legacy:
            src = os.path.join(base, f)
            dst = os.path.join(target_dir, f)
            if os.path.exists(dst):
                continue
            try:
                shutil.move(src, dst)
                moved += 1
            except Exception:
                pass
        if moved:
            self._emit_log(f"기존 크롤 결과 {moved}개를 '{bid}' 계정으로 이동했습니다")

    def _save_generated_posts(self):
        data = []
        for item in self._generated_posts:
            data.append({
                "place": item["place"],
                "content": item["content"],
                "posted": item.get("posted", False),
            })
        with open(self._get_posts_file(), "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        try:
            self._cleanup_orphan_image_folders()
        except Exception:
            pass

    def _load_generated_posts(self) -> list:
        filepath = self._get_posts_file()
        if os.path.exists(filepath):
            try:
                with open(filepath, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return []

    def _generate_posts(self):
        if not self._check_api_keys():
            return
        # 크롤 결과 파일에서 스냅샷 로드 (크롤이 동시에 돌아도 안 꼬이게)
        snap = self._load_crawled()
        if snap:
            self.crawled_data = snap
        # 업체 선택
        selected = self._show_posting_selector()
        if not selected:
            return

        cfg = load_config()
        provider = "GPT"  # GPT 전용 고정
        api_keys = [k for k in cfg.get("gpt_key_list", []) if k]
        if not api_keys:
            QMessageBox.critical(self, "오류", "GPT API 키를 설정해주세요.")
            return

        api_key = api_keys[0]
        _lkm = self.cfg.get("last_keyword_by_account", {}) or {}
        keyword = self.keyword_input.currentText().strip() or getattr(self, 'last_keyword', '') or _lkm.get(self._current_blog_id(), "") or self.cfg.get("last_keyword", "")
        # 크롤된 데이터에서 키워드 자동 복원 (앞키워드 첫 항목이 "{지역}{업종}" 형식)
        if not keyword and selected:
            first = selected[0]
            fk = first.get("front_keywords", "")
            if fk:
                keyword = fk.split(",")[0].strip()
                self._emit_log(f"크롤 데이터에서 키워드 복원: '{keyword}'")
        total = len(selected)

        self._emit_status("포스트 생성 중...", "#8b5cf6")
        self._emit_log(f"포스트 생성 시작: {total}개 업체")

        self._generating_data = {"selected": selected, "keyword": keyword, "api_key": api_key, "provider": provider}
        self._generated_posts = self._load_generated_posts()

        self.is_generating = True
        self.stop_flag = False

        def _worker():
            from concurrent.futures import ThreadPoolExecutor, as_completed
            import threading
            lock = threading.Lock()
            done_count = [0]
            new_posts = []

            pix_keys = [k for k in cfg.get("pixabay_key_list", []) if k]

            def _gen_one(idx, place):
                if self.stop_flag:
                    return
                name = place.get("name", "")
                try:
                    content = generate_content(
                        provider=provider,
                        api_key=api_key,
                        place=place,
                        keyword=keyword,
                        prompt_override=getattr(self, "_override_prompt_name", None),
                    )
                    if self.stop_flag:
                        return

                    # 이미지: 네이버 실사 1장 + Pixabay 나머지 (중복 방지)
                    img_paths = []
                    if pix_keys:
                        try:
                            biz = (place.get("category") or "").strip() or keyword
                            img_count = int(content.get("image_count", 3) or 3)
                            pkey = self._place_key(place)
                            safe_name = "".join(c for c in (pkey[0] + "_" + pkey[1]) if c.isalnum() or c in "-_")[:80] or "unknown"
                            persist_dir = os.path.join(
                                os.path.dirname(__file__), "saved_images",
                                self._account_key(), safe_name
                            )
                            os.makedirs(persist_dir, exist_ok=True)
                            for _f in os.listdir(persist_dir):
                                try: os.remove(os.path.join(persist_dir, _f))
                                except: pass

                            # 1) 실사 1장 (place_id 있을 때만)
                            real_paths = []
                            pid = (place.get("place_id") or "").strip()
                            if pid:
                                try:
                                    import real_photos as _rp
                                    from selenium import webdriver as _wd
                                    from selenium.webdriver.chrome.options import Options as _Opt
                                    from selenium.webdriver.chrome.service import Service as _Svc
                                    from webdriver_manager.chrome import ChromeDriverManager as _CM
                                    _opt = _Opt()
                                    _opt.add_argument("--headless=new")
                                    _opt.add_argument("--disable-blink-features=AutomationControlled")
                                    _opt.add_argument("--no-sandbox")
                                    _opt.add_argument("--disable-gpu")
                                    _opt.add_argument("--window-size=1920,1080")
                                    _opt.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36")
                                    _opt.add_experimental_option("excludeSwitches", ["enable-automation"])
                                    _drv = _wd.Chrome(service=_Svc(_CM().install()), options=_opt)
                                    _drv.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
                                        "source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
                                    })
                                    try:
                                        real_paths = _rp.pick_real_photos(
                                            _drv, pid, place.get("name", ""),
                                            target_real=1, out_dir=persist_dir,
                                            emit_log=self._emit_log,
                                        )
                                    finally:
                                        _drv.quit()
                                except Exception as _e:
                                    self._emit_log(f"실사 수집 실패 ({name}): {_e}")
                            img_paths.extend(real_paths)

                            # 2) Pixabay 폴백 — 중복 방지 ID 파일 계정별로
                            need = max(0, img_count - len(img_paths))
                            if need > 0:
                                try:
                                    import image_handler as _ih
                                    _used_file = os.path.join(
                                        os.path.dirname(__file__), "saved_images",
                                        self._account_key(), "used_pixabay_ids.json",
                                    )
                                    _ih.configure_used_ids_file(_used_file)
                                    self._emit_log(f"  Pixabay 보충 {need}장 (업종: {biz})")
                                    _tmp = _ih.download_images(pix_keys[0], biz, need)
                                    import shutil as _sh
                                    for _i, _p in enumerate(_tmp):
                                        _dst = os.path.join(persist_dir, f"pix_{_i+1}.jpg")
                                        try:
                                            _sh.copyfile(_p, _dst)
                                            img_paths.append(_dst)
                                        except Exception:
                                            pass
                                except Exception as _e:
                                    self._emit_log(f"Pixabay 실패 ({name}): {_e}")
                        except Exception as _ie:
                            self._emit_log(f"이미지 저장 실패 ({name}): {_ie}")

                    content["image_paths"] = img_paths

                    with lock:
                        new_posts.append({"place": place, "content": content, "posted": False})
                        done_count[0] += 1
                        self._emit_log(f"[{done_count[0]}/{total}] '{name}' 생성 완료 (이미지 {len(img_paths)}장)")
                        self._emit_status(f"생성 {done_count[0]}/{total}", "#8b5cf6")
                except Exception as e:
                    with lock:
                        done_count[0] += 1
                        self._emit_log(f"[{done_count[0]}/{total}] '{name}' 생성 실패: {e}")

            max_workers = 5
            with ThreadPoolExecutor(max_workers=max_workers) as ex:
                futures = [ex.submit(_gen_one, i, p) for i, p in enumerate(selected, 1)]
                for f in as_completed(futures):
                    if self.stop_flag:
                        for fu in futures:
                            fu.cancel()
                        break

            # 병합 + 중복 제거 (같은 업체명+주소는 새 것으로 교체, 1개만 유지)
            def _k(p):
                pl = p.get("place", {})
                return (pl.get("name", ""), pl.get("address", "") or pl.get("jibun_address", ""))
            merged_map = {}  # key → post
            # 1) 기존 포스트 먼저 (오래된 것)
            for p in self._generated_posts:
                merged_map[_k(p)] = p
            # 2) 새로 생성한 것으로 덮어쓰기
            for p in new_posts:
                merged_map[_k(p)] = p
            self._generated_posts = list(merged_map.values())
            self.is_generating = False

            if self.stop_flag:
                self._emit_log(f"포스트 생성 중단됨 ({len(new_posts)}/{total}개 완료)")
                self._emit_status("중단됨", "#ef4444")
            else:
                self._emit_status("생성 완료", "#22c55e")
                self._emit_log(f"포스트 생성 완료: {len(self._generated_posts)}/{total}개")
            # 이미지 갯수 평준화 (부족 → 추가 다운로드, 초과 → 제거)
            if not self.stop_flag and pix_keys:
                try:
                    self._normalize_post_images(pix_keys[0])
                except Exception as _e:
                    self._emit_log(f"이미지 평준화 오류: {_e}")
            self._save_generated_posts()

            # 메인 스레드에서 미리보기 열기
            from PySide6.QtCore import QMetaObject, Q_ARG
            QMetaObject.invokeMethod(self, "_show_post_preview", Qt.QueuedConnection)

        threading.Thread(target=_worker, daemon=True).start()

    @Slot()
    def _show_post_preview(self):
        """생성된 포스트 미리보기"""
        if not self._generated_posts:
            QMessageBox.information(self, "결과", "생성된 포스트가 없습니다.")
            return

        from PySide6.QtWidgets import QTextEdit, QLineEdit

        dlg = QDialog(self)
        dlg.setWindowTitle("생성된 포스트 미리보기")
        dlg.resize(1000, 700)

        layout = QVBoxLayout(dlg)

        # 상단
        top = QHBoxLayout()
        select_all_cb = QCheckBox("전체 선택")
        count_label = QLabel(f"0 / {len(self._generated_posts)}개 선택됨")
        count_label.setStyleSheet("font-weight: bold; font-size: 13px;")
        top.addWidget(select_all_cb)
        top.addStretch()
        top.addWidget(count_label)

        btn_delete = QPushButton("선택 삭제")
        btn_delete.setStyleSheet("background: #ef4444; color: white; border: none; border-radius: 8px; font-size: 13px; padding: 8px 15px;")
        btn_delete.setCursor(Qt.PointingHandCursor)
        btn_post_selected = QPushButton("선택 포스트 생성")
        btn_post_selected.setStyleSheet("background: #4a6cf7; color: white; border: none; border-radius: 8px; font-size: 13px; font-weight: bold; padding: 8px 20px;")
        btn_post_selected.setCursor(Qt.PointingHandCursor)
        btn_close = QPushButton("닫기")
        btn_close.setStyleSheet("padding: 8px 20px;")
        btn_close.clicked.connect(dlg.reject)
        top.addWidget(btn_delete)
        top.addWidget(btn_post_selected)
        top.addWidget(btn_close)
        layout.addLayout(top)

        # 스크롤 영역
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        post_checkboxes = []
        post_editors = []
        title_editors = []

        for idx, item in enumerate(self._generated_posts):
            place = item["place"]
            content = item["content"]

            # 카드
            card = QFrame()
            card.setStyleSheet("background: white; border: 1px solid #e2e8f0; border-radius: 8px; padding: 10px; margin-bottom: 5px;")
            card_layout = QVBoxLayout(card)
            card_layout.setSpacing(5)

            # 헤더: 체크박스 + 업체명 + 저장 버튼
            hdr = QHBoxLayout()
            cb = QCheckBox(place.get("name", ""))
            cb.setStyleSheet("font-weight: bold; font-size: 13px;")
            post_checkboxes.append(cb)
            hdr.addWidget(cb)
            hdr.addStretch()

            btn_save = QPushButton("저장")
            btn_save.setStyleSheet("background: #22c55e; color: white; border: none; border-radius: 6px; padding: 5px 15px;")
            btn_save.setCursor(Qt.PointingHandCursor)
            hdr.addWidget(btn_save)
            card_layout.addLayout(hdr)

            # 제목 (편집 가능)
            title_edit = QLineEdit()
            title_edit.setText(content.get("title", ""))
            title_edit.setStyleSheet("color: #4a6cf7; font-size: 12px; padding: 5px; border: 1px solid #e2e8f0; border-radius: 4px;")
            title_editors.append(title_edit)
            card_layout.addWidget(title_edit)

            # 본문 (편집 가능)
            editor = QTextEdit()
            editor.setPlainText(content.get("body", ""))
            editor.setMaximumHeight(200)
            editor.setStyleSheet("border: 1px solid #e2e8f0; border-radius: 4px; padding: 5px; font-size: 11px;")
            post_editors.append(editor)
            card_layout.addWidget(editor)

            # 수집 데이터 표시
            addr = place.get("address", "")
            jibun = place.get("jibun_address", "")
            if jibun:
                full_addr = f"{addr} {jibun}" if addr else jibun
            else:
                full_addr = addr
            info_text = f"주소: {full_addr}  |  카테고리: {place.get('category', '')}  |  근처역: {place.get('nearby_station', '')}  |  태그: {', '.join(content.get('tags', []))}"
            info_label = QLabel(info_text)
            info_label.setStyleSheet("color: #64748b; font-size: 10px; padding: 3px 0; background: #f8fafc; border-radius: 4px; padding: 5px;")
            info_label.setWordWrap(True)
            card_layout.addWidget(info_label)

            # 저장 버튼 기능
            def make_save(i, te, ed):
                def save():
                    self._generated_posts[i]["content"]["title"] = te.text()
                    self._generated_posts[i]["content"]["body"] = ed.toPlainText()
                    self._save_generated_posts()
                    QMessageBox.information(dlg, "저장", "저장되었습니다.")
                return save

            btn_save.clicked.connect(make_save(idx, title_edit, editor))

            scroll_layout.addWidget(card)

        scroll_layout.addStretch()
        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)

        # 선택 카운트 업데이트
        def update_count():
            cnt = sum(1 for cb in post_checkboxes if cb.isChecked())
            count_label.setText(f"{cnt} / {len(self._generated_posts)}개 선택됨")

        for cb in post_checkboxes:
            cb.stateChanged.connect(lambda: update_count())

        # 선택 삭제
        def delete_selected():
            reply = QMessageBox.question(dlg, "삭제 확인", "선택된 포스트를 삭제하시겠습니까?",
                                          QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply != QMessageBox.Yes:
                return
            # 역순으로 삭제
            cards = [scroll_layout.itemAt(i).widget() for i in range(scroll_layout.count()) if scroll_layout.itemAt(i).widget()]
            for i in range(len(post_checkboxes) - 1, -1, -1):
                if post_checkboxes[i].isChecked():
                    self._generated_posts.pop(i)
                    post_checkboxes.pop(i)
                    post_editors.pop(i)
                    title_editors.pop(i)
                    widget = cards[i]
                    scroll_layout.removeWidget(widget)
                    widget.deleteLater()
            self._save_generated_posts()
            update_count()
            count_label.setText(f"0 / {len(self._generated_posts)}개 선택됨")

        btn_delete.clicked.connect(delete_selected)

        # 전체 선택
        def toggle_all(state):
            checked = select_all_cb.isChecked()
            for cb in post_checkboxes:
                cb.setChecked(checked)
            update_count()
        select_all_cb.stateChanged.connect(toggle_all)

        # 선택 포스팅
        def post_selected():
            self.posting_targets = []
            for i, cb in enumerate(post_checkboxes):
                if cb.isChecked():
                    self._generated_posts[i]["content"]["title"] = title_editors[i].text()
                    self._generated_posts[i]["content"]["body"] = post_editors[i].toPlainText()
                    self.posting_targets.append(self._generated_posts[i])
            if not self.posting_targets:
                QMessageBox.warning(dlg, "경고", "선택된 포스트가 없습니다.")
                return
            dlg.accept()
            self._post_generated()

        btn_post_selected.clicked.connect(post_selected)
        dlg.exec()

    def _post_generated(self):
        """미리보기에서 선택된 포스트를 블로그에 포스팅"""
        if not self.posting_targets:
            return

        cfg = load_config()
        from config import get_active_account
        account = get_active_account(cfg)
        if not account.get("naver_id") or not account.get("naver_pw"):
            QMessageBox.critical(self, "오류", "네이버 계정 정보를 설정해주세요.")
            return

        total = len(self.posting_targets)
        interval_sec = self._get_interval_seconds()
        if getattr(self, "random_interval_2h", None) and self.random_interval_2h.isChecked():
            interval_label = "랜덤 109~133분"
        elif getattr(self, "random_interval_3h", None) and self.random_interval_3h.isChecked():
            interval_label = "랜덤 169~193분"
        else:
            h = int(self.interval_hour.currentText())
            m = int(self.interval_min.currentText())
            interval_label = f"{h}시간 {m}분"
        acc_idx = cfg.get("active_account", 0) + 1

        reply = QMessageBox.question(self, "포스팅 확인",
            f"[아이디 {acc_idx}] {account['naver_id']}\n{total}개 포스팅\n간격: {interval_label}\n\n시작할까요?")
        if reply != QMessageBox.Yes:
            return

        keyword = self.keyword_input.currentText().strip()
        pix_keys = [k for k in cfg.get("pixabay_key_list", []) if k]
        self.is_posting = True
        self.stop_flag = False

        def _worker():
            poster = None
            try:
                self._emit_status("포스팅 중...", "#8b5cf6")
                # 메인 프로그램 창 위치에 브라우저 띄우기
                _g = self.geometry()
                poster = NaverBlogPoster(
                    naver_id=account["naver_id"],
                    naver_pw=account["naver_pw"],
                    blog_id=account["blog_id"],
                    window_x=_g.x(),
                    window_y=_g.y(),
                )
                self._active_posters.append(poster)
                self._emit_log("브라우저 시작...")
                poster.start_browser()
                self._emit_log("네이버 로그인 중...")
                if not poster.login():
                    self._emit_log("로그인 실패!")
                    return
                self._emit_log("로그인 성공")

                import datetime as _dt
                base_time = _dt.datetime.now()
                first_immediate = getattr(self, 'first_post_immediate', True)
                running_dt = base_time
                for i, item in enumerate(self.posting_targets, 1):
                    if self.stop_flag:
                        self._emit_log("포스팅 중단됨")
                        break

                    # 첫 포스트: 즉시 발행 (옵션에 따라) / 나머지: 예약 발행
                    if i == 1 and first_immediate:
                        schedule_time = None  # 즉시 발행
                    else:
                        # 매 포스트마다 _get_interval_seconds() 호출 → 랜덤 모드면 매번 다른 간격
                        running_dt = running_dt + _dt.timedelta(seconds=self._get_interval_seconds())
                        sched_dt = running_dt
                        # 10분 단위 반올림
                        minute = (sched_dt.minute // 10) * 10
                        sched_dt = sched_dt.replace(minute=minute, second=0, microsecond=0)
                        if sched_dt <= _dt.datetime.now():
                            sched_dt += _dt.timedelta(minutes=10)
                        schedule_time = sched_dt.strftime("%Y-%m-%d %H:%M")
                        self._emit_log(f"[{i}/{total}] 예약 시간: {schedule_time}")

                    place = item["place"]
                    content = item["content"]
                    name = place.get("name", "")
                    self._emit_log(f"[{i}/{total}] '{name}' 포스팅 중...")
                    self._emit_status(f"포스팅 {i}/{total}", "#8b5cf6")

                    # 포스트 생성 시 저장된 이미지 우선 사용 (재다운로드 방지)
                    img_paths = []
                    saved_paths = content.get("image_paths", []) or []
                    saved_paths = [p for p in saved_paths if p and os.path.exists(p)]
                    if saved_paths:
                        img_paths = saved_paths
                        self._emit_log(f"저장된 이미지 {len(img_paths)}장 사용")
                    elif pix_keys:
                        try:
                            biz = (place.get("category") or "").strip() or keyword
                            self._emit_log(f"이미지 재검색 업종: '{biz}' (저장본 없음)")
                            img_paths = download_images(
                                pix_keys[0], biz,
                                content.get("image_count", 3)
                            )
                            self._emit_log(f"이미지 {len(img_paths)}장 다운로드")
                        except Exception as e:
                            self._emit_log(f"이미지 다운로드 실패: {e}")
                    else:
                        self._emit_log("Pixabay API 키 미설정 - 이미지 없이 포스팅")

                    body = content.get("body", "")

                    # 제목의 잘못된 지역명 교체 (검색 키워드의 구 ≠ 실제 주소의 구인 경우)
                    actual_address = place.get("address", "") or place.get("jibun_address", "")
                    import re
                    actual_gu_match = re.search(r"([가-힣]+구)", actual_address)
                    if actual_gu_match:
                        actual_gu = actual_gu_match.group(1)
                        title = content.get("title", "")
                        wrong_gus = re.findall(r"([가-힣]+구)", title)
                        for wg in wrong_gus:
                            if wg != actual_gu:
                                title = title.replace(wg, actual_gu)
                                self._emit_log(f"제목 지역 교체: {wg} → {actual_gu}")
                        # 동 이름도 확인
                        actual_dong_match = re.search(r"([가-힣]+동)", actual_address)
                        if actual_dong_match:
                            actual_dong = actual_dong_match.group(1)
                            wrong_dongs = re.findall(r"([가-힣]+동)", title)
                            for wd in wrong_dongs:
                                if wd != actual_dong and wd != actual_gu:
                                    title = title.replace(wd, actual_dong)
                        content["title"] = title

                    # 이미지가 있는데 본문에 [이미지] 마커가 없으면 균등 분포로 자동 삽입
                    if img_paths and "[이미지]" not in body:
                        paragraphs = [p for p in body.split("\n\n") if p.strip()]
                        if len(paragraphs) >= 2:
                            step = max(1, len(paragraphs) // (len(img_paths) + 1))
                            new_parts = []
                            img_used = 0
                            for idx, p in enumerate(paragraphs):
                                new_parts.append(p)
                                if img_used < len(img_paths) and (idx + 1) % step == 0 and idx < len(paragraphs) - 1:
                                    new_parts.append("[이미지]")
                                    img_used += 1
                            body = "\n\n".join(new_parts)
                            content["body"] = body
                            self._emit_log(f"본문에 [이미지] 마커 {img_used}개 자동 삽입")

                    try:
                        success = poster.write_post(
                            title=content["title"],
                            body=content["body"],
                            tags=content["tags"],
                            image_paths=img_paths,
                            category=account.get("blog_category", ""),
                            place_name=name,
                            place_address=place.get("address", "") or place.get("jibun_address", ""),
                            schedule_time=schedule_time,
                        )
                        msg = "완료!" if success else "실패"
                        self._emit_log(f"[{i}/{total}] '{name}' {msg}")
                        # 브라우저 세션 확인 — 죽었으면 즉시 중단
                        try:
                            _ = poster.driver.current_url
                        except Exception as _se:
                            self._emit_log(f"!!! 브라우저 세션이 종료됨: {str(_se)[:120]}. 포스팅 중단")
                            self.stop_flag = True
                            break
                        if success:
                            # 포스팅 완료 시: 생성된 포스트만 제거 (업체는 F7 목록에 유지)
                            _k = self._place_key(place)
                            self._generated_posts = [
                                gp for gp in self._generated_posts
                                if self._place_key(gp.get("place", {})) != _k
                            ]
                            self._save_generated_posts()
                            self._emit_log(f"'{place.get('name','')}' 포스팅 완료 → 생성된 포스트 제거")
                    except Exception as e:
                        self._emit_log(f"'{name}' 오류: {e}")

                if not self.stop_flag:
                    self._emit_log(f"전체 포스팅 완료! ({total}개)")
                self._emit_status("완료", "#22c55e")

            except Exception as e:
                self._emit_log(f"오류: {e}")
                self._emit_status("오류", "#ef4444")
            finally:
                self.is_posting = False
                if poster:
                    try: self._active_posters.remove(poster)
                    except Exception: pass
                    poster.close()

        threading.Thread(target=_worker, daemon=True).start()

    # ── 포스팅 업체 선택 ──
    @staticmethod
    def _gu_of_place(p: dict) -> str:
        """주소에서 '구/군' 추출 — '강남구' 같은 단일 토큰 반환.
        '서울특별시' 같은 시 토큰은 건너뛰고 구/군 우선."""
        import re as _re
        addr = (p.get("address", "") or "") + " " + (p.get("jibun_address", "") or "")
        m = _re.search(r"([가-힣]+[구군])", addr)
        if m:
            return m.group(1)
        m = _re.search(r"([가-힣]+시)", addr)
        return m.group(1) if m else "지역 미상"

    def _load_all_crawl_results(self) -> dict:
        """현재 계정의 logs 폴더의 모든 크롤링 결과를 (구 × 업종)별로 집계.
        여러 크롤 파일에 걸쳐 같은 (구, 업종)은 한 그룹으로 합치고 (업체명, 주소) 기준 중복 제거.
        라벨에 가장 최근 크롤 날짜 표시."""
        from places_crawler import load_results
        log_dir = self._get_logs_dir()
        # (gu, keyword) → {"items": list, "seen": set, "latest": str}
        buckets: dict = {}
        if not os.path.exists(log_dir):
            return {}
        files = sorted(
            [f for f in os.listdir(log_dir) if f.endswith(".json")],
            reverse=True
        )
        for f in files:
            try:
                raw = load_results(os.path.join(log_dir, f))
                keyword = (raw.get("keyword") or "").strip()
                items = raw.get("items", [])
                if not items:
                    continue
                date_raw = f.replace("places_", "").replace(".json", "")
                try:
                    if len(date_raw) >= 13:
                        date_part = f"{date_raw[:4]}-{date_raw[4:6]}-{date_raw[6:8]} {date_raw[9:11]}:{date_raw[11:13]}"
                    else:
                        date_part = date_raw
                except Exception:
                    date_part = date_raw
                # 업종명 정규화 — 지역 토큰 제거 (앞쪽)
                biz = keyword
                _parts = keyword.split()
                if len(_parts) >= 2:
                    biz = _parts[-1]
                for p in items:
                    gu = self._gu_of_place(p)
                    key = (gu, biz)
                    b = buckets.get(key)
                    if b is None:
                        b = {"items": [], "seen": set(), "latest": date_part}
                        buckets[key] = b
                    n = (p.get("name") or "").strip()
                    a = (p.get("address") or p.get("jibun_address") or "").strip()
                    sig = (n, a)
                    if n and sig not in b["seen"]:
                        b["seen"].add(sig)
                        b["items"].append(p)
                    # 파일은 최신→과거 순으로 처리 중이므로 첫 등장이 최신
            except Exception:
                continue
        # 라벨 생성 — 구별로 정렬
        groups = {}
        for (gu, biz), b in sorted(buckets.items(), key=lambda x: (x[0][0], x[0][1])):
            label = f"[{gu}] {biz} ({len(b['items'])}개) · {b['latest']}"
            groups[label] = b["items"]
        return groups

    def _show_posting_selector(self) -> list:
        # 새로고침 시 보존된 체크/편집 상태 (한 번 적용 후 폐기)
        _preserved_checks = getattr(self, "_selector_preserved_checks", None) or set()
        _preserved_edits = getattr(self, "_selector_preserved_edits", None) or {}
        self._selector_preserved_checks = None
        self._selector_preserved_edits = None
        # 모든 크롤링 결과 로드
        all_groups = self._load_all_crawl_results()
        # 현재 메모리에 있는 것도 추가 — 구별로 분할해서 표시
        if self.crawled_data:
            _sub_cur = {}
            for _p in self.crawled_data:
                _sub_cur.setdefault(self._gu_of_place(_p), []).append(_p)
            _now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            if len(_sub_cur) > 1:
                for _gu, _its in sorted(_sub_cur.items()):
                    all_groups[f"[현재/{_gu}] ({len(_its)}개) · {_now}"] = _its
            else:
                all_groups[f"현재 크롤링 결과 ({len(self.crawled_data)}개) · {_now}"] = self.crawled_data

        if not all_groups:
            QMessageBox.warning(self, "경고", "크롤링 결과가 없습니다.")
            return []

        # 생성된 포스트 로드해서 (업체명+주소) → post 매핑
        _gen_posts = self._load_generated_posts()
        _post_map = {}
        for _gp in _gen_posts:
            _pl = _gp.get("place", {})
            _key = (_pl.get("name", ""), _pl.get("address", "") or _pl.get("jibun_address", ""))
            _post_map[_key] = _gp

        dlg = QDialog(self)
        dlg.setWindowTitle("포스팅할 업체 선택")
        dlg.resize(950, 600)

        layout = QVBoxLayout(dlg)

        # 상단
        top = QHBoxLayout()
        btn_all = QPushButton("전체 선택")
        btn_all.setStyleSheet("padding: 6px 15px;")
        btn_none = QPushButton("전체 해제")
        btn_none.setStyleSheet("padding: 6px 15px;")
        count_label = QLabel("0개 선택됨")
        count_label.setStyleSheet("font-weight: bold; font-size: 13px;")
        top.addWidget(btn_all)
        top.addWidget(btn_none)
        top.addStretch()
        top.addWidget(count_label)
        layout.addLayout(top)

        # 트리뷰
        tree = QTreeWidget()
        tree.setHeaderLabels(["업체명", "업체주소", "카테고리", "근처역", "앞 키워드", "태그", "픽사베이 키워드"])
        tree.setColumnWidth(0, 320)
        tree.setColumnWidth(1, 220)
        tree.setColumnWidth(2, 120)
        tree.setColumnWidth(3, 120)
        tree.setColumnWidth(4, 180)
        tree.setColumnWidth(5, 140)
        tree.setColumnWidth(6, 180)
        tree.setAlternatingRowColors(True)
        tree.setStyleSheet("""
            QTreeWidget { font-size: 12px; }
            QTreeWidget::item { padding: 3px 0; }
            alternate-background-color: #f5f5f5;
        """)

        # 그룹별 트리 구성
        all_items = []  # (QTreeWidgetItem, place_data)
        for group_name, places in all_groups.items():
            if not places:
                continue
            parent = QTreeWidgetItem(tree)
            parent.setText(0, group_name)
            parent.setFlags(parent.flags() | Qt.ItemIsUserCheckable)
            parent.setCheckState(0, Qt.Unchecked)
            parent.setExpanded(False)

            posted_count = 0
            child_count = 0

            for p in places:
                addr = p.get("address", "")
                jibun = p.get("jibun_address", "")
                if jibun:
                    full_addr = f"{addr} {jibun}" if addr else jibun
                else:
                    full_addr = addr

                child = QTreeWidgetItem(parent)
                child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
                child.setCheckState(0, Qt.Unchecked)

                # 상태 표시등: 생성된 포스트가 있는지 + 업로드 완료 여부
                _k = (p.get("name", ""), p.get("address", "") or p.get("jibun_address", ""))
                _gp = _post_map.get(_k)
                is_generated = _gp is not None
                is_done = bool(_gp and _gp.get("posted", False))
                if is_done:
                    child.setText(0, "[완] " + p.get("name", ""))
                    child.setForeground(0, QColor("#3b82f6"))
                    posted_count += 1
                elif is_generated:
                    child.setText(0, "●  " + p.get("name", ""))
                    child.setForeground(0, QColor("#22c55e"))
                    posted_count += 1
                else:
                    child.setText(0, "●  " + p.get("name", ""))
                    child.setForeground(0, QColor("#ef4444"))

                child.setText(1, full_addr)
                child.setText(2, p.get("category", ""))
                child.setText(3, p.get("nearby_station", ""))
                child.setText(4, p.get("front_keywords", ""))
                child.setText(5, p.get("tags", ""))
                child.setText(6, p.get("pixabay_keywords", ""))
                child.setFlags(child.flags() | Qt.ItemIsEditable | Qt.ItemIsUserCheckable)
                # 새로고침 전 체크/편집 상태 복원
                _pkey = (p.get("name", ""), p.get("address", "") or p.get("jibun_address", ""))
                if _pkey in _preserved_checks:
                    child.setCheckState(0, Qt.Checked)
                _e = _preserved_edits.get(_pkey)
                if _e:
                    if "front_keywords" in _e: child.setText(4, _e["front_keywords"])
                    if "tags" in _e: child.setText(5, _e["tags"])
                    if "pixabay_keywords" in _e: child.setText(6, _e["pixabay_keywords"])
                all_items.append((child, p))
                child_count += 1

            # 폴더 상태등: 모두 초록 / 모두 빨강 / 혼합 주황
            if child_count > 0:
                if posted_count == child_count:
                    parent.setForeground(0, QColor("#22c55e"))
                    parent.setText(0, "●  " + group_name)
                elif posted_count == 0:
                    parent.setForeground(0, QColor("#ef4444"))
                    parent.setText(0, "●  " + group_name)
                else:
                    parent.setForeground(0, QColor("#f59e0b"))
                    parent.setText(0, "●  " + group_name)

        def update_count():
            cnt = sum(1 for item, _ in all_items if item.checkState(0) == Qt.Checked)
            count_label.setText(f"{cnt}개 선택됨")

        def on_item_changed(item, col):
            if col != 0:
                return
            # 부모 체크 → 자식 전부 체크/해제
            if item.childCount() > 0:
                state = item.checkState(0)
                for i in range(item.childCount()):
                    item.child(i).setCheckState(0, state)
            update_count()

        tree.itemChanged.connect(on_item_changed)

        def select_all():
            tree.blockSignals(True)
            for i in range(tree.topLevelItemCount()):
                parent = tree.topLevelItem(i)
                parent.setCheckState(0, Qt.Checked)
                for j in range(parent.childCount()):
                    parent.child(j).setCheckState(0, Qt.Checked)
            tree.blockSignals(False)
            update_count()

        def select_none():
            tree.blockSignals(True)
            for i in range(tree.topLevelItemCount()):
                parent = tree.topLevelItem(i)
                parent.setCheckState(0, Qt.Unchecked)
                for j in range(parent.childCount()):
                    parent.child(j).setCheckState(0, Qt.Unchecked)
            tree.blockSignals(False)
            update_count()

        btn_all.clicked.connect(select_all)
        btn_none.clicked.connect(select_none)

        layout.addWidget(tree)

        # 하단
        bottom = QHBoxLayout()

        # 삭제 버튼들 — 영구 저장 (deleted_keys_*.json) + generated_posts 동기화
        def _persist_delete(places_to_delete: list):
            if not places_to_delete:
                return
            keys = {self._place_key(p) for p in places_to_delete}
            # logs 파일에서 즉시 영구 삭제
            self._purge_places_from_logs(places_to_delete)
            # generated_posts에도 있으면 제거
            changed = False
            new_gp = []
            current = self._load_generated_posts()
            for gp in current:
                if self._place_key(gp.get("place", {})) in keys:
                    changed = True
                    continue
                new_gp.append(gp)
            if changed:
                self._generated_posts = new_gp
                self._save_generated_posts()

        def delete_selected():
            # 체크된 places 수집
            checked_places = []
            for it, p in all_items:
                if it.checkState(0) == Qt.Checked:
                    checked_places.append(p)
            if not checked_places:
                QMessageBox.information(dlg, "안내", "선택된 항목이 없습니다.")
                return
            reply = QMessageBox.question(dlg, "삭제 확인", f"{len(checked_places)}개 항목을 삭제하시겠습니까?\n(F8에서도 같이 제거됩니다)",
                                          QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply != QMessageBox.Yes:
                return
            _persist_delete(checked_places)
            # UI에서 제거
            tree.blockSignals(True)
            to_remove_idx = []
            for i in range(tree.topLevelItemCount()):
                parent = tree.topLevelItem(i)
                for j in range(parent.childCount() - 1, -1, -1):
                    child = parent.child(j)
                    if child.checkState(0) == Qt.Checked:
                        all_items[:] = [(it, p) for it, p in all_items if it is not child]
                        parent.removeChild(child)
                if parent.childCount() == 0:
                    to_remove_idx.append(i)
            for idx in reversed(to_remove_idx):
                tree.takeTopLevelItem(idx)
            tree.blockSignals(False)
            update_count()
            self._emit_log(f"{len(checked_places)}개 항목 삭제 (영구 반영)")

        def delete_all():
            all_places = [p for _, p in all_items]
            if not all_places:
                return
            reply = QMessageBox.question(dlg, "전체 삭제 확인", f"현재 표시된 전체 {len(all_places)}개 항목을 삭제하시겠습니까?\n(F8에서도 같이 제거됩니다)",
                                          QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply != QMessageBox.Yes:
                return
            _persist_delete(all_places)
            tree.blockSignals(True)
            tree.clear()
            all_items.clear()
            tree.blockSignals(False)
            update_count()
            self._emit_log(f"{len(all_places)}개 전체 삭제 (영구 반영)")

        btn_del_sel = QPushButton("선택 삭제")
        btn_del_sel.setStyleSheet("background: #ef4444; color: white; border: none; border-radius: 8px; padding: 8px 15px;")
        btn_del_sel.setCursor(Qt.PointingHandCursor)
        btn_del_sel.clicked.connect(delete_selected)

        btn_del_all = QPushButton("전체 삭제")
        btn_del_all.setStyleSheet("background: #dc2626; color: white; border: none; border-radius: 8px; padding: 8px 15px;")
        btn_del_all.setCursor(Qt.PointingHandCursor)
        btn_del_all.clicked.connect(delete_all)

        btn_refresh = QPushButton("새로고침")
        btn_refresh.setStyleSheet("background: #64748b; color: white; border: none; border-radius: 8px; padding: 8px 15px;")
        btn_refresh.setCursor(Qt.PointingHandCursor)
        def do_refresh():
            # 현재 체크/편집 상태 보존 → 재오픈 시 복원
            _ck = set()
            _ed = {}
            for it, p in all_items:
                _k = (p.get("name", ""), p.get("address", "") or p.get("jibun_address", ""))
                if it.checkState(0) == Qt.Checked:
                    _ck.add(_k)
                _e = {}
                if it.text(4) != (p.get("front_keywords") or ""): _e["front_keywords"] = it.text(4)
                if it.text(5) != (p.get("tags") or ""): _e["tags"] = it.text(5)
                if it.text(6) != (p.get("pixabay_keywords") or ""): _e["pixabay_keywords"] = it.text(6)
                if _e: _ed[_k] = _e
            self._selector_preserved_checks = _ck
            self._selector_preserved_edits = _ed
            dlg.done(2)  # 특수 코드: 새로고침
        btn_refresh.clicked.connect(do_refresh)

        bottom.addWidget(btn_del_sel)
        bottom.addWidget(btn_del_all)
        bottom.addWidget(btn_refresh)
        bottom.addStretch()

        btn_cancel = QPushButton("취소")
        btn_cancel.setStyleSheet("padding: 8px 20px;")
        btn_cancel.clicked.connect(dlg.reject)

        # 프롬프트 선택 드롭다운 (취소 옆)
        try:
            import json as _pj
            _prompts_path = os.path.join(os.path.dirname(__file__), "prompts.json")
            with open(_prompts_path, "r", encoding="utf-8") as _pf:
                _prompts_dict = _pj.load(_pf)
            _prompt_keys = list(_prompts_dict.keys())
        except Exception:
            _prompt_keys = []
        prompt_combo = QComboBox()
        prompt_combo.addItem("(자동 선택)")
        for k in _prompt_keys:
            prompt_combo.addItem(k)
        prompt_combo.setStyleSheet("padding: 6px 10px; font-size: 12px;")
        prompt_lbl = QLabel("프롬프트:")
        prompt_lbl.setStyleSheet("font-size: 12px; color: #475569;")

        btn_ok = QPushButton("포스트 생성")
        btn_ok.setStyleSheet("background: #8b5cf6; color: white; border: none; border-radius: 8px; font-size: 13px; font-weight: bold; padding: 10px 25px;")
        btn_ok.setCursor(Qt.PointingHandCursor)

        def _on_ok():
            sel = prompt_combo.currentText()
            self._override_prompt_name = sel if sel and sel != "(자동 선택)" else None
            dlg.accept()
        btn_ok.clicked.connect(_on_ok)

        bottom.addWidget(btn_cancel)
        bottom.addWidget(prompt_lbl)
        bottom.addWidget(prompt_combo)
        bottom.addWidget(btn_ok)
        layout.addLayout(bottom)

        _rc = dlg.exec()
        if _rc == 2:
            # 새로고침: 다이얼로그 재오픈
            return self._show_posting_selector()
        if _rc == QDialog.Accepted:
            selected = []
            for item, p in all_items:
                if item.checkState(0) == Qt.Checked:
                    # 수정된 값 반영
                    p["front_keywords"] = item.text(4)
                    p["tags"] = item.text(5)
                    p["pixabay_keywords"] = item.text(6)
                    selected.append(p)
            return selected
        return []

    # ── 블로그 포스팅 ──
    def _start_blog_posting(self):
        # 크롤링 결과 불러오기 (현재 없으면 logs에서 로드)
        if not self.crawled_data:
            log_dir = self._get_logs_dir()
            if os.path.exists(log_dir):
                files = sorted(
                    [f for f in os.listdir(log_dir) if f.endswith(".json")],
                    reverse=True
                )
                if files:
                    from places_crawler import load_results
                    filepath = os.path.join(log_dir, files[0])
                    raw = load_results(filepath)
                    self.crawled_data = raw.get("items", []) if isinstance(raw, dict) else raw

        if not self.crawled_data:
            QMessageBox.warning(self, "경고", "먼저 크롤링을 실행해주세요.")
            return

        # 업체 선택 다이얼로그
        selected = self._show_posting_selector()
        if not selected:
            return

        cfg = load_config()
        provider = "GPT"  # GPT 전용 고정
        api_keys = [k for k in cfg.get("gpt_key_list", []) if k]
        if not api_keys:
            QMessageBox.critical(self, "오류", "GPT API 키를 설정해주세요.")
            return
        api_key = api_keys[0]
        from config import get_active_account
        account = get_active_account(cfg)
        if not account.get("naver_id") or not account.get("naver_pw"):
            QMessageBox.critical(self, "오류", "네이버 계정 정보를 설정해주세요.")
            return
        if self.is_posting:
            return

        total = len(selected)
        interval_sec = self._get_interval_seconds()
        if getattr(self, "random_interval_2h", None) and self.random_interval_2h.isChecked():
            interval_label = "랜덤 109~133분"
        elif getattr(self, "random_interval_3h", None) and self.random_interval_3h.isChecked():
            interval_label = "랜덤 169~193분"
        else:
            h = int(self.interval_hour.currentText())
            m = int(self.interval_min.currentText())
            interval_label = f"{h}시간 {m}분"
        acc_idx = cfg.get("active_account", 0) + 1

        reply = QMessageBox.question(self, "포스팅 확인",
            f"[아이디 {acc_idx}] {account['naver_id']}\n선택된 {total}개 업체 포스팅\n간격: {interval_label}\n\n시작할까요?")
        if reply != QMessageBox.Yes:
            return

        keyword = self.keyword_input.currentText().strip()
        self.is_posting = True
        self.stop_flag = False
        self.posting_targets = selected

        def _worker():
            poster = None
            try:
                self._emit_status("포스팅 중...", "#8b5cf6")
                # 메인 프로그램 창 위치에 브라우저 띄우기
                _g = self.geometry()
                poster = NaverBlogPoster(
                    naver_id=account["naver_id"],
                    naver_pw=account["naver_pw"],
                    blog_id=account["blog_id"],
                    window_x=_g.x(),
                    window_y=_g.y(),
                )
                self._active_posters.append(poster)
                self._emit_log("브라우저 시작...")
                poster.start_browser()

                self._emit_log("네이버 로그인 중...")
                if not poster.login():
                    self._emit_log("로그인 실패!")
                    return

                self._emit_log("로그인 성공")

                for i, place in enumerate(self.posting_targets, 1):
                    if self.stop_flag:
                        self._emit_log("포스팅 중단됨")
                        break

                    name = place.get("name", "")
                    self._emit_log(f"[{i}/{total}] '{name}' 글 생성 중...")
                    self._emit_status(f"포스팅 {i}/{total}", "#8b5cf6")

                    try:
                        content = generate_content(
                            provider=cfg["ai_provider"],
                            api_key=api_key,
                            place=place, keyword=keyword
                        )
                    except Exception as e:
                        self._emit_log(f"'{name}' 글 생성 실패: {e}")
                        continue

                    img_paths = []
                    pix_keys = [k for k in cfg.get("pixabay_key_list", []) if k]
                    if pix_keys:
                        biz = (place.get("category") or "").strip() or keyword
                        img_paths = download_images(
                            pix_keys[0], biz,
                            content.get("image_count", 3)
                        )

                    self._emit_log(f"[{i}/{total}] '{name}' 포스팅 중...")

                    try:
                        success = poster.write_post(
                            title=content["title"], body=content["body"],
                            tags=content["tags"], image_paths=img_paths,
                            category=account.get("blog_category", ""),
                        )
                        msg = "완료!" if success else "실패"
                        self._emit_log(f"[{i}/{total}] '{name}' {msg}")
                    except Exception as e:
                        self._emit_log(f"'{name}' 오류: {e}")

                    if i < total and not self.stop_flag:
                        # 매 포스트마다 새 random 간격 (체크된 경우)
                        this_interval = self._get_interval_seconds()
                        if this_interval <= 0:
                            continue
                        remaining = this_interval
                        while remaining > 0 and not self.stop_flag:
                            m_left, s_left = divmod(remaining, 60)
                            self._emit_status(f"대기 {m_left}분 {s_left}초", "#94a3b8")
                            import time as _time
                            wait = min(60, remaining)
                            _time.sleep(wait)
                            remaining -= wait

                if not self.stop_flag:
                    self._emit_log(f"전체 포스팅 완료! ({total}개)")

                self._emit_status("완료", "#22c55e")

            except Exception as e:
                self._emit_log(f"오류: {e}")
                self._emit_status("오류", "#ef4444")
            finally:
                self.is_posting = False
                if poster:
                    try: self._active_posters.remove(poster)
                    except Exception: pass
                    poster.close()

        threading.Thread(target=_worker, daemon=True).start()


LOGIN_CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "login_cache.json")


def _load_login_cache():
    try:
        with open(LOGIN_CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_login_cache(data):
    try:
        with open(LOGIN_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


class LoginDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("로그인")
        self.setFixedSize(340, 290)
        self.user = None  # 로그인 성공 시 유저 dict

        layout = QVBoxLayout(self)
        layout.setContentsMargins(25, 20, 25, 20)

        title = QLabel("블로그마스터")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #1e293b;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        layout.addSpacing(12)

        self.id_edit = QLineEdit()
        self.id_edit.setPlaceholderText("아이디")
        self.id_edit.setStyleSheet("padding: 8px; font-size: 13px;")
        layout.addWidget(self.id_edit)

        self.pw_edit = QLineEdit()
        self.pw_edit.setEchoMode(QLineEdit.Password)
        self.pw_edit.setPlaceholderText("비밀번호")
        self.pw_edit.setStyleSheet("padding: 8px; font-size: 13px;")
        self.pw_edit.returnPressed.connect(self._try_login)
        layout.addWidget(self.pw_edit)

        chk_row = QHBoxLayout()
        self.chk_save_id = QCheckBox("아이디 저장")
        self.chk_save_pw = QCheckBox("비밀번호 저장")
        self.chk_save_id.setStyleSheet("font-size: 11px; color: #64748b;")
        self.chk_save_pw.setStyleSheet("font-size: 11px; color: #64748b;")
        chk_row.addWidget(self.chk_save_id)
        chk_row.addWidget(self.chk_save_pw)
        chk_row.addStretch()
        layout.addLayout(chk_row)

        self.msg = QLabel("")
        self.msg.setStyleSheet("color: #ef4444; font-size: 11px;")
        layout.addWidget(self.msg)

        btn_row = QHBoxLayout()
        btn_login = QPushButton("로그인")
        btn_login.setStyleSheet("background: #4a6cf7; color: white; border: none; border-radius: 6px; padding: 8px 20px; font-weight: bold;")
        btn_login.setCursor(Qt.PointingHandCursor)
        btn_login.clicked.connect(self._try_login)
        btn_cancel = QPushButton("종료")
        btn_cancel.setStyleSheet("padding: 8px 20px;")
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(btn_login)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

        layout.addSpacing(4)
        signup_row = QHBoxLayout()
        signup_row.addStretch()
        self.signup_link = QLabel('<a href="#" style="color:#64748b; text-decoration:none;">회원가입</a>')
        self.signup_link.setStyleSheet("font-size: 11px; color: #64748b;")
        self.signup_link.setCursor(Qt.PointingHandCursor)
        self.signup_link.linkActivated.connect(lambda: self._open_register())
        signup_row.addWidget(self.signup_link)
        signup_row.addStretch()
        layout.addLayout(signup_row)

        self._load_cached()

    def _load_cached(self):
        cache = _load_login_cache()
        if cache.get("save_id"):
            self.chk_save_id.setChecked(True)
            self.id_edit.setText(cache.get("id", ""))
        if cache.get("save_pw"):
            self.chk_save_pw.setChecked(True)
            self.pw_edit.setText(cache.get("pw", ""))

    def _open_signup(self):
        dlg = SignupDialog(self)
        if dlg.exec() == QDialog.Accepted and dlg.username:
            self.id_edit.setText(dlg.username)
            self.pw_edit.setFocus()
            self.msg.setStyleSheet("color: #22c55e; font-size: 11px;")
            self.msg.setText("가입 완료. 로그인해주세요.")

    def _open_register(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("회원가입")
        dlg.setFixedSize(400, 680)
        layout = QVBoxLayout(dlg)

        title = QLabel("회원가입")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #1e293b;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        layout.addSpacing(8)

        dup_checked = {"phone": False, "email": False}
        fields = {}

        for label, key, placeholder, is_pw in [
            ("아이디 *", "uid", "영문/숫자", False),
            ("비밀번호 *", "pw", "비밀번호", True),
            ("비밀번호 확인 *", "pw2", "비밀번호 재입력", True),
            ("이름 *", "name", "홍길동", False),
            ("생년월일 *", "birth", "1990-01-01", False),
            ("연락처 *", "phone", "010-1234-5678", False),
            ("이메일 *", "email", "example@email.com", False),
            ("추천인", "referrer", "추천인 아이디 (선택)", False),
        ]:
            row = QHBoxLayout()
            row.addWidget(QLabel(label))
            row.addStretch()
            # 중복확인 버튼 (연락처/이메일)
            if key in ("phone", "email"):
                btn_dup = QPushButton("중복확인")
                btn_dup.setStyleSheet("background: #64748b; color: white; border: none; border-radius: 4px; padding: 3px 8px; font-size: 11px;")
                btn_dup.setCursor(Qt.PointingHandCursor)
                btn_dup.setProperty("field_key", key)
                row.addWidget(btn_dup)
            layout.addLayout(row)
            entry = QLineEdit()
            entry.setPlaceholderText(placeholder)
            entry.setStyleSheet("padding: 6px;")
            if is_pw:
                entry.setEchoMode(QLineEdit.Password)
            if key == "birth":
                entry.setMaxLength(10)
                def _dash_date(t, e=entry):
                    digits = t.replace("-", "")
                    if not digits.isdigit(): return
                    e.blockSignals(True)
                    digits = digits[:8]
                    if len(digits) >= 6: r = f"{digits[:4]}-{digits[4:6]}-{digits[6:]}"
                    elif len(digits) >= 4: r = f"{digits[:4]}-{digits[4:]}"
                    else: r = digits
                    e.setText(r); e.setCursorPosition(len(r))
                    e.blockSignals(False)
                entry.textChanged.connect(_dash_date)
            if key == "phone":
                entry.setMaxLength(13)
                def _dash_phone(t, e=entry):
                    digits = t.replace("-", "")
                    if not digits.isdigit(): return
                    e.blockSignals(True)
                    digits = digits[:11]
                    if len(digits) >= 8: r = f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
                    elif len(digits) >= 3: r = f"{digits[:3]}-{digits[3:]}"
                    else: r = digits
                    e.setText(r); e.setCursorPosition(len(r))
                    e.blockSignals(False)
                entry.textChanged.connect(_dash_phone)
                entry.textChanged.connect(lambda: dup_checked.update({"phone": False}))
            if key == "email":
                entry.textChanged.connect(lambda: dup_checked.update({"email": False}))
            layout.addWidget(entry)
            fields[key] = entry
            # 중복확인 연결
            if key in ("phone", "email"):
                def _check_dup(checked, k=key):
                    val = fields[k].text().strip()
                    if not val:
                        QMessageBox.warning(dlg, "경고", "값을 입력해주세요.")
                        return
                    from users import load_users
                    users = load_users()
                    taken = any(u.get(k, "") == val for u in users.values())
                    if taken:
                        QMessageBox.warning(dlg, "중복", f"이미 등록된 {('연락처' if k=='phone' else '이메일')}입니다.")
                        dup_checked[k] = False
                    else:
                        QMessageBox.information(dlg, "확인", "사용 가능합니다.")
                        dup_checked[k] = True
                btn_dup.clicked.connect(_check_dup)

        msg = QLabel("")
        msg.setStyleSheet("color: #ef4444; font-size: 11px;")
        layout.addWidget(msg)

        btn_row = QHBoxLayout()
        btn_ok = QPushButton("가입하기")
        btn_ok.setStyleSheet("background: #22c55e; color: white; border: none; border-radius: 6px; padding: 8px 20px; font-weight: bold;")
        btn_cancel = QPushButton("취소")
        btn_cancel.setStyleSheet("padding: 8px 20px;")
        btn_cancel.clicked.connect(dlg.reject)
        btn_row.addWidget(btn_ok)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

        def do_register():
            uid = fields["uid"].text().strip()
            pw = fields["pw"].text()
            name = fields["name"].text().strip()
            birth = fields["birth"].text().strip()
            phone = fields["phone"].text().strip()
            email = fields["email"].text().strip()
            referrer = fields["referrer"].text().strip()
            pw2 = fields["pw2"].text()
            if not uid or not pw or not name or not birth or not phone or not email:
                msg.setText("필수 항목을 모두 입력해주세요.")
                return
            if pw != pw2:
                msg.setText("비밀번호가 일치하지 않습니다.")
                return
            if not dup_checked["phone"]:
                msg.setText("연락처 중복확인을 해주세요.")
                return
            if not dup_checked["email"]:
                msg.setText("이메일 중복확인을 해주세요.")
                return
            import datetime as _dt
            expire_date = (_dt.date.today() + _dt.timedelta(days=30)).strftime("%Y-%m-%d")
            from users import create_user
            if not create_user(uid, pw, role="user", expires=expire_date, max_accounts=3,
                             name=name, birth=birth, phone=phone, referrer=referrer):
                msg.setText("이미 존재하는 아이디입니다.")
                return
            from users import update_user
            update_user(uid, email=email)
            self.id_edit.setText(uid)
            self.pw_edit.setFocus()
            self.msg.setStyleSheet("color: #22c55e; font-size: 11px;")
            self.msg.setText("가입 완료! 로그인해주세요.")
            QMessageBox.information(dlg, "가입 완료",
                f"가입을 축하드립니다!\n\n아이디: {uid}\n무료 체험: 30일 ({expire_date}까지)\n네이버 계정: 3개 사용 가능\n\nAPI 키는 설정에서 직접 발급받아 입력해주세요.")
            dlg.accept()

        btn_ok.clicked.connect(do_register)
        dlg.exec()

    def _try_login(self):
        from users import verify, is_expired
        import datetime as _dt
        uid = self.id_edit.text().strip()
        pw = self.pw_edit.text()
        user = verify(uid, pw)
        if not user:
            self.msg.setStyleSheet("color: #ef4444; font-size: 11px;")
            self.msg.setText("아이디 또는 비밀번호가 틀렸습니다.")
            return
        if is_expired(user):
            self.msg.setStyleSheet(u"color: #ef4444; font-size: 11px;")
            self.msg.setText(f"사용 기간 만료 ({user.get('expires','')})")
            return
        cache = {
            "save_id": self.chk_save_id.isChecked(),
            "save_pw": self.chk_save_pw.isChecked(),
            "id": uid if self.chk_save_id.isChecked() else "",
            "pw": pw if self.chk_save_pw.isChecked() else "",
        }
        _save_login_cache(cache)
        self.user = user
        self.accept()


class SignupDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("회원가입")
        self.setFixedSize(340, 300)
        self.username = None

        layout = QVBoxLayout(self)
        layout.setContentsMargins(25, 20, 25, 20)

        title = QLabel("회원가입")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #1e293b;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        layout.addSpacing(10)

        self.id_edit = QLineEdit()
        self.id_edit.setPlaceholderText("아이디 (3자 이상)")
        self.id_edit.setStyleSheet("padding: 8px; font-size: 13px;")
        layout.addWidget(self.id_edit)

        self.pw_edit = QLineEdit()
        self.pw_edit.setEchoMode(QLineEdit.Password)
        self.pw_edit.setPlaceholderText("비밀번호 (4자 이상)")
        self.pw_edit.setStyleSheet("padding: 8px; font-size: 13px;")
        layout.addWidget(self.pw_edit)

        self.pw2_edit = QLineEdit()
        self.pw2_edit.setEchoMode(QLineEdit.Password)
        self.pw2_edit.setPlaceholderText("비밀번호 확인")
        self.pw2_edit.setStyleSheet("padding: 8px; font-size: 13px;")
        self.pw2_edit.returnPressed.connect(self._try_signup)
        layout.addWidget(self.pw2_edit)

        self.msg = QLabel("")
        self.msg.setStyleSheet("color: #ef4444; font-size: 11px;")
        self.msg.setWordWrap(True)
        layout.addWidget(self.msg)

        btn_row = QHBoxLayout()
        btn_ok = QPushButton("가입")
        btn_ok.setStyleSheet("background: #22c55e; color: white; border: none; border-radius: 6px; padding: 8px 20px; font-weight: bold;")
        btn_ok.setCursor(Qt.PointingHandCursor)
        btn_ok.clicked.connect(self._try_signup)
        btn_cancel = QPushButton("취소")
        btn_cancel.setStyleSheet("padding: 8px 20px;")
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(btn_ok)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

    def _try_signup(self):
        from users import create_user, load_users
        uid = self.id_edit.text().strip()
        pw = self.pw_edit.text()
        pw2 = self.pw2_edit.text()
        if len(uid) < 3:
            self.msg.setText("아이디는 3자 이상이어야 합니다.")
            return
        if len(pw) < 4:
            self.msg.setText("비밀번호는 4자 이상이어야 합니다.")
            return
        if pw != pw2:
            self.msg.setText("비밀번호가 일치하지 않습니다.")
            return
        if uid in load_users():
            self.msg.setText("이미 존재하는 아이디입니다.")
            return
        expires = (datetime.date.today() + datetime.timedelta(days=30)).strftime("%Y-%m-%d")
        if not create_user(uid, pw, role="user", expires=expires):
            self.msg.setText("가입 실패. 다시 시도해주세요.")
            return
        self.username = uid
        self.accept()


class ExpiresEditor(QWidget):
    """만료일 편집기 — 달력 팝업 + +30/+90/+180/+365 + 무제한"""
    def __init__(self, expires_str: str = "", parent=None):
        super().__init__(parent)
        from PySide6.QtWidgets import QSizePolicy
        layout = QHBoxLayout(self)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(2)

        self.line = QLineEdit()
        self.line.setPlaceholderText("무제한")
        self.line.setText(expires_str or "")
        self.line.setMinimumWidth(110)
        self.line.setStyleSheet("padding: 2px 4px; font-size: 12px;")
        layout.addWidget(self.line)

        btn_cal = QPushButton("📅")
        btn_cal.setFixedWidth(28)
        btn_cal.setStyleSheet("padding: 2px;")
        btn_cal.setCursor(Qt.PointingHandCursor)
        btn_cal.clicked.connect(self._open_calendar)
        layout.addWidget(btn_cal)

        for days in (30, 90, 180, 365):
            b = QPushButton(f"+{days}")
            b.setFixedWidth(38)
            b.setStyleSheet("padding: 2px; font-size: 10px;")
            b.setCursor(Qt.PointingHandCursor)
            b.clicked.connect(lambda _, d=days: self._add_days(d))
            layout.addWidget(b)

        btn_inf = QPushButton("무제한")
        btn_inf.setFixedWidth(54)
        btn_inf.setStyleSheet("padding: 2px; font-size: 10px;")
        btn_inf.setCursor(Qt.PointingHandCursor)
        btn_inf.clicked.connect(lambda: self.line.setText(""))
        layout.addWidget(btn_inf)
        layout.addStretch()

    def _open_calendar(self):
        from PySide6.QtWidgets import QCalendarWidget, QDialogButtonBox
        from PySide6.QtCore import QDate
        dlg = QDialog(self)
        dlg.setWindowTitle("만료일 선택")
        v = QVBoxLayout(dlg)
        cal = QCalendarWidget()
        txt = self.line.text().strip()
        if txt:
            d = QDate.fromString(txt, "yyyy-MM-dd")
            if d.isValid():
                cal.setSelectedDate(d)
        v.addWidget(cal)
        bbox = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bbox.accepted.connect(dlg.accept)
        bbox.rejected.connect(dlg.reject)
        v.addWidget(bbox)
        if dlg.exec() == QDialog.Accepted:
            self.line.setText(cal.selectedDate().toString("yyyy-MM-dd"))

    def _add_days(self, days: int):
        from PySide6.QtCore import QDate
        txt = self.line.text().strip()
        base = QDate.currentDate()
        if txt:
            d = QDate.fromString(txt, "yyyy-MM-dd")
            if d.isValid():
                base = d
        self.line.setText(base.addDays(days).toString("yyyy-MM-dd"))

    def get_value(self) -> str:
        return self.line.text().strip()


class YesNoRadio(QWidget):
    """예/아니오 라디오 그룹 (셀 위젯용)"""
    def __init__(self, checked: bool = False, parent=None):
        super().__init__(parent)
        from PySide6.QtWidgets import QRadioButton, QButtonGroup
        h = QHBoxLayout(self)
        h.setContentsMargins(6, 0, 6, 0)
        h.setSpacing(8)
        self.yes = QRadioButton("예")
        self.no = QRadioButton("아니오")
        self.group = QButtonGroup(self)
        self.group.setExclusive(True)
        self.group.addButton(self.yes)
        self.group.addButton(self.no)
        (self.yes if checked else self.no).setChecked(True)
        h.addWidget(self.yes)
        h.addWidget(self.no)
        h.addStretch()

    def is_yes(self) -> bool:
        return self.yes.isChecked()


class AdminDialog(QDialog):
    """관리자 전용 — 사용자 계정 + 기간 관리"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("관리자 메뉴")
        self.resize(1000, 460)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("사용자 관리 (아이디 / 역할 / 사용 만료일, 빈값=무제한)"))

        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["아이디", "역할", "계정 수", "만료일", "API키 부여", "네이버ID 초기화", "비밀번호 재설정"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setColumnWidth(0, 130)
        self.table.setColumnWidth(1, 90)
        self.table.setColumnWidth(2, 80)
        self.table.setColumnWidth(3, 380)
        self.table.setColumnWidth(4, 130)
        self.table.setColumnWidth(5, 200)
        self.table.verticalHeader().setDefaultSectionSize(36)
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #e2e8f0;
                selection-background-color: #e2e8f0;
                selection-color: #000;
                background: white;
            }
            QTableWidget::item:selected {
                background-color: #e2e8f0;
                color: #000;
            }
        """)
        layout.addWidget(self.table)

        self._reload()

        bottom = QHBoxLayout()
        btn_add = QPushButton("+ 사용자 추가")
        btn_add.setStyleSheet("background: #22c55e; color: white; border: none; border-radius: 6px; padding: 8px 14px;")
        btn_add.clicked.connect(self._add_user)
        btn_del = QPushButton("선택 삭제")
        btn_del.setStyleSheet("background: #ef4444; color: white; border: none; border-radius: 6px; padding: 8px 14px;")
        btn_del.clicked.connect(self._del_user)
        btn_save = QPushButton("저장")
        btn_save.setStyleSheet("background: #4a6cf7; color: white; border: none; border-radius: 6px; padding: 8px 20px; font-weight: bold;")
        btn_save.clicked.connect(self._save_all)
        btn_close = QPushButton("닫기")
        btn_close.setStyleSheet("padding: 8px 20px;")
        btn_close.clicked.connect(self.reject)
        bottom.addWidget(btn_add)
        bottom.addWidget(btn_del)
        bottom.addStretch()
        bottom.addWidget(btn_save)
        bottom.addWidget(btn_close)
        layout.addLayout(bottom)

    def _reload(self):
        from users import load_users
        users = load_users()
        self.table.setRowCount(len(users))
        for row, (uid, u) in enumerate(sorted(users.items())):
            it_id = QTableWidgetItem(uid)
            if uid == "admin":
                it_id.setFlags(it_id.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 0, it_id)

            # 역할 드롭다운
            role_combo = QComboBox()
            role_combo.addItems(["user", "admin"])
            role_combo.setCurrentText(u.get("role", "user") if u.get("role") in ("user", "admin") else "user")
            role_combo.setStyleSheet("padding: 2px 6px;")
            self.table.setCellWidget(row, 1, role_combo)

            # 계정 수 드롭다운 (3/6/9)
            acc_combo = QComboBox()
            acc_combo.addItems(["3", "6", "9"])
            cur_max = str(u.get("max_accounts", 3))
            if cur_max in ("3", "6", "9"):
                acc_combo.setCurrentText(cur_max)
            else:
                acc_combo.setCurrentText("3")
            acc_combo.setStyleSheet("padding: 2px 6px;")
            self.table.setCellWidget(row, 2, acc_combo)

            # 만료일 편집기
            exp_editor = ExpiresEditor(u.get("expires", ""))
            self.table.setCellWidget(row, 3, exp_editor)

            # API키 부여 라디오 (예/아니오)
            api_toggle = YesNoRadio(checked=bool(u.get("shared_api_keys")))
            if uid == "admin":
                api_toggle.setEnabled(False)
            self.table.setCellWidget(row, 4, api_toggle)

            # 네이버ID 초기화: 슬롯 선택 콤보 + 초기화 버튼
            reset_w = QWidget()
            reset_lay = QHBoxLayout(reset_w)
            reset_lay.setContentsMargins(4, 0, 4, 0)
            reset_lay.setSpacing(4)
            slot_combo = QComboBox()
            slot_combo.addItems([f"슬롯 {i+1}" for i in range(9)])
            slot_combo.setStyleSheet("padding: 2px 4px; font-size: 11px;")
            reset_btn = QPushButton("초기화")
            reset_btn.setStyleSheet("background: #ef4444; color: white; border: none; border-radius: 4px; padding: 4px 10px; font-size: 11px;")
            reset_btn.setCursor(Qt.PointingHandCursor)
            reset_btn.clicked.connect(lambda _, _uid=uid, _cb=slot_combo: self._reset_naver_slot(_uid, _cb.currentIndex()))
            reset_all_btn = QPushButton("전체 초기화")
            reset_all_btn.setStyleSheet("background: #b91c1c; color: white; border: none; border-radius: 4px; padding: 4px 10px; font-size: 11px; font-weight: bold;")
            reset_all_btn.setCursor(Qt.PointingHandCursor)
            reset_all_btn.clicked.connect(lambda _, _uid=uid: self._reset_naver_all(_uid))
            reset_lay.addWidget(slot_combo)
            reset_lay.addWidget(reset_btn)
            reset_lay.addWidget(reset_all_btn)
            self.table.setCellWidget(row, 5, reset_w)

            self.table.setItem(row, 6, QTableWidgetItem(""))  # 비번 재설정 입력칸

    def _reset_naver_all(self, uid: str):
        """관리자: 특정 사용자의 모든 슬롯 네이버 ID 데이터 + 잠금 전체 초기화"""
        reply = QMessageBox.question(
            self, "네이버 ID 전체 초기화",
            f"'{uid}'의 모든 슬롯(1~9) 네이버 계정 데이터를 초기화하고 잠금을 전부 해제할까요?"
        )
        if reply != QMessageBox.Yes:
            return
        try:
            import json as _json
            from config import CONFIG_FILE as _CF
            from users import update_user as _upd
            raw = {}
            if os.path.exists(_CF):
                raw = _json.load(open(_CF, encoding="utf-8"))
            abu = dict(raw.get("accounts_by_user") or {})
            empty_accs = [{"blog_id": "", "naver_id": "", "naver_pw": "", "blog_category": ""} for _ in range(9)]
            abu[uid] = empty_accs
            raw["accounts_by_user"] = abu
            with open(_CF, "w", encoding="utf-8") as f:
                _json.dump(raw, f, ensure_ascii=False, indent=2)
            _upd(uid, accounts=empty_accs, locked_naver_ids=[])
            QMessageBox.information(self, "완료", f"'{uid}' 전체 슬롯 초기화됨 (9개 슬롯 + 잠금 해제).")
        except Exception as e:
            QMessageBox.critical(self, "실패", f"전체 초기화 실패: {e}")

    def _reset_naver_slot(self, uid: str, slot_idx: int):
        """관리자: 특정 사용자의 특정 슬롯 네이버 ID 데이터 초기화 + 잠금 해제"""
        reply = QMessageBox.question(
            self, "네이버 ID 초기화",
            f"'{uid}'의 슬롯 {slot_idx+1}번 네이버 계정 데이터를 초기화할까요?\n(잠금도 해제되어 다시 등록 가능)"
        )
        if reply != QMessageBox.Yes:
            return
        try:
            import json as _json
            from config import CONFIG_FILE as _CF
            from users import load_users as _lu, update_user as _upd
            # 1) 로컬 config의 그 슬롯 비우기
            raw = {}
            if os.path.exists(_CF):
                raw = _json.load(open(_CF, encoding="utf-8"))
            abu = dict(raw.get("accounts_by_user") or {})
            accs = list(abu.get(uid) or [])
            removed_nid = ""
            if 0 <= slot_idx < len(accs):
                removed_nid = (accs[slot_idx].get("naver_id") or "").strip().lower()
                accs[slot_idx] = {"blog_id": "", "naver_id": "", "naver_pw": "", "blog_category": ""}
                abu[uid] = accs
                raw["accounts_by_user"] = abu
                with open(_CF, "w", encoding="utf-8") as f:
                    _json.dump(raw, f, ensure_ascii=False, indent=2)
            # 2) Firebase users.accounts + locked_naver_ids 갱신
            ue = _lu().get(uid, {})
            locked = set((s or "").strip().lower() for s in ue.get("locked_naver_ids", []))
            if removed_nid in locked:
                locked.discard(removed_nid)
            _upd(uid, accounts=accs, locked_naver_ids=sorted(locked))
            QMessageBox.information(self, "완료", f"'{uid}' 슬롯 {slot_idx+1}번 초기화됨.")
        except Exception as e:
            QMessageBox.critical(self, "실패", f"초기화 실패: {e}")

    def _add_user(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("회원가입")
        dlg.setFixedSize(420, 560)
        layout = QVBoxLayout(dlg)

        fields = {}
        for label, key, placeholder, required in [
            ("아이디 *", "uid", "영문/숫자", True),
            ("비밀번호 *", "pw", "비밀번호", True),
            ("이름 *", "name", "홍길동", True),
            ("생년월일 *", "birth", "1990-01-01", True),
            ("연락처 *", "phone", "010-1234-5678", True),
            ("추천인", "referrer", "추천인 아이디 (없으면 비워두세요)", False),
        ]:
            layout.addWidget(QLabel(label))
            entry = QLineEdit()
            entry.setPlaceholderText(placeholder)
            entry.setStyleSheet("padding: 6px;")
            if key == "pw":
                entry.setEchoMode(QLineEdit.Password)
            if key == "birth":
                entry.setMaxLength(10)
                entry.textChanged.connect(lambda t, e=entry: self._auto_dash_date(e, t))
            if key == "phone":
                entry.setMaxLength(13)
                entry.textChanged.connect(lambda t, e=entry: self._auto_dash_phone(e, t))
            layout.addWidget(entry)
            fields[key] = entry

        btn_row = QHBoxLayout()
        btn_ok = QPushButton("가입")
        btn_ok.setMinimumWidth(100)
        btn_ok.setStyleSheet("background: #4a6cf7; color: white; border: none; border-radius: 6px; padding: 8px 20px; font-weight: bold; font-size: 13px;")
        btn_cancel = QPushButton("취소")
        btn_cancel.setMinimumWidth(100)
        btn_cancel.setStyleSheet("padding: 8px 20px; font-size: 13px;")
        btn_cancel.clicked.connect(dlg.reject)
        btn_row.addWidget(btn_ok)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

        def do_register():
            uid = fields["uid"].text().strip()
            pw = fields["pw"].text()
            name = fields["name"].text().strip()
            birth = fields["birth"].text().strip()
            phone = fields["phone"].text().strip()
            referrer = fields["referrer"].text().strip()
            if not uid or not pw or not name or not birth or not phone:
                QMessageBox.warning(dlg, "경고", "필수 항목을 모두 입력해주세요.")
                return
            from users import create_user
            if not create_user(uid, pw, role="user", expires="",
                             name=name, birth=birth, phone=phone, referrer=referrer):
                QMessageBox.warning(dlg, "경고", "이미 존재하는 아이디입니다.")
                return
            QMessageBox.information(dlg, "완료", f"'{uid}' 가입 완료!")
            dlg.accept()
            self._reload()

        btn_ok.clicked.connect(do_register)
        dlg.exec()

    def _auto_dash_date(self, entry, text):
        """생년월일 자동 하이픈: 19900101 → 1990-01-01"""
        digits = text.replace("-", "")
        if not digits.isdigit():
            return
        entry.blockSignals(True)
        if len(digits) > 8:
            digits = digits[:8]
        if len(digits) >= 6:
            result = f"{digits[:4]}-{digits[4:6]}-{digits[6:]}"
        elif len(digits) >= 4:
            result = f"{digits[:4]}-{digits[4:]}"
        else:
            result = digits
        entry.setText(result)
        entry.setCursorPosition(len(result))
        entry.blockSignals(False)

    def _auto_dash_phone(self, entry, text):
        """연락처 자동 하이픈: 01012345678 → 010-1234-5678"""
        digits = text.replace("-", "")
        if not digits.isdigit():
            return
        entry.blockSignals(True)
        if len(digits) > 11:
            digits = digits[:11]
        if len(digits) >= 8:
            result = f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
        elif len(digits) >= 3:
            result = f"{digits[:3]}-{digits[3:]}"
        else:
            result = digits
        entry.setText(result)
        entry.setCursorPosition(len(result))
        entry.blockSignals(False)

    def _del_user(self):
        row = self.table.currentRow()
        if row < 0:
            return
        uid = self.table.item(row, 0).text()
        if uid == "admin":
            QMessageBox.warning(self, "경고", "admin 계정은 삭제할 수 없습니다.")
            return
        if QMessageBox.question(self, "확인", f"'{uid}' 삭제?") != QMessageBox.Yes:
            return
        from users import delete_user
        delete_user(uid)
        self._reload()

    def _save_all(self):
        from users import update_user, load_users
        from config import load_config
        admin_cfg = load_config()
        admin_keys = {
            k: [v for v in admin_cfg.get(k, []) if v]
            for k in ("gpt_key_list", "gemini_key_list", "pixabay_key_list")
        }
        existing_users = load_users()
        for row in range(self.table.rowCount()):
            uid = self.table.item(row, 0).text().strip()
            role_widget = self.table.cellWidget(row, 1)
            role = role_widget.currentText() if role_widget else "user"
            acc_widget = self.table.cellWidget(row, 2)
            max_acc = int(acc_widget.currentText()) if acc_widget else 3
            exp_widget = self.table.cellWidget(row, 3)
            expires = exp_widget.get_value() if exp_widget else ""
            api_widget = self.table.cellWidget(row, 4)
            api_on = api_widget.is_yes() if api_widget else False
            new_pw_item = self.table.item(row, 6)
            new_pw = (new_pw_item.text() if new_pw_item else "").strip()
            if role not in ("user", "admin"):
                role = "user"
            if uid == "admin":
                shared = None
            elif api_on:
                shared = {k: v for k, v in admin_keys.items() if v}
            else:
                shared = {}
            update_user(
                uid,
                password=(new_pw if new_pw else None),
                role=role,
                expires=expires,
                max_accounts=max_acc,
                shared_api_keys=shared,
            )
        QMessageBox.information(self, "완료", "저장되었습니다.")
        self._reload()



def _ver_tuple(v: str):
    """버전 문자열을 튜플로 (예: '1.0.10' -> (1,0,10))"""
    out = []
    for x in (v or "").split("."):
        try:
            out.append(int(x))
        except ValueError:
            out.append(0)
    return tuple(out)


def _check_previous_update_result():
    """직전 인앱 패치 결과 로그 확인 — 실패면 사용자에게 알림"""
    import tempfile
    log = os.path.join(tempfile.gettempdir(), "blogmaster_update", "update_result.log")
    if not os.path.exists(log):
        return
    try:
        with open(log, "r", encoding="cp949", errors="ignore") as f:
            content = f.read().strip()
        os.remove(log)
        if content and not content.startswith("OK"):
            QMessageBox.warning(None, "업데이트 결과", f"직전 자동업데이트 실패:\n{content}")
    except Exception:
        pass


def _check_and_offer_update_pre_login(parent=None):
    """로그인 전 자동업데이트 — urllib + 부모 위젯 필수(QMessageBox None parent segfault 회피)"""
    _check_previous_update_result()
    def _log(msg):
        try:
            from app_paths import data_file
            with open(data_file("update_check.log"), "a", encoding="utf-8") as _f:
                import datetime as _dt
                _f.write(f"[{_dt.datetime.now()}] {msg}\n")
        except Exception:
            pass
    _log(f"check start, current={APP_VERSION}")
    try:
        import urllib.request as _urlreq
        import json as _json
        import ssl as _ssl
        # frozen exe에서 SSL 인증서 찾기 실패 대응 → certifi CA 번들 명시 + requests 폴백
        data = None
        try:
            import certifi as _certifi
            ctx = _ssl.create_default_context(cafile=_certifi.where())
        except Exception:
            ctx = _ssl.create_default_context()
        req = _urlreq.Request(
            "https://api.github.com/repos/kingth0506/BlogMaster/releases/latest",
            headers={"User-Agent": "BlogMaster", "Accept": "application/json"}
        )
        try:
            with _urlreq.urlopen(req, timeout=15, context=ctx) as resp:
                data = _json.loads(resp.read().decode("utf-8"))
        except Exception as e:
            _log(f"urllib fail, trying requests: {e}")
            # 2차 시도: requests (certifi 자동 사용)
            try:
                import requests as _req
                r = _req.get(
                    "https://api.github.com/repos/kingth0506/BlogMaster/releases/latest",
                    headers={"User-Agent": "BlogMaster", "Accept": "application/json"},
                    timeout=15,
                )
                r.raise_for_status()
                data = r.json()
            except Exception as e2:
                _log(f"requests fail too: {e2}")
                # 네트워크 완전 실패 시 수동 다운로드 안내 (한 번만)
                try:
                    reply = QMessageBox.question(
                        parent, "업데이트 확인 실패",
                        "인터넷 연결 또는 인증서 문제로 업데이트를 확인하지 못했습니다.\n"
                        "브라우저에서 최신 버전을 직접 다운로드하시겠습니까?",
                        QMessageBox.Yes | QMessageBox.No
                    )
                    if reply == QMessageBox.Yes:
                        import webbrowser
                        webbrowser.open("https://github.com/kingth0506/BlogMaster/releases/latest")
                except Exception:
                    pass
                return
        latest = (data.get("tag_name") or "").lstrip("v")
        _log(f"latest={latest}, current={APP_VERSION}")
        if not latest or _ver_tuple(latest) <= _ver_tuple(APP_VERSION):
            _log("up to date, no dialog")
            return
        try:
            reply = QMessageBox.question(
                parent, "업데이트",
                f"새 버전(v{latest})이 있습니다.\n지금 업데이트하시겠습니까?",
                QMessageBox.Yes | QMessageBox.No
            )
        except Exception as e:
            _log(f"QMessageBox fail: {e}")
            return
        if reply != QMessageBox.Yes:
            _log("user declined")
            return
        _log("user accepted, starting update")
        _do_in_app_update()
    except Exception as e:
        _log(f"outer exception: {e}")


def _do_in_app_update():
    """변경 파일만 다운(zip) → 임시 폴더 압축 해제 → updater.bat이 메인 종료 대기 후 교체 + 재시작"""
    import requests as _req
    import tempfile, zipfile, shutil, subprocess
    from PySide6.QtWidgets import QProgressDialog

    if getattr(sys, 'frozen', False):
        install_dir = os.path.dirname(os.path.abspath(sys.executable))
        exe_name = os.path.basename(sys.executable)
    else:
        install_dir = os.path.dirname(os.path.abspath(__file__))
        exe_name = "BlogMaster.exe"

    work_dir = os.path.join(tempfile.gettempdir(), "blogmaster_update")
    if os.path.exists(work_dir):
        shutil.rmtree(work_dir, ignore_errors=True)
    os.makedirs(work_dir, exist_ok=True)
    staging = os.path.join(work_dir, "staging")
    zip_path = os.path.join(work_dir, "BlogMaster_files.zip")

    url = "https://github.com/kingth0506/BlogMaster/releases/latest/download/BlogMaster_files.zip"

    progress = UpdateProgressDialog()
    progress.show()
    QApplication.processEvents()

    try:
        with _req.get(url, stream=True, timeout=600) as rr:
            rr.raise_for_status()
            total = int(rr.headers.get('Content-Length', 0))
            downloaded = 0
            with open(zip_path, "wb") as f:
                for chunk in rr.iter_content(chunk_size=65536):
                    if progress.wasCanceled():
                        progress.close()
                        return
                    f.write(chunk)
                    downloaded += len(chunk)
                    if total:
                        pct = int(downloaded * 100 / total)
                        mb_d = downloaded / (1024 * 1024)
                        mb_t = total / (1024 * 1024)
                        progress.setProgress(pct, f"{mb_d:.1f} / {mb_t:.1f} MB")
                    QApplication.processEvents()
        progress.setLabelText("압축 해제 중...")
        QApplication.processEvents()
        with zipfile.ZipFile(zip_path, 'r') as z:
            z.extractall(staging)
        progress.setProgress(100, "완료")
    except Exception as e:
        progress.close()
        QMessageBox.warning(None, "업데이트 실패", f"다운로드/해제 실패:\n{e}")
        return

    progress.close()

    bat_path = os.path.join(work_dir, "updater.bat")
    install_dir_w = install_dir.replace("/", "\\")
    staging_w = staging.replace("/", "\\")
    log_path = os.path.join(work_dir, "update_result.log").replace("/", "\\")
    tl_tmp = os.path.join(work_dir, "tl.txt").replace("/", "\\")
    bat = (
        "@echo off\r\n"
        "setlocal\r\n"
        f"set LOG=\"{log_path}\"\r\n"
        ":wait\r\n"
        "timeout /t 1 /nobreak >nul 2>nul\r\n"
        f"tasklist /FI \"IMAGENAME eq {exe_name}\" /NH > \"{tl_tmp}\" 2>nul\r\n"
        f"findstr /I /C:\"{exe_name}\" \"{tl_tmp}\" >nul 2>nul\r\n"
        "if %errorlevel%==0 goto wait\r\n"
        f"del /Q \"{tl_tmp}\" 2>nul\r\n"
        f"robocopy \"{staging_w}\" \"{install_dir_w}\" /E /R:3 /W:1 /NFL /NDL /NJH /NJS /NC /NS /NP >nul 2>nul\r\n"
        "if errorlevel 8 (\r\n"
        f"  echo ROBOCOPY_FAIL errorlevel=%errorlevel% > %LOG%\r\n"
        "  exit /b 1\r\n"
        ")\r\n"
        f"if not exist \"{install_dir_w}\\{exe_name}\" (\r\n"
        f"  echo MISSING_EXE after_xcopy > %LOG%\r\n"
        "  exit /b 2\r\n"
        ")\r\n"
        f"del /Q \"{install_dir_w}\\NaverBlogAuto.exe\" 2>nul\r\n"
        "del /Q \"%USERPROFILE%\\Desktop\\NaverBlogAuto.lnk\" 2>nul\r\n"
        "del /Q \"%USERPROFILE%\\OneDrive\\Desktop\\NaverBlogAuto.lnk\" 2>nul\r\n"
        "del /Q \"%APPDATA%\\Microsoft\\Windows\\Start Menu\\Programs\\NaverBlogAuto\\NaverBlogAuto.lnk\" 2>nul\r\n"
        "echo OK > %LOG%\r\n"
        f"start \"\" \"{install_dir_w}\\{exe_name}\"\r\n"
        "del /Q \"%~f0\"\r\n"
    )
    with open(bat_path, "w", encoding="cp949") as f:
        f.write(bat)

    # cmd 완전 숨김: VBS 런처로 bat을 hidden(WindowStyle=0) 실행 → 어떤 cmd 창도 안 뜸
    vbs_path = os.path.join(work_dir, "run_hidden.vbs").replace("/", "\\")
    # VBS는 큰따옴표를 두 번 써서 escape: "" → 실제 큰따옴표 1개
    vbs_content = (
        'Set WshShell = CreateObject("WScript.Shell")\r\n'
        f'WshShell.Run "cmd /c ""{bat_path}""", 0, False\r\n'
    )
    with open(vbs_path, "w", encoding="cp949") as f:
        f.write(vbs_content)
    CREATE_NO_WINDOW = 0x08000000
    si = subprocess.STARTUPINFO()
    si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    si.wShowWindow = 0
    subprocess.Popen(
        ["wscript.exe", vbs_path],
        creationflags=CREATE_NO_WINDOW,
        startupinfo=si,
        close_fds=True,
    )
    sys.exit(0)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    # Windows 다크모드 무시하고 전역 라이트 팔레트 강제
    try:
        from PySide6.QtGui import QPalette, QColor
        app.setStyle("Fusion")
        _pal = QPalette()
        _pal.setColor(QPalette.Window, QColor("#ffffff"))
        _pal.setColor(QPalette.WindowText, QColor("#1e293b"))
        _pal.setColor(QPalette.Base, QColor("#ffffff"))
        _pal.setColor(QPalette.AlternateBase, QColor("#f1f5f9"))
        _pal.setColor(QPalette.ToolTipBase, QColor("#ffffff"))
        _pal.setColor(QPalette.ToolTipText, QColor("#1e293b"))
        _pal.setColor(QPalette.Text, QColor("#1e293b"))
        _pal.setColor(QPalette.Button, QColor("#f1f5f9"))
        _pal.setColor(QPalette.ButtonText, QColor("#1e293b"))
        _pal.setColor(QPalette.BrightText, QColor("#ef4444"))
        _pal.setColor(QPalette.Link, QColor("#3b82f6"))
        _pal.setColor(QPalette.Highlight, QColor("#3b82f6"))
        _pal.setColor(QPalette.HighlightedText, QColor("#ffffff"))
        _pal.setColor(QPalette.PlaceholderText, QColor("#94a3b8"))
        app.setPalette(_pal)
    except Exception:
        pass
    app.setStyleSheet(STYLE)
    app.setFont(QFont("맑은 고딕", 10))
    # 트레이로 숨길 때 앱이 종료되지 않도록
    app.setQuitOnLastWindowClosed(False)

    # 로그인 다이얼로그 먼저 생성 → 자동업데이트의 parent로 사용 (None parent segfault 회피)
    login = LoginDialog()
    _check_and_offer_update_pre_login(parent=login)
    if login.exec() != QDialog.Accepted or not login.user:
        sys.exit(0)

    # 로그인 유저 컨텍스트 설정 (MainWindow 생성 전에 필수 — load_config가 이걸로 유저별 accounts 분기)
    import config as _cfg
    _cfg.set_current_user(login.user.get("username", "admin"))

    window = MainWindow()
    window.apply_user_session(login.user)
    window.show()
    window.raise_()
    window.activateWindow()

    sys.exit(app.exec())
